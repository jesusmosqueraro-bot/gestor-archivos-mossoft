"""Pruebas de Costos de Inventario: un activo puede marcarse 'propio' (con costo_compra, pago
único) o 'alquilado' (con costo_alquiler_mensual, pago recurrente) — nunca las dos cosas a la
vez, el campo que no aplica se descarta al guardar. Cubre creación/edición, la agregación de
totales que se muestra en /tickets/inventario y su reflejo en el Tablero Ejecutivo.

También cubre un bug real reportado por Tomás: en producción (Postgres), 'costo_compra' y
'costo_alquiler_mensual' son columnas NUMERIC(14,2) y psycopg2 las devuelve como
decimal.Decimal, NUNCA como float. En cuanto CUALQUIER activo quedaba con un costo guardado,
/tickets/inventario reventaba con 500 ("TypeError: unsupported operand type(s) for +=: 'float'
and 'decimal.Decimal'") al sumar los totales, y el modal de editar tampoco podía serializar ese
activo a JSON (Decimal no es serializable). Las pruebas de este archivo corren sobre SQLite (esa
columna es REAL, vuelve como float) y por eso nunca detectaron esto — las pruebas de abajo
simulan explícitamente el Decimal que sí llega desde Postgres."""
import decimal


def _crear_activo_directo(app, nombre='30001', tipo_costo=None, costo_compra=None, costo_alquiler_mensual=None):
    conn, db_type = app.get_db()
    cur = conn.cursor()
    q = ("INSERT INTO activos_inventario (nombre, tipo_activo, estado, tipo_costo, costo_compra, "
         "costo_alquiler_mensual, fecha_creacion, creado_por) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id"
         if db_type == 'postgres' else
         "INSERT INTO activos_inventario (nombre, tipo_activo, estado, tipo_costo, costo_compra, "
         "costo_alquiler_mensual, fecha_creacion, creado_por) VALUES (?, ?, ?, ?, ?, ?, ?, ?)")
    cur.execute(q, (nombre, 'Portátil', 'Disponible', tipo_costo, costo_compra, costo_alquiler_mensual,
                     '2026-09-04 09:00:00', 'admin'))
    activo_id = cur.fetchone()[0] if db_type == 'postgres' else cur.lastrowid
    conn.commit()
    conn.close()
    return activo_id


def test_crear_activo_propio_guarda_costo_de_compra(admin_session, app):
    r = admin_session.post('/tickets/inventario/nuevo', data={
        'nombre': '30001', 'tipo_activo': 'Portátil', 'estado': 'Disponible',
        'tipo_costo': 'propio', 'costo_compra': '2500000',
    })

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT tipo_costo, costo_compra, costo_alquiler_mensual FROM activos_inventario WHERE nombre = ?", ('30001',))
    fila = cur.fetchone()
    conn.close()
    assert fila == ('propio', 2500000.0, None)


def test_crear_activo_alquilado_guarda_costo_mensual(admin_session, app):
    r = admin_session.post('/tickets/inventario/nuevo', data={
        'nombre': '30002', 'tipo_activo': 'Portátil', 'estado': 'Disponible',
        'tipo_costo': 'alquilado', 'costo_alquiler_mensual': '180000',
    })

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT tipo_costo, costo_compra, costo_alquiler_mensual FROM activos_inventario WHERE nombre = ?", ('30002',))
    fila = cur.fetchone()
    conn.close()
    assert fila == ('alquilado', None, 180000.0)


def test_crear_activo_alquilado_ignora_costo_de_compra_enviado_por_error(admin_session, app):
    """Si el formulario manda ambos campos (por ejemplo, el usuario cambió de opción sin borrar
    el otro campo), solo se guarda el que corresponde a 'tipo_costo' — el otro queda en None y no
    se contabiliza."""
    r = admin_session.post('/tickets/inventario/nuevo', data={
        'nombre': '30003', 'tipo_activo': 'Portátil', 'estado': 'Disponible',
        'tipo_costo': 'alquilado', 'costo_alquiler_mensual': '200000', 'costo_compra': '9999999',
    })

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT costo_compra, costo_alquiler_mensual FROM activos_inventario WHERE nombre = ?", ('30003',))
    fila = cur.fetchone()
    conn.close()
    assert fila == (None, 200000.0)


def test_crear_activo_con_costo_no_numerico_no_falla_y_queda_en_none(admin_session, app):
    r = admin_session.post('/tickets/inventario/nuevo', data={
        'nombre': '30004', 'tipo_activo': 'Portátil', 'estado': 'Disponible',
        'tipo_costo': 'propio', 'costo_compra': 'no-es-un-numero',
    })

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT costo_compra FROM activos_inventario WHERE nombre = ?", ('30004',))
    assert cur.fetchone()[0] is None
    conn.close()


def test_crear_activo_con_tipo_costo_invalido_no_guarda_costos(admin_session, app):
    r = admin_session.post('/tickets/inventario/nuevo', data={
        'nombre': '30005', 'tipo_activo': 'Portátil', 'estado': 'Disponible',
        'tipo_costo': 'algo-raro', 'costo_compra': '1000000',
    })

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT tipo_costo, costo_compra FROM activos_inventario WHERE nombre = ?", ('30005',))
    assert cur.fetchone() == (None, None)
    conn.close()


def test_editar_activo_actualiza_el_costo(admin_session, app):
    activo_id = _crear_activo_directo(app, nombre='30006', tipo_costo='propio', costo_compra=1000000)

    r = admin_session.post(f'/tickets/inventario/{activo_id}/editar', data={
        'nombre': '30006', 'tipo_activo': 'Portátil', 'estado': 'Disponible',
        'tipo_costo': 'propio', 'costo_compra': '1500000',
    })

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT costo_compra FROM activos_inventario WHERE id = ?", (activo_id,))
    assert cur.fetchone()[0] == 1500000.0
    conn.close()


def test_editar_activo_cambia_de_propio_a_alquilado_limpia_costo_de_compra(admin_session, app):
    activo_id = _crear_activo_directo(app, nombre='30007', tipo_costo='propio', costo_compra=1000000)

    r = admin_session.post(f'/tickets/inventario/{activo_id}/editar', data={
        'nombre': '30007', 'tipo_activo': 'Portátil', 'estado': 'Disponible',
        'tipo_costo': 'alquilado', 'costo_alquiler_mensual': '250000',
    })

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT tipo_costo, costo_compra, costo_alquiler_mensual FROM activos_inventario WHERE id = ?", (activo_id,))
    assert cur.fetchone() == ('alquilado', None, 250000.0)
    conn.close()


def test_inventario_muestra_los_totales_de_costos(admin_session, app):
    _crear_activo_directo(app, nombre='30008', tipo_costo='propio', costo_compra=1000000)
    _crear_activo_directo(app, nombre='30009', tipo_costo='propio', costo_compra=2000000)
    _crear_activo_directo(app, nombre='30010', tipo_costo='alquilado', costo_alquiler_mensual=150000)
    # Activo sin tipo_costo definido: no debe sumar a ningún total.
    _crear_activo_directo(app, nombre='30011')

    texto = admin_session.get('/tickets/inventario').get_data(as_text=True)

    assert '3,000,000' in texto  # total de compra: 1,000,000 + 2,000,000
    assert '150,000' in texto    # total de alquiler mensual
    assert '1,800,000' in texto  # proyección anual: 150,000 * 12


def test_totales_costos_inventario_ignora_activos_sin_tipo_costo(app):
    import app as arkiv
    activos = [
        {'tipo_costo': 'propio', 'costo_compra': 500000, 'costo_alquiler_mensual': None},
        {'tipo_costo': None, 'costo_compra': None, 'costo_alquiler_mensual': None},
        {'tipo_costo': 'alquilado', 'costo_compra': None, 'costo_alquiler_mensual': 100000},
    ]
    totales = arkiv._totales_costos_inventario(activos)
    assert totales['total_compra'] == 500000
    assert totales['cantidad_propios'] == 1
    assert totales['total_alquiler_mensual'] == 100000
    assert totales['cantidad_alquilados'] == 1
    assert totales['total_alquiler_anual'] == 1200000


def test_decimal_a_float_convierte_decimal_de_postgres(app):
    """_decimal_a_float() es lo que se llama justo al leer costo_compra/costo_alquiler_mensual de
    la base — con esto, un Decimal de Postgres se vuelve float antes de tocar cualquier cálculo o
    plantilla; None se mantiene como None (activo sin costo guardado)."""
    assert app._decimal_a_float(decimal.Decimal('1500000.00')) == 1500000.0
    assert isinstance(app._decimal_a_float(decimal.Decimal('1500000.00')), float)
    assert app._decimal_a_float(None) is None


def test_totales_costos_inventario_no_revienta_con_decimal_de_postgres(app):
    """Reproduce exactamente el 500 que reportó Tomás: en Postgres estos valores llegan como
    decimal.Decimal, no como float. Sumar un Decimal contra el acumulador float (0.0) con '+='
    lanzaba TypeError sin que nada lo atajara — bastaba con que UN activo tuviera un costo
    guardado para tumbar /tickets/inventario en cada carga."""
    activos = [
        {'tipo_costo': 'propio', 'costo_compra': decimal.Decimal('2500000.00'), 'costo_alquiler_mensual': None},
        {'tipo_costo': 'alquilado', 'costo_compra': None, 'costo_alquiler_mensual': decimal.Decimal('180000.00')},
    ]

    totales = app._totales_costos_inventario(activos)

    assert totales['total_compra'] == 2500000.0
    assert totales['total_alquiler_mensual'] == 180000.0
    assert totales['total_alquiler_anual'] == 2160000.0


def test_modal_inventario_incluye_los_campos_de_costo(admin_session):
    texto = admin_session.get('/tickets/inventario').get_data(as_text=True)
    assert 'name="tipo_costo"' in texto
    assert 'name="costo_compra"' in texto
    assert 'name="costo_alquiler_mensual"' in texto


def test_tablero_ejecutivo_muestra_los_totales_de_costos_de_inventario(admin_session, app):
    _crear_activo_directo(app, nombre='30012', tipo_costo='propio', costo_compra=3000000)
    _crear_activo_directo(app, nombre='30013', tipo_costo='alquilado', costo_alquiler_mensual=200000)

    texto = admin_session.get('/tablero-ejecutivo').get_data(as_text=True)

    assert 'Costos de Inventario' in texto
    assert '3,000,000' in texto
    assert '200,000' in texto


def test_tabla_inventario_muestra_el_costo_por_articulo(admin_session, app):
    """Tomás pidió ver el costo POR ARTÍCULO (mensual y anual si es alquilado, o el valor de
    compra si es propio) directamente en la tabla de Inventario — antes solo se veía el TOTAL
    agregado de todo el inventario en las tarjetas de arriba, nunca el de cada activo."""
    _crear_activo_directo(app, nombre='30014', tipo_costo='propio', costo_compra=2500000)
    _crear_activo_directo(app, nombre='30015', tipo_costo='alquilado', costo_alquiler_mensual=150000)
    _crear_activo_directo(app, nombre='30016')  # sin tipo_costo definido

    texto = admin_session.get('/tickets/inventario').get_data(as_text=True)

    assert 'Propio' in texto
    assert '2,500,000' in texto
    assert 'Alquiler' in texto
    assert '150,000' in texto
    assert '/mes' in texto
    assert '1,800,000' in texto  # proyección anual del activo alquilado: 150,000 * 12
    assert '/año' in texto


def test_csv_exportar_incluye_costo_de_alquiler_anual_por_articulo(admin_session, app):
    _crear_activo_directo(app, nombre='30017', tipo_costo='alquilado', costo_alquiler_mensual=100000)

    r = admin_session.get('/tickets/inventario/exportar_csv')
    texto = r.get_data(as_text=True)

    assert 'COSTO ALQUILER ANUAL' in texto
    assert '1200000' in texto  # 100,000 * 12, sin separador de miles en el CSV crudo


def test_xlsx_exportar_incluye_costo_de_alquiler_anual_por_articulo(admin_session, app):
    import io
    import openpyxl

    _crear_activo_directo(app, nombre='30018', tipo_costo='alquilado', costo_alquiler_mensual=100000)

    r = admin_session.get('/tickets/inventario/exportar_xlsx')
    wb = openpyxl.load_workbook(io.BytesIO(r.get_data()))
    ws = wb.active

    encabezados = [c.value for c in ws[1]]
    assert 'Costo alquiler anual' in encabezados
    col_anual = encabezados.index('Costo alquiler anual') + 1
    valores_columna = [ws.cell(row=fila, column=col_anual).value for fila in range(2, ws.max_row + 1)]
    assert 1200000 in valores_columna


def test_pdf_exportar_no_falla_con_activo_alquilado(admin_session, app):
    """Regresión: el cálculo de 'costo_alquiler_mensual * 12' dentro del PDF no debe romper la
    exportación cuando el valor viene de Postgres como Decimal (aquí simulado con float normal,
    ver test_totales_costos_inventario_no_revienta_con_decimal_de_postgres para el caso Decimal)."""
    _crear_activo_directo(app, nombre='30019', tipo_costo='alquilado', costo_alquiler_mensual=180000)

    r = admin_session.get('/tickets/inventario/exportar_pdf')

    assert r.status_code == 200
    assert r.headers['Content-Type'] == 'application/pdf'
