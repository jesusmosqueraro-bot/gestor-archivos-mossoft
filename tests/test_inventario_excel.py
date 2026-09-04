"""Pruebas de carga masiva (Excel) y exportación (CSV/Excel/PDF) del Inventario de Activos.
La plantilla descargable, el parser de importación y las exportaciones comparten el mismo orden
de columnas (COLUMNAS_INVENTARIO_XLSX) y las mismas reglas de validación que crear_activo."""
import io

import openpyxl
import pytest


def _crear_area(app, nombre='Sistemas'):
    conn, db_type = app.get_db()
    cur = conn.cursor()
    q = ("INSERT INTO ticket_configuraciones (tipo, nombre, estado) VALUES (%s, %s, 'activo')"
         if db_type == 'postgres' else
         "INSERT INTO ticket_configuraciones (tipo, nombre, estado) VALUES (?, ?, 'activo')")
    cur.execute(q, ('area', nombre))
    conn.commit()
    conn.close()
    return nombre


def _crear_sede(app, nombre='Sede Principal'):
    conn, db_type = app.get_db()
    cur = conn.cursor()
    q = ("INSERT INTO ticket_configuraciones (tipo, nombre, estado) VALUES (%s, %s, 'activo')"
         if db_type == 'postgres' else
         "INSERT INTO ticket_configuraciones (tipo, nombre, estado) VALUES (?, ?, 'activo')")
    cur.execute(q, ('sede', nombre))
    conn.commit()
    conn.close()
    return nombre


def _crear_activo_directo(app, nombre='40001'):
    conn, db_type = app.get_db()
    cur = conn.cursor()
    q = ("INSERT INTO activos_inventario (nombre, tipo_activo, estado, fecha_creacion, creado_por) "
         "VALUES (%s, %s, %s, %s, %s)" if db_type == 'postgres' else
         "INSERT INTO activos_inventario (nombre, tipo_activo, estado, fecha_creacion, creado_por) "
         "VALUES (?, ?, ?, ?, ?)")
    cur.execute(q, (nombre, 'Portátil', 'Disponible', '2026-09-04 09:00:00', 'admin'))
    conn.commit()
    conn.close()


def _libro_xlsx(filas, encabezados=None):
    """Arma un .xlsx en memoria con encabezado (fila 1) + las filas de datos dadas, en el mismo
    orden de columnas que espera importar_inventario_xlsx."""
    import app as arkiv
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(encabezados or arkiv.COLUMNAS_INVENTARIO_XLSX)
    for fila in filas:
        ws.append(fila)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_plantilla_xlsx_descarga_con_encabezados_y_fila_de_ejemplo(admin_session):
    r = admin_session.get('/tickets/inventario/plantilla_xlsx')
    assert r.status_code == 200
    assert r.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    primera_fila = [c.value for c in ws[1]]
    assert primera_fila[0] == 'Placa'
    assert ws.max_row >= 2  # trae una fila de ejemplo


def test_importar_xlsx_crea_los_activos_validos(admin_session, app):
    _crear_sede(app, 'Sede Principal')
    _crear_area(app, 'Sistemas')

    buffer = _libro_xlsx([
        ['40010', 'Portátil', 'Dell', 'Latitude', 'SN-1', 'Disponible', '', 'Sede Principal',
         'Sistemas', '', 'propio', 2000000, '', 'Equipo nuevo'],
        ['40011', 'Portátil', 'HP', 'ProBook', 'SN-2', 'Disponible', '', '', '', '', 'alquilado',
         '', 90000, ''],
    ])

    r = admin_session.post('/tickets/inventario/importar_xlsx', data={
        'archivo': (buffer, 'inventario.xlsx')
    }, content_type='multipart/form-data', follow_redirects=False)

    assert r.status_code == 302
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT tipo_costo, costo_compra FROM activos_inventario WHERE nombre = ?", ('40010',))
    assert cur.fetchone() == ('propio', 2000000.0)
    cur.execute("SELECT tipo_costo, costo_alquiler_mensual, sede, area FROM activos_inventario WHERE nombre = ?", ('40011',))
    fila = cur.fetchone()
    conn.close()
    assert fila[0] == 'alquilado'
    assert fila[1] == 90000.0
    assert fila[2] is None  # 'Sede Principal' fue borrado en este caso (celda vacía)
    assert fila[3] is None


def test_importar_xlsx_omite_placas_duplicadas_con_las_existentes(admin_session, app):
    _crear_activo_directo(app, nombre='40020')
    buffer = _libro_xlsx([
        ['40020', 'Portátil', '', '', '', 'Disponible', '', '', '', '', '', '', '', ''],
        ['40021', 'Portátil', '', '', '', 'Disponible', '', '', '', '', '', '', '', ''],
    ])

    r = admin_session.post('/tickets/inventario/importar_xlsx', data={
        'archivo': (buffer, 'inventario.xlsx')
    }, content_type='multipart/form-data', follow_redirects=True)

    assert r.status_code == 200
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM activos_inventario WHERE nombre = '40020'")
    assert cur.fetchone()[0] == 1  # no se duplicó
    cur.execute("SELECT COUNT(*) FROM activos_inventario WHERE nombre = '40021'")
    assert cur.fetchone()[0] == 1  # este sí se creó
    conn.close()
    texto = r.get_data(as_text=True)
    assert 'omitieron' in texto.lower() or 'ya existe' in texto.lower()


def test_importar_xlsx_omite_placas_repetidas_dentro_del_mismo_archivo(admin_session, app):
    buffer = _libro_xlsx([
        ['40030', 'Portátil', '', '', '', 'Disponible', '', '', '', '', '', '', '', ''],
        ['40030', 'Portátil', '', '', '', 'Disponible', '', '', '', '', '', '', '', ''],
    ])

    admin_session.post('/tickets/inventario/importar_xlsx', data={
        'archivo': (buffer, 'inventario.xlsx')
    }, content_type='multipart/form-data')

    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM activos_inventario WHERE nombre = '40030'")
    assert cur.fetchone()[0] == 1
    conn.close()


def test_importar_xlsx_ignora_filas_en_blanco_sin_contarlas_como_error(admin_session, app):
    buffer = _libro_xlsx([
        ['40040', 'Portátil', '', '', '', 'Disponible', '', '', '', '', '', '', '', ''],
        [None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ['40041', 'Portátil', '', '', '', 'Disponible', '', '', '', '', '', '', '', ''],
    ])

    r = admin_session.post('/tickets/inventario/importar_xlsx', data={
        'archivo': (buffer, 'inventario.xlsx')
    }, content_type='multipart/form-data', follow_redirects=True)

    texto = r.get_data(as_text=True)
    assert 'se cargaron 2' in texto.lower()


def test_importar_xlsx_rechaza_archivo_sin_extension_xlsx(admin_session):
    r = admin_session.post('/tickets/inventario/importar_xlsx', data={
        'archivo': (io.BytesIO(b'no es un excel'), 'inventario.csv')
    }, content_type='multipart/form-data', follow_redirects=True)

    assert 'xlsx' in r.get_data(as_text=True).lower()


def test_importar_xlsx_requiere_rol_operativo(sesion_usuario):
    buffer = _libro_xlsx([['40050', 'Portátil', '', '', '', 'Disponible', '', '', '', '', '', '', '', '']])
    r = sesion_usuario.post('/tickets/inventario/importar_xlsx', data={
        'archivo': (buffer, 'inventario.xlsx')
    }, content_type='multipart/form-data')
    assert r.status_code in (302, 403)


def test_exportar_csv_incluye_los_activos_y_sus_costos(admin_session, app):
    _crear_activo_directo(app, nombre='40060')
    r = admin_session.get('/tickets/inventario/exportar_csv')
    assert r.status_code == 200
    assert 'text/csv' in r.headers['Content-Type']
    texto = r.get_data(as_text=True)
    assert '40060' in texto


def test_exportar_xlsx_respeta_el_filtro_de_estado(admin_session, app):
    _crear_activo_directo(app, nombre='40070')
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("UPDATE activos_inventario SET estado = 'Perdido' WHERE nombre = ?", ('40070',))
    conn.commit()
    conn.close()
    _crear_activo_directo(app, nombre='40071')

    r = admin_session.get('/tickets/inventario/exportar_xlsx?estado=Perdido')
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    placas = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert '40070' in placas
    assert '40071' not in placas


def test_exportar_pdf_devuelve_un_pdf_valido(admin_session, app):
    _crear_activo_directo(app, nombre='40080')
    r = admin_session.get('/tickets/inventario/exportar_pdf')
    assert r.status_code == 200
    assert r.headers['Content-Type'] == 'application/pdf'
    assert r.data[:4] == b'%PDF'


def test_modal_inventario_incluye_los_enlaces_de_importar_exportar(admin_session):
    texto = admin_session.get('/tickets/inventario').get_data(as_text=True)
    assert 'importar_xlsx' in texto
    assert 'exportar_csv' in texto
    assert 'exportar_xlsx' in texto
    assert 'exportar_pdf' in texto
    assert 'plantilla_xlsx' in texto
