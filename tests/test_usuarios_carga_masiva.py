"""Pruebas de la Carga Masiva de usuarios (Gestión de Usuarios): Tomás pidió poder cargar los
datos de varias personas de una sola vez, con rol 'estandar' siempre y sin excepción, para que
el sistema le otorgue a cada una su propio usuario y contraseña temporal de ingreso — dejando
como único paso manual posterior (para un agente o administrador) ajustar el rol y el estado de
cada cuenta si hace falta. Reutiliza las MISMAS reglas de validación que el alta individual
(_crear_usuario_interno) — ver tests/test_inventario_usuario_rapido.py para esas reglas en
detalle; aquí se cubre sobre todo el comportamiento de la carga por archivo: qué filas se omiten,
que el rol siempre queda en 'estandar', y quién puede usar esta ruta."""
import io

import openpyxl


def _crear_especialidad(app, nombre='Auxiliar Administrativo'):
    conn, db_type = app.get_db()
    cur = conn.cursor()
    q = ("INSERT INTO especialidades_catalogo (nombre, estado) VALUES (%s, 'activo')"
         if db_type == 'postgres' else
         "INSERT INTO especialidades_catalogo (nombre, estado) VALUES (?, 'activo')")
    cur.execute(q, (nombre,))
    conn.commit()
    conn.close()
    return nombre


def _archivo_usuarios(filas):
    """Arma un .xlsx en memoria con el mismo formato de COLUMNAS_USUARIOS_XLSX — encabezado más
    una fila por tupla en 'filas' (primer_nombre, segundo_nombre, primer_apellido,
    segundo_apellido, email, telefono, cedula, especialidad)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Primer nombre', 'Segundo nombre', 'Primer apellido', 'Segundo apellido',
               'Correo electrónico', 'Teléfono', 'Cédula', 'Especialidad / Cargo'])
    for fila in filas:
        ws.append(list(fila))
    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


def test_carga_masiva_crea_usuarios_con_rol_estandar(admin_session, app):
    especialidad = _crear_especialidad(app)
    archivo = _archivo_usuarios([
        ('Laura', '', 'Gómez', 'Ramírez', 'laura.gomez@preventivaips.com.co', '3001111111', '10001', especialidad),
        ('Pedro', '', 'Ruiz', '', 'pedro.ruiz@preventivaips.com.co', '3002222222', '10002', especialidad),
    ])

    r = admin_session.post('/usuarios/importar_xlsx', data={'archivo': (archivo, 'usuarios.xlsx')},
                            content_type='multipart/form-data', follow_redirects=True)

    assert r.status_code == 200
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT rol, especialidad FROM usuarios WHERE cedula = ?", ('10001',))
    fila1 = cur.fetchone()
    cur.execute("SELECT rol, especialidad FROM usuarios WHERE cedula = ?", ('10002',))
    fila2 = cur.fetchone()
    conn.close()
    assert fila1 == ('estandar', especialidad)
    assert fila2 == ('estandar', especialidad)


def test_carga_masiva_omite_fila_sin_campos_obligatorios_pero_sigue_con_las_demas(admin_session, app):
    especialidad = _crear_especialidad(app)
    archivo = _archivo_usuarios([
        ('', '', '', '', '', '', '', ''),  # fila totalmente vacía: se ignora sin contar error
        ('Sin', '', 'Apellido', '', '', '', '', especialidad),  # sin correo: falla la validación
        ('Carla', '', 'Nieto', '', 'carla.nieto@preventivaips.com.co', '', '10003', especialidad),
    ])

    r = admin_session.post('/usuarios/importar_xlsx', data={'archivo': (archivo, 'usuarios.xlsx')},
                            content_type='multipart/form-data', follow_redirects=True)
    texto = r.get_data(as_text=True)

    assert 'Se crearon 1 usuario' in texto
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE cedula = ?", ('10003',))
    assert cur.fetchone()[0] == 1
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE nombre LIKE 'Sin Apellido%'")
    assert cur.fetchone()[0] == 0
    conn.close()


def test_carga_masiva_omite_cedula_repetida_dentro_del_mismo_archivo(admin_session, app):
    especialidad = _crear_especialidad(app)
    archivo = _archivo_usuarios([
        ('Ana', '', 'Torres', '', 'ana.torres@preventivaips.com.co', '', '10004', especialidad),
        ('Ana', '', 'Torres', '', 'ana.torres.otra@preventivaips.com.co', '', '10004', especialidad),
    ])

    r = admin_session.post('/usuarios/importar_xlsx', data={'archivo': (archivo, 'usuarios.xlsx')},
                            content_type='multipart/form-data', follow_redirects=True)
    texto = r.get_data(as_text=True)

    assert 'repetida más de una vez en este archivo' in texto
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE cedula = ?", ('10004',))
    assert cur.fetchone()[0] == 1
    conn.close()


def test_carga_masiva_omite_cedula_ya_registrada_en_la_base(admin_session, app, crear_usuario):
    especialidad = _crear_especialidad(app)
    crear_usuario(cedula='10005')
    archivo = _archivo_usuarios([
        ('Otra', '', 'Persona', '', 'otra.persona@preventivaips.com.co', '', '10005', especialidad),
    ])

    r = admin_session.post('/usuarios/importar_xlsx', data={'archivo': (archivo, 'usuarios.xlsx')},
                            content_type='multipart/form-data', follow_redirects=True)
    texto = r.get_data(as_text=True)

    assert 'Ya existe un usuario registrado con la cédula 10005' in texto


def test_carga_masiva_ignora_cualquier_columna_de_rol_y_siempre_usa_estandar(admin_session, app):
    """La plantilla no tiene columna de rol a propósito — pero por si alguien agrega una a mano
    y le escribe 'admin' o 'agente', la ruta no la lee en absoluto: el rol queda fijo."""
    especialidad = _crear_especialidad(app)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Primer nombre', 'Segundo nombre', 'Primer apellido', 'Segundo apellido',
               'Correo electrónico', 'Teléfono', 'Cédula', 'Especialidad / Cargo', 'Rol'])
    ws.append(['Marco', '', 'Silva', '', 'marco.silva@preventivaips.com.co', '', '10006', especialidad, 'admin'])
    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)

    admin_session.post('/usuarios/importar_xlsx', data={'archivo': (salida, 'usuarios.xlsx')},
                        content_type='multipart/form-data', follow_redirects=True)

    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT rol FROM usuarios WHERE cedula = ?", ('10006',))
    assert cur.fetchone()[0] == 'estandar'
    conn.close()


def test_plantilla_usuarios_xlsx_trae_las_columnas_esperadas(admin_session):
    r = admin_session.get('/usuarios/plantilla_xlsx')

    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.get_data()))
    encabezados = [c.value for c in wb.active[1]]
    assert encabezados == ['Primer nombre', 'Segundo nombre', 'Primer apellido', 'Segundo apellido',
                           'Correo electrónico', 'Teléfono', 'Cédula', 'Especialidad / Cargo']


def test_estandar_no_puede_usar_la_carga_masiva_de_usuarios(client, sesion_usuario, app):
    especialidad = _crear_especialidad(app)
    archivo = _archivo_usuarios([
        ('Nadie', '', 'Debería', '', 'nadie@preventivaips.com.co', '', '10007', especialidad),
    ])

    r = client.post('/usuarios/importar_xlsx', data={'archivo': (archivo, 'usuarios.xlsx')},
                     content_type='multipart/form-data')

    assert r.status_code in (302, 403)
    conn, db_type = app.get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE cedula = ?", ('10007',))
    assert cur.fetchone()[0] == 0
    conn.close()
