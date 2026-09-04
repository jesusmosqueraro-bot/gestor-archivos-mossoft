import os
import calendar
import math
import uuid
import random
import sqlite3
import urllib.request
import urllib.parse
import json
import unicodedata
import re
import io
import csv
import gzip
import threading
import base64
import hashlib
import hmac
import traceback
import time
import decimal
import secrets
import pyotp
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, jsonify, stream_with_context, send_file
from werkzeug.security import generate_password_hash, check_password_hash

# 📧 Validación de formato de correo electrónico del lado del SERVIDOR. Hasta ahora los
# formularios de Gestión de Usuarios solo se apoyaban en el atributo HTML type="email" del
# navegador (validación de cliente, fácil de saltarse enviando el formulario a mano); esto la
# complementa con una revisión real en el backend, siguiendo la recomendación de pruebas de
# seguridad de entrada revisada con el equipo (estructura básica usuario@dominio.tld). No
# pretende cubrir el 100% de la RFC 5322 (eso es innecesariamente estricto para este caso de
# uso) — solo rechazar valores que claramente no son un correo.
REGEX_EMAIL = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

def _email_tiene_formato_valido(email):
    return bool(email) and bool(REGEX_EMAIL.match(email.strip()))

# flask_wtf seguro (protección CSRF real)
try:
    from flask_wtf import CSRFProtect
    from flask_wtf.csrf import CSRFError
except Exception:
    CSRFProtect = None
    CSRFError = Exception

# 🔒 Hallazgo QA H-08 (30/08/2026): Flask-Limiter para poner un tope de peticiones por minuto
# en los endpoints públicos más sensibles (login, recuperación de clave, verificación 2FA).
# Antes no existía ningún límite a nivel de aplicación, así que un script podía probar
# contraseñas o códigos de 6 dígitos tan rápido como el servidor aguantara.
try:
    from flask_limiter import Limiter
except Exception:
    Limiter = None

# psycopg2 seguro para Render
try:
    import psycopg2
except Exception:
    psycopg2 = None

# cryptography para cifrado real de la bóveda de credenciales
try:
    from cryptography.fernet import Fernet, InvalidToken
except Exception:
    Fernet = None
    InvalidToken = Exception

import cloudinary
import cloudinary.uploader

# requests seguro
try:
    import requests
except Exception:
    requests = None

# bleach: limpia el HTML que produce el editor de texto enriquecido (Quill) antes de
# guardarlo, para que nadie pueda inyectar <script>, atributos onerror/onclick, etc. en la
# descripción de un ticket, un comentario o un comunicado. Si por algo no está instalado, se
# cae a texto plano (nunca se guarda HTML sin filtrar).
try:
    import bleach
    from bleach.css_sanitizer import CSSSanitizer
except Exception:
    bleach = None
    CSSSanitizer = None

# anthropic: SDK oficial de Claude, usado solo para sugerir la clasificación (tipo/categoría/
# prioridad) de un ticket nuevo — ver _clasificar_ticket_con_ia(). Si no está instalado, o si
# no hay ANTHROPIC_API_KEY configurada en las variables de entorno de Render, el botón
# "Clasificar con IA" simplemente no se muestra (ver _ia_clasificacion_disponible()); nunca
# tumba la app ni bloquea la creación manual de tickets.
try:
    import anthropic
except Exception:
    anthropic = None

app = Flask(__name__)
# 🔐 SECRET_KEY: nunca debe tener un valor real escrito en el código (quedaría expuesto en GitHub).
# Si no está seteada en las variables de entorno de Render, se genera una aleatoria en cada arranque.
# Esto no causa fricción extra: las sesiones ya se invalidan en cada reinicio por SERVER_INSTANCE_ID (ver abajo).
_SECRET_KEY_ENV = os.environ.get('SECRET_KEY')
if not _SECRET_KEY_ENV:
    print("⚠️ SECRET_KEY no configurada en variables de entorno de Render: se generó una aleatoria solo para esta instancia. Agrega SECRET_KEY en Render para persistirla.")
app.secret_key = _SECRET_KEY_ENV or base64.urlsafe_b64encode(os.urandom(32)).decode('utf-8')

SERVER_INSTANCE_ID = str(uuid.uuid4())
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=25)

# 🍪 Banderas explícitas de la cookie de sesión (antes quedaban en los valores por
# defecto de Flask). SECURE: el navegador nunca la envía por HTTP sin cifrar (Render
# solo sirve por HTTPS, así que no afecta nada). HTTPONLY: JavaScript en el navegador
# no puede leerla, ni siquiera vía un XSS. SAMESITE=Lax: no se envía en peticiones
# disparadas desde otros sitios, lo que además mitiga CSRF de rebote.
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'


@app.after_request
def _agregar_cabeceras_seguridad(response):
    # 🛡️ Cabeceras HTTP de seguridad estándar, ausentes hasta ahora. No cambian
    # el comportamiento visible de la app; reducen la superficie de ataque del
    # navegador (clickjacking, MIME-sniffing, fuga de referrer, etc.).
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    response.headers['Strict-Transport-Security'] = 'max-age=63072000; includeSubDomains'
    # CSP armada específicamente con los orígenes externos que la app realmente usa
    # hoy (Tailwind CDN, Font Awesome/cdnjs, Google Fonts, Cloudinary, reCAPTCHA).
    # 'unsafe-inline' se mantiene porque las plantillas usan scripts/estilos en línea
    # (onclick=, <script> por página); quitarlo requeriría reescribir todas las
    # plantillas a un esquema de nonces, algo que no se hizo en este cambio para no
    # arriesgar romper funcionalidad fuera de esta ronda de correcciones.
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com https://cdnjs.cloudflare.com "
        "https://www.google.com https://www.gstatic.com; "
        "style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
        "font-src 'self' data: https://fonts.gstatic.com https://cdnjs.cloudflare.com; "
        "img-src 'self' data: https://res.cloudinary.com; "
        "media-src 'self' https://res.cloudinary.com; "
        "connect-src 'self' https://www.google.com; "
        # 🩹 'self' y res.cloudinary.com: el visor de archivos (index.html) inserta un
        # <iframe src="URL_DE_CLOUDINARY"> para previsualizar PDFs sin descargarlos.
        # Sin res.cloudinary.com aquí, el navegador bloqueaba ese iframe por CSP y el
        # botón de vista previa de un PDF se veía en blanco (no se había notado porque
        # el resto de la app no depende de esto). 'self' queda por si el día de mañana
        # se enruta algún iframe a través de una ruta propia (como /pdf_proxy).
        "frame-src 'self' https://www.google.com https://res.cloudinary.com; "
        "object-src 'none'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        "frame-ancestors 'self';"
    )
    # 🔙 Sin caché en páginas de la app (todo menos /static/): evita que el botón
    # "atrás" del navegador muestre una página ya autenticada tomada de la caché
    # local (o del historial bfcache) sin volver a pasar por el servidor. Sin esto,
    # tras cerrar sesión o con la sesión expirada, "atrás" podía mostrar contenido
    # protegido sin pedir clave de nuevo. Con estas cabeceras, el navegador siempre
    # vuelve a pedir la página al servidor, que valida la sesión y redirige a
    # /login si ya no es válida (hallazgo QA, corregido).
    if not request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# 🛡️ Protección CSRF real vía Flask-WTF. Todas las plantillas con formularios POST ya
# incluyen (o se les agregó) el campo oculto csrf_token(); CSRFProtect valida ese token en
# cada POST/PUT/PATCH/DELETE y registra csrf_token() como global de Jinja automáticamente.
# Las dos únicas llamadas fetch() que no envían el token (incrementar_vista/descarga, meros
# contadores de vistas/descargas sin impacto sensible) quedan exentas explícitamente donde
# están definidas más abajo.
if CSRFProtect:
    csrf = CSRFProtect(app)

    @app.errorhandler(CSRFError)
    def _manejar_csrf_invalido(e):
        print(f"⚠️ CSRF inválido/expirado: {getattr(e, 'description', e)}")
        return redirect(request.referrer or url_for('index'))
else:
    # Si por alguna razón flask_wtf no está instalado, no tumbamos la app: se cae de vuelta
    # al placeholder inofensivo (sin protección CSRF real) en lugar de un 500 en cada página.
    print("⚠️ flask_wtf no está instalado: la protección CSRF real está desactivada. Agrega Flask-WTF a requirements.txt.")
    class _CsrfExemptDummy:
        def exempt(self, f):
            return f
    csrf = _CsrfExemptDummy()

    @app.context_processor
    def _inyectar_csrf_token_placeholder():
        return dict(csrf_token=lambda: '')

# 🔒 Hallazgo QA H-08: límites de peticiones por minuto. Usa _obtener_ip_cliente (misma función
# que ya usa el resto de la app para IP real detrás del proxy de Render vía X-Forwarded-For) en
# vez de la IP cruda de Flask-Limiter, así el límite es por visitante real y no por el balanceador.
#
# 📦 Almacenamiento de los contadores: por defecto en memoria del proceso — con los 2 workers de
# gunicorn de este servicio (ver Procfile) cada uno cuenta por su lado, así que el tope real puede
# llegar a ser hasta el doble del configurado. Sigue siendo una barrera muy superior a no tener
# ningún límite, pero deja de ser exacto en cuanto haya más de un worker/instancia.
#
# Si en Render se agrega un addon de Redis y se define la variable de entorno RATELIMIT_STORAGE_URI
# (o REDIS_URL, la que Render suele exponer para su addon "Key Value") con algo como
# "redis://usuario:clave@host:puerto", los contadores pasan a vivir ahí — compartidos entre todos
# los workers/instancias, sin ese margen de doble conteo. No hace falta tocar código para activarlo,
# solo agregar la variable de entorno; sin ella, sigue funcionando igual que hasta ahora (memoria).
_RATELIMIT_STORAGE_URI = os.environ.get('RATELIMIT_STORAGE_URI') or os.environ.get('REDIS_URL') or "memory://"

if Limiter:
    try:
        limiter = Limiter(
            key_func=lambda: _obtener_ip_cliente(),
            app=app,
            storage_uri=_RATELIMIT_STORAGE_URI,
            default_limits=[],
            # 🛟 in_memory_fallback_enabled: si Redis está configurado pero en algún momento no
            # responde (se cae el addon, URL mal escrita, problema de red puntual), Flask-Limiter
            # sigue limitando usando memoria local en vez de dejar pasar todo sin control
            # (fail-open) o tumbar la petición con un error 500. Es puramente un respaldo — no
            # reemplaza a Redis como fuente de verdad mientras esté disponible.
            in_memory_fallback_enabled=True,
        )
        if _RATELIMIT_STORAGE_URI != "memory://":
            print("🔒 Flask-Limiter usando almacenamiento compartido (Redis) para los límites de peticiones.")
    except Exception as e:
        # 🛟 Si la URL de Redis está mal escrita o hay cualquier otro problema al construir el
        # limiter (no la conexión en sí, que se valida perezosamente en cada petición — de eso se
        # encarga in_memory_fallback_enabled de arriba): no tiene sentido que la app entera falle
        # por esto. Se cae a memoria (el comportamiento de siempre) y sigue arrancando normal.
        print(f"⚠️ No se pudo inicializar Flask-Limiter con almacenamiento compartido ({e}); usando memoria como respaldo.")
        limiter = Limiter(
            key_func=lambda: _obtener_ip_cliente(),
            app=app,
            storage_uri="memory://",
            default_limits=[],
        )
else:
    print("⚠️ flask_limiter no está instalado: no habrá límite de peticiones por minuto. Agrega Flask-Limiter a requirements.txt.")
    class _LimiterExemptDummy:
        def limit(self, *args, **kwargs):
            def decorador(f):
                return f
            return decorador
    limiter = _LimiterExemptDummy()

# 🇨🇴 ZONA HORARIA COLOMBIA CON FALLBACK SEGURO
try:
    ZONA_HORARIA_COLOMBIA = ZoneInfo("America/Bogota")
except Exception:
    ZONA_HORARIA_COLOMBIA = timezone(timedelta(hours=-5))

def obtener_fecha_actual():
    # ⚠️ Formato ISO 8601 (AAAA-MM-DD HH:MM:SS): Postgres lo interpreta correctamente
    # sin importar su configuración de DateStyle. El formato anterior "DD/MM/AAAA hh:mm AM/PM"
    # (ej: "21/08/2026 05:51 PM") fallaba con "date/time field value out of range" en Neon
    # cada vez que el día del mes era mayor a 12, porque Postgres lo interpretaba como MM/DD/AAAA.
    try:
        return datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def normalizar(texto):
    if not texto: return ""
    texto = unicodedata.normalize('NFD', str(texto))
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return texto.lower().strip()

def _usuario_desde_nombre(texto):
    """Deja un nombre/apellido apto para nombre de usuario: sin tildes, minúsculas, solo
    letras y números (usa el mismo criterio de `normalizar` y además quita espacios/guiones,
    para apellidos compuestos como 'De la Cruz')."""
    return re.sub(r'[^a-z0-9]', '', normalizar(texto))

def _generar_username_unico(primer_nombre, primer_apellido, segundo_nombre='', segundo_apellido=''):
    """Genera un nombre de usuario a partir del nombre real de la persona, con la estructura
    pedida: primer nombre + primer apellido. Si ya existe, prueba estructuras alternativas
    (agregando la inicial del segundo apellido, la inicial del segundo nombre, o solo la
    inicial del primer nombre) y, como último recurso, agrega un número — hasta encontrar
    una que esté libre."""
    pn = _usuario_desde_nombre(primer_nombre)
    sn = _usuario_desde_nombre(segundo_nombre)
    pa = _usuario_desde_nombre(primer_apellido)
    sa = _usuario_desde_nombre(segundo_apellido)

    candidatos = []
    if pn and pa:
        candidatos.append(pn + pa)
        if sa:
            candidatos.append(pn + pa + sa[0])
        if sn:
            candidatos.append(pn + sn[0] + pa)
        candidatos.append(pn[0] + pa)

    vistos = set()
    candidatos_unicos = []
    for c in candidatos:
        if c and c not in vistos:
            vistos.add(c)
            candidatos_unicos.append(c)
    if not candidatos_unicos:
        candidatos_unicos = ['usuario']

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "SELECT id FROM usuarios WHERE LOWER(usuario) = LOWER(%s)" if db_type == 'postgres' else "SELECT id FROM usuarios WHERE LOWER(usuario) = LOWER(?)"
        for candidato in candidatos_unicos:
            cursor.execute(q, (candidato,))
            if not cursor.fetchone():
                return candidato
        # Todas las estructuras anteriores ya existen: se agrega un número incremental a la
        # primera opción hasta encontrar una libre.
        base = candidatos_unicos[0]
        for n in range(2, 1000):
            candidato = f"{base}{n}"
            cursor.execute(q, (candidato,))
            if not cursor.fetchone():
                return candidato
    finally:
        conn.close()
    # Fallback extremo (no debería alcanzarse en la práctica).
    return f"{candidatos_unicos[0]}{int(datetime.now(ZONA_HORARIA_COLOMBIA).timestamp())}"

# 🔐 CIFRADO REAL DE LA BÓVEDA (Fernet / AES-128-CBC + HMAC)
# La clave SOLO debe vivir en la variable de entorno ENCRYPTION_KEY de Render.
# Genera una con: python3 -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"
_ENCRYPTION_KEY = os.environ.get('ENCRYPTION_KEY')
_fernet = None
if Fernet and _ENCRYPTION_KEY:
    try:
        _fernet = Fernet(_ENCRYPTION_KEY.encode('utf-8'))
    except Exception as _e:
        print(f"⚠️ ENCRYPTION_KEY inválida, revisa la variable de entorno: {_e}")
        _fernet = None

def _encriptar_xor_legacy(texto, clave):
    """Compatibilidad de solo-lectura con el cifrado XOR viejo, para poder migrar datos existentes."""
    bytes_texto = texto.encode('utf-8')
    return bytes([b ^ clave[i % len(clave)] for i, b in enumerate(bytes_texto)])

def desencriptar_texto_legacy_xor(texto_cifrado):
    """Descifra valores guardados con el XOR viejo (antes de Fernet). Solo para migración."""
    if not texto_cifrado:
        return ""
    try:
        clave = app.secret_key.encode('utf-8')
        bytes_cifrados = base64.b64decode(texto_cifrado.encode('utf-8'))
        descifrado = bytes([b ^ clave[i % len(clave)] for i, b in enumerate(bytes_cifrados)])
        return descifrado.decode('utf-8')
    except Exception:
        return ""

def encriptar_texto(texto):
    if not texto:
        return ""
    if not _fernet:
        # Sin ENCRYPTION_KEY configurada no ciframos en claro por error: mejor fallar visible.
        raise RuntimeError("ENCRYPTION_KEY no configurada en las variables de entorno de Render.")
    return _fernet.encrypt(texto.encode('utf-8')).decode('utf-8')

def _migrar_credencial_a_fernet(credencial_id, password_plano):
    """Re-guarda con Fernet (AES real) una credencial de la bóveda que todavía
    estaba cifrada con el XOR viejo. Migra de forma transparente, un registro
    a la vez, cada vez que alguien la lee — igual que ya se hace con los
    passwords de usuarios en texto plano."""
    if not _fernet or credencial_id is None:
        return
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        nueva_cifrada = encriptar_texto(password_plano)
        query = "UPDATE credenciales SET password_cifrada = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE credenciales SET password_cifrada = ? WHERE id = ?"
        cursor.execute(query, (nueva_cifrada, credencial_id))
        conn.commit()
        conn.close()
        print(f"🔐 Credencial ID {credencial_id} migrada de cifrado XOR (débil) a Fernet/AES.")
    except Exception as e:
        print(f"Error migrando credencial {credencial_id} a Fernet: {e}")

def desencriptar_texto(texto_cifrado, credencial_id=None):
    """Descifra un valor de la bóveda. Si se pasa credencial_id y el valor
    resulta estar todavía en el XOR viejo (no es un token Fernet válido),
    se re-cifra con Fernet y se re-guarda de una vez en la base de datos."""
    if not texto_cifrado:
        return ""
    if _fernet:
        try:
            return _fernet.decrypt(texto_cifrado.encode('utf-8')).decode('utf-8')
        except (InvalidToken, Exception):
            pass  # No es un token Fernet válido: puede ser un valor viejo cifrado con XOR. Seguimos abajo.
    # Sin ENCRYPTION_KEY configurada (o el valor no era un token Fernet válido): intentamos
    # el XOR viejo de todas formas, ya que ese cifrado no depende de _fernet en absoluto.
    # Antes esta rama solo se intentaba cuando _fernet SÍ existía, dejando sin poder leerse
    # cualquier valor viejo si ENCRYPTION_KEY no estaba configurada.
    legacy = desencriptar_texto_legacy_xor(texto_cifrado)
    if legacy:
        if _fernet:
            _migrar_credencial_a_fernet(credencial_id, legacy)
        return legacy
    return "⚠️ No se pudo descifrar (revisa ENCRYPTION_KEY)"


def _parsear_dias_rotacion(valor):
    """Convierte el campo de formulario 'rotacion_dias' (texto, puede venir vacío) a un entero
    positivo o None (sin política de rotación para esa credencial)."""
    try:
        dias = int(str(valor).strip())
        return dias if dias > 0 else None
    except (TypeError, ValueError):
        return None


def _normalizar_etiquetas(valor):
    """Convierte el campo de formulario 'etiquetas' (texto libre separado por comas) a una
    cadena limpia lista para guardar: recorta espacios, descarta vacías y quita duplicados
    (sin importar mayúsculas/minúsculas), conservando el orden de aparición."""
    if not valor:
        return ''
    vistas = set()
    limpias = []
    for cruda in str(valor).split(','):
        etiqueta = cruda.strip()
        clave = etiqueta.lower()
        if etiqueta and clave not in vistas:
            vistas.add(clave)
            limpias.append(etiqueta)
    return ', '.join(limpias)


def _lista_etiquetas(etiquetas_texto):
    """Convierte la cadena guardada en 'etiquetas' de vuelta a una lista, para mostrarla como
    chips en la interfaz o para armar el filtro de etiquetas disponibles."""
    if not etiquetas_texto:
        return []
    return [e.strip() for e in etiquetas_texto.split(',') if e.strip()]


def _serializar_campos_personalizados(nombres, valores):
    """Bóveda Fase 2 — arma la lista de campos personalizados (nombre/valor) enviada desde el
    formulario (como dos listas paralelas, una por cada fila que el usuario agregó con JS) y la
    convierte a JSON listo para cifrar. Descarta filas sin nombre; conserva el orden capturado."""
    campos = []
    for nombre, valor in zip(nombres or [], valores or []):
        nombre = (nombre or '').strip()
        valor = (valor or '').strip()
        if nombre:
            campos.append({'nombre': nombre, 'valor': valor})
    return json.dumps(campos, ensure_ascii=False) if campos else ''


def _campos_personalizados_desde_json(texto_json):
    """Inverso de _serializar_campos_personalizados: interpreta el JSON ya descifrado y regresa
    la lista de {nombre, valor}. Nunca deja tumbar la vista si el JSON quedó corrupto/vacío."""
    if not texto_json:
        return []
    try:
        campos = json.loads(texto_json)
        if isinstance(campos, list):
            return [{'nombre': c.get('nombre', ''), 'valor': c.get('valor', '')} for c in campos if isinstance(c, dict)]
    except (TypeError, ValueError):
        pass
    return []

# ☁️ CLOUDINARY
cloudinary.config(
    cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key=os.environ.get('CLOUDINARY_API_KEY'),
    api_secret=os.environ.get('CLOUDINARY_API_SECRET')
)

# 📦 EXTENSIONES PERMITIDAS (INCLUYE COMPRIMIDOS Y DOCUMENTOS DE OFFICE)
ALLOWED_EXTENSIONS = {
    'png', 'jpg', 'jpeg', 'gif', 'webp',
    'pdf', 'txt', 'docx', 'xlsx', 'pptx',
    'mp4', 'mov', 'webm', 'avi',
    'zip', 'rar', '7z', 'tar', 'gz'
}
# 📦 LÍMITE DE TAMAÑO DE SUBIDA: 500 MB, con margen para videos de 4-10 minutos en buena
# calidad (un video de 4 min en 1080p suele pesar 100-300 MB según la compresión). Esto es
# el tope del lado de Flask/Render; el otro tope posible es el plan de Cloudinary (ver abajo).
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024

# 📧 URL DE TU GOOGLE APPS SCRIPT OFICIAL (PUERTO 443 HTTPS - SIN BLOQUEOS DE RENDER)
GMAIL_SCRIPT_URL = os.environ.get('GMAIL_SCRIPT_URL', "https://script.google.com/macros/s/AKfycbw81Lhduv86Y5vNWZ6hQ3XoHcRwmdIacKsRxBMfEJQ0uKlVPQx8zEiy-uEQoVWhjNuc/exec")

# 🔑 CLAVE SECRETA DE RECAPTCHA V2
# Nunca debe tener un valor real escrito en el código. Debe venir SIEMPRE de la variable
# de entorno RECAPTCHA_SECRET_KEY en Render (usa el valor que ya tenías funcionando).
RECAPTCHA_SECRET_KEY = os.environ.get('RECAPTCHA_SECRET_KEY', '')
if not RECAPTCHA_SECRET_KEY:
    print("⚠️ RECAPTCHA_SECRET_KEY no configurada en variables de entorno de Render: el login fallará hasta que la agregues.")

DATABASE_URL = os.environ.get('DATABASE_URL')

# 🔒 Endpoints a los que SÍ puede entrar un usuario marcado con "debe cambiar su contraseña",
# aunque todavía no la haya cambiado: la propia página/acción de cambio, cerrar sesión, y los
# archivos estáticos (CSS/JS/imágenes) que esa página necesita para verse bien.
ENDPOINTS_PERMITIDOS_CAMBIO_PASSWORD_OBLIGATORIO = {'cambiar_password_perfil', 'logout', 'static'}

# 🔒 Hallazgo QA H-05 (30/08/2026): admin y agente manejan datos de toda la organización
# (bóveda de accesos, gestor de archivos, base de datos), así que su 2FA ya no es opcional.
# Mismo mecanismo que ENDPOINTS_PERMITIDOS_CAMBIO_PASSWORD_OBLIGATORIO: mientras
# session['debe_activar_2fa'] esté encendido, cualquier ruta que no sea la propia activación
# de 2FA (o cerrar sesión) redirige a /perfil/2fa. Se incluye cambiar_password_perfil por si
# una cuenta recién creada por un admin tiene AMBAS pendientes a la vez (esa, al ir primero en
# validar_instancia_y_sesion, se resuelve antes de llegar aquí).
ENDPOINTS_PERMITIDOS_2FA_OBLIGATORIO = {'perfil_2fa', 'perfil_2fa_confirmar', 'cambiar_password_perfil', 'logout', 'static'}

@app.before_request
def validar_instancia_y_sesion():
    session.permanent = True
    if session.get('logged_in'):
        if session.get('instance_id') != SERVER_INSTANCE_ID:
            session.clear()
            return redirect(url_for('login', expirado='1'))
        if session.get('debe_cambiar_password') and request.endpoint not in ENDPOINTS_PERMITIDOS_CAMBIO_PASSWORD_OBLIGATORIO:
            return redirect(url_for('cambiar_password_perfil'))
        if session.get('debe_activar_2fa') and request.endpoint not in ENDPOINTS_PERMITIDOS_2FA_OBLIGATORIO:
            return redirect(url_for('perfil_2fa', obligatorio='1'))

def get_db():
    if DATABASE_URL and psycopg2:
        url = DATABASE_URL.replace("postgres://", "postgresql://", 1) if DATABASE_URL.startswith("postgres://") else DATABASE_URL
        conn = psycopg2.connect(url)
        return conn, 'postgres'
    else:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        DB_NAME = os.path.join(BASE_DIR, "gestor.db")
        conn = sqlite3.connect(DB_NAME)
        return conn, 'sqlite'

def init_db():
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        if db_type == 'postgres':
            cursor.execute('''CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY, usuario VARCHAR(100) UNIQUE NOT NULL, password_hash VARCHAR(255) NOT NULL, correo VARCHAR(200) NOT NULL, rol VARCHAR(50) NOT NULL DEFAULT 'estandar', estado VARCHAR(20) NOT NULL DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS galerias (
                id VARCHAR(50) PRIMARY KEY, titulo VARCHAR(200) NOT NULL, descripcion TEXT, fecha_subida VARCHAR(100), categoria VARCHAR(100) DEFAULT 'General', area VARCHAR(100) DEFAULT 'General', tipo VARCHAR(100) DEFAULT 'Instructivo', tags TEXT DEFAULT '', vistas INTEGER DEFAULT 0, descargas INTEGER DEFAULT 0, estado VARCHAR(50) DEFAULT 'activo'
            )''')
            # 👁️ Quién ha visto cada instructivo (UNIQUE por usuario: guarda la fecha de la
            # PRIMERA vista, igual que comunicados_leidos — no es un log de cada vista repetida).
            # 'vistas' en 'galerias' sigue siendo el contador de TODAS las vistas (no cambia);
            # esta tabla es aparte, solo para poder mostrar/exportar quién y cuándo.
            cursor.execute('''CREATE TABLE IF NOT EXISTS galerias_vistas (
                id SERIAL PRIMARY KEY, galeria_id VARCHAR(50) REFERENCES galerias(id) ON DELETE CASCADE, usuario VARCHAR(100) NOT NULL, fecha VARCHAR(100) NOT NULL, UNIQUE(galeria_id, usuario)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS archivos (
                id SERIAL PRIMARY KEY, galeria_id VARCHAR(50) REFERENCES galerias(id) ON DELETE CASCADE, filename TEXT, url_archivo TEXT NOT NULL DEFAULT '', nombre_original VARCHAR(255) NOT NULL DEFAULT '', estado VARCHAR(50) DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS logs (
                id SERIAL PRIMARY KEY, usuario VARCHAR(100) NOT NULL, accion VARCHAR(100) NOT NULL, detalles TEXT, fecha VARCHAR(100) NOT NULL
            )''')
            # 📧 Bitácora de envíos de correo (tickets y recuperación de clave). NUNCA guarda el
            # cuerpo del mensaje ni el código de verificación — solo asunto, destinatario, tipo
            # y si se logró enviar, para poder auditar entregas sin exponer datos sensibles.
            cursor.execute('''CREATE TABLE IF NOT EXISTS correos_log (
                id SERIAL PRIMARY KEY, fecha VARCHAR(100) NOT NULL, destinatario VARCHAR(200) NOT NULL, asunto VARCHAR(255) NOT NULL, tipo VARCHAR(30) NOT NULL, estado VARCHAR(20) NOT NULL, detalle_error TEXT
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales (
                id SERIAL PRIMARY KEY, titulo VARCHAR(150) NOT NULL, url_acceso TEXT, usuario_acceso VARCHAR(150) NOT NULL, password_cifrada TEXT NOT NULL, area VARCHAR(100) DEFAULT 'General', notas TEXT, fecha_creacion VARCHAR(100) NOT NULL, estado VARCHAR(50) DEFAULT 'activo'
            )''')
            # 📜 Historial de contraseñas anteriores de cada credencial (Bóveda Fase 2): se guarda
            # una fila cada vez que editar_credencial() cambia la clave de verdad, ANTES de
            # sobrescribirla, para poder consultar (con su propio audit-log) qué clave se usaba
            # antes. Solo aplica a items tipo 'credencial' (no a notas seguras).
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales_historial (
                id SERIAL PRIMARY KEY, credencial_id INTEGER NOT NULL REFERENCES credenciales(id) ON DELETE CASCADE, password_cifrada TEXT NOT NULL, fecha_cambio VARCHAR(100) NOT NULL, cambiado_por VARCHAR(100)
            )''')
            # 🔐 Bóveda Fase 3 — comparticiones puntuales de un ítem 'personal' (ver columna
            # visibilidad en credenciales) con otro miembro del equipo. Un ítem 'equipo' (el
            # comportamiento de siempre) no necesita filas aquí: lo ve todo admin/agente.
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales_compartidas (
                id SERIAL PRIMARY KEY, credencial_id INTEGER NOT NULL REFERENCES credenciales(id) ON DELETE CASCADE, usuario VARCHAR(100) NOT NULL, fecha_compartido VARCHAR(100) NOT NULL, compartido_por VARCHAR(100)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS comunicados (
                id SERIAL PRIMARY KEY, titulo VARCHAR(200) NOT NULL, contenido TEXT NOT NULL, nivel VARCHAR(50) DEFAULT 'info', fijado INTEGER DEFAULT 0, imagen_url TEXT DEFAULT '', estado VARCHAR(50) DEFAULT 'activo', fecha VARCHAR(100) NOT NULL, autor VARCHAR(100) NOT NULL
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets (
                id SERIAL PRIMARY KEY, titulo VARCHAR(200) NOT NULL, descripcion TEXT NOT NULL, tipo VARCHAR(20) DEFAULT 'Incidente', categoria VARCHAR(50) DEFAULT 'Otro', prioridad VARCHAR(20) DEFAULT 'Media', estado VARCHAR(20) DEFAULT 'Abierto', creado_por VARCHAR(100) NOT NULL, asignado_a VARCHAR(100), fecha_creacion VARCHAR(100) NOT NULL, fecha_actualizacion VARCHAR(100) NOT NULL
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets_comentarios (
                id SERIAL PRIMARY KEY, ticket_id INTEGER REFERENCES tickets(id) ON DELETE CASCADE, autor VARCHAR(100) NOT NULL, mensaje TEXT NOT NULL, tipo VARCHAR(20) DEFAULT 'comentario', fecha VARCHAR(100) NOT NULL
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets_adjuntos (
                id SERIAL PRIMARY KEY, ticket_id INTEGER REFERENCES tickets(id) ON DELETE CASCADE, comentario_id INTEGER REFERENCES tickets_comentarios(id) ON DELETE CASCADE, url TEXT NOT NULL, nombre_original VARCHAR(255) NOT NULL, subido_por VARCHAR(100) NOT NULL, fecha VARCHAR(100) NOT NULL
            )''')
            # ✅ Tareas asociadas a un ticket: subtareas internas del equipo de soporte para
            # repartir el trabajo de un mismo caso entre varias personas (inspirado en la
            # pestaña "Tareas" de Aranda Service Desk, simplificado a lo que Arkiv realmente
            # necesita). Solo las ve/gestiona el equipo operativo (agente/admin) — ver
            # agente_o_admin_required en las rutas /tickets/<id>/tareas/*.
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets_tareas (
                id SERIAL PRIMARY KEY, ticket_id INTEGER REFERENCES tickets(id) ON DELETE CASCADE, asunto VARCHAR(200) NOT NULL, descripcion TEXT, responsable VARCHAR(100), estado VARCHAR(20) DEFAULT 'pendiente', fecha_limite VARCHAR(20), creado_por VARCHAR(100) NOT NULL, fecha_creacion VARCHAR(100) NOT NULL, fecha_completada VARCHAR(100)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS conocimiento_articulos (
                id SERIAL PRIMARY KEY, titulo VARCHAR(200) NOT NULL, descripcion TEXT, url_documento TEXT NOT NULL, nombre_archivo VARCHAR(255) NOT NULL, vistas INTEGER DEFAULT 0, creado_por VARCHAR(100) NOT NULL, fecha_creacion VARCHAR(100) NOT NULL, estado VARCHAR(20) DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS ticket_configuraciones (
                id SERIAL PRIMARY KEY, tipo VARCHAR(20) NOT NULL, nombre VARCHAR(150) NOT NULL, estado VARCHAR(20) DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS activos_inventario (
                id SERIAL PRIMARY KEY, nombre VARCHAR(200) NOT NULL, tipo_activo VARCHAR(100) DEFAULT 'Otro', marca VARCHAR(100), modelo VARCHAR(100), numero_serie VARCHAR(150), estado VARCHAR(30) DEFAULT 'Disponible', asignado_a VARCHAR(100), sede VARCHAR(100), area VARCHAR(100), observaciones TEXT, fecha_creacion VARCHAR(100) NOT NULL, creado_por VARCHAR(100) NOT NULL, eliminado INTEGER DEFAULT 0
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS inventario_adjuntos (
                id SERIAL PRIMARY KEY, activo_id INTEGER REFERENCES activos_inventario(id) ON DELETE CASCADE, url TEXT NOT NULL, nombre_original VARCHAR(255) NOT NULL, subido_por VARCHAR(100) NOT NULL, fecha VARCHAR(100) NOT NULL
            )''')
            # 🪪 Certificación de devolución de activos: cuando Gestión Humana o TI confirma que
            # un colaborador ya devolvió el PC/activo que tenía asignado, queda una fila aquí
            # (quién lo confirmó, cuándo, con qué observaciones) — es el "certificado" que se
            # puede exportar, y a la vez lo que destraba la baja de su cuenta en Gestión de
            # Usuarios (ver toggle_estado_usuario / _activos_pendientes_devolucion).
            cursor.execute('''CREATE TABLE IF NOT EXISTS inventario_devoluciones (
                id SERIAL PRIMARY KEY, activo_id INTEGER REFERENCES activos_inventario(id) ON DELETE CASCADE, colaborador VARCHAR(150) NOT NULL, confirmado_por VARCHAR(100) NOT NULL, fecha VARCHAR(100) NOT NULL, observaciones TEXT
            )''')
            # 🏥 Acta de recibido para activos BIOMÉDICOS entregados a domicilio (bombas de
            # infusión y equipos similares que se le llevan a un paciente a su casa): a diferencia
            # de 'firma_asignacion_url' en activos_inventario (que solo guarda la ÚLTIMA firma y
            # se sobreescribe en cada reasignación), aquí queda UN REGISTRO HISTÓRICO por cada
            # entrega — quién recibió (paciente o cuidador), su documento, la dirección donde
            # quedó el equipo y su firma — para tener con qué responder si el equipo se pierde o
            # se daña, aunque el activo ya se haya vuelto a asignar a alguien más después.
            cursor.execute('''CREATE TABLE IF NOT EXISTS actas_recibido_biomedico (
                id SERIAL PRIMARY KEY, activo_id INTEGER NOT NULL REFERENCES activos_inventario(id) ON DELETE CASCADE, nombre_responsable VARCHAR(200) NOT NULL, relacion_responsable VARCHAR(30) NOT NULL, documento_responsable VARCHAR(50), telefono_contacto VARCHAR(30), direccion_entrega TEXT NOT NULL, firma_url TEXT NOT NULL, fecha VARCHAR(100) NOT NULL, creado_por VARCHAR(100) NOT NULL
            )''')
            # 🖼️ Panel de marca del login (imagen o video corto que rota junto al formulario de
            # acceso). Administrable desde Novedades y Comunicados → "Fondo de Login" — ver
            # _fondo_login_activo(). 'orden' decide en qué secuencia rotan los activos; 'estado'
            # deja pausar un archivo sin borrarlo (y perder su historial/URL de Cloudinary).
            cursor.execute('''CREATE TABLE IF NOT EXISTS login_fondo_media (
                id SERIAL PRIMARY KEY, tipo VARCHAR(20) NOT NULL, url TEXT NOT NULL, public_id VARCHAR(200), orden INTEGER DEFAULT 0, estado VARCHAR(20) DEFAULT 'activo', fecha_creacion VARCHAR(100) NOT NULL, creado_por VARCHAR(100) NOT NULL
            )''')
            # 🗂️ Catálogo administrable de Tipos de activo (Portátil, Impresora, Servidor...),
            # inspirado en el módulo de Solvyx: cada tipo tiene una key estable, una etiqueta
            # visible, un ícono (nombre de ícono de Font Awesome, sin el prefijo 'fa-') y un
            # orden de despliegue. Antes era una lista fija en el código (TIPOS_ACTIVO);
            # ahora el equipo puede agregar/reordenar/desactivar tipos sin tocar código.
            cursor.execute('''CREATE TABLE IF NOT EXISTS tipos_activo_catalogo (
                id SERIAL PRIMARY KEY, key VARCHAR(50) NOT NULL, etiqueta VARCHAR(150) NOT NULL, icono VARCHAR(50) DEFAULT 'box', orden INTEGER DEFAULT 0, estado VARCHAR(20) DEFAULT 'activo'
            )''')
            # 🔁 Historial de reemplazos de activos (Reemplazar activo / Trazabilidad, visto en
            # Solvyx): cada fila conecta un activo "anterior" con el activo que lo reemplazó,
            # con el motivo, notas libres y qué pasó con el activo anterior. Reconstruyendo la
            # cadena (activo_nuevo_id de una fila == activo_anterior_id de la siguiente) se arma
            # la trazabilidad completa de un equipo a través de sus reemplazos sucesivos.
            cursor.execute('''CREATE TABLE IF NOT EXISTS activos_reemplazos (
                id SERIAL PRIMARY KEY, activo_anterior_id INTEGER NOT NULL REFERENCES activos_inventario(id) ON DELETE CASCADE, activo_nuevo_id INTEGER REFERENCES activos_inventario(id) ON DELETE SET NULL, motivo VARCHAR(50) NOT NULL, notas TEXT, fecha_reemplazo VARCHAR(100) NOT NULL, estado_anterior_resultante VARCHAR(30) NOT NULL, creado_por VARCHAR(100) NOT NULL, fecha_creacion VARCHAR(100) NOT NULL
            )''')
            # 👁️ Registro de qué usuario ya leyó cada comunicado (se marca al ver el muro de
            # comunicados o el comunicado fijado en la bienvenida) — para saber quién falta.
            cursor.execute('''CREATE TABLE IF NOT EXISTS comunicados_leidos (
                id SERIAL PRIMARY KEY, comunicado_id INTEGER REFERENCES comunicados(id) ON DELETE CASCADE, usuario VARCHAR(100) NOT NULL, fecha VARCHAR(100) NOT NULL, UNIQUE(comunicado_id, usuario)
            )''')
            # 🔔 Notificaciones internas (campanita): se generan en los mismos puntos donde ya
            # sale un correo (ticket creado, comentado, cambio de estado).
            cursor.execute('''CREATE TABLE IF NOT EXISTS notificaciones (
                id SERIAL PRIMARY KEY, usuario VARCHAR(100) NOT NULL, tipo VARCHAR(50) DEFAULT 'ticket', mensaje TEXT NOT NULL, url TEXT DEFAULT '', leida INTEGER DEFAULT 0, estado VARCHAR(20) DEFAULT 'activa', fecha VARCHAR(100) NOT NULL
            )''')
            # 🆔 Catálogo de aplicativos/herramientas para los que se crean credenciales a los
            # colaboradores (KUBAPP, SAMI, Moodle, Wolkvox, Correo, Solvyx...). Administrable:
            # el equipo de soporte puede agregar o desactivar aplicativos sin tocar código.
            cursor.execute('''CREATE TABLE IF NOT EXISTS aplicativos_catalogo (
                id SERIAL PRIMARY KEY, nombre VARCHAR(150) NOT NULL, estado VARCHAR(20) DEFAULT 'activo'
            )''')
            # 🩺 Catálogo de especialidades/áreas para los usuarios de Arkiv (Medicina General,
            # Enfermería, Odontología, Administrativo...). Administrable desde /usuarios: el
            # equipo admin puede agregar nuevas especialidades sin tocar código, igual que con
            # el catálogo de aplicativos.
            cursor.execute('''CREATE TABLE IF NOT EXISTS especialidades_catalogo (
                id SERIAL PRIMARY KEY, nombre VARCHAR(150) NOT NULL, estado VARCHAR(20) DEFAULT 'activo'
            )''')
            # 🪪 Altas y bajas de credenciales de colaboradores: un registro por cada aplicativo
            # que se le habilita a una persona (no por colaborador), para poder deshabilitar el
            # acceso a un aplicativo puntual sin afectar los demás que tenga esa misma persona.
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales_colaboradores (
                id SERIAL PRIMARY KEY, colaborador VARCHAR(200) NOT NULL, aplicativo VARCHAR(150) NOT NULL, password_cifrada TEXT NOT NULL, fecha_creacion VARCHAR(100), fecha_solicitud VARCHAR(100), analista_gestiona VARCHAR(150), solicitado_por VARCHAR(150), capacitado_por VARCHAR(150), medio_envio VARCHAR(30), estado VARCHAR(20) DEFAULT 'activo', fecha_deshabilitacion VARCHAR(100), deshabilitado_por VARCHAR(150), fecha_registro VARCHAR(100) NOT NULL, registrado_por VARCHAR(150) NOT NULL
            )''')
            # 🔐 Códigos de respaldo (recuperación) de la verificación en dos pasos (2FA): se
            # generan 10 de un solo uso al activar el 2FA, se guardan SOLO su hash (nunca el
            # código en claro) y sirven para entrar si la persona pierde su app autenticadora.
            cursor.execute('''CREATE TABLE IF NOT EXISTS totp_codigos_respaldo (
                id SERIAL PRIMARY KEY, usuario VARCHAR(100) NOT NULL, codigo_hash VARCHAR(255) NOT NULL, usado INTEGER DEFAULT 0, fecha_creacion VARCHAR(100) NOT NULL, fecha_uso VARCHAR(100)
            )''')
            # 📋 Plantillas de solicitud: administradas por el equipo de soporte (/tickets/plantillas)
            # para acelerar la creación de tickets recurrentes — al elegir una en "Nueva Solicitud"
            # se prellenan tipo, título, categoría, prioridad, área, sede y descripción, todo
            # editable antes de enviar.
            cursor.execute('''CREATE TABLE IF NOT EXISTS ticket_plantillas (
                id SERIAL PRIMARY KEY, nombre VARCHAR(150) NOT NULL, tipo VARCHAR(20) DEFAULT 'Incidente', categoria VARCHAR(50), prioridad VARCHAR(20) DEFAULT 'Media', area VARCHAR(100), sede VARCHAR(100), titulo VARCHAR(200) NOT NULL, descripcion TEXT NOT NULL, estado VARCHAR(20) DEFAULT 'activo', creado_por VARCHAR(100) NOT NULL, fecha_creacion VARCHAR(100) NOT NULL
            )''')
            # 📅 Vencimiento de documentos por empleado (certificaciones, cursos, exámenes
            # médicos ocupacionales, licencias...): un registro por documento/colaborador,
            # independiente de los instructivos institucionales de 'galerias'. 'alerta_nivel'
            # guarda el último aviso ya enviado ('proximo_a_vencer'/'vencido') para no repetir
            # la misma notificación — ver _revisar_alertas_vencimientos().
            cursor.execute('''CREATE TABLE IF NOT EXISTS documentos_empleado (
                id SERIAL PRIMARY KEY, usuario VARCHAR(100) NOT NULL, tipo_documento VARCHAR(150) NOT NULL, titulo VARCHAR(200) NOT NULL, descripcion TEXT, url_archivo TEXT DEFAULT '', nombre_original VARCHAR(255) DEFAULT '', fecha_emision VARCHAR(20), fecha_vencimiento VARCHAR(20), alerta_nivel VARCHAR(20), estado VARCHAR(20) DEFAULT 'activo', creado_por VARCHAR(100) NOT NULL, fecha_creacion VARCHAR(100) NOT NULL
            )''')
            conn.commit()

            for col_query in [
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS categoria VARCHAR(100) DEFAULT 'General';",
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS tipo VARCHAR(100) DEFAULT 'Instructivo';",
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS tags TEXT DEFAULT '';",
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS vistas INTEGER DEFAULT 0;",
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS descargas INTEGER DEFAULT 0;",
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS estado VARCHAR(50) DEFAULT 'activo';",
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS area VARCHAR(100) DEFAULT 'General';",
                # 👁️ Visibilidad del instructivo: 'todos' (lo ve cualquier usuario logueado) o
                # 'admin' (solo lo ven las cuentas Admin/Agente). Por defecto 'todos', para no
                # ocultar de golpe instructivos que ya existían antes de esta función.
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS visibilidad VARCHAR(20) DEFAULT 'todos';",
                "ALTER TABLE archivos ADD COLUMN IF NOT EXISTS estado VARCHAR(50) DEFAULT 'activo';",
                "ALTER TABLE archivos ADD COLUMN IF NOT EXISTS url_archivo TEXT DEFAULT '';",
                "ALTER TABLE archivos ADD COLUMN IF NOT EXISTS nombre_original VARCHAR(255) DEFAULT '';",
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS estado VARCHAR(50) DEFAULT 'activo';",
                # 🔐 Auditoría de accesos: en qué credencial quedó registrada cada consulta/copia
                # de clave (ver _revelar_credencial); NULL para el resto de acciones del log.
                "ALTER TABLE logs ADD COLUMN IF NOT EXISTS credencial_id INTEGER;",
                # 🔁 Política de rotación de contraseñas (opcional, por credencial): cada cuántos
                # días se debería cambiar, cuándo se rotó por última vez, y cuándo se avisó (una
                # sola vez por ciclo) que ya toca — ver _revisar_recordatorios_rotacion().
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS rotacion_dias INTEGER;",
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS fecha_ultima_rotacion VARCHAR(100);",
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS rotacion_recordatorio_fecha VARCHAR(100);",
                # 🏷️ Etiquetas múltiples (además de 'area', que sigue siendo la categoría única de
                # siempre): texto separado por comas, normalizado en crear_credencial/
                # editar_credencial (ver _normalizar_etiquetas). Permite organizar más fino sin
                # romper el filtro por categoría que ya existe.
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS etiquetas TEXT;",
                # 📝 Bóveda Fase 2 — 'nota_segura' (contenido_seguro cifrado, sin usuario/clave
                # reales: se guardan valores placeholder en usuario_acceso/password_cifrada para
                # no romper sus columnas NOT NULL) y campos personalizados (JSON cifrado, lista de
                # {nombre, valor} por credencial). tipo_item distingue una credencial normal de
                # una nota segura; el resto de la app sigue asumiendo 'credencial' si no está.
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS tipo_item VARCHAR(20) DEFAULT 'credencial';",
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS contenido_seguro TEXT;",
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS campos_personalizados TEXT;",
                # 🔐 Bóveda Fase 3 — bóvedas personales: 'propietario' es quien creó el ítem;
                # 'visibilidad' es 'equipo' (todo admin/agente lo ve, el comportamiento de
                # siempre — y el valor por defecto, así que nada existente cambia) o 'personal'
                # (solo el dueño, quienes tengan comparticiones puntuales en
                # credenciales_compartidas, y los admin —que conservan acceso de auditoría—).
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS propietario VARCHAR(100);",
                "ALTER TABLE credenciales ADD COLUMN IF NOT EXISTS visibilidad VARCHAR(20) DEFAULT 'equipo';",
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS estado VARCHAR(20) DEFAULT 'activo';",
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nombre VARCHAR(200);",
                # 📞 Teléfono de contacto del usuario: opcional, se usa para prellenar el número
                # de contacto al crear un ticket y para que soporte tenga cómo ubicarlo.
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS telefono VARCHAR(50);",
                # 🎨 Preferencia de tema (claro/oscuro) de cada usuario: se guarda en su cuenta
                # (no solo en el navegador) para que lo siga a donde inicie sesión.
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS tema VARCHAR(20) DEFAULT 'oscuro';",
                # 🪪 Cédula/documento de identidad: opcional, no es única (no todos los usuarios
                # existentes la tienen todavía). Sirve para buscar y asociar rápidamente a la
                # persona correcta al asignarle un activo del Inventario, sin tener que escribir
                # su nombre completo de memoria.
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS cedula VARCHAR(30);",
                # 🩺 Especialidad/área del usuario (Medicina General, Enfermería, Administrativo...),
                # elegida de especialidades_catalogo. Obligatoria para cuentas nuevas; las que ya
                # existían quedan en NULL hasta que se editen.
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS especialidad VARCHAR(150);",
                # 🔒 Forzar cambio de contraseña: se marca en TRUE cuando un admin crea la cuenta
                # o le reasigna la contraseña a otra persona (contraseña temporal conocida por el
                # admin), para obligar a que la cambien en su próximo inicio de sesión antes de
                # poder usar el resto de Arkiv. Se limpia a FALSE apenas la cambian.
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS debe_cambiar_password BOOLEAN DEFAULT FALSE;",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS tipo VARCHAR(20) DEFAULT 'Incidente';",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sla_respuesta_limite VARCHAR(100);",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sla_resolucion_limite VARCHAR(100);",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sla_respuesta_cumplida VARCHAR(100);",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sla_resolucion_cumplida VARCHAR(100);",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sla_modificaciones INTEGER DEFAULT 0;",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS calificacion INTEGER;",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS calificacion_fecha VARCHAR(100);",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS area VARCHAR(100);",
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sede VARCHAR(100);",
                # 📞 Número de contacto para ESTE ticket en particular: puede ser distinto al
                # teléfono registrado en el perfil del usuario (p. ej. reporta desde la
                # extensión de otra persona). Se prellena con el del perfil pero es editable.
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS telefono_contacto VARCHAR(50);",
                # 🔔 Último nivel de alerta de SLA ya avisado para este ticket ('proximo_a_vencer'
                # o 'vencido') — evita reenviar el mismo aviso en cada visita a la lista de
                # tickets; se limpia cuando se extiende el SLA para que pueda volver a avisar.
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sla_alerta_nivel VARCHAR(20);",
                # 🗑️ Baja lógica: para tickets de prueba/capacitación o creados por error que el
                # equipo de soporte quiera sacar de las listas por completo (a diferencia de
                # 'Cancelado', que es un estado visible pensado para dejar rastro de auditoría).
                # Solo el super-admin puede activarla (ver eliminar_ticket()).
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS eliminado INTEGER DEFAULT 0;",
                # 🔗 Activo de Inventario al que se refiere este ticket (opcional). Permite, desde
                # el activo, ver el historial completo de solicitudes que se le han abierto.
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS activo_id INTEGER REFERENCES activos_inventario(id) ON DELETE SET NULL;",
                # ⏸️ Fecha desde la que el ticket entró en estado 'Pendiente' (pausa de SLA). Se
                # usa para congelar la barra de progreso y, al salir de Pendiente, para calcular
                # cuántas horas estuvo pausado y correr el límite de resolución esa misma cantidad.
                # NULL cuando el ticket nunca ha estado pausado o ya se reanudó.
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS sla_pausado_desde VARCHAR(100);",
                # 🔗 Ticket ya Resuelto/Cerrado que el usuario asocia a este (p. ej. un caso nuevo
                # relacionado con uno ya cerrado, o un duplicado que hace referencia al original).
                # Puramente informativo, no afecta el flujo ni el SLA del ticket referenciado.
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS ticket_relacionado_id INTEGER REFERENCES tickets(id) ON DELETE SET NULL;",
                # 🙋 Usuario real para quien es la solicitud, cuando quien la crea es un agente
                # actuando en nombre de otra persona (p. ej. sube un PQRS a nombre de un usuario).
                # Cuando es NULL, el beneficiario es 'creado_por' (el caso normal: el usuario crea
                # su propio ticket). Se usa para decidir quién puede calificar el servicio, de modo
                # que el agente que resuelve el caso no pueda calificarse a sí mismo.
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS solicitante_real VARCHAR(150);",
                # 📢 Fecha en que se envió (una sola vez) el recordatorio automático de lectura
                # pendiente de este comunicado — ver _revisar_recordatorios_lectura(). NULL
                # mientras no se haya enviado ninguno (automático o manual) todavía.
                "ALTER TABLE comunicados ADD COLUMN IF NOT EXISTS recordatorio_enviado_fecha VARCHAR(100);",
                # 📍 Dirección física y usuario responsable de cada Sede (solo aplican cuando
                # tipo = 'sede'; para 'area'/'categoria' quedan en NULL sin usarse).
                "ALTER TABLE ticket_configuraciones ADD COLUMN IF NOT EXISTS direccion TEXT;",
                "ALTER TABLE ticket_configuraciones ADD COLUMN IF NOT EXISTS responsable VARCHAR(100);",
                # 🕵️ Historial de sesiones: IP y dispositivo (navegador/SO detectados a partir del
                # User-Agent) de cada inicio/cierre de sesión — ver registrar_log() y la ruta
                # /perfil/historial-sesiones. NULL para el resto de acciones del log (no aplica).
                "ALTER TABLE logs ADD COLUMN IF NOT EXISTS ip VARCHAR(100);",
                "ALTER TABLE logs ADD COLUMN IF NOT EXISTS dispositivo VARCHAR(255);",
                # 🔐 Verificación en dos pasos (2FA/TOTP): secreto TOTP de la cuenta (NULL hasta
                # que se active) y si ya quedó confirmado/activo. Ver rutas /perfil/2fa*.
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS totp_secret VARCHAR(64);",
                # 🔒 Hallazgo QA H-07 (30/08/2026): el secreto TOTP se guardaba en texto plano.
                # A partir de ahora se cifra con Fernet (igual que las contraseñas de la bóveda,
                # ver encriptar_texto/desencriptar_texto) — un token Fernet no cabe en
                # VARCHAR(64), así que se amplía la columna a TEXT para las instalaciones que
                # ya la tenían creada con el tamaño viejo. _migrar_totp_secret_a_fernet() migra
                # de forma transparente cada secreto existente en texto plano la próxima vez
                # que se usa (login 2FA o vista de /perfil/2fa).
                "ALTER TABLE usuarios ALTER COLUMN totp_secret TYPE TEXT;",
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS totp_habilitado BOOLEAN DEFAULT FALSE;",
                # 🖼️ Foto de perfil (autoservicio, ver /perfil): URL de Cloudinary de la imagen que
                # la propia persona subió, o NULL si nunca subió una (se muestra un ícono genérico).
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS foto_perfil TEXT;",
                # 📅 Vencimiento de documentos institucionales (Instructivos y Archivos): fecha
                # opcional en la que el instructivo/documento deja de estar vigente (ej. una
                # política o certificado con renovación periódica). 'alerta_vencimiento_nivel'
                # guarda el último aviso ya enviado, igual que sla_alerta_nivel en Tickets — ver
                # _revisar_alertas_vencimientos().
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS fecha_vencimiento VARCHAR(20);",
                "ALTER TABLE galerias ADD COLUMN IF NOT EXISTS alerta_vencimiento_nivel VARCHAR(20);",
                # 🔒 Bloqueo automático por intentos fallidos de inicio de sesión (distinto del
                # bloqueo manual que ya maneja 'estado'): 'intentos_fallidos_login' cuenta los
                # fallos de contraseña consecutivos y se reinicia en cada acierto;
                # 'bloqueado_por_intentos' se enciende al llegar a UMBRAL_INTENTOS_FALLIDOS_LOGIN
                # y bloquea el acceso aunque la clave sea correcta, hasta que un admin lo
                # desbloquee (ver desbloquear_intentos_usuario) o la propia persona recupere su
                # clave por correo (ver validar_codigo).
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS intentos_fallidos_login INTEGER DEFAULT 0;",
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS bloqueado_por_intentos BOOLEAN DEFAULT FALSE;",
                # 🗑️ Papelera de notificaciones: 'activa' = visible en la campanita, 'archivada' =
                # movida a la papelera (se conserva, pero deja de contar para el badge de no
                # leídas y de aparecer en la lista principal) hasta que la persona la restaure o
                # la elimine definitivamente desde ahí. Ver /notificaciones/<id>/archivar,
                # /restaurar, /eliminar y /notificaciones/papelera/vaciar.
                "ALTER TABLE notificaciones ADD COLUMN IF NOT EXISTS estado VARCHAR(20) DEFAULT 'activa';",
                # 🏢 Área del activo (Sistemas, Urgencias, Facturación, etc.) — igual que 'sede',
                # se valida contra el catálogo de Áreas ya administrado en /tickets/configuracion
                # (ver _config_ticket_lista('area')); ayuda a ubicar con más precisión un equipo
                # de cómputo o biomédico dentro de la sede donde está.
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS area VARCHAR(100);",
                # 🚚 Catálogo de Proveedores (mismo mecanismo que Área/Sede: filas de
                # ticket_configuraciones con tipo = 'proveedor'). 'nit' y 'razon_social' solo
                # tienen contenido real para tipo = 'proveedor'; para área/sede/categoría quedan
                # en NULL sin usarse — ver _config_ticket_lista().
                "ALTER TABLE ticket_configuraciones ADD COLUMN IF NOT EXISTS nit VARCHAR(50);",
                "ALTER TABLE ticket_configuraciones ADD COLUMN IF NOT EXISTS razon_social VARCHAR(200);",
                # 🚚 Proveedor del activo (quién lo vendió/suministró) — se valida contra el
                # catálogo de Proveedores de /tickets/configuracion, igual que 'area' y 'sede'.
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS proveedor VARCHAR(100);",
                # 💰 Costos de Inventario: 'tipo_costo' distingue si el activo es propio (se pagó
                # una vez, 'costo_compra') o alquilado (se paga mes a mes, 'costo_alquiler_mensual')
                # — ambos campos conviven en la misma fila porque un activo solo es una cosa u
                # otra a la vez, nunca las dos; el que no aplica queda en NULL y no se contabiliza
                # (ver _totales_costos_inventario). Sin 'tipo_costo' definido (activos ya
                # existentes antes de este cambio) el activo simplemente no entra en los totales.
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS tipo_costo VARCHAR(20);",
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS costo_compra NUMERIC(14,2);",
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS costo_alquiler_mensual NUMERIC(14,2);",
                # ✍️ Firma digital: se captura una sola vez al crear la cuenta (dibujada a mano o
                # una imagen subida, ver _subir_firma_desde_dataurl) y queda en 'usuarios.firma'
                # como URL de Cloudinary. Al asignar un activo a esa persona, esa MISMA firma se
                # copia a 'activos_inventario.firma_asignacion_url' (ver crear_activo/editar_activo)
                # como constancia de aceptación — sin tener que volver a firmar en cada asignación.
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS firma TEXT;",
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS firma_asignacion_url TEXT;",
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS firma_asignacion_fecha VARCHAR(20);",
                # 🏥 Marca manual (no depende del catálogo de Tipos de activo) para identificar
                # equipos biomédicos que se entregan a domicilio (bombas de infusión, etc.) — ver
                # actas_recibido_biomedico. Al asignar uno de estos activos, el modal pide además
                # la dirección de entrega y la firma del paciente/cuidador que lo recibió.
                "ALTER TABLE activos_inventario ADD COLUMN IF NOT EXISTS es_biomedico BOOLEAN DEFAULT FALSE;"
            ]:
                try:
                    cursor.execute(col_query)
                    conn.commit()
                except Exception:
                    conn.rollback()

        else:
            cursor.execute('''CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT, usuario TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL, correo TEXT NOT NULL, rol TEXT NOT NULL DEFAULT 'estandar', estado TEXT NOT NULL DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS galerias (
                id TEXT PRIMARY KEY, titulo TEXT NOT NULL, descripcion TEXT, fecha_subida TEXT, categoria TEXT DEFAULT 'General', area TEXT DEFAULT 'General', tipo TEXT DEFAULT 'Instructivo', tags TEXT DEFAULT '', vistas INTEGER DEFAULT 0, descargas INTEGER DEFAULT 0, estado TEXT DEFAULT 'activo'
            )''')
            # 👁️ Quién ha visto cada instructivo. Ver comentario equivalente en la rama de Postgres.
            cursor.execute('''CREATE TABLE IF NOT EXISTS galerias_vistas (
                id INTEGER PRIMARY KEY AUTOINCREMENT, galeria_id TEXT NOT NULL, usuario TEXT NOT NULL, fecha TEXT NOT NULL, FOREIGN KEY(galeria_id) REFERENCES galerias(id) ON DELETE CASCADE, UNIQUE(galeria_id, usuario)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS archivos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, galeria_id TEXT, filename TEXT, url_archivo TEXT NOT NULL DEFAULT '', nombre_original TEXT NOT NULL DEFAULT '', estado TEXT DEFAULT 'activo', FOREIGN KEY(galeria_id) REFERENCES galerias(id) ON DELETE CASCADE
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT, usuario TEXT NOT NULL, accion TEXT NOT NULL, detalles TEXT, fecha TEXT NOT NULL
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS correos_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT NOT NULL, destinatario TEXT NOT NULL, asunto TEXT NOT NULL, tipo TEXT NOT NULL, estado TEXT NOT NULL, detalle_error TEXT
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales (
                id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, url_acceso TEXT, usuario_acceso TEXT NOT NULL, password_cifrada TEXT NOT NULL, area TEXT DEFAULT 'General', notas TEXT, fecha_creacion VARCHAR(100) NOT NULL, estado TEXT DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales_historial (
                id INTEGER PRIMARY KEY AUTOINCREMENT, credencial_id INTEGER NOT NULL, password_cifrada TEXT NOT NULL, fecha_cambio TEXT NOT NULL, cambiado_por TEXT, FOREIGN KEY(credencial_id) REFERENCES credenciales(id) ON DELETE CASCADE
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales_compartidas (
                id INTEGER PRIMARY KEY AUTOINCREMENT, credencial_id INTEGER NOT NULL, usuario TEXT NOT NULL, fecha_compartido TEXT NOT NULL, compartido_por TEXT, FOREIGN KEY(credencial_id) REFERENCES credenciales(id) ON DELETE CASCADE
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS comunicados (
                id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, contenido TEXT NOT NULL, nivel TEXT DEFAULT 'info', fijado INTEGER DEFAULT 0, imagen_url TEXT DEFAULT '', estado TEXT DEFAULT 'activo', fecha TEXT NOT NULL, autor TEXT NOT NULL
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, descripcion TEXT NOT NULL, tipo TEXT DEFAULT 'Incidente', categoria TEXT DEFAULT 'Otro', prioridad TEXT DEFAULT 'Media', estado TEXT DEFAULT 'Abierto', creado_por TEXT NOT NULL, asignado_a TEXT, fecha_creacion TEXT NOT NULL, fecha_actualizacion TEXT NOT NULL
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets_comentarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT, ticket_id INTEGER, autor TEXT NOT NULL, mensaje TEXT NOT NULL, tipo TEXT DEFAULT 'comentario', fecha TEXT NOT NULL, FOREIGN KEY(ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets_adjuntos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, ticket_id INTEGER, comentario_id INTEGER, url TEXT NOT NULL, nombre_original TEXT NOT NULL, subido_por TEXT NOT NULL, fecha TEXT NOT NULL, FOREIGN KEY(ticket_id) REFERENCES tickets(id) ON DELETE CASCADE, FOREIGN KEY(comentario_id) REFERENCES tickets_comentarios(id) ON DELETE CASCADE
            )''')
            # ✅ Tareas asociadas a un ticket. Ver comentario equivalente en la rama de Postgres.
            cursor.execute('''CREATE TABLE IF NOT EXISTS tickets_tareas (
                id INTEGER PRIMARY KEY AUTOINCREMENT, ticket_id INTEGER, asunto TEXT NOT NULL, descripcion TEXT, responsable TEXT, estado TEXT DEFAULT 'pendiente', fecha_limite TEXT, creado_por TEXT NOT NULL, fecha_creacion TEXT NOT NULL, fecha_completada TEXT, FOREIGN KEY(ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS conocimiento_articulos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, descripcion TEXT, url_documento TEXT NOT NULL, nombre_archivo TEXT NOT NULL, vistas INTEGER DEFAULT 0, creado_por TEXT NOT NULL, fecha_creacion TEXT NOT NULL, estado TEXT DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS ticket_configuraciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT NOT NULL, nombre TEXT NOT NULL, estado TEXT DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS activos_inventario (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, tipo_activo TEXT DEFAULT 'Otro', marca TEXT, modelo TEXT, numero_serie TEXT, estado TEXT DEFAULT 'Disponible', asignado_a TEXT, sede TEXT, area TEXT, observaciones TEXT, fecha_creacion TEXT NOT NULL, creado_por TEXT NOT NULL, eliminado INTEGER DEFAULT 0
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS inventario_adjuntos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, activo_id INTEGER, url TEXT NOT NULL, nombre_original TEXT NOT NULL, subido_por TEXT NOT NULL, fecha TEXT NOT NULL, FOREIGN KEY(activo_id) REFERENCES activos_inventario(id) ON DELETE CASCADE
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS inventario_devoluciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT, activo_id INTEGER NOT NULL, colaborador TEXT NOT NULL, confirmado_por TEXT NOT NULL, fecha TEXT NOT NULL, observaciones TEXT, FOREIGN KEY(activo_id) REFERENCES activos_inventario(id) ON DELETE CASCADE
            )''')
            # 🏥 Acta de recibido para activos biomédicos entregados a domicilio. Ver comentario
            # equivalente en la rama de Postgres.
            cursor.execute('''CREATE TABLE IF NOT EXISTS actas_recibido_biomedico (
                id INTEGER PRIMARY KEY AUTOINCREMENT, activo_id INTEGER NOT NULL, nombre_responsable TEXT NOT NULL, relacion_responsable TEXT NOT NULL, documento_responsable TEXT, telefono_contacto TEXT, direccion_entrega TEXT NOT NULL, firma_url TEXT NOT NULL, fecha TEXT NOT NULL, creado_por TEXT NOT NULL, FOREIGN KEY(activo_id) REFERENCES activos_inventario(id) ON DELETE CASCADE
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS login_fondo_media (
                id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT NOT NULL, url TEXT NOT NULL, public_id TEXT, orden INTEGER DEFAULT 0, estado TEXT DEFAULT 'activo', fecha_creacion TEXT NOT NULL, creado_por TEXT NOT NULL
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS tipos_activo_catalogo (
                id INTEGER PRIMARY KEY AUTOINCREMENT, key TEXT NOT NULL, etiqueta TEXT NOT NULL, icono TEXT DEFAULT 'box', orden INTEGER DEFAULT 0, estado TEXT DEFAULT 'activo'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS activos_reemplazos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, activo_anterior_id INTEGER NOT NULL, activo_nuevo_id INTEGER, motivo TEXT NOT NULL, notas TEXT, fecha_reemplazo TEXT NOT NULL, estado_anterior_resultante TEXT NOT NULL, creado_por TEXT NOT NULL, fecha_creacion TEXT NOT NULL, FOREIGN KEY(activo_anterior_id) REFERENCES activos_inventario(id) ON DELETE CASCADE, FOREIGN KEY(activo_nuevo_id) REFERENCES activos_inventario(id)
            )''')
            # 👁️ Registro de qué usuario ya leyó cada comunicado (se marca al ver el muro de
            # comunicados o el comunicado fijado en la bienvenida) — para saber quién falta.
            cursor.execute('''CREATE TABLE IF NOT EXISTS comunicados_leidos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, comunicado_id INTEGER NOT NULL, usuario TEXT NOT NULL, fecha TEXT NOT NULL, FOREIGN KEY(comunicado_id) REFERENCES comunicados(id) ON DELETE CASCADE, UNIQUE(comunicado_id, usuario)
            )''')
            # 🔔 Notificaciones internas (campanita): se generan en los mismos puntos donde ya
            # sale un correo (ticket creado, comentado, cambio de estado).
            cursor.execute('''CREATE TABLE IF NOT EXISTS notificaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT, usuario TEXT NOT NULL, tipo TEXT DEFAULT 'ticket', mensaje TEXT NOT NULL, url TEXT DEFAULT '', leida INTEGER DEFAULT 0, estado TEXT DEFAULT 'activa', fecha TEXT NOT NULL
            )''')
            # 🆔 Catálogo de aplicativos/herramientas para los que se crean credenciales a los
            # colaboradores (KUBAPP, SAMI, Moodle, Wolkvox, Correo, Solvyx...). Administrable:
            # el equipo de soporte puede agregar o desactivar aplicativos sin tocar código.
            cursor.execute('''CREATE TABLE IF NOT EXISTS aplicativos_catalogo (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, estado TEXT DEFAULT 'activo'
            )''')
            # 🩺 Catálogo de especialidades/áreas para los usuarios de Arkiv. Ver comentario
            # equivalente en la rama de Postgres.
            cursor.execute('''CREATE TABLE IF NOT EXISTS especialidades_catalogo (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, estado TEXT DEFAULT 'activo'
            )''')
            # 🪪 Altas y bajas de credenciales de colaboradores: un registro por cada aplicativo
            # que se le habilita a una persona (no por colaborador), para poder deshabilitar el
            # acceso a un aplicativo puntual sin afectar los demás que tenga esa misma persona.
            cursor.execute('''CREATE TABLE IF NOT EXISTS credenciales_colaboradores (
                id INTEGER PRIMARY KEY AUTOINCREMENT, colaborador TEXT NOT NULL, aplicativo TEXT NOT NULL, password_cifrada TEXT NOT NULL, fecha_creacion TEXT, fecha_solicitud TEXT, analista_gestiona TEXT, solicitado_por TEXT, capacitado_por TEXT, medio_envio TEXT, estado TEXT DEFAULT 'activo', fecha_deshabilitacion TEXT, deshabilitado_por TEXT, fecha_registro TEXT NOT NULL, registrado_por TEXT NOT NULL
            )''')
            # 🔐 Códigos de respaldo (recuperación) del 2FA. Ver comentario equivalente en la
            # rama de Postgres.
            cursor.execute('''CREATE TABLE IF NOT EXISTS totp_codigos_respaldo (
                id INTEGER PRIMARY KEY AUTOINCREMENT, usuario TEXT NOT NULL, codigo_hash TEXT NOT NULL, usado INTEGER DEFAULT 0, fecha_creacion TEXT NOT NULL, fecha_uso TEXT
            )''')
            # 📋 Plantillas de solicitud. Ver comentario equivalente en la rama de Postgres.
            cursor.execute('''CREATE TABLE IF NOT EXISTS ticket_plantillas (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, tipo TEXT DEFAULT 'Incidente', categoria TEXT, prioridad TEXT DEFAULT 'Media', area TEXT, sede TEXT, titulo TEXT NOT NULL, descripcion TEXT NOT NULL, estado TEXT DEFAULT 'activo', creado_por TEXT NOT NULL, fecha_creacion TEXT NOT NULL
            )''')
            # 📅 Vencimiento de documentos por empleado. Ver comentario equivalente en la rama
            # de Postgres.
            cursor.execute('''CREATE TABLE IF NOT EXISTS documentos_empleado (
                id INTEGER PRIMARY KEY AUTOINCREMENT, usuario TEXT NOT NULL, tipo_documento TEXT NOT NULL, titulo TEXT NOT NULL, descripcion TEXT, url_archivo TEXT DEFAULT '', nombre_original TEXT DEFAULT '', fecha_emision TEXT, fecha_vencimiento TEXT, alerta_nivel TEXT, estado TEXT DEFAULT 'activo', creado_por TEXT NOT NULL, fecha_creacion TEXT NOT NULL
            )''')

            for col_sql in ["categoria", "tipo", "tags", "vistas", "descargas", "estado"]:
                try:
                    cursor.execute(f"ALTER TABLE galerias ADD COLUMN {col_sql} TEXT DEFAULT 'activo';")
                    conn.commit()
                except Exception:
                    pass
            try:
                cursor.execute("ALTER TABLE archivos ADD COLUMN estado TEXT DEFAULT 'activo';")
                conn.commit()
            except Exception:
                pass
            try:
                cursor.execute("ALTER TABLE credenciales ADD COLUMN estado TEXT DEFAULT 'activo';")
                conn.commit()
            except Exception:
                pass
            for col_credencial_sql in [
                "ALTER TABLE logs ADD COLUMN credencial_id INTEGER;",
                "ALTER TABLE credenciales ADD COLUMN rotacion_dias INTEGER;",
                "ALTER TABLE credenciales ADD COLUMN fecha_ultima_rotacion TEXT;",
                "ALTER TABLE credenciales ADD COLUMN rotacion_recordatorio_fecha TEXT;",
                "ALTER TABLE credenciales ADD COLUMN etiquetas TEXT;",
                "ALTER TABLE credenciales ADD COLUMN tipo_item TEXT DEFAULT 'credencial';",
                "ALTER TABLE credenciales ADD COLUMN contenido_seguro TEXT;",
                "ALTER TABLE credenciales ADD COLUMN campos_personalizados TEXT;",
                "ALTER TABLE credenciales ADD COLUMN propietario TEXT;",
                "ALTER TABLE credenciales ADD COLUMN visibilidad TEXT DEFAULT 'equipo';"
            ]:
                try:
                    cursor.execute(col_credencial_sql)
                    conn.commit()
                except Exception:
                    pass
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN estado TEXT DEFAULT 'activo';")
                conn.commit()
            except Exception:
                pass
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN nombre TEXT;")
                conn.commit()
            except Exception:
                pass
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN telefono TEXT;")
                conn.commit()
            except Exception:
                pass
            try:
                # 🎨 Preferencia de tema (claro/oscuro) de cada usuario: se guarda en su cuenta
                # (no solo en el navegador) para que lo siga a donde inicie sesión.
                cursor.execute("ALTER TABLE usuarios ADD COLUMN tema TEXT DEFAULT 'oscuro';")
                conn.commit()
            except Exception:
                pass
            try:
                # 🪪 Cédula/documento de identidad: opcional, no única. Ver comentario equivalente
                # en la rama de Postgres.
                cursor.execute("ALTER TABLE usuarios ADD COLUMN cedula TEXT;")
                conn.commit()
            except Exception:
                pass
            try:
                # 🩺 Especialidad/área del usuario. Ver comentario equivalente en la rama de Postgres.
                cursor.execute("ALTER TABLE usuarios ADD COLUMN especialidad TEXT;")
                conn.commit()
            except Exception:
                pass
            try:
                # 🔒 Forzar cambio de contraseña. Ver comentario equivalente en la rama de Postgres.
                cursor.execute("ALTER TABLE usuarios ADD COLUMN debe_cambiar_password INTEGER DEFAULT 0;")
                conn.commit()
            except Exception:
                pass
            for col_ticket_sql in [
                "ALTER TABLE tickets ADD COLUMN tipo TEXT DEFAULT 'Incidente';",
                "ALTER TABLE tickets ADD COLUMN sla_respuesta_limite TEXT;",
                "ALTER TABLE tickets ADD COLUMN sla_resolucion_limite TEXT;",
                "ALTER TABLE tickets ADD COLUMN sla_respuesta_cumplida TEXT;",
                "ALTER TABLE tickets ADD COLUMN sla_resolucion_cumplida TEXT;",
                "ALTER TABLE tickets ADD COLUMN sla_modificaciones INTEGER DEFAULT 0;",
                "ALTER TABLE tickets ADD COLUMN calificacion INTEGER;",
                "ALTER TABLE tickets ADD COLUMN calificacion_fecha TEXT;",
                "ALTER TABLE tickets ADD COLUMN area TEXT;",
                "ALTER TABLE tickets ADD COLUMN sede TEXT;",
                "ALTER TABLE tickets ADD COLUMN telefono_contacto TEXT;",
                "ALTER TABLE tickets ADD COLUMN sla_alerta_nivel TEXT;",
                "ALTER TABLE tickets ADD COLUMN eliminado INTEGER DEFAULT 0;",
                "ALTER TABLE tickets ADD COLUMN activo_id INTEGER REFERENCES activos_inventario(id);",
                "ALTER TABLE tickets ADD COLUMN sla_pausado_desde TEXT;",
                "ALTER TABLE tickets ADD COLUMN ticket_relacionado_id INTEGER REFERENCES tickets(id);",
                "ALTER TABLE tickets ADD COLUMN solicitante_real TEXT;"
            ]:
                try:
                    cursor.execute(col_ticket_sql)
                    conn.commit()
                except Exception:
                    pass
            for col_comunicado_sql in [
                "ALTER TABLE comunicados ADD COLUMN recordatorio_enviado_fecha TEXT;"
            ]:
                try:
                    cursor.execute(col_comunicado_sql)
                    conn.commit()
                except Exception:
                    pass
            try:
                cursor.execute("ALTER TABLE galerias ADD COLUMN visibilidad TEXT DEFAULT 'todos';")
                conn.commit()
            except Exception:
                pass
            for col_config_sql in [
                "ALTER TABLE ticket_configuraciones ADD COLUMN direccion TEXT;",
                "ALTER TABLE ticket_configuraciones ADD COLUMN responsable TEXT;"
            ]:
                try:
                    cursor.execute(col_config_sql)
                    conn.commit()
                except Exception:
                    pass
            # 🕵️ Historial de sesiones: IP y dispositivo detectados de cada inicio/cierre de
            # sesión. Ver comentario equivalente en la rama de Postgres.
            for col_log_sql in [
                "ALTER TABLE logs ADD COLUMN ip TEXT;",
                "ALTER TABLE logs ADD COLUMN dispositivo TEXT;"
            ]:
                try:
                    cursor.execute(col_log_sql)
                    conn.commit()
                except Exception:
                    pass
            # 🔐 Verificación en dos pasos (2FA/TOTP). Ver comentario equivalente en la rama de
            # Postgres.
            for col_totp_sql in [
                "ALTER TABLE usuarios ADD COLUMN totp_secret TEXT;",
                "ALTER TABLE usuarios ADD COLUMN totp_habilitado INTEGER DEFAULT 0;",
                "ALTER TABLE usuarios ADD COLUMN foto_perfil TEXT;"
            ]:
                try:
                    cursor.execute(col_totp_sql)
                    conn.commit()
                except Exception:
                    pass
            # 📅 Vencimiento de documentos institucionales. Ver comentario equivalente en la
            # rama de Postgres.
            for col_venc_sql in [
                "ALTER TABLE galerias ADD COLUMN fecha_vencimiento TEXT;",
                "ALTER TABLE galerias ADD COLUMN alerta_vencimiento_nivel TEXT;"
            ]:
                try:
                    cursor.execute(col_venc_sql)
                    conn.commit()
                except Exception:
                    pass
            # 🔒 Bloqueo automático por intentos fallidos de inicio de sesión. Ver comentario
            # equivalente en la rama de Postgres.
            for col_bloqueo_sql in [
                "ALTER TABLE usuarios ADD COLUMN intentos_fallidos_login INTEGER DEFAULT 0;",
                "ALTER TABLE usuarios ADD COLUMN bloqueado_por_intentos INTEGER DEFAULT 0;"
            ]:
                try:
                    cursor.execute(col_bloqueo_sql)
                    conn.commit()
                except Exception:
                    pass
            # 🗑️ Papelera de notificaciones. Ver comentario equivalente en la rama de Postgres.
            for col_notif_estado_sql in [
                "ALTER TABLE notificaciones ADD COLUMN estado TEXT DEFAULT 'activa';"
            ]:
                try:
                    cursor.execute(col_notif_estado_sql)
                    conn.commit()
                except Exception:
                    pass
            # 🏢 Área del activo. Ver comentario equivalente en la rama de Postgres.
            for col_area_activo_sql in [
                "ALTER TABLE activos_inventario ADD COLUMN area TEXT;"
            ]:
                try:
                    cursor.execute(col_area_activo_sql)
                    conn.commit()
                except Exception:
                    pass
            # 🚚 Catálogo de Proveedores y proveedor del activo. Ver comentario equivalente en
            # la rama de Postgres.
            for col_proveedor_sql in [
                "ALTER TABLE ticket_configuraciones ADD COLUMN nit TEXT;",
                "ALTER TABLE ticket_configuraciones ADD COLUMN razon_social TEXT;",
                "ALTER TABLE activos_inventario ADD COLUMN proveedor TEXT;"
            ]:
                try:
                    cursor.execute(col_proveedor_sql)
                    conn.commit()
                except Exception:
                    pass
            # 💰 Costos de Inventario. Ver comentario equivalente en la rama de Postgres.
            for col_costo_sql in [
                "ALTER TABLE activos_inventario ADD COLUMN tipo_costo TEXT;",
                "ALTER TABLE activos_inventario ADD COLUMN costo_compra REAL;",
                "ALTER TABLE activos_inventario ADD COLUMN costo_alquiler_mensual REAL;"
            ]:
                try:
                    cursor.execute(col_costo_sql)
                    conn.commit()
                except Exception:
                    pass
            # ✍️ Firma digital. Ver comentario equivalente en la rama de Postgres.
            for col_firma_sql in [
                "ALTER TABLE usuarios ADD COLUMN firma TEXT;",
                "ALTER TABLE activos_inventario ADD COLUMN firma_asignacion_url TEXT;",
                "ALTER TABLE activos_inventario ADD COLUMN firma_asignacion_fecha TEXT;"
            ]:
                try:
                    cursor.execute(col_firma_sql)
                    conn.commit()
                except Exception:
                    pass
            # 🏥 Activos biomédicos entregados a domicilio. Ver comentario equivalente en la
            # rama de Postgres.
            for col_biomedico_sql in [
                "ALTER TABLE activos_inventario ADD COLUMN es_biomedico INTEGER DEFAULT 0;"
            ]:
                try:
                    cursor.execute(col_biomedico_sql)
                    conn.commit()
                except Exception:
                    pass

        # 📇 ÍNDICES — hasta ahora la única tabla con un índice real era 'usuarios' (por su
        # UNIQUE en 'usuario'); todo lo demás dependía de recorrer la tabla entera en cada
        # consulta. Con pocos cientos de filas eso no se nota, pero 'logs', 'tickets' y
        # 'notificaciones' crecen sin parar con el uso diario, y ya hay consultas que filtran
        # justo por estas columnas (ver ver_logs, ver_tickets, la campanita). La sintaxis
        # CREATE INDEX IF NOT EXISTS es idéntica en Postgres y sqlite, así que esta única
        # lista sirve para ambos motores — por eso vive fuera del if/else de arriba. Cada
        # sentencia va en su propio try/except: si alguna tabla no existe todavía (instalación
        # nueva a mitad de arranque) o el índice ya está creado, simplemente se ignora y se
        # sigue con las demás.
        for indice_sql in [
            "CREATE INDEX IF NOT EXISTS idx_logs_usuario ON logs (usuario);",
            "CREATE INDEX IF NOT EXISTS idx_logs_fecha ON logs (fecha);",
            "CREATE INDEX IF NOT EXISTS idx_logs_accion ON logs (accion);",
            "CREATE INDEX IF NOT EXISTS idx_logs_credencial_id ON logs (credencial_id);",
            "CREATE INDEX IF NOT EXISTS idx_correos_log_fecha ON correos_log (fecha);",
            "CREATE INDEX IF NOT EXISTS idx_credenciales_area ON credenciales (area);",
            "CREATE INDEX IF NOT EXISTS idx_credenciales_estado ON credenciales (estado);",
            "CREATE INDEX IF NOT EXISTS idx_credenciales_propietario ON credenciales (propietario);",
            "CREATE INDEX IF NOT EXISTS idx_tickets_estado ON tickets (estado);",
            "CREATE INDEX IF NOT EXISTS idx_tickets_creado_por ON tickets (creado_por);",
            "CREATE INDEX IF NOT EXISTS idx_tickets_asignado_a ON tickets (asignado_a);",
            "CREATE INDEX IF NOT EXISTS idx_tickets_categoria ON tickets (categoria);",
            "CREATE INDEX IF NOT EXISTS idx_tickets_comentarios_ticket_id ON tickets_comentarios (ticket_id);",
            "CREATE INDEX IF NOT EXISTS idx_tickets_adjuntos_ticket_id ON tickets_adjuntos (ticket_id);",
            "CREATE INDEX IF NOT EXISTS idx_tickets_tareas_ticket_id ON tickets_tareas (ticket_id);",
            # 🔎 Para la cola "Mis Tareas" (WHERE responsable = ? sobre TODA la tabla, sin
            # filtrar primero por ticket_id como sí hace el detalle del ticket).
            "CREATE INDEX IF NOT EXISTS idx_tickets_tareas_responsable ON tickets_tareas (responsable);",
            "CREATE INDEX IF NOT EXISTS idx_activos_inventario_estado ON activos_inventario (estado);",
            "CREATE INDEX IF NOT EXISTS idx_activos_inventario_asignado_a ON activos_inventario (asignado_a);",
            "CREATE INDEX IF NOT EXISTS idx_notificaciones_usuario_estado ON notificaciones (usuario, estado);",
            "CREATE INDEX IF NOT EXISTS idx_notificaciones_usuario_leida ON notificaciones (usuario, leida);",
            "CREATE INDEX IF NOT EXISTS idx_credenciales_colaboradores_colaborador ON credenciales_colaboradores (colaborador);",
            "CREATE INDEX IF NOT EXISTS idx_credenciales_colaboradores_estado ON credenciales_colaboradores (estado);",
            "CREATE INDEX IF NOT EXISTS idx_galerias_vistas_galeria_id ON galerias_vistas (galeria_id);",
            "CREATE INDEX IF NOT EXISTS idx_inventario_devoluciones_activo_id ON inventario_devoluciones (activo_id);",
            "CREATE INDEX IF NOT EXISTS idx_login_fondo_media_estado ON login_fondo_media (estado);",
            "CREATE INDEX IF NOT EXISTS idx_documentos_empleado_usuario ON documentos_empleado (usuario);",
            "CREATE INDEX IF NOT EXISTS idx_documentos_empleado_estado ON documentos_empleado (estado);",
            "CREATE INDEX IF NOT EXISTS idx_totp_codigos_respaldo_usuario ON totp_codigos_respaldo (usuario);",
            "CREATE INDEX IF NOT EXISTS idx_comunicados_estado ON comunicados (estado);",
            "CREATE INDEX IF NOT EXISTS idx_conocimiento_articulos_estado ON conocimiento_articulos (estado);",
        ]:
            try:
                cursor.execute(indice_sql)
                conn.commit()
            except Exception:
                conn.rollback()

        # Siembra las categorías de ticket por defecto la primera vez (instalación nueva o
        # actualización desde una versión sin `ticket_configuraciones`), para que la lista no
        # aparezca vacía. Áreas y Sedes NO se siembran: son específicas de la organización y
        # las define el equipo de soporte desde /tickets/configuracion.
        q_count_cat = "SELECT COUNT(*) FROM ticket_configuraciones WHERE tipo = %s" if db_type == 'postgres' else "SELECT COUNT(*) FROM ticket_configuraciones WHERE tipo = ?"
        cursor.execute(q_count_cat, ('categoria',))
        if cursor.fetchone()[0] == 0:
            q_seed_cat = "INSERT INTO ticket_configuraciones (tipo, nombre) VALUES (%s, %s)" if db_type == 'postgres' else "INSERT INTO ticket_configuraciones (tipo, nombre) VALUES (?, ?)"
            for nombre_cat in ['Hardware', 'Software', 'Acceso/Credenciales', 'Red/Internet', 'Otro']:
                cursor.execute(q_seed_cat, ('categoria', nombre_cat))
            conn.commit()

        # 🆔 Siembra el catálogo de aplicativos con los que ya se manejan hoy, solo la primera
        # vez (instalación nueva o actualización desde una versión sin esta tabla). Después de
        # esto, el catálogo se administra por completo desde el módulo (agregar/desactivar).
        cursor.execute("SELECT COUNT(*) FROM aplicativos_catalogo")
        if cursor.fetchone()[0] == 0:
            q_seed_app = "INSERT INTO aplicativos_catalogo (nombre) VALUES (%s)" if db_type == 'postgres' else "INSERT INTO aplicativos_catalogo (nombre) VALUES (?)"
            for nombre_app in ['KUBAPP', 'SAMI', 'Moodle', 'Wolkvox', 'Correo (cPanel / Outlook)', 'Solvyx']:
                cursor.execute(q_seed_app, (nombre_app,))
            conn.commit()

        # 🗂️ Siembra el catálogo de Tipos de activo con los mismos 8 tipos que antes vivían
        # fijos en el código (TIPOS_ACTIVO), solo la primera vez. De ahí en adelante se
        # administra por completo desde /tickets/inventario/tipos (agregar, reordenar,
        # desactivar) sin volver a tocar código, igual que aplicativos/especialidades.
        cursor.execute("SELECT COUNT(*) FROM tipos_activo_catalogo")
        if cursor.fetchone()[0] == 0:
            q_seed_tipo = "INSERT INTO tipos_activo_catalogo (key, etiqueta, icono, orden) VALUES (%s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO tipos_activo_catalogo (key, etiqueta, icono, orden) VALUES (?, ?, ?, ?)"
            tipos_semilla = [
                ('DESKTOP', 'Computador de Escritorio', 'desktop'),
                ('LAPTOP', 'Portátil', 'laptop'),
                ('PRINTER', 'Impresora', 'print'),
                ('MONITOR', 'Monitor', 'display'),
                ('PHONE', 'Teléfono/Celular', 'mobile-screen'),
                ('SERVER', 'Servidor', 'server'),
                ('NETWORK', 'Red (Switch/Router/AP)', 'network-wired'),
                ('OTHER', 'Otro', 'box'),
            ]
            for orden_i, (key_tipo, etiqueta_tipo, icono_tipo) in enumerate(tipos_semilla):
                cursor.execute(q_seed_tipo, (key_tipo, etiqueta_tipo, icono_tipo, orden_i))
            conn.commit()

        # 🩺 Sincroniza el catálogo de especialidades con la lista real de Preventiva. A
        # diferencia del seed de aplicativos (que solo corre si la tabla está vacía), esta
        # sincronización corre SIEMPRE en cada arranque, pero es idempotente: cada nombre se
        # cruza contra lo que ya exista (sin importar mayúsculas/acentos) y solo se agrega si
        # todavía no está — así no duplica lo que el equipo ya había cargado a mano desde el
        # módulo (p. ej. "AUX. ADMINISTRATIVO", "PSICOLOGIA", "MEDICINA GENERAL"...). Quedan
        # afuera las que ya cubre una especialidad existente en otra forma (p. ej. "Psicólogo"
        # ya lo cubre "Psicología", "Facturación" ya lo cubre "Aux. Facturación").
        especialidades_a_sincronizar = [
            'JEFE DE ENFERMERIA', 'INFECTOLOGO', 'TRABAJADORA SOCIAL', 'BACTERIOLOGA ADMINISTRATIVA',
            'ESPECIALISTA EN INFECTOLOGIA PEDIATRICA', 'AUXILIAR CONTABLE', 'INTERNISTA E INFECTOLOGO',
            'TERAPIA RESPIRATORIA', 'FONOAUDIOLOGA', 'FISIOTERAPEUTA', 'AUXILIAR EN RADIOLOGIA ORAL',
            'MEDICINA DOMICILIARIA', 'NEUROPSICOLOGIA', 'NUTRICIONISTA'
        ]
        q_check_esp = "SELECT id FROM especialidades_catalogo WHERE UPPER(nombre) = UPPER(%s)" if db_type == 'postgres' else "SELECT id FROM especialidades_catalogo WHERE UPPER(nombre) = UPPER(?)"
        q_ins_esp = "INSERT INTO especialidades_catalogo (nombre) VALUES (%s)" if db_type == 'postgres' else "INSERT INTO especialidades_catalogo (nombre) VALUES (?)"
        for nombre_esp in especialidades_a_sincronizar:
            cursor.execute(q_check_esp, (nombre_esp,))
            if not cursor.fetchone():
                cursor.execute(q_ins_esp, (nombre_esp,))
        conn.commit()

        cursor.execute("SELECT COUNT(*) FROM usuarios")
        if cursor.fetchone()[0] == 0:
            # Solo se ejecuta si la tabla usuarios está realmente vacía (instalación nueva).
            # La contraseña inicial se define por variable de entorno; si no está seteada,
            # se genera una aleatoria y se imprime UNA vez en los logs de Render para que la copies.
            pass_inicial = os.environ.get('ADMIN_PASSWORD_INICIAL')
            if not pass_inicial:
                pass_inicial = base64.urlsafe_b64encode(os.urandom(9)).decode('utf-8')
                print(f"🔑 Usuario admin creado. Contraseña inicial generada (cámbiala tras iniciar sesión): {pass_inicial}")
            query_admin = "INSERT INTO usuarios (usuario, password_hash, correo, rol) VALUES (%s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO usuarios (usuario, password_hash, correo, rol) VALUES (?, ?, ?, ?)"
            cursor.execute(query_admin, ('admin', generate_password_hash(pass_inicial), 'notificacionesarkiv@gmail.com', 'admin'))

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error inicializando base de datos: {e}")

init_db()

def registrar_log(usuario, accion, detalles="", credencial_id=None, ip=None, dispositivo=None):
    """credencial_id es opcional: solo se usa para marcar qué entradas del log son consultas a
    la bóveda de Credenciales (ver _revelar_credencial), y así poder filtrar/armar la auditoría
    de accesos por credencial sin tener que parsear el texto libre de 'detalles'.
    ip/dispositivo son opcionales: solo se pasan desde login()/logout() para alimentar
    /perfil/historial-sesiones (ver _obtener_ip_cliente/_detectar_dispositivo más abajo); el
    resto de acciones del log los deja en NULL, que es justo lo que se espera de ellas."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        fecha_actual = obtener_fecha_actual()
        query = "INSERT INTO logs (usuario, accion, detalles, fecha, credencial_id, ip, dispositivo) VALUES (%s, %s, %s, %s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO logs (usuario, accion, detalles, fecha, credencial_id, ip, dispositivo) VALUES (?, ?, ?, ?, ?, ?, ?)"
        cursor.execute(query, (usuario, accion, detalles, fecha_actual, credencial_id, ip, dispositivo))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error registrando log: {e}")

# 🕵️ HISTORIAL DE SESIONES — helpers usados por login()/logout() para registrar desde dónde se
# conectó cada cuenta (ver /perfil/historial-sesiones más abajo). No son librerías de
# fingerprinting: solo reconocen los patrones más comunes de User-Agent para mostrar algo
# legible ("Chrome en Windows") en vez del header crudo.
def _obtener_ip_cliente():
    """Preferimos X-Forwarded-For (la IP real del visitante, que agrega el proxy de Render)
    porque request.remote_addr, detrás de ese proxy, siempre muestra la IP interna del balanceador."""
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr or ''

def _detectar_dispositivo(user_agent):
    if not user_agent:
        return "Dispositivo desconocido"

    if re.search(r'Windows', user_agent):
        so = "Windows"
    elif re.search(r'iPhone', user_agent):
        so = "iPhone"
    elif re.search(r'iPad', user_agent):
        so = "iPad"
    elif re.search(r'Android', user_agent):
        so = "Android"
    elif re.search(r'Mac OS X', user_agent):
        so = "Mac"
    elif re.search(r'Linux', user_agent):
        so = "Linux"
    else:
        so = None

    if re.search(r'Edg/', user_agent):
        navegador = "Edge"
    elif re.search(r'OPR/|Opera', user_agent):
        navegador = "Opera"
    elif re.search(r'Chrome/', user_agent) and 'Chromium' not in user_agent:
        navegador = "Chrome"
    elif re.search(r'Firefox/', user_agent):
        navegador = "Firefox"
    elif re.search(r'Safari/', user_agent) and 'Chrome/' not in user_agent:
        navegador = "Safari"
    else:
        navegador = None

    if navegador and so:
        return f"{navegador} en {so}"
    if navegador:
        return navegador
    if so:
        return so
    return user_agent[:60]

# 🔐 VERIFICACIÓN EN DOS PASOS (2FA/TOTP) — helpers usados por /perfil/2fa*, login() y
# /login/2fa. El secreto TOTP y los códigos de respaldo (de un solo uso) permiten iniciar
# sesión incluso si la persona pierde acceso a su app autenticadora (Google Authenticator,
# Authy, etc.) mientras aún conserva alguno de sus 10 códigos de respaldo.
NOMBRE_EMISOR_2FA = "Arkiv - Preventiva"


def _normalizar_codigo_respaldo(codigo):
    """Los códigos de respaldo se muestran como 'XXXX-XXXX' pero se aceptan con o sin guion,
    en mayúsculas o minúsculas: se normalizan antes de hashear/comparar para que ambos formatos
    funcionen igual."""
    return re.sub(r'[^A-Z0-9]', '', (codigo or '').upper())


def _generar_codigos_respaldo_2fa(usuario):
    """Genera 10 códigos de respaldo nuevos para 'usuario', reemplazando cualquier lote previo
    (al activar el 2FA por primera vez o al regenerarlos manualmente). Devuelve la lista de
    códigos EN CLARO — es la única vez que existen en texto plano; en la base de datos solo se
    guarda su hash, igual que las contraseñas."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_del = "DELETE FROM totp_codigos_respaldo WHERE usuario = %s" if db_type == 'postgres' else "DELETE FROM totp_codigos_respaldo WHERE usuario = ?"
        cursor.execute(q_del, (usuario,))

        codigos_planos = []
        fecha_actual = obtener_fecha_actual()
        q_ins = "INSERT INTO totp_codigos_respaldo (usuario, codigo_hash, usado, fecha_creacion) VALUES (%s, %s, 0, %s)" if db_type == 'postgres' else "INSERT INTO totp_codigos_respaldo (usuario, codigo_hash, usado, fecha_creacion) VALUES (?, ?, 0, ?)"
        for _ in range(10):
            crudo = secrets.token_hex(4).upper()  # 8 caracteres hexadecimales
            codigo_formateado = f"{crudo[:4]}-{crudo[4:]}"
            codigos_planos.append(codigo_formateado)
            cursor.execute(q_ins, (usuario, generate_password_hash(crudo), fecha_actual))

        conn.commit()
        conn.close()
        return codigos_planos
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"⚠️ Error generando códigos de respaldo 2FA: {e}")
        return []


def _verificar_codigo_respaldo_2fa(usuario, codigo_ingresado):
    """Verifica un código de respaldo de un solo uso. Si es válido lo marca como usado (no se
    puede reutilizar) y devuelve True; en cualquier otro caso devuelve False."""
    codigo_normalizado = _normalizar_codigo_respaldo(codigo_ingresado)
    if not codigo_normalizado:
        return False

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT id, codigo_hash FROM totp_codigos_respaldo WHERE usuario = %s AND usado = 0" if db_type == 'postgres' else "SELECT id, codigo_hash FROM totp_codigos_respaldo WHERE usuario = ? AND usado = 0"
        cursor.execute(q_sel, (usuario,))
        filas = cursor.fetchall()

        for codigo_id, codigo_hash in filas:
            if check_password_hash(codigo_hash, codigo_normalizado):
                fecha_actual = obtener_fecha_actual()
                q_upd = "UPDATE totp_codigos_respaldo SET usado = 1, fecha_uso = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE totp_codigos_respaldo SET usado = 1, fecha_uso = ? WHERE id = ?"
                cursor.execute(q_upd, (fecha_actual, codigo_id))
                conn.commit()
                conn.close()
                return True

        conn.close()
        return False
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"⚠️ Error verificando código de respaldo 2FA: {e}")
        return False


# 🔒 Hallazgo QA H-07 (30/08/2026): el secreto TOTP (totp_secret) se guardaba en texto plano en
# la tabla usuarios — cualquiera con acceso a la base de datos (o al Gestor de BD, /admin/db)
# podía leerlo y clonar la verificación en dos pasos de cualquier cuenta. Ahora se cifra con
# Fernet, igual que las contraseñas guardadas en la bóveda (ver encriptar_texto/desencriptar_texto
# más arriba). _migrar_totp_secret_a_fernet migra en el momento cualquier secreto que todavía
# esté en texto plano de antes de este cambio, la primera vez que se usa para iniciar sesión.
def _migrar_totp_secret_a_fernet(usuario, secreto_plano):
    """Re-guarda cifrado con Fernet un totp_secret que todavía estaba en texto plano."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q_upd = "UPDATE usuarios SET totp_secret = %s WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET totp_secret = ? WHERE usuario = ?"
        cursor.execute(q_upd, (encriptar_texto(secreto_plano), usuario))
        conn.commit()
        conn.close()
        print(f"🔐 totp_secret de '{usuario}' migrado de texto plano a Fernet/AES.")
    except Exception as e:
        print(f"⚠️ Error migrando totp_secret de '{usuario}' a Fernet: {e}")


def _descifrar_totp_secret(usuario, secreto_guardado):
    """Descifra el totp_secret leído de la base de datos. Un secreto TOTP en base32 (el formato
    en que se genera y se guardaba antes de este cambio) nunca es un token Fernet válido, así
    que un fallo al descifrar se interpreta como 'todavía en texto plano' y se migra de una vez
    (en vez de fallar el intento de login)."""
    if not secreto_guardado:
        return None
    if not _fernet:
        return secreto_guardado
    try:
        return _fernet.decrypt(secreto_guardado.encode('utf-8')).decode('utf-8')
    except (InvalidToken, Exception):
        _migrar_totp_secret_a_fernet(usuario, secreto_guardado)
        return secreto_guardado


def _desactivar_2fa_cuenta(usuario):
    """Apaga el 2FA de una cuenta: limpia el secreto TOTP y borra sus códigos de respaldo. Lo
    usan tanto /perfil/2fa/desactivar (el propio usuario, con su contraseña) como el admin desde
    Gestión de Usuarios (recuperación por pérdida de dispositivo)."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_upd = "UPDATE usuarios SET totp_secret = NULL, totp_habilitado = FALSE WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET totp_secret = NULL, totp_habilitado = 0 WHERE usuario = ?"
        cursor.execute(q_upd, (usuario,))
        q_del = "DELETE FROM totp_codigos_respaldo WHERE usuario = %s" if db_type == 'postgres' else "DELETE FROM totp_codigos_respaldo WHERE usuario = ?"
        cursor.execute(q_del, (usuario,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"⚠️ Error desactivando 2FA: {e}")
        return False

# 📧 Bitácora de correos enviados (ver /logs/correos). Se llama desde enviar_correo_ticket y
# enviar_correo_recuperacion, casi siempre dentro de un hilo aparte (threading.Thread), por eso
# abre su propia conexión en vez de reutilizar una del hilo principal.
# 🛡️ NUNCA recibe ni guarda el cuerpo del correo ni el código de verificación — solo asunto,
# destinatario, tipo y estado — para que este log se pueda mostrar a agentes/admins sin
# exponer datos sensibles de recuperación de contraseña.
def registrar_correo_log(destinatario, asunto, tipo, estado, detalle_error=None):
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        fecha_actual = obtener_fecha_actual()
        query = "INSERT INTO correos_log (fecha, destinatario, asunto, tipo, estado, detalle_error) VALUES (%s, %s, %s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO correos_log (fecha, destinatario, asunto, tipo, estado, detalle_error) VALUES (?, ?, ?, ?, ?, ?)"
        cursor.execute(query, (fecha_actual, destinatario, asunto, tipo, estado, detalle_error))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error registrando log de correo: {e}")

def _migrar_password_a_hash(usuario, password_plano):
    """Re-guarda con hash una contraseña que todavía estaba en texto plano."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        nuevo_hash = generate_password_hash(password_plano)
        query = "UPDATE usuarios SET password_hash = %s WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET password_hash = ? WHERE usuario = ?"
        cursor.execute(query, (nuevo_hash, usuario))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error migrando password a hash para {usuario}: {e}")

# 🔒 LONGITUD MÍNIMA DE CONTRASEÑA DE CUENTA (hallazgo QA H-02, 30/08/2026)
# Antes solo se exigía en /perfil/cambiar_password (autoservicio); un admin creando o editando
# una cuenta, o alguien recuperando su clave por correo, podía dejar contraseñas más cortas.
# Esta única constante y su helper se reutilizan en los 4 puntos donde se asigna una contraseña
# de cuenta, para que la política sea consistente sin importar quién la establezca.
LONGITUD_MINIMA_PASSWORD_CUENTA = 8


def _password_cuenta_valida(password_plano):
    """True si 'password_plano' cumple la longitud mínima exigida para contraseñas de cuenta
    (login, recuperación, alta/edición de usuario por un admin). No evalúa fortaleza ni la
    compara contra contraseñas comunes — eso es un análisis aparte, propio de la bóveda
    (ver _password_es_debil)."""
    return bool(password_plano) and len(password_plano) >= LONGITUD_MINIMA_PASSWORD_CUENTA


# 🔒 BLOQUEO AUTOMÁTICO POR INTENTOS FALLIDOS DE INICIO DE SESIÓN
# Es un candado aparte del 'estado' (activo/inactivo), que sigue siendo del resorte exclusivo
# de un admin (ver toggle_estado_usuario): este otro lo enciende el propio sistema de login,
# sin intervención humana, tras varios fallos de contraseña seguidos — y se apaga por
# cualquiera de dos caminos: un admin lo desbloquea manualmente desde Gestión de Usuarios
# (ver desbloquear_intentos_usuario), o la propia persona recupera el acceso cambiando su
# contraseña por el flujo de correo (ver validar_codigo). Cualquiera de los dos reinicia el
# contador para que fallos ya superados no se acumulen hacia un futuro bloqueo.
UMBRAL_INTENTOS_FALLIDOS_LOGIN = 3

def _registrar_intento_fallido_login(usuario):
    """Suma un intento fallido de contraseña para 'usuario'. Si con este llega (o ya venía)
    al umbral, deja/mantiene la cuenta bloqueada. Devuelve (bloqueada, intentos_restantes):
    'bloqueada' es True si la cuenta queda (o ya estaba) bloqueada por intentos;
    'intentos_restantes' es cuántos fallos más admite antes de bloquearse (0 si ya está
    bloqueada) — se usa para mostrar el contador de intentos en la pantalla de login."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q_sel = "SELECT COALESCE(intentos_fallidos_login, 0), COALESCE(bloqueado_por_intentos, FALSE) FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT COALESCE(intentos_fallidos_login, 0), COALESCE(bloqueado_por_intentos, 0) FROM usuarios WHERE usuario = ?"
        cursor.execute(q_sel, (usuario,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return False, UMBRAL_INTENTOS_FALLIDOS_LOGIN

        intentos_actuales = (row[0] or 0) + 1
        ya_estaba_bloqueada = bool(row[1])

        if ya_estaba_bloqueada:
            # Ya estaba bloqueada de antes: solo se deja constancia del intento nuevo, sin
            # volver a "bloquear" ni repetir el registro de auditoría del bloqueo.
            q_upd = "UPDATE usuarios SET intentos_fallidos_login = %s WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET intentos_fallidos_login = ? WHERE usuario = ?"
            cursor.execute(q_upd, (intentos_actuales, usuario))
            conn.commit()
            conn.close()
            return True, 0

        if intentos_actuales >= UMBRAL_INTENTOS_FALLIDOS_LOGIN:
            q_upd = "UPDATE usuarios SET intentos_fallidos_login = %s, bloqueado_por_intentos = TRUE WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET intentos_fallidos_login = ?, bloqueado_por_intentos = 1 WHERE usuario = ?"
            cursor.execute(q_upd, (intentos_actuales, usuario))
            conn.commit()
            conn.close()
            registrar_log(usuario, "Cuenta Bloqueada por Intentos", f"Se bloqueó la cuenta tras {intentos_actuales} intentos fallidos consecutivos de inicio de sesión.", ip=_obtener_ip_cliente(), dispositivo=_detectar_dispositivo(request.headers.get('User-Agent', '')))
            return True, 0

        q_upd = "UPDATE usuarios SET intentos_fallidos_login = %s WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET intentos_fallidos_login = ? WHERE usuario = ?"
        cursor.execute(q_upd, (intentos_actuales, usuario))
        conn.commit()
        conn.close()
        return False, (UMBRAL_INTENTOS_FALLIDOS_LOGIN - intentos_actuales)
    except Exception as e:
        print(f"⚠️ Error registrando intento fallido de login para {usuario}: {e}")
        return False, UMBRAL_INTENTOS_FALLIDOS_LOGIN

def _resetear_intentos_fallidos_login(usuario):
    """Limpia el contador de intentos fallidos (sin tocar 'bloqueado_por_intentos': ese solo
    lo apaga un admin o el flujo de recuperación de clave). Se llama tras una contraseña
    correcta, para que fallos viejos ya superados no se acumulen hacia un futuro bloqueo."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "UPDATE usuarios SET intentos_fallidos_login = 0 WHERE usuario = %s AND COALESCE(intentos_fallidos_login, 0) != 0" if db_type == 'postgres' else "UPDATE usuarios SET intentos_fallidos_login = 0 WHERE usuario = ? AND COALESCE(intentos_fallidos_login, 0) != 0"
        cursor.execute(q, (usuario,))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error reiniciando intentos fallidos de login para {usuario}: {e}")

def verificar_recaptcha(response_token):
    if not response_token: return False
    url = "https://www.google.com/recaptcha/api/siteverify"
    data = urllib.parse.urlencode({'secret': RECAPTCHA_SECRET_KEY, 'response': response_token}).encode('utf-8')
    try:
        req = urllib.request.Request(url, data=data)
        with urllib.request.urlopen(req, timeout=5) as response:
            return json.loads(response.read().decode('utf-8')).get('success', False)
    except Exception as e:
        return False

def archivo_permitido(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'): return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('rol') != 'admin': return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# 🛡️ Más estricto que admin_required: exige la cuenta LITERAL 'admin' (super-admin), no
# cualquier usuario con rol 'admin'. Usado en el Gestor de Base de Datos y los Respaldos —
# ambos dan acceso directo a los datos crudos de toda la organización (incluidas otras
# cuentas admin), así que quedan fuera del alcance de agentes y de admins comunes.
def superadmin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('username') != 'admin': return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# 🧑‍💼 Rol "Agente": analista de soporte TI que resuelve PQRS/Tickets. Tiene el mismo nivel
# de acceso operativo que Admin en (casi) todos los módulos — Gestor de Archivos, Comunicados,
# Tickets/Soporte TI completo, Bóveda de Accesos, Papelera y Auditoría — con excepciones que
# quedan exclusivas del rol 'admin' (Gestión de Usuarios, @admin_required) o, más estricto
# todavía, de la cuenta super-admin literal 'admin' (Gestor de Base de Datos y Respaldos,
# @superadmin_required — ni siquiera otros admins los ven).
ROLES_CON_ACCESO_OPERATIVO = ('admin', 'agente')

def agente_o_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('rol') not in ROLES_CON_ACCESO_OPERATIVO: return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# 🪪 Rol "Gestión Humana": rol nuevo y acotado, distinto de ROLES_CON_ACCESO_OPERATIVO
# (admin/agente) a propósito — NO hereda el resto de permisos de soporte TI (Tickets,
# Bóveda de Accesos, Papelera, Auditoría, etc.), solo puede certificar la devolución de
# activos del Inventario antes de que se pueda dar de baja la cuenta de un colaborador
# (ver certificacion_devoluciones / confirmar_devolucion_activo). Admin y Agente conservan
# también esta certificación, porque ya administran el Inventario por completo.
ROLES_CERTIFICACION_DEVOLUCION = ('admin', 'agente', 'gestion_humana')

def certificacion_devolucion_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('rol') not in ROLES_CERTIFICACION_DEVOLUCION: return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# 🗄️ MÓDULO ADMINISTRADOR DE BASE DE DATOS (LECTURA + CONSOLA SQL LIBRE)
@app.route('/admin/db', methods=['GET', 'POST'])
@login_required
@admin_required
@superadmin_required
def visor_db():
    tabla_seleccionada = request.args.get('tabla', 'usuarios')
    q_sql = request.form.get('sql', '').strip() or request.args.get('sql', '').strip()
    
    tablas_permitidas = [
        'usuarios', 'galerias', 'archivos', 'logs', 'credenciales', 'comunicados',
        'tickets', 'tickets_comentarios', 'tickets_adjuntos', 'conocimiento_articulos',
        'ticket_configuraciones', 'activos_inventario', 'inventario_devoluciones', 'aplicativos_catalogo',
        'credenciales_colaboradores', 'login_fondo_media'
    ]
    if tabla_seleccionada not in tablas_permitidas:
        tabla_seleccionada = 'usuarios'

    columnas = []
    registros = []
    mensaje_exito = None
    error_sql = None

    # 🛡️ Todo el módulo (ver tablas y ejecutar SQL libre) está restringido a la cuenta
    # super-admin ('admin') por @superadmin_required — ver ese decorador. Antes solo se
    # restringía aquí la ejecución de SQL personalizado y la vista de tablas quedaba
    # abierta a cualquier admin; ahora ni siquiera se llega a esta función sin ser
    # super-admin, así que ya no hace falta ese chequeo adicional.

    conn, db_type = get_db()
    cursor = conn.cursor()

    try:
        if q_sql:
            # 📝 Hallazgo QA H-04 (consola SQL libre, riesgo aceptado con control
            # compensatorio): antes solo quedaba rastro en la bitácora de las sentencias
            # que modificaban datos (INSERT/UPDATE/DELETE/etc.), y solo si terminaban
            # bien. Un SELECT manual —la forma más simple de exfiltrar datos por esta
            # consola sin dejar huella— no generaba ningún registro, y una sentencia que
            # fallara (por ejemplo, un intento de inyección mal formado) tampoco. Ahora
            # se registra CADA sentencia que se intenta ejecutar aquí, éxito o error.
            es_lectura = q_sql.lower().startswith('select')
            try:
                cursor.execute(q_sql)
                if es_lectura:
                    registros = cursor.fetchall()
                    if cursor.description:
                        columnas = [desc[0] for desc in cursor.description]
                    registrar_log(session['username'], "Consulta SQL Manual", f"SQL: {q_sql[:300]}")
                else:
                    conn.commit()
                    mensaje_exito = f"Sentencia SQL ejecutada con éxito. Filas afectadas: {cursor.rowcount}"
                    registrar_log(session['username'], "Ejecución SQL Manual", f"SQL: {q_sql[:300]}")
                    cursor.execute(f"SELECT * FROM {tabla_seleccionada} LIMIT 100")
                    registros = cursor.fetchall()
                    if cursor.description:
                        columnas = [desc[0] for desc in cursor.description]
            except Exception as e_sql:
                conn.rollback()
                registrar_log(session['username'], "Error en SQL Manual", f"SQL: {q_sql[:300]} | Error: {str(e_sql)[:200]}")
                raise
        else:
            cursor.execute(f"SELECT * FROM {tabla_seleccionada} LIMIT 100")
            registros = cursor.fetchall()
            if cursor.description:
                columnas = [desc[0] for desc in cursor.description]
    except Exception as e:
        conn.rollback()
        error_sql = str(e)
    finally:
        conn.close()

    return render_template(
        'admin_db.html',
        tabla=tabla_seleccionada,
        tablas=tablas_permitidas,
        columnas=columnas,
        registros=registros,
        sql=q_sql,
        exito=mensaje_exito,
        error=error_sql,
        es_superadmin=(session.get('username') == 'admin')
    )


# 💾 MÓDULO DE RESPALDOS DE BASE DE DATOS ------------------------------------------------
# Vuelca todas las tablas (datos y metadatos — los ARCHIVOS en sí ya viven aparte, en
# Cloudinary) a un archivo JSON. Dos vías: un botón manual "Generar y descargar ahora"
# (esta sección), y un hilo en segundo plano que genera uno automático cada día (ver
# _respaldo_diario_automatico más abajo). Ambos guardan el archivo en RESPALDOS_DIR, que
# debe apuntar a un disco PERSISTENTE de Render (Mount Path /var/data) — sin eso, cualquier
# archivo escrito en el propio servidor se pierde en el siguiente despliegue.
RESPALDOS_DIR = os.environ.get('RESPALDOS_DIR', '/var/data/respaldos')
RESPALDOS_RETENCION_DIAS = 30  # Antigüedad máxima de los respaldos AUTOMÁTICOS antes de borrarlos solos.

TABLAS_RESPALDO = [
    'usuarios', 'galerias', 'archivos', 'logs', 'credenciales', 'comunicados',
    'comunicados_leidos', 'notificaciones', 'tickets', 'tickets_comentarios',
    'tickets_adjuntos', 'conocimiento_articulos', 'ticket_configuraciones',
    'activos_inventario', 'inventario_adjuntos', 'inventario_devoluciones', 'aplicativos_catalogo',
    'credenciales_colaboradores', 'login_fondo_media', 'actas_recibido_biomedico'
]


def _valor_respaldo_serializable(v):
    """json.dumps no sabe convertir algunos tipos que psycopg2/sqlite3 entregan tal cual
    (datetime, Decimal) — esto los vuelve texto/número plano antes de serializar, sin tocar
    el resto."""
    if isinstance(v, (datetime,)):
        return v.isoformat()
    if isinstance(v, decimal.Decimal):
        return float(v)
    return v


def _generar_respaldo_datos():
    """Vuelca todas las tablas de TABLAS_RESPALDO a un diccionario serializable — el
    contenido real del archivo de respaldo. Si una tabla puntual falla (p. ej. todavía no
    existe en una instalación muy vieja), se omite y se sigue con las demás: un respaldo
    parcial es mejor que ninguno."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    respaldo = {
        'generado': datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d %H:%M:%S'),
        'motor': db_type,
        'tablas': {}
    }
    for tabla in TABLAS_RESPALDO:
        try:
            cursor.execute(f"SELECT * FROM {tabla}")
            columnas = [d[0] for d in cursor.description]
            filas = cursor.fetchall()
            respaldo['tablas'][tabla] = {
                'columnas': columnas,
                'filas': [[_valor_respaldo_serializable(v) for v in fila] for fila in filas]
            }
        except Exception as e:
            print(f"⚠️ Error respaldando tabla '{tabla}': {e}")
    conn.close()
    return respaldo


def _guardar_respaldo_en_disco(prefijo='manual'):
    """Genera el respaldo y lo guarda como JSON en RESPALDOS_DIR. Devuelve la ruta completa
    del archivo creado, o None si no se pudo escribir (p. ej. el disco persistente de Render
    todavía no está montado — la carpeta ni siquiera se puede crear)."""
    try:
        os.makedirs(RESPALDOS_DIR, exist_ok=True)
    except Exception as e:
        print(f"⚠️ No se pudo crear/acceder a la carpeta de respaldos '{RESPALDOS_DIR}': {e}")
        return None
    datos = _generar_respaldo_datos()
    nombre = f"{prefijo}_{datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d_%H%M%S')}.json"
    ruta = os.path.join(RESPALDOS_DIR, nombre)
    try:
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False)
        return ruta
    except Exception as e:
        print(f"⚠️ Error escribiendo el respaldo en '{ruta}': {e}")
        return None


# 📧 Copia de los respaldos fuera de Render: además de quedar guardados en el disco
# persistente (RESPALDOS_DIR), cada respaldo (manual o automático) se envía también por
# correo a RESPALDO_EMAIL_DESTINO, reutilizando el mismo webhook de Apps Script que ya usan
# los demás correos del sistema (ver GMAIL_SCRIPT_URL, más abajo). Así, si algo le pasa a la
# cuenta de Render o al disco persistente mismo, sigue existiendo una copia totalmente
# independiente. El archivo se comprime con gzip antes de codificarlo en base64 (los
# respaldos son JSON — comprimen muy bien) para no acercarse al límite de adjuntos de Gmail.
#
# ⚠️ El Apps Script detrás de GMAIL_SCRIPT_URL hoy solo sabe recibir {para, asunto, cuerpo} —
# hay que agregarle el manejo de 'adjunto_nombre' / 'adjunto_tipo' / 'adjunto_base64' para que
# el adjunto realmente llegue (instrucciones aparte). Mientras eso no se haga, el correo sale
# pero sin el archivo adjunto.
RESPALDO_EMAIL_DESTINO = os.environ.get('RESPALDO_EMAIL_DESTINO', 'notificacionesarkiv@gmail.com')
RESPALDO_EMAIL_MAX_MB = 20  # Margen bajo el límite real de Gmail (25 MB), pensando en el overhead de base64.


def _enviar_respaldo_por_correo(ruta_archivo):
    """Comprime el respaldo ya guardado en disco y lo envía como adjunto a
    RESPALDO_EMAIL_DESTINO. No borra ni modifica el archivo original en RESPALDOS_DIR — esto
    es solo una copia adicional fuera de Render. Se llama siempre en un hilo aparte (ver los
    dos puntos de llamada: el botón manual y el hilo del respaldo diario) para no demorar la
    respuesta ni el propio ciclo de respaldo si el correo tarda o falla."""
    if not RESPALDO_EMAIL_DESTINO:
        return False
    nombre_original = os.path.basename(ruta_archivo)
    asunto = f"Respaldo Arkiv - {nombre_original}"
    try:
        with open(ruta_archivo, 'rb') as f:
            contenido = f.read()
        comprimido = gzip.compress(contenido)
        tamano_mb = len(comprimido) / (1024 * 1024)
        if tamano_mb > RESPALDO_EMAIL_MAX_MB:
            print(f"⚠️ Respaldo '{nombre_original}' pesa {tamano_mb:.1f} MB comprimido — supera el límite de "
                  f"{RESPALDO_EMAIL_MAX_MB} MB para enviarlo por correo. Se guardó en el disco pero NO se envió copia por correo.")
            registrar_correo_log(RESPALDO_EMAIL_DESTINO, asunto, 'respaldo', 'omitido',
                                  f"Adjunto de {tamano_mb:.1f} MB supera el límite de {RESPALDO_EMAIL_MAX_MB} MB")
            return False
        adjunto_base64 = base64.b64encode(comprimido).decode('ascii')
        cuerpo = (
            "Respaldo automático de la base de datos de Arkiv.\n\n"
            f"Archivo: {nombre_original}.gz\n"
            f"Tamaño comprimido: {tamano_mb:.2f} MB\n\n"
            "Este correo se genera automáticamente como copia de seguridad fuera de Render.\n"
            "---\nEquipo de Soporte - ARKIV System"
        )
        payload = {
            "para": RESPALDO_EMAIL_DESTINO,
            "asunto": asunto,
            "cuerpo": cuerpo,
            "adjunto_nombre": f"{nombre_original}.gz",
            "adjunto_tipo": "application/gzip",
            "adjunto_base64": adjunto_base64,
        }
        if requests:
            res = requests.post(GMAIL_SCRIPT_URL, json=payload, timeout=30)
            print(f"✅ Copia del respaldo '{nombre_original}' enviada por correo a {RESPALDO_EMAIL_DESTINO}. Status: {res.status_code} | Respuesta del script: {res.text[:300]!r}")
        else:
            data_json = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(GMAIL_SCRIPT_URL, data=data_json, headers={'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=30) as response:
                response.read()
            print(f"✅ Copia del respaldo '{nombre_original}' enviada por correo a {RESPALDO_EMAIL_DESTINO} vía urllib.")
        registrar_correo_log(RESPALDO_EMAIL_DESTINO, asunto, 'respaldo', 'enviado')
        return True
    except Exception as e:
        print(f"⚠️ Error enviando copia por correo del respaldo '{nombre_original}': {e}")
        registrar_correo_log(RESPALDO_EMAIL_DESTINO, asunto, 'respaldo', 'error', str(e)[:300])
        return False


def _listar_respaldos():
    """Lista los archivos de respaldo ya guardados en RESPALDOS_DIR, más recientes primero.
    Devuelve [] si la carpeta no existe todavía (nunca se ha generado un respaldo, o el
    disco persistente no está montado)."""
    try:
        if not os.path.isdir(RESPALDOS_DIR):
            return []
        items = []
        for nombre in os.listdir(RESPALDOS_DIR):
            if not nombre.endswith('.json'):
                continue
            ruta = os.path.join(RESPALDOS_DIR, nombre)
            try:
                stat = os.stat(ruta)
                items.append({
                    'nombre': nombre,
                    'tipo': 'Automático (diario)' if nombre.startswith('auto_') else 'Manual',
                    'tamano_kb': round(stat.st_size / 1024, 1),
                    'fecha': datetime.fromtimestamp(stat.st_mtime, tz=ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception:
                continue
        items.sort(key=lambda x: x['fecha'], reverse=True)
        return items
    except Exception as e:
        print(f"⚠️ Error listando respaldos: {e}")
        return []


def _limpiar_respaldos_viejos():
    """Borra respaldos AUTOMÁTICOS ('auto_...') con más de RESPALDOS_RETENCION_DIAS de
    antigüedad, para que el disco no se llene solo con el tiempo. Los respaldos MANUALES
    nunca se borran solos — quien los generó decide cuándo quitarlos desde la página."""
    try:
        if not os.path.isdir(RESPALDOS_DIR):
            return
        limite = datetime.now(ZONA_HORARIA_COLOMBIA) - timedelta(days=RESPALDOS_RETENCION_DIAS)
        for nombre in os.listdir(RESPALDOS_DIR):
            if not nombre.startswith('auto_') or not nombre.endswith('.json'):
                continue
            ruta = os.path.join(RESPALDOS_DIR, nombre)
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(ruta), tz=ZONA_HORARIA_COLOMBIA)
                if mtime < limite:
                    os.remove(ruta)
                    print(f"🧹 Respaldo automático vencido eliminado: {nombre}")
            except Exception:
                pass
    except Exception as e:
        print(f"⚠️ Error limpiando respaldos viejos: {e}")


def _respaldo_diario_automatico():
    """Hilo en segundo plano (arranca una vez al cargar la app): cada hora revisa si ya se
    generó el respaldo automático de HOY y, si no, lo genera. El propio nombre del archivo
    del día actúa como candado (se crea con modo exclusivo 'x', que falla si ya existe) —
    así, aunque Render corra 2 procesos gunicorn de esta misma app en paralelo (como está
    configurado), solo uno de ellos termina generando el respaldo cada día."""
    time.sleep(30)  # Pequeña espera para no competir con el arranque del propio servidor.
    while True:
        try:
            os.makedirs(RESPALDOS_DIR, exist_ok=True)
            hoy = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d')
            marcador = os.path.join(RESPALDOS_DIR, f"auto_{hoy}.json")
            try:
                fh = os.open(marcador, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                os.close(fh)
            except FileExistsError:
                pass  # Ya se generó hoy (por este proceso o por otro worker) — nada que hacer.
            else:
                datos = _generar_respaldo_datos()
                with open(marcador, 'w', encoding='utf-8') as f:
                    json.dump(datos, f, ensure_ascii=False)
                print(f"✅ Respaldo automático diario generado: {marcador}")
                threading.Thread(target=_enviar_respaldo_por_correo, args=(marcador,), daemon=True).start()
                _limpiar_respaldos_viejos()
        except Exception as e:
            print(f"⚠️ Error en el hilo de respaldo automático: {e}")
        time.sleep(3600)  # Revisa cada hora si ya cambió el día.


if os.environ.get('DESHABILITAR_RESPALDO_AUTOMATICO') != '1':
    threading.Thread(target=_respaldo_diario_automatico, daemon=True).start()


@app.route('/admin/respaldos')
@login_required
@admin_required
@superadmin_required
def ver_respaldos():
    disco_disponible = os.path.isdir(RESPALDOS_DIR) or _crear_dir_respaldos_silencioso()
    return render_template('respaldos.html', respaldos=_listar_respaldos(), respaldos_dir=RESPALDOS_DIR, disco_disponible=disco_disponible, respaldo_email_destino=RESPALDO_EMAIL_DESTINO)


def _crear_dir_respaldos_silencioso():
    try:
        os.makedirs(RESPALDOS_DIR, exist_ok=True)
        return True
    except Exception:
        return False


@app.route('/admin/respaldos/generar', methods=['POST'])
@login_required
@admin_required
@superadmin_required
def generar_respaldo():
    """Botón "Generar y descargar ahora": crea un respaldo manual, lo guarda en el disco
    (queda listado igual que los automáticos) y lo entrega de inmediato como descarga."""
    ruta = _guardar_respaldo_en_disco(prefijo='manual')
    if not ruta:
        flash("No se pudo generar el respaldo: la carpeta de respaldos no está disponible (¿el disco persistente de Render ya está montado en /var/data?).", "error")
        return redirect(url_for('ver_respaldos'))
    registrar_log(session.get('username'), "Respaldo de Base de Datos", f"Respaldo manual generado: {os.path.basename(ruta)}")
    threading.Thread(target=_enviar_respaldo_por_correo, args=(ruta,), daemon=True).start()
    return send_file(ruta, as_attachment=True, download_name=os.path.basename(ruta))


@app.route('/admin/respaldos/descargar/<nombre>')
@login_required
@admin_required
@superadmin_required
def descargar_respaldo(nombre):
    nombre_seguro = os.path.basename(nombre)
    ruta = os.path.join(RESPALDOS_DIR, nombre_seguro)
    if not nombre_seguro.endswith('.json') or not os.path.isfile(ruta):
        flash("Ese archivo de respaldo ya no existe.", "error")
        return redirect(url_for('ver_respaldos'))
    return send_file(ruta, as_attachment=True, download_name=nombre_seguro)


@app.route('/admin/respaldos/eliminar/<nombre>', methods=['POST'])
@login_required
@admin_required
@superadmin_required
def eliminar_respaldo(nombre):
    nombre_seguro = os.path.basename(nombre)
    ruta = os.path.join(RESPALDOS_DIR, nombre_seguro)
    if nombre_seguro.endswith('.json') and os.path.isfile(ruta):
        try:
            os.remove(ruta)
            registrar_log(session.get('username'), "Respaldo de Base de Datos", f"Respaldo eliminado manualmente: {nombre_seguro}")
        except Exception as e:
            print(f"⚠️ Error eliminando respaldo '{nombre_seguro}': {e}")
    return redirect(url_for('ver_respaldos'))


# 📢 MÓDULO MURO DE COMUNICADOS
@app.route('/comunicados')
@login_required
def ver_comunicados():
    pestana = request.args.get('tab', 'activos')
    q_busqueda = request.args.get('q', '').strip().lower()
    
    conn, db_type = get_db()
    cursor = conn.cursor()
    
    estado_filtro = 'activo' if pestana == 'activos' else 'archivado'
    
    try:
        query = "SELECT id, titulo, contenido, nivel, fijado, imagen_url, estado, fecha, autor FROM comunicados WHERE estado = %s ORDER BY fijado DESC, id DESC" if db_type == 'postgres' else "SELECT id, titulo, contenido, nivel, fijado, imagen_url, estado, fecha, autor FROM comunicados WHERE estado = ? ORDER BY fijado DESC, id DESC"
        cursor.execute(query, (estado_filtro,))
        rows = cursor.fetchall()
    except Exception as e:
        print(f"Error consultando comunicados: {e}")
        rows = []

    conn.close()

    # 👤 "Publicado por" muestra el nombre/alias real del autor, no su usuario de inicio de
    # sesión crudo (p. ej. 'analistati' en vez de 'Ana Lisboa T.').
    nombres_usuarios = _mapa_nombres_usuarios()

    comunicados = []
    for r in rows:
        c_id, titulo, contenido, nivel, fijado, img_url, estado, fecha, autor = r
        texto_full = f"{titulo} {contenido} {autor}".lower()
        if not q_busqueda or q_busqueda in texto_full:
            comunicados.append({
                'id': c_id,
                'titulo': titulo,
                'contenido': contenido,
                'nivel': nivel,
                'fijado': fijado,
                'imagen_url': img_url,
                'estado': estado,
                'fecha': fecha,
                'autor': _nombre_para_mostrar(autor, nombres_usuarios)
            })

    # 👁️ Ver el muro de Comunicados marca como "leídos" todos los que están activos (no los
    # archivados: esos ya son historial, no algo pendiente por leer).
    usuario_actual = session.get('username')
    if pestana == 'activos':
        for c in comunicados:
            _marcar_comunicado_leido(c['id'], usuario_actual)

    # 👁️ Para soporte/admin, se muestra cuántos usuarios (de los activos) ya leyeron cada
    # comunicado — útil para políticas de lectura obligatoria.
    es_soporte = session.get('rol') in ROLES_CON_ACCESO_OPERATIVO
    if es_soporte:
        # 📢 Recordatorio automático de lectura pendiente — ver _revisar_recordatorios_lectura().
        _revisar_recordatorios_lectura()
    if es_soporte and comunicados:
        conteos = _conteo_lecturas_comunicados([c['id'] for c in comunicados])
        total_usuarios = _total_usuarios_activos()
        for c in comunicados:
            c['leidos'] = conteos.get(c['id'], 0)
            c['total_usuarios'] = total_usuarios

    return render_template('comunicados.html', comunicados=comunicados, pestana=pestana, q_busqueda=q_busqueda, rol=session.get('rol'), es_soporte=es_soporte)

@app.route('/comunicados/crear', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_comunicado():
    titulo = request.form.get('titulo', '').strip()
    # 📝 El contenido llega como HTML del editor de texto enriquecido (Quill) — se limpia
    # ANTES de guardarlo, para no confiar en lo que manda el navegador.
    contenido = _sanitizar_html_enriquecido(request.form.get('contenido', '').strip())
    nivel = request.form.get('nivel', 'info').strip()
    # ⚠️ Usar un bool nativo de Python (no 0/1 literal): la columna "fijado" en Neon
    # es de tipo BOOLEAN, y Postgres rechaza "column is of type boolean but expression
    # is of type integer" si se le pasa un entero. psycopg2 adapta True/False
    # correctamente a boolean, y sqlite3 los guarda igual de bien como 0/1.
    fijado = (request.form.get('fijado') == 'on')
    imagen = request.files.get('imagen')
    
    imagen_url = ""
    if imagen and archivo_permitido(imagen.filename):
        try:
            upload_result = cloudinary.uploader.upload(
                imagen, 
                resource_type="image",
                use_filename=True,
                unique_filename=True
            )
            imagen_url = upload_result.get('secure_url', '')
        except Exception as e:
            print(f"Error subiendo imagen de comunicado: {e}")

    if titulo and contenido and not _html_esta_vacio(contenido):
        fecha_act = obtener_fecha_actual()
        autor = session.get('username', 'Admin')

        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            q_ins = "INSERT INTO comunicados (titulo, contenido, nivel, fijado, imagen_url, estado, fecha, autor) VALUES (%s, %s, %s, %s, %s, 'activo', %s, %s)" if db_type == 'postgres' else "INSERT INTO comunicados (titulo, contenido, nivel, fijado, imagen_url, estado, fecha, autor) VALUES (?, ?, ?, ?, ?, 'activo', ?, ?)"
            cursor.execute(q_ins, (titulo, contenido, nivel, fijado, imagen_url, fecha_act, autor))
            conn.commit()
            registrar_log(autor, "Publicación de Comunicado", f"Nuevo comunicado: '{titulo}' [{nivel}]")
        except Exception as e:
            conn.rollback()
            print(f"Error creando comunicado: {e}")
        conn.close()

    return redirect(url_for('ver_comunicados'))

@app.route('/comunicados/editar/<int:com_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_comunicado(com_id):
    conn, db_type = get_db()
    cursor = conn.cursor()

    try:
        q_sel = "SELECT imagen_url FROM comunicados WHERE id = %s" if db_type == 'postgres' else "SELECT imagen_url FROM comunicados WHERE id = ?"
        cursor.execute(q_sel, (com_id,))
        row = cursor.fetchone()

        if not row:
            conn.close()
            return redirect(url_for('ver_comunicados'))

        imagen_url_actual = row[0] or ""

        titulo = request.form.get('titulo', '').strip()
        # 📝 El contenido llega como HTML del editor de texto enriquecido (Quill) — se limpia
        # ANTES de guardarlo, para no confiar en lo que manda el navegador.
        contenido = _sanitizar_html_enriquecido(request.form.get('contenido', '').strip())
        nivel = request.form.get('nivel', 'info').strip()
        fijado = (request.form.get('fijado') == 'on')
        imagen = request.files.get('imagen')

        if not titulo or not contenido or _html_esta_vacio(contenido):
            conn.close()
            return redirect(url_for('ver_comunicados'))

        # Si se adjunta una nueva imagen válida, se reemplaza; si no, se conserva la actual.
        imagen_url = imagen_url_actual
        if imagen and imagen.filename and archivo_permitido(imagen.filename):
            try:
                upload_result = cloudinary.uploader.upload(
                    imagen,
                    resource_type="image",
                    use_filename=True,
                    unique_filename=True
                )
                imagen_url = upload_result.get('secure_url', imagen_url_actual)
            except Exception as e:
                print(f"Error subiendo nueva imagen de comunicado {com_id}: {e}")

        q_upd = "UPDATE comunicados SET titulo = %s, contenido = %s, nivel = %s, fijado = %s, imagen_url = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE comunicados SET titulo = ?, contenido = ?, nivel = ?, fijado = ?, imagen_url = ? WHERE id = ?"
        cursor.execute(q_upd, (titulo, contenido, nivel, fijado, imagen_url, com_id))
        conn.commit()
        registrar_log(session.get('username'), "Edición de Comunicado", f"Comunicado '{titulo}' (ID {com_id}) actualizado")
    except Exception as e:
        conn.rollback()
        print(f"Error editando comunicado {com_id}: {e}")

    conn.close()
    return redirect(url_for('ver_comunicados'))

@app.route('/comunicados/archivar/<int:com_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def archivar_comunicado(com_id):
    conn, db_type = get_db()
    cursor = conn.cursor()

    try:
        q_sel = "SELECT estado, titulo FROM comunicados WHERE id = %s" if db_type == 'postgres' else "SELECT estado, titulo FROM comunicados WHERE id = ?"
        cursor.execute(q_sel, (com_id,))
        row = cursor.fetchone()

        if row:
            nuevo_estado = 'archivado' if row[0] == 'activo' else 'activo'
            # "fijado = 0" literal fallaba en Postgres contra la columna BOOLEAN real; "false" funciona en ambos motores.
            q_upd = "UPDATE comunicados SET estado = %s, fijado = false WHERE id = %s" if db_type == 'postgres' else "UPDATE comunicados SET estado = ?, fijado = false WHERE id = ?"
            cursor.execute(q_upd, (nuevo_estado, com_id))
            conn.commit()
            registrar_log(session['username'], "Cambio Estado Comunicado", f"Comunicado '{row[1]}' movido a {nuevo_estado}")
    except Exception as e:
        conn.rollback()
        print(f"Error archivando/reactivando comunicado {com_id}: {e}")

    conn.close()
    return redirect(url_for('ver_comunicados'))

@app.route('/comunicados/eliminar/<int:com_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_comunicado(com_id):
    # El propio botón en comunicados.html pregunta "¿Enviar este comunicado a la papelera?",
    # así que esto debe ser un envío a la papelera (estado='eliminado'), no un borrado
    # permanente inmediato — igual que ya se hace con instructivos y credenciales.
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM comunicados WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM comunicados WHERE id = ?"
        cursor.execute(q_sel, (com_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else f"ID {com_id}"

        q_upd = "UPDATE comunicados SET estado = 'eliminado' WHERE id = %s" if db_type == 'postgres' else "UPDATE comunicados SET estado = 'eliminado' WHERE id = ?"
        cursor.execute(q_upd, (com_id,))
        conn.commit()
        registrar_log(session['username'], "Eliminación de Comunicado", f"Se envió a la papelera el comunicado '{titulo}'")
    except Exception as e:
        conn.rollback()
        print(f"Error enviando comunicado {com_id} a la papelera: {e}")

    conn.close()
    # 🧭 Antes esto volvía al muro de Comunicados, donde el aviso simplemente desaparecía sin
    # ninguna confirmación visible de a dónde fue — eso hacía parecer que se había borrado para
    # siempre. Ahora se redirige directo a la Papelera con la pestaña de Comunicados ya abierta,
    # para que quede claro que sigue ahí y se puede restaurar.
    return redirect(url_for('ver_papelera', tab='comunicados'))


# 🖼️ FONDO DE LOGIN — panel de marca (imagen/video corto, uno o varios rotando) que se
# muestra junto al formulario de /login. Vive bajo Novedades y Comunicados (mismo permiso,
# agente_o_admin_required) porque es contenido institucional igual que los comunicados,
# aunque su tabla (login_fondo_media) y su plantilla son propias — ver _fondo_login_activo().
@app.route('/comunicados/fondo_login')
@login_required
@agente_o_admin_required
def ver_fondo_login():
    conn, db_type = get_db()
    cursor = conn.cursor()
    items = []
    try:
        cursor.execute("SELECT id, tipo, url, orden, estado, fecha_creacion, creado_por FROM login_fondo_media ORDER BY orden ASC, id ASC")
        items = [{'id': r[0], 'tipo': r[1], 'url': r[2], 'orden': r[3], 'estado': r[4],
                  'fecha_creacion': r[5], 'creado_por': r[6]} for r in cursor.fetchall()]
    except Exception as e:
        print(f"⚠️ Error consultando el fondo de login: {e}")
    conn.close()
    return render_template('fondo_login.html', items=items)


@app.route('/comunicados/fondo_login/subir', methods=['POST'])
@login_required
@agente_o_admin_required
def subir_fondo_login():
    archivo = request.files.get('archivo')
    tipo, url, public_id, error = _subir_fondo_login(archivo)
    if error:
        flash(error, "error")
        return redirect(url_for('ver_fondo_login'))

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COALESCE(MAX(orden), -1) FROM login_fondo_media")
        siguiente_orden = cursor.fetchone()[0] + 1
        fecha_act = obtener_fecha_actual()
        q = ("INSERT INTO login_fondo_media (tipo, url, public_id, orden, estado, fecha_creacion, creado_por) VALUES (%s, %s, %s, %s, 'activo', %s, %s)"
             if db_type == 'postgres' else
             "INSERT INTO login_fondo_media (tipo, url, public_id, orden, estado, fecha_creacion, creado_por) VALUES (?, ?, ?, ?, 'activo', ?, ?)")
        cursor.execute(q, (tipo, url, public_id, siguiente_orden, fecha_act, session.get('username')))
        conn.commit()
        registrar_log(session.get('username'), "Fondo de Login", f"Se agregó un(a) {tipo} nuevo al panel de marca del login ('{archivo.filename}')")
        flash("Archivo agregado al fondo de login.", "exito")
    except Exception as e:
        conn.rollback()
        print(f"Error guardando el fondo de login: {e}")
        flash("El archivo se subió, pero no se pudo guardar el registro. Intenta de nuevo.", "error")
    conn.close()
    return redirect(url_for('ver_fondo_login'))


@app.route('/comunicados/fondo_login/<int:item_id>/toggle', methods=['POST'])
@login_required
@agente_o_admin_required
def toggle_fondo_login(item_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    ph = '%s' if db_type == 'postgres' else '?'
    try:
        cursor.execute(f"SELECT estado FROM login_fondo_media WHERE id = {ph}", (item_id,))
        row = cursor.fetchone()
        if row:
            nuevo_estado = 'inactivo' if row[0] == 'activo' else 'activo'
            cursor.execute(f"UPDATE login_fondo_media SET estado = {ph} WHERE id = {ph}", (nuevo_estado, item_id))
            conn.commit()
            registrar_log(session.get('username'), "Fondo de Login", f"Se marcó el archivo #{item_id} como '{nuevo_estado}'")
    except Exception as e:
        conn.rollback()
        print(f"Error cambiando estado del fondo de login {item_id}: {e}")
    conn.close()
    return redirect(url_for('ver_fondo_login'))


@app.route('/comunicados/fondo_login/<int:item_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_fondo_login(item_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    ph = '%s' if db_type == 'postgres' else '?'
    try:
        cursor.execute(f"SELECT tipo, public_id FROM login_fondo_media WHERE id = {ph}", (item_id,))
        row = cursor.fetchone()
        if row:
            tipo, public_id = row
            cursor.execute(f"DELETE FROM login_fondo_media WHERE id = {ph}", (item_id,))
            conn.commit()
            registrar_log(session.get('username'), "Fondo de Login", f"Se eliminó del panel de marca del login el archivo #{item_id}")
            if public_id:
                try:
                    cloudinary.uploader.destroy(public_id, resource_type=('video' if tipo == 'video' else 'image'))
                except Exception as e:
                    print(f"⚠️ No se pudo borrar de Cloudinary el archivo de fondo de login #{item_id} ({public_id}): {e}")
    except Exception as e:
        conn.rollback()
        print(f"Error eliminando el fondo de login {item_id}: {e}")
    conn.close()
    return redirect(url_for('ver_fondo_login'))


@app.route('/comunicados/fondo_login/<int:item_id>/mover/<direccion>', methods=['POST'])
@login_required
@agente_o_admin_required
def mover_fondo_login(item_id, direccion):
    """Sube o baja un archivo en el orden de rotación, intercambiando su 'orden' con el del
    vecino inmediato — mismo patrón que el reordenamiento de Tipos de Activo del Inventario."""
    if direccion not in ('subir', 'bajar'):
        return redirect(url_for('ver_fondo_login'))
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, orden FROM login_fondo_media ORDER BY orden ASC, id ASC")
        filas = cursor.fetchall()
        posicion = next((i for i, f in enumerate(filas) if f[0] == item_id), None)
        vecino_pos = posicion - 1 if direccion == 'subir' else posicion + 1
        if posicion is not None and 0 <= vecino_pos < len(filas):
            id_actual, orden_actual = filas[posicion]
            id_vecino, orden_vecino = filas[vecino_pos]
            ph = '%s' if db_type == 'postgres' else '?'
            cursor.execute(f"UPDATE login_fondo_media SET orden = {ph} WHERE id = {ph}", (orden_vecino, id_actual))
            cursor.execute(f"UPDATE login_fondo_media SET orden = {ph} WHERE id = {ph}", (orden_actual, id_vecino))
            conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Error reordenando el fondo de login: {e}")
    conn.close()
    return redirect(url_for('ver_fondo_login'))

@app.route('/comunicados/<int:com_id>/lecturas')
@login_required
@agente_o_admin_required
def lecturas_comunicado(com_id):
    """JSON con quién (de las cuentas activas) ya leyó este comunicado y quién falta —
    usado por el modal de "Ver lecturas" en el muro de Comunicados. Incluye la cédula de quien
    leyó (cuando la tiene registrada) para poder identificarlo sin ambigüedad al exportar."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT usuario, COALESCE(nombre, usuario), cedula FROM usuarios WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY usuario ASC")
        todos = {r[0]: {'nombre': r[1], 'cedula': r[2] or ''} for r in cursor.fetchall()}

        q = "SELECT usuario, fecha FROM comunicados_leidos WHERE comunicado_id = %s" if db_type == 'postgres' else "SELECT usuario, fecha FROM comunicados_leidos WHERE comunicado_id = ?"
        cursor.execute(q, (com_id,))
        leidos = {r[0]: r[1] for r in cursor.fetchall()}
        conn.close()

        leyeron = [{'usuario': u, 'nombre': todos[u]['nombre'], 'cedula': todos[u]['cedula'], 'fecha': f} for u, f in leidos.items() if u in todos]
        faltan = [{'usuario': u, 'nombre': info['nombre']} for u, info in todos.items() if u not in leidos]
        leyeron.sort(key=lambda x: x['fecha'], reverse=True)
        faltan.sort(key=lambda x: x['nombre'].lower())
        return {'leyeron': leyeron, 'faltan': faltan, 'total': len(todos)}
    except Exception as e:
        conn.close()
        print(f"⚠️ Error listando lecturas del comunicado {com_id}: {e}")
        return {'leyeron': [], 'faltan': [], 'total': 0}


@app.route('/comunicados/<int:com_id>/lecturas/exportar_csv')
@login_required
@agente_o_admin_required
def exportar_lecturas_comunicado(com_id):
    """CSV de quién ha leído este comunicado (nombre, cédula, usuario, fecha) — quien falta por
    leer no tiene fecha de lectura, así que no aplica a este export (ver panel de Cumplimiento
    para el detalle de quién falta)."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_tit = "SELECT titulo FROM comunicados WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM comunicados WHERE id = ?"
        cursor.execute(q_tit, (com_id,))
        fila_tit = cursor.fetchone()
        titulo = fila_tit[0] if fila_tit else str(com_id)

        q = (
            "SELECT COALESCE(u.nombre, cl.usuario), u.cedula, cl.usuario, cl.fecha "
            "FROM comunicados_leidos cl LEFT JOIN usuarios u ON u.usuario = cl.usuario "
            "WHERE cl.comunicado_id = %s ORDER BY cl.fecha DESC"
        ) if db_type == 'postgres' else (
            "SELECT COALESCE(u.nombre, cl.usuario), u.cedula, cl.usuario, cl.fecha "
            "FROM comunicados_leidos cl LEFT JOIN usuarios u ON u.usuario = cl.usuario "
            "WHERE cl.comunicado_id = ? ORDER BY cl.fecha DESC"
        )
        cursor.execute(q, (com_id,))
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error exportando lecturas del comunicado {com_id}: {e}")
        filas = []
    conn.close()

    registrar_log(session.get('username'), "Exportación de Lecturas de Comunicado", f"Exportó las lecturas de '{titulo}'")
    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y%m%d_%H%M")
    nombre_limpio = re.sub(r'[^A-Za-z0-9_-]+', '_', titulo)[:60]
    return _exportar_csv_personas(f"Arkiv_Lecturas_{nombre_limpio}_{fecha_filename}.csv", filas)


@app.route('/comunicados/cumplimiento')
@login_required
@agente_o_admin_required
def cumplimiento_comunicados():
    """Panel de cumplimiento de lectura: de un vistazo, qué % de las cuentas activas ya leyó
    cada comunicado activo y quién falta — sin tener que abrir el modal de cada uno por
    separado en el muro de Comunicados."""
    _revisar_recordatorios_lectura()
    detalle = _detalle_cumplimiento_comunicados()
    return render_template(
        'comunicados_cumplimiento.html', detalle=detalle,
        horas_recordatorio=HORAS_RECORDATORIO_LECTURA_COMUNICADO
    )


@app.route('/comunicados/<int:com_id>/recordatorio', methods=['POST'])
@login_required
@agente_o_admin_required
def enviar_recordatorio_comunicado(com_id):
    """Envía ahora mismo (a demanda) el recordatorio de lectura pendiente para un comunicado
    puntual — no hay que esperar a que se cumplan las horas del aviso automático."""
    detalle = _detalle_cumplimiento_comunicados()
    comunicado = next((c for c in detalle if c['id'] == com_id), None)
    if comunicado:
        cantidad = _enviar_recordatorio_lectura(comunicado)
        registrar_log(session.get('username'), "Recordatorio de Lectura Enviado", f"Comunicado '{comunicado['titulo']}' (ID {com_id}): recordatorio a {cantidad} usuario(s)")
    return redirect(url_for('cumplimiento_comunicados'))


@app.route('/restaurar_comunicado/<int:com_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def restaurar_comunicado(com_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM comunicados WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM comunicados WHERE id = ?"
        cursor.execute(q_sel, (com_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else f"ID {com_id}"

        q_upd = "UPDATE comunicados SET estado = 'activo' WHERE id = %s" if db_type == 'postgres' else "UPDATE comunicados SET estado = 'activo' WHERE id = ?"
        cursor.execute(q_upd, (com_id,))
        conn.commit()
        registrar_log(session['username'], "Restauración de Comunicado", f"Se restauró el comunicado '{titulo}' desde la papelera.")
    except Exception as e:
        conn.rollback()
        print(f"Error restaurando comunicado {com_id}: {e}")

    conn.close()
    return redirect(url_for('ver_papelera'))

@app.route('/destruir_comunicado/<int:com_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def destruir_comunicado(com_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM comunicados WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM comunicados WHERE id = ?"
        cursor.execute(q_sel, (com_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else f"ID {com_id}"

        q_del = "DELETE FROM comunicados WHERE id = %s" if db_type == 'postgres' else "DELETE FROM comunicados WHERE id = ?"
        cursor.execute(q_del, (com_id,))
        conn.commit()
        registrar_log(session['username'], "Eliminación Permanente", f"Se destruyó permanentemente el comunicado '{titulo}'.")
    except Exception as e:
        conn.rollback()
        print(f"Error destruyendo comunicado {com_id}: {e}")

    conn.close()
    return redirect(url_for('ver_papelera'))

# 🎫 MÓDULO DE SOPORTE TI (TICKETS / SOLICITUDES INTERNAS)
CATEGORIAS_TICKET = ['Hardware', 'Software', 'Acceso/Credenciales', 'Red/Internet', 'Otro']
PRIORIDADES_TICKET = ['Baja', 'Media', 'Alta', 'Urgente']
ESTADOS_TICKET = ['Abierto', 'En Proceso', 'Pendiente', 'Resuelto', 'Cerrado', 'Cancelado']
TIPOS_TICKET = ['Incidente', 'Requerimiento']
# Metadatos de cada tipo para pintar las tarjetas de selección y las insignias (inspirado
# en la mesa de ayuda externa que ya usa la organización: distingue "algo se rompió" de
# "necesito algo nuevo" desde el primer paso de creación).
TIPOS_TICKET_INFO = {
    'Incidente': {
        'icono': 'fa-triangle-exclamation',
        'descripcion': 'Algo no funciona, está roto o te bloquea para trabajar.'
    },
    'Requerimiento': {
        'icono': 'fa-clipboard-list',
        'descripcion': 'Solicitas algo nuevo: acceso, equipo, software o información.'
    }
}
MAX_ADJUNTOS_TICKET = 5

# 📦 INVENTARIO DE ACTIVOS DE TI
ESTADOS_ACTIVO = ['Disponible', 'Asignado', 'Mantenimiento', 'Baja', 'Perdido']
TIPOS_ACTIVO = ['Computador de Escritorio', 'Portátil', 'Impresora', 'Monitor', 'Teléfono/Celular', 'Servidor', 'Red (Switch/Router/AP)', 'Otro']
ICONOS_TIPO_ACTIVO = ['desktop', 'laptop', 'print', 'display', 'mobile-screen', 'server', 'network-wired', 'box',
                      'tablet', 'keyboard', 'headphones', 'camera', 'video', 'wifi', 'hard-drive', 'database',
                      'microchip', 'plug', 'tv', 'phone', 'box-archive', 'shield-halved', 'briefcase']
MOTIVOS_REEMPLAZO_ACTIVO = [
    {'clave': 'Equipo dañado', 'icono': 'screwdriver-wrench', 'descripcion': 'No funciona o requiere reparación mayor'},
    {'clave': 'Renovación', 'icono': 'arrows-rotate', 'descripcion': 'Reemplazo por uno más nuevo o mejor'},
    {'clave': 'Pérdida', 'icono': 'lock', 'descripcion': 'El equipo se extravió'},
    {'clave': 'Robo', 'icono': 'user-secret', 'descripcion': 'Hurto reportado a las autoridades'},
    {'clave': 'Fin de vida útil', 'icono': 'calendar-xmark', 'descripcion': 'Equipo obsoleto, se da de baja'},
    {'clave': 'Reasignación', 'icono': 'right-left', 'descripcion': 'Cambio de usuario, sin daño'},
    {'clave': 'Otro', 'icono': 'comment', 'descripcion': 'Especifica el motivo en las notas'},
]
ESTADOS_RESULTANTES_REEMPLAZO = [
    {'valor': 'Disponible', 'etiqueta': 'En bodega — puede reasignarse'},
    {'valor': 'Mantenimiento', 'etiqueta': 'En mantenimiento — pendiente reparar'},
    {'valor': 'Baja', 'etiqueta': 'Dado de baja — definitivamente'},
    {'valor': 'Perdido', 'etiqueta': 'Perdido'},
]


def _activos_pendientes_devolucion(nombre_colaborador, usuario_login=None):
    """Activos del Inventario que siguen en estado 'Asignado' a nombre de este colaborador
    (por nombre completo o por su usuario de login), pendientes de que Gestión Humana o TI
    certifiquen su devolución (ver /inventario/certificacion_devoluciones). 'asignado_a' es
    texto libre —igual que el resto del módulo de Inventario— así que la coincidencia es por
    subcadena, no exacta, para no dejar pasar por alto una devolución real solo porque el
    nombre se escribió con una variación menor."""
    objetivos = [t for t in [(nombre_colaborador or '').strip().lower(), (usuario_login or '').strip().lower()] if t]
    if not objetivos:
        return []
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nombre, tipo_activo, asignado_a FROM activos_inventario WHERE estado = 'Asignado' AND COALESCE(eliminado, 0) = 0")
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error consultando activos pendientes de devolución: {e}")
        filas = []
    conn.close()
    pendientes = []
    for activo_id, nombre_activo, tipo_activo, asignado_a in filas:
        asignado_norm = (asignado_a or '').strip().lower()
        if asignado_norm and any(obj in asignado_norm or asignado_norm in obj for obj in objetivos):
            pendientes.append({'id': activo_id, 'nombre': nombre_activo, 'tipo_activo': tipo_activo, 'asignado_a': asignado_a})
    return pendientes

# ⏱️ SLA (Acuerdos de Nivel de Servicio): horas máximas de "primera respuesta" (sacar el
# ticket de 'Abierto') y de "resolución" (llegar a 'Resuelto'/'Cerrado') según la prioridad.
# Son valores de partida razonables para un equipo de soporte interno; se pueden ajustar
# aquí si el equipo de TI define otros tiempos formales más adelante.
SLA_HORAS_POR_PRIORIDAD = {
    'Urgente': {'respuesta': 2, 'resolucion': 24},
    'Alta': {'respuesta': 4, 'resolucion': 48},
    'Media': {'respuesta': 8, 'resolucion': 96},
    'Baja': {'respuesta': 24, 'resolucion': 168},
}
MAX_MODIFICACIONES_SLA = 2
FORMATO_FECHA_TICKET = "%Y-%m-%d %H:%M:%S"

# 🙂 Calificación de satisfacción (1 a 5 estrellas) que el solicitante puede dejar una sola
# vez, cuando su ticket ya está Resuelto o Cerrado — igual que en la mesa de ayuda externa.
CALIFICACION_MIN = 1
CALIFICACION_MAX = 5

# ⏳ Umbral para distinguir "Vigente" de "Próximo a vencer" en los filtros de la lista: si
# queda menos del 20% del tiempo total de resolución (o menos de 4 horas, lo que sea mayor),
# se considera que el ticket está por vencer.
UMBRAL_PROXIMO_A_VENCER_PORCENTAJE = 0.2
UMBRAL_PROXIMO_A_VENCER_HORAS_MIN = 4


def _puede_ver_ticket(creado_por):
    """Un ticket solo lo puede ver quien lo creó o cualquier cuenta con rol admin (equipo de soporte TI)."""
    return session.get('rol') in ROLES_CON_ACCESO_OPERATIVO or session.get('username') == creado_por


def _estados_disponibles_ticket(estado_actual):
    """Estados que se pueden elegir desde el estado ACTUAL del ticket — usada tanto para
    filtrar el desplegable de 'Gestionar Solicitud' en ticket_detalle.html (ver ver_ticket) como
    para validar el cambio del lado del servidor (ver actualizar_ticket), así ambos quedan
    siempre sincronizados con una sola regla.

    No se puede saltar directo a 'Resuelto' sin que el ticket haya pasado antes por
    'En Proceso' (guardando ese cambio primero), ni a 'Cerrado' sin haber pasado antes por
    'Resuelto' — pedido explícito de Tomás para no permitir cerrar/resolver casos que nunca se
    empezaron a atender. Una vez 'Cerrado' el ticket queda bloqueado: debe ser el último estado,
    sin más cambios posibles desde este formulario."""
    if estado_actual == 'Cerrado':
        return ['Cerrado']
    disponibles = []
    for e in ESTADOS_TICKET:
        if e == 'Resuelto' and estado_actual not in ('En Proceso', 'Resuelto'):
            continue
        if e == 'Cerrado' and estado_actual not in ('Resuelto', 'Cerrado'):
            continue
        disponibles.append(e)
    return disponibles


def _parsear_fecha_ticket(fecha_str):
    """Convierte el formato de fecha usado en todo el módulo ('YYYY-MM-DD HH:MM:SS') a
    datetime. Devuelve None si el valor está vacío o no tiene el formato esperado."""
    if not fecha_str:
        return None
    try:
        return datetime.strptime(fecha_str, FORMATO_FECHA_TICKET)
    except Exception:
        return None


def _calcular_limite_sla(fecha_base_str, horas):
    """Suma 'horas' a una fecha base (en el formato del módulo) y devuelve el resultado en
    ese mismo formato. Si la fecha base no se puede interpretar, usa el momento actual."""
    base = _parsear_fecha_ticket(fecha_base_str) or datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)
    return (base + timedelta(hours=horas)).strftime(FORMATO_FECHA_TICKET)


def _calcular_sla_ticket(ticket):
    """Calcula, a partir de los campos crudos del ticket, el estado de cumplimiento del SLA
    de RESOLUCIÓN (el que se muestra de forma prominente en la interfaz, como hace la mesa
    de ayuda externa con su 'Compromiso con el usuario'). Devuelve un dict listo para pintar:
    {'estado': 'en_tiempo'|'vencido'|'cumplido'|'incumplido'|'sin_datos', 'texto': ..., 'color': ...}
    """
    # ⏸️ Mientras el ticket está en 'Pendiente' el conteo de SLA queda en pausa (puede que el
    # agente esté esperando información del usuario y no sea justo que el caso se "queme" por
    # eso). Se corta aquí, antes de comparar contra el límite, para que no aparezca ni vencido
    # ni próximo a vencer mientras dure la pausa — ver también _progreso_ticket() y
    # _revisar_alertas_sla(), que respetan el mismo estado.
    if ticket.get('estado') == 'Pendiente':
        return {'estado': 'pausado', 'texto': 'SLA en pausa (Pendiente)', 'color': 'slate'}

    limite = _parsear_fecha_ticket(ticket.get('sla_resolucion_limite'))
    if not limite:
        return {'estado': 'sin_datos', 'texto': 'Sin datos de SLA', 'color': 'slate'}

    cumplida = _parsear_fecha_ticket(ticket.get('sla_resolucion_cumplida'))
    ahora = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)

    if cumplida:
        if cumplida <= limite:
            return {'estado': 'cumplido', 'texto': 'Resuelto en tiempo', 'color': 'emerald'}
        return {'estado': 'incumplido', 'texto': 'Resuelto fuera de tiempo', 'color': 'rose'}

    if ahora > limite:
        return {'estado': 'vencido', 'texto': 'SLA vencido', 'color': 'rose'}

    restante = limite - ahora
    if restante.days >= 1:
        texto = f"Quedan ~{restante.days} día(s)"
    else:
        horas_restantes = max(1, restante.seconds // 3600)
        texto = f"Quedan ~{horas_restantes} hora(s)"
    return {'estado': 'en_tiempo', 'texto': texto, 'color': 'cyan'}


def _bucket_cumplimiento_ticket(ticket):
    """Agrupa el ticket en una de las categorías que se muestran como pestañas filtrables en
    la lista (inspirado en la mesa de ayuda externa): 'cerrado' (ya Resuelto/Cerrado, sin
    importar si llegó a tiempo o no), 'vencido' (abierto y ya pasó la fecha límite),
    'proximo_a_vencer' (abierto y le queda poco tiempo) o 'vigente' (abierto, con margen)."""
    if ticket.get('estado') in ('Resuelto', 'Cerrado', 'Cancelado'):
        return 'cerrado'

    sla = ticket.get('sla') or _calcular_sla_ticket(ticket)
    if sla.get('estado') == 'vencido':
        return 'vencido'
    if sla.get('estado') != 'en_tiempo':
        # 'sin_datos' (no debería pasar para un ticket abierto, pero por seguridad) se trata
        # como vigente en vez de ocultarlo de todas las pestañas.
        return 'vigente'

    limite = _parsear_fecha_ticket(ticket.get('sla_resolucion_limite'))
    ahora = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)
    horas_totales = SLA_HORAS_POR_PRIORIDAD.get(ticket.get('prioridad'), SLA_HORAS_POR_PRIORIDAD['Media'])['resolucion']
    umbral_horas = max(UMBRAL_PROXIMO_A_VENCER_HORAS_MIN, horas_totales * UMBRAL_PROXIMO_A_VENCER_PORCENTAJE)
    horas_restantes = (limite - ahora).total_seconds() / 3600 if limite else 0
    return 'proximo_a_vencer' if horas_restantes <= umbral_horas else 'vigente'


def _progreso_ticket(ticket):
    """Calcula qué porcentaje del tiempo total de SLA de resolución ya transcurrió (o
    transcurrió hasta que se resolvió), como una barra de avance visual en la lista de
    tickets. No depende de nueva información en la BD: se deriva de fechas ya existentes."""
    creado = _parsear_fecha_ticket(ticket.get('fecha_creacion'))
    if not creado:
        return 0
    horas_totales = SLA_HORAS_POR_PRIORIDAD.get(ticket.get('prioridad'), SLA_HORAS_POR_PRIORIDAD['Media'])['resolucion']
    if horas_totales <= 0:
        return 0

    cumplida = _parsear_fecha_ticket(ticket.get('sla_resolucion_cumplida'))
    pausado_desde = _parsear_fecha_ticket(ticket.get('sla_pausado_desde'))
    if cumplida and ticket.get('estado') in ('Resuelto', 'Cerrado'):
        fin = cumplida
    elif ticket.get('estado') == 'Pendiente' and pausado_desde:
        # ⏸️ Congela la barra en el momento en que entró a Pendiente, en vez de dejarla seguir
        # avanzando hacia el 100% mientras el SLA está en pausa (ver _calcular_sla_ticket()).
        fin = pausado_desde
    else:
        fin = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)

    horas_transcurridas = (fin - creado).total_seconds() / 3600
    return max(0, min(100, round(horas_transcurridas / horas_totales * 100)))


def _revisar_alertas_sla():
    """Revisa los tickets abiertos/en proceso y avisa (campanita + correo) la primera vez que
    uno entra en 'Próximo a vencer' o se escala a 'Vencido' — así el equipo no depende de estar
    mirando la lista para enterarse. No hay un scheduler aparte corriendo en Render, así que
    esto se llama de forma perezosa (best-effort, nunca debe tumbar la página) cada vez que
    alguien del equipo de soporte abre la lista de tickets o los indicadores. 'sla_alerta_nivel'
    guarda el último nivel ya avisado por ticket para no repetir el mismo aviso una y otra vez;
    se limpia a NULL cuando se extiende el SLA (ver extender_sla_ticket) para que pueda volver
    a avisar si se vuelve a acercar la nueva fecha límite."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, titulo, tipo, prioridad, estado, asignado_a, fecha_creacion, "
            "sla_resolucion_limite, sla_resolucion_cumplida, sla_alerta_nivel FROM tickets "
            "WHERE estado NOT IN ('Resuelto', 'Cerrado', 'Cancelado', 'Pendiente') AND COALESCE(eliminado, 0) = 0"
        )
        filas = cursor.fetchall()

        for (ticket_id, titulo, tipo, prioridad, estado, asignado_a, fecha_creacion,
             sla_limite, sla_cumplida, nivel_previo) in filas:
            t = {
                'estado': estado, 'prioridad': prioridad, 'fecha_creacion': fecha_creacion,
                'sla_resolucion_limite': sla_limite, 'sla_resolucion_cumplida': sla_cumplida,
            }
            bucket = _bucket_cumplimiento_ticket(t)
            if bucket not in ('proximo_a_vencer', 'vencido'):
                continue
            if bucket == nivel_previo or (bucket == 'proximo_a_vencer' and nivel_previo == 'vencido'):
                continue  # ya se avisó este nivel (o uno peor) — no repetir ni "des-escalar"

            q_upd = ("UPDATE tickets SET sla_alerta_nivel = %s WHERE id = %s" if db_type == 'postgres'
                      else "UPDATE tickets SET sla_alerta_nivel = ? WHERE id = ?")
            cursor.execute(q_upd, (bucket, ticket_id))
            conn.commit()

            codigo = _codigo_ticket(tipo or 'Incidente', ticket_id, fecha_creacion)
            url_ticket = url_for('ver_ticket', ticket_id=ticket_id)
            equipo = [] if asignado_a else _equipo_soporte_activo()
            destinatarios = [asignado_a] if asignado_a else [m['usuario'] for m in equipo]

            if bucket == 'vencido':
                mensaje = f"⛔ SLA vencido en la solicitud {codigo}: '{titulo}'"
                asunto = f"[Arkiv] SLA vencido — solicitud {codigo}"
                cuerpo_estado = "ya superó"
            else:
                mensaje = f"⏳ La solicitud {codigo} está próxima a vencer su SLA: '{titulo}'"
                asunto = f"[Arkiv] SLA próximo a vencer — solicitud {codigo}"
                cuerpo_estado = "está por superar"
            cuerpo = (
                f"La solicitud de soporte {codigo} ('{titulo}') {cuerpo_estado} su tiempo límite de resolución.\n\n"
                f"Ingresa a Arkiv, módulo Solicitudes TI, para revisarla.\n\n---\nArkiv"
            )

            crear_notificacion_para_varios(destinatarios, mensaje, url=url_ticket)
            if asignado_a:
                correo_dest = _correo_de_usuario(asignado_a)
                if correo_dest:
                    threading.Thread(target=enviar_correo_ticket, args=(correo_dest, asunto, cuerpo)).start()
            else:
                for miembro in equipo:
                    if miembro['correo']:
                        threading.Thread(target=enviar_correo_ticket, args=(miembro['correo'], asunto, cuerpo)).start()

            # 🚨 Escalamiento a supervisor: si el ticket ya está VENCIDO (no solo "próximo a
            # vencer") y tiene un agente puntual asignado, el aviso de arriba solo llegó a ese
            # agente — quien ya lo tenía y, evidentemente, no alcanzó a resolverlo a tiempo. Se
            # escala avisando también a todos los admins activos (los "supervisores" de Arkiv,
            # no hay un rol distinto todavía) para que puedan intervenir, reasignar o priorizar.
            # Si el ticket no tiene agente asignado, los admins ya estaban en 'equipo' arriba —
            # no hace falta escalar dos veces.
            if bucket == 'vencido' and asignado_a:
                admins = [a for a in _admins_activos() if a['usuario'] != asignado_a]
                if admins:
                    mensaje_escalado = f"🚨 Escalamiento SLA: la solicitud {codigo} sigue vencida (asignada a {asignado_a}): '{titulo}'"
                    asunto_escalado = f"[Arkiv] Escalamiento SLA — solicitud {codigo} vencida"
                    cuerpo_escalado = (
                        f"La solicitud de soporte {codigo} ('{titulo}') superó su tiempo límite de resolución "
                        f"y sigue asignada a '{asignado_a}' sin resolverse.\n\n"
                        f"Se escala a tu cuenta como administrador para que la revises, reasignes o priorices "
                        f"según corresponda.\n\nIngresa a Arkiv, módulo Solicitudes TI, para revisarla.\n\n---\nArkiv"
                    )
                    crear_notificacion_para_varios([a['usuario'] for a in admins], mensaje_escalado, url=url_ticket)
                    for admin in admins:
                        if admin['correo']:
                            threading.Thread(target=enviar_correo_ticket, args=(admin['correo'], asunto_escalado, cuerpo_escalado)).start()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error revisando alertas de SLA: {e}")


def _config_ticket_lista(tipo_config):
    """Devuelve [{'id', 'nombre', 'direccion', 'responsable', 'nit', 'razon_social'}] de las
    Áreas, Sedes, Categorías o Proveedores activos configurados por el equipo de soporte en
    /tickets/configuracion (tipo_config: 'area' | 'sede' | 'categoria' | 'proveedor').
    'direccion'/'responsable' solo tienen contenido real para tipo_config == 'sede' (y
    'responsable' también para 'area'/'categoria'); 'nit'/'razon_social' solo para
    tipo_config == 'proveedor'. Para el resto de combinaciones quedan en None.
    Si la tabla no existe todavía o falla la consulta, devuelve una lista vacía en vez de
    romper la página que la llama."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT id, nombre, direccion, responsable, nit, razon_social FROM ticket_configuraciones WHERE tipo = %s AND estado = 'activo' ORDER BY nombre ASC" if db_type == 'postgres' else "SELECT id, nombre, direccion, responsable, nit, razon_social FROM ticket_configuraciones WHERE tipo = ? AND estado = 'activo' ORDER BY nombre ASC"
        cursor.execute(q, (tipo_config,))
        filas = [{'id': r[0], 'nombre': r[1], 'direccion': r[2], 'responsable': r[3], 'nit': r[4], 'razon_social': r[5]} for r in cursor.fetchall()]
        conn.close()
        return filas
    except Exception as e:
        print(f"⚠️ Error listando configuración de tickets ('{tipo_config}'): {e}")
        return []


def _plantillas_ticket_activas():
    """Plantillas de solicitud activas (administradas por el equipo de soporte en
    /tickets/plantillas), usadas para prellenar 'Nueva Solicitud' — ver ver_tickets()."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, nombre, tipo, categoria, prioridad, area, sede, titulo, descripcion FROM ticket_plantillas WHERE estado = 'activo' ORDER BY nombre ASC")
        filas = [
            {
                'id': r[0], 'nombre': r[1], 'tipo': r[2] or 'Incidente', 'categoria': r[3] or '',
                'prioridad': r[4] or 'Media', 'area': r[5] or '', 'sede': r[6] or '',
                'titulo': r[7], 'descripcion': r[8]
            }
            for r in cursor.fetchall()
        ]
        conn.close()
        return filas
    except Exception as e:
        print(f"⚠️ Error listando plantillas de ticket: {e}")
        return []


def _correo_de_usuario(username):
    """Busca el correo registrado de un usuario por su nombre de cuenta. Devuelve None si no
    existe o si algo falla (nunca debe tumbar el flujo del ticket que lo llama)."""
    if not username:
        return None
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT correo FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT correo FROM usuarios WHERE usuario = ?"
        cursor.execute(q, (username,))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else None
    except Exception as e:
        print(f"⚠️ Error buscando correo de '{username}': {e}")
        return None


def _info_usuario(username):
    """Trae el perfil completo (nombre, correo, teléfono, rol) de un usuario por su cuenta de
    inicio de sesión. Se usa para mostrar la mayor cantidad de información posible del
    solicitante en el detalle de un ticket. Devuelve None si no existe o si algo falla."""
    if not username:
        return None
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT usuario, nombre, correo, telefono, rol FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT usuario, nombre, correo, telefono, rol FROM usuarios WHERE usuario = ?"
        cursor.execute(q, (username,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            return None
        return {'usuario': row[0], 'nombre': row[1], 'correo': row[2], 'telefono': row[3], 'rol': row[4]}
    except Exception as e:
        print(f"⚠️ Error buscando perfil de '{username}': {e}")
        return None


def _equipo_soporte_activo():
    """Devuelve la lista de cuentas activas con rol 'admin' o 'agente' — el 'equipo de
    soporte' que se notifica cuando se crea un ticket nuevo (salvo que quede asignado
    automáticamente a alguien puntual). Devuelve [] si algo falla."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT usuario, correo FROM usuarios WHERE estado = 'activo' AND rol IN ('admin', 'agente')" if db_type == 'postgres' else "SELECT usuario, correo FROM usuarios WHERE estado = 'activo' AND rol IN ('admin', 'agente')"
        cursor.execute(q)
        filas = [{'usuario': r[0], 'correo': r[1]} for r in cursor.fetchall()]
        conn.close()
        return filas
    except Exception as e:
        print(f"⚠️ Error listando equipo de soporte: {e}")
        return []


def _admins_activos():
    """Devuelve las cuentas activas con rol 'admin' — los 'supervisores' a quienes se
    escala un ticket cuyo SLA ya venció y sigue sin resolverse (ver _revisar_alertas_sla).
    Devuelve [] si algo falla."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT usuario, correo FROM usuarios WHERE estado = 'activo' AND rol = 'admin'")
        filas = [{'usuario': r[0], 'correo': r[1]} for r in cursor.fetchall()]
        conn.close()
        return filas
    except Exception as e:
        print(f"⚠️ Error listando administradores activos: {e}")
        return []


def crear_notificacion(usuario, mensaje, url='', tipo='ticket'):
    """Crea una notificación interna (campanita) para un usuario puntual. Nunca debe tumbar
    el flujo que la llama: cualquier error se registra en consola y se ignora."""
    if not usuario:
        return
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        fecha = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d %H:%M:%S')
        q = "INSERT INTO notificaciones (usuario, tipo, mensaje, url, fecha) VALUES (%s, %s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO notificaciones (usuario, tipo, mensaje, url, fecha) VALUES (?, ?, ?, ?, ?)"
        cursor.execute(q, (usuario, tipo, mensaje, url, fecha))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error creando notificación para '{usuario}': {e}")


def crear_notificacion_para_varios(usuarios, mensaje, url='', tipo='ticket'):
    """Igual que crear_notificacion pero para una lista de usuarios (p. ej. todo el equipo
    de soporte), sin duplicar destinatarios."""
    for u in set(u for u in (usuarios or []) if u):
        crear_notificacion(u, mensaje, url=url, tipo=tipo)


# 📅 VENCIMIENTO DE DOCUMENTOS (institucionales en 'galerias' y por empleado en
# 'documentos_empleado') -----------------------------------------------------------------
# Umbral de días antes del vencimiento en el que un documento pasa a "Próximo a vencer".
UMBRAL_DIAS_VENCIMIENTO_PROXIMO = 30


def _bucket_vencimiento(fecha_vencimiento_str):
    """Clasifica una fecha de vencimiento (texto 'YYYY-MM-DD') en None (sin vencimiento o
    vigente y lejos de vencer), 'proximo_a_vencer' (dentro del umbral) o 'vencido' (ya pasó).
    Nunca lanza excepción: una fecha vacía o mal formada se trata como "sin vencimiento"."""
    if not fecha_vencimiento_str:
        return None
    try:
        fecha_venc = datetime.strptime(str(fecha_vencimiento_str)[:10], '%Y-%m-%d')
    except Exception:
        return None
    hoy = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)
    dias_restantes = (fecha_venc - hoy).total_seconds() / 86400
    if dias_restantes < 0:
        return 'vencido'
    if dias_restantes <= UMBRAL_DIAS_VENCIMIENTO_PROXIMO:
        return 'proximo_a_vencer'
    return None


def _revisar_alertas_vencimientos():
    """Revisa documentos institucionales (galerias) y por empleado (documentos_empleado) con
    fecha de vencimiento, y avisa (campanita + correo) la primera vez que uno entra en
    'Próximo a vencer' o se escala a 'Vencido' — mismo patrón perezoso que _revisar_alertas_sla
    (no hay scheduler aparte en Render, así que esto se llama cada vez que alguien visita el
    listado de Instructivos o el panel de Vencimientos). El nivel ya avisado se guarda por
    documento para no repetir el mismo aviso una y otra vez."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()

        # --- Documentos institucionales (galerias) ---
        try:
            cursor.execute(
                "SELECT id, titulo, fecha_vencimiento, alerta_vencimiento_nivel FROM galerias "
                "WHERE COALESCE(estado, 'activo') != 'eliminado' AND fecha_vencimiento IS NOT NULL AND fecha_vencimiento != ''"
            )
            filas_galerias = cursor.fetchall()
        except Exception as e:
            print(f"⚠️ Error listando galerías con vencimiento: {e}")
            filas_galerias = []

        equipo = None
        for galeria_id, titulo, fecha_venc, nivel_previo in filas_galerias:
            bucket = _bucket_vencimiento(fecha_venc)
            if bucket not in ('proximo_a_vencer', 'vencido'):
                continue
            if bucket == nivel_previo or (bucket == 'proximo_a_vencer' and nivel_previo == 'vencido'):
                continue

            q_upd = ("UPDATE galerias SET alerta_vencimiento_nivel = %s WHERE id = %s" if db_type == 'postgres'
                      else "UPDATE galerias SET alerta_vencimiento_nivel = ? WHERE id = ?")
            cursor.execute(q_upd, (bucket, galeria_id))
            conn.commit()

            if equipo is None:
                equipo = _equipo_soporte_activo()
            url_doc = url_for('index')
            if bucket == 'vencido':
                mensaje = f"⛔ El instructivo '{titulo}' ya venció (fecha límite {fecha_venc})"
                asunto = f"[Arkiv] Documento vencido — '{titulo}'"
                cuerpo_estado = "ya superó"
            else:
                mensaje = f"⏳ El instructivo '{titulo}' está próximo a vencer (fecha límite {fecha_venc})"
                asunto = f"[Arkiv] Documento próximo a vencer — '{titulo}'"
                cuerpo_estado = "está por superar"
            cuerpo = (
                f"El instructivo/documento '{titulo}' {cuerpo_estado} su fecha de vencimiento ({fecha_venc}).\n\n"
                f"Ingresa a Arkiv, módulo Instructivos y Archivos, para revisarlo y renovarlo si corresponde.\n\n---\nArkiv"
            )
            crear_notificacion_para_varios([m['usuario'] for m in equipo], mensaje, url=url_doc, tipo='vencimiento')
            for miembro in equipo:
                if miembro['correo']:
                    threading.Thread(target=enviar_correo_ticket, args=(miembro['correo'], asunto, cuerpo)).start()

        # --- Documentos por empleado (documentos_empleado) ---
        try:
            cursor.execute(
                "SELECT id, usuario, titulo, tipo_documento, fecha_vencimiento, alerta_nivel FROM documentos_empleado "
                "WHERE COALESCE(estado, 'activo') = 'activo' AND fecha_vencimiento IS NOT NULL AND fecha_vencimiento != ''"
            )
            filas_empleado = cursor.fetchall()
        except Exception as e:
            print(f"⚠️ Error listando documentos de empleado con vencimiento: {e}")
            filas_empleado = []

        admins_cache = None
        for doc_id, usuario_doc, titulo, tipo_doc, fecha_venc, nivel_previo in filas_empleado:
            bucket = _bucket_vencimiento(fecha_venc)
            if bucket not in ('proximo_a_vencer', 'vencido'):
                continue
            if bucket == nivel_previo or (bucket == 'proximo_a_vencer' and nivel_previo == 'vencido'):
                continue

            q_upd = ("UPDATE documentos_empleado SET alerta_nivel = %s WHERE id = %s" if db_type == 'postgres'
                      else "UPDATE documentos_empleado SET alerta_nivel = ? WHERE id = ?")
            cursor.execute(q_upd, (bucket, doc_id))
            conn.commit()

            if admins_cache is None:
                admins_cache = _equipo_soporte_activo()
            url_doc = url_for('ver_vencimientos')
            if bucket == 'vencido':
                mensaje = f"⛔ El documento '{titulo}' ({tipo_doc}) de {usuario_doc} ya venció (fecha límite {fecha_venc})"
                asunto = f"[Arkiv] Documento vencido — '{titulo}'"
                cuerpo_estado = "ya superó"
            else:
                mensaje = f"⏳ El documento '{titulo}' ({tipo_doc}) de {usuario_doc} está próximo a vencer (fecha límite {fecha_venc})"
                asunto = f"[Arkiv] Documento próximo a vencer — '{titulo}'"
                cuerpo_estado = "está por superar"
            cuerpo = (
                f"El documento '{titulo}' ({tipo_doc}) {cuerpo_estado} su fecha de vencimiento ({fecha_venc}).\n\n"
                f"Ingresa a Arkiv, módulo Vencimiento de Documentos, para revisarlo y gestionar su renovación.\n\n---\nArkiv"
            )
            destinatarios = set([usuario_doc] + [m['usuario'] for m in admins_cache])
            crear_notificacion_para_varios(list(destinatarios), mensaje, url=url_doc, tipo='vencimiento')
            correo_empleado = _correo_de_usuario(usuario_doc)
            if correo_empleado:
                threading.Thread(target=enviar_correo_ticket, args=(correo_empleado, asunto, cuerpo)).start()
            for miembro in admins_cache:
                if miembro['correo']:
                    threading.Thread(target=enviar_correo_ticket, args=(miembro['correo'], asunto, cuerpo)).start()

        conn.close()
    except Exception as e:
        print(f"⚠️ Error revisando alertas de vencimiento de documentos: {e}")


def _subir_archivo_a_cloudinary(file):
    """Sube un único archivo a Cloudinary aplicando las mismas reglas que /subir y
    /editar_galeria según su extensión (video/pdf/comprimido/imagen). Devuelve (url,
    nombre_original) o (None, None) si el archivo no viene, no es válido, o falla la subida."""
    if not file or not file.filename or not archivo_permitido(file.filename):
        return None, None
    try:
        ext = file.filename.rsplit('.', 1)[1].lower()
        if ext in ['mp4', 'mov', 'webm', 'avi']:
            upload_result = cloudinary.uploader.upload_large(
                file.stream, resource_type="video", filename=file.filename,
                use_filename=True, unique_filename=True, chunk_size=6000000, timeout=600
            )
        elif ext == 'pdf':
            upload_result = cloudinary.uploader.upload(
                file, resource_type="image", format="pdf",
                use_filename=True, unique_filename=True, timeout=60
            )
        elif ext in ['zip', 'rar', '7z', 'tar', 'gz', 'txt', 'docx', 'xlsx', 'pptx']:
            upload_result = cloudinary.uploader.upload_large(
                file.stream, resource_type="raw", filename=file.filename,
                use_filename=True, unique_filename=True, chunk_size=6000000, timeout=600
            )
        else:
            upload_result = cloudinary.uploader.upload(
                file, resource_type="image", use_filename=True, unique_filename=True, timeout=60
            )
        return upload_result['secure_url'], file.filename
    except Exception as e:
        print(f"⚠️ Error subiendo el archivo '{file.filename}' a Cloudinary: {e}")
        return None, None


# 🖼️ FOTO DE PERFIL (autoservicio, ver /perfil) — deliberadamente más estricta que
# _subir_archivo_a_cloudinary: solo formatos de imagen real (nada de video/zip/docx como
# "foto de perfil") y un tamaño máximo bajo, porque esto lo sube cualquier usuario autenticado
# sin revisión de un admin.
EXTENSIONES_FOTO_PERFIL = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
TAMANO_MAXIMO_FOTO_PERFIL = 5 * 1024 * 1024  # 5 MB

def _subir_foto_perfil(file):
    """Sube una foto de perfil a Cloudinary, recortada a 400×400 centrada en el rostro si se
    detecta uno. Devuelve (url, None) si todo sale bien, o (None, mensaje_error) si el archivo
    no es una imagen válida, supera el tamaño máximo, o falla la subida."""
    if not file or not file.filename:
        return None, None
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in EXTENSIONES_FOTO_PERFIL:
        return None, 'Formato de imagen no permitido. Usa JPG, PNG, GIF o WEBP.'
    file.stream.seek(0, os.SEEK_END)
    tamano = file.stream.tell()
    file.stream.seek(0)
    if tamano > TAMANO_MAXIMO_FOTO_PERFIL:
        return None, 'La imagen no puede superar 5 MB.'
    try:
        upload_result = cloudinary.uploader.upload(
            file, resource_type="image", use_filename=True, unique_filename=True, timeout=60,
            transformation=[{'width': 400, 'height': 400, 'crop': 'fill', 'gravity': 'face'}]
        )
        return upload_result['secure_url'], None
    except Exception as e:
        print(f"⚠️ Error subiendo foto de perfil: {e}")
        return None, 'No se pudo subir la imagen. Intenta de nuevo.'


# ✍️ FIRMA DIGITAL (dibujada a mano o subida como imagen, ver static/js/firma-digital.js) — se
# captura al crear la cuenta (Gestión de Usuarios y el alta rápida desde Inventario) y se
# reutiliza al asignar un activo a esa persona (ver crear_activo/editar_activo), sin tener que
# volver a firmar cada vez. Llega desde el navegador como un data URL completo
# ('data:image/png;base64,...'), tanto si se dibujó en el <canvas> como si se subió un archivo
# (firma-digital.js convierte ambos casos al mismo formato antes de mandarlos).
TAMANO_MAXIMO_FIRMA = 3 * 1024 * 1024  # 3 MB — generoso para una firma o una foto de una hoja firmada.

def _subir_firma_desde_dataurl(data_url):
    """Sube una firma capturada en el navegador a Cloudinary. Devuelve (url, error): error es
    None si todo salió bien, o si 'data_url' viene vacío (la firma es opcional — no pasar nada
    no es un error)."""
    if not data_url:
        return None, None
    if not isinstance(data_url, str) or not data_url.startswith('data:image/'):
        return None, 'Formato de firma no reconocido.'
    try:
        _cabecera, contenido_b64 = data_url.split(',', 1)
    except ValueError:
        return None, 'Formato de firma no reconocido.'
    if (len(contenido_b64) * 3 / 4) > TAMANO_MAXIMO_FIRMA:
        return None, 'La firma no puede superar 3 MB.'
    try:
        upload_result = cloudinary.uploader.upload(
            data_url, resource_type="image", use_filename=False, unique_filename=True, timeout=60
        )
        return upload_result['secure_url'], None
    except Exception as e:
        print(f"⚠️ Error subiendo firma digital: {e}")
        return None, 'No se pudo guardar la firma. Intenta de nuevo.'


# 🖼️ FONDO DE LOGIN (panel de marca que rota junto al formulario de acceso, ver
# _fondo_login_activo/login_fondo_media) — deliberadamente más estricto que
# _subir_archivo_a_cloudinary: solo imagen o video corto (nada de PDF/zip/docx, esto se
# muestra en la pantalla PÚBLICA de inicio de sesión) y con un tope de tamaño bajo, para no
# volver lento el login — es lo primero que carga cualquier persona, cada vez.
EXTENSIONES_FONDO_LOGIN_IMAGEN = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
EXTENSIONES_FONDO_LOGIN_VIDEO = {'mp4', 'mov', 'webm'}
TAMANO_MAXIMO_FONDO_LOGIN = 60 * 1024 * 1024  # 60 MB: de sobra para un video corto de marca bien comprimido

def _subir_fondo_login(file):
    """Sube una imagen o video corto para el panel de marca del login. Devuelve
    (tipo, url, public_id, error) — tipo es 'imagen' o 'video'; si algo no es válido, los
    tres primeros vienen en None y 'error' trae el mensaje para mostrar."""
    if not file or not file.filename:
        return None, None, None, 'No se recibió ningún archivo.'
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext in EXTENSIONES_FONDO_LOGIN_VIDEO:
        tipo = 'video'
    elif ext in EXTENSIONES_FONDO_LOGIN_IMAGEN:
        tipo = 'imagen'
    else:
        return None, None, None, 'Formato no permitido. Usa una imagen (JPG, PNG, GIF, WEBP) o un video corto (MP4, MOV, WEBM).'
    file.stream.seek(0, os.SEEK_END)
    tamano = file.stream.tell()
    file.stream.seek(0)
    if tamano > TAMANO_MAXIMO_FONDO_LOGIN:
        return None, None, None, 'El archivo no puede superar 60 MB — entre más liviano, más rápido carga el login para todos.'
    try:
        if tipo == 'video':
            upload_result = cloudinary.uploader.upload_large(
                file.stream, resource_type="video", filename=file.filename,
                use_filename=True, unique_filename=True, chunk_size=6000000, timeout=600
            )
        else:
            upload_result = cloudinary.uploader.upload(
                file, resource_type="image", use_filename=True, unique_filename=True, timeout=60
            )
        return tipo, upload_result['secure_url'], upload_result.get('public_id'), None
    except Exception as e:
        print(f"⚠️ Error subiendo archivo de fondo de login '{file.filename}': {e}")
        return None, None, None, 'No se pudo subir el archivo a Cloudinary. Intenta de nuevo.'


def _marcar_comunicado_leido(comunicado_id, usuario):
    """Registra que 'usuario' ya vio este comunicado (muro de Comunicados o el fijado en la
    bienvenida). Si ya estaba marcado, no hace nada — la tabla tiene un UNIQUE(comunicado_id,
    usuario) para evitar duplicados; el intento repetido se ignora silenciosamente."""
    if not comunicado_id or not usuario:
        return
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        fecha = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d %H:%M:%S')
        if db_type == 'postgres':
            q = "INSERT INTO comunicados_leidos (comunicado_id, usuario, fecha) VALUES (%s, %s, %s) ON CONFLICT (comunicado_id, usuario) DO NOTHING"
        else:
            q = "INSERT OR IGNORE INTO comunicados_leidos (comunicado_id, usuario, fecha) VALUES (?, ?, ?)"
        cursor.execute(q, (comunicado_id, usuario, fecha))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error marcando comunicado {comunicado_id} como leído por '{usuario}': {e}")


def _conteo_lecturas_comunicados(ids_comunicados):
    """Devuelve {comunicado_id: cantidad_de_usuarios_que_lo_leyeron} para una lista de ids —
    usado en el muro de Comunicados para que soporte/admin vea cuántos ya lo vieron."""
    if not ids_comunicados:
        return {}
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        placeholders = ','.join(['%s' if db_type == 'postgres' else '?'] * len(ids_comunicados))
        q = f"SELECT comunicado_id, COUNT(*) FROM comunicados_leidos WHERE comunicado_id IN ({placeholders}) GROUP BY comunicado_id"
        cursor.execute(q, tuple(ids_comunicados))
        conteos = {r[0]: r[1] for r in cursor.fetchall()}
        conn.close()
        return conteos
    except Exception as e:
        print(f"⚠️ Error contando lecturas de comunicados: {e}")
        return {}


def _total_usuarios_activos():
    """Cantidad de cuentas activas — el denominador para mostrar 'X de Y ya lo leyeron'."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM usuarios WHERE COALESCE(estado, 'activo') = 'activo'")
        total = cursor.fetchone()[0]
        conn.close()
        return total or 0
    except Exception:
        return 0


# ⏳ Cuántas horas después de publicado se envía (una sola vez) el recordatorio automático de
# lectura pendiente a quien todavía no haya leído un comunicado activo.
HORAS_RECORDATORIO_LECTURA_COMUNICADO = 48


def _detalle_cumplimiento_comunicados():
    """Para cada comunicado activo (no archivado ni eliminado), cuántos usuarios activos ya lo
    leyeron y cuáles faltan — la base tanto del panel de Cumplimiento como de los recordatorios
    automáticos/manuales. Devuelve una lista ordenada del más reciente al más antiguo."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, titulo, nivel, fecha, autor, recordatorio_enviado_fecha FROM comunicados WHERE estado = 'activo' ORDER BY id DESC")
        comunicados_rows = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando comunicados para cumplimiento: {e}")
        conn.close()
        return []

    try:
        cursor.execute("SELECT usuario, COALESCE(nombre, usuario), correo FROM usuarios WHERE COALESCE(estado, 'activo') = 'activo'")
        activos = [{'usuario': r[0], 'nombre': r[1], 'correo': r[2]} for r in cursor.fetchall()]
    except Exception as e:
        print(f"⚠️ Error listando usuarios activos para cumplimiento: {e}")
        activos = []

    leidos_por_comunicado = {}
    try:
        ids = [r[0] for r in comunicados_rows]
        if ids:
            placeholders = ','.join(['%s' if db_type == 'postgres' else '?'] * len(ids))
            cursor.execute(f"SELECT comunicado_id, usuario FROM comunicados_leidos WHERE comunicado_id IN ({placeholders})", tuple(ids))
            for com_id, usuario in cursor.fetchall():
                leidos_por_comunicado.setdefault(com_id, set()).add(usuario)
    except Exception as e:
        print(f"⚠️ Error listando lecturas para cumplimiento: {e}")

    conn.close()

    total_activos = len(activos)
    resultado = []
    for com_id, titulo, nivel, fecha, autor, recordatorio_fecha in comunicados_rows:
        leidos_usuarios = leidos_por_comunicado.get(com_id, set())
        faltan = [u for u in activos if u['usuario'] not in leidos_usuarios]
        resultado.append({
            'id': com_id, 'titulo': titulo, 'nivel': nivel, 'fecha': fecha, 'autor': autor,
            'recordatorio_enviado_fecha': recordatorio_fecha,
            'leidos': len(leidos_usuarios), 'total': total_activos,
            'porcentaje': round((len(leidos_usuarios) / total_activos) * 100) if total_activos else 100,
            'faltan': faltan
        })
    return resultado


def _enviar_recordatorio_lectura(comunicado, marcar_enviado=True):
    """Notifica (campanita + correo) a cada usuario en comunicado['faltan'] que le falta leer
    ese comunicado, y opcionalmente marca recordatorio_enviado_fecha para que el chequeo
    automático no lo vuelva a enviar. Nunca debe tumbar la página que lo llama."""
    if not comunicado.get('faltan'):
        return 0
    try:
        url_comunicado = url_for('ver_comunicados')
        mensaje = f"📢 Te falta leer el comunicado '{comunicado['titulo']}'"
        asunto = f"[Arkiv] Recordatorio: te falta leer '{comunicado['titulo']}'"
        cuerpo = (
            f"Hola,\n\nTe falta leer el comunicado '{comunicado['titulo']}' publicado en el muro de "
            f"Novedades y Comunicados de Arkiv.\n\nIngresa a Arkiv para revisarlo.\n\n---\nArkiv"
        )
        for u in comunicado['faltan']:
            crear_notificacion(u['usuario'], mensaje, url=url_comunicado, tipo='comunicado')
            if u.get('correo'):
                threading.Thread(target=enviar_correo_ticket, args=(u['correo'], asunto, cuerpo)).start()

        if marcar_enviado:
            fecha_act = obtener_fecha_actual()
            conn, db_type = get_db()
            cursor = conn.cursor()
            q = "UPDATE comunicados SET recordatorio_enviado_fecha = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE comunicados SET recordatorio_enviado_fecha = ? WHERE id = ?"
            cursor.execute(q, (fecha_act, comunicado['id']))
            conn.commit()
            conn.close()
        return len(comunicado['faltan'])
    except Exception as e:
        print(f"⚠️ Error enviando recordatorio de lectura del comunicado {comunicado.get('id')}: {e}")
        return 0


def _revisar_recordatorios_lectura():
    """Recordatorio automático (una sola vez por comunicado): si pasaron más de
    HORAS_RECORDATORIO_LECTURA_COMUNICADO horas desde que se publicó un comunicado activo y
    todavía faltan usuarios por leerlo, se les avisa por campanita/correo. Se llama de forma
    perezosa (sin scheduler externo, ver _revisar_alertas_sla para el mismo patrón en Tickets)
    cada vez que el equipo de soporte visita el muro de Comunicados o el panel de Cumplimiento —
    nunca debe tumbar esa página si algo falla."""
    try:
        ahora = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)
        for c in _detalle_cumplimiento_comunicados():
            if c['recordatorio_enviado_fecha']:
                continue  # ya se envió (automático o manual) — no repetir
            fecha_pub = _parsear_fecha_ticket(c['fecha'])  # mismo formato 'YYYY-MM-DD HH:MM:SS'
            if not fecha_pub or (ahora - fecha_pub).total_seconds() < HORAS_RECORDATORIO_LECTURA_COMUNICADO * 3600:
                continue
            if not c['faltan']:
                continue
            _enviar_recordatorio_lectura(c)
    except Exception as e:
        print(f"⚠️ Error revisando recordatorios de lectura de comunicados: {e}")


# 📝 Editor de texto enriquecido (Quill): negrilla, cursiva, subrayado, resaltado de color y
# listas en la descripción de tickets, sus comentarios/respuestas, y el contenido de los
# Comunicados. Quill entrega HTML (p. ej. "<p><strong>hola</strong> <span style=\"background-
# color: rgb(255,255,0)\">urgente</span></p>"), así que ese HTML se limpia ANTES de guardarlo
# — nunca se confía en lo que llega del navegador — permitiendo solo un puñado de etiquetas y
# el estilo de color/resaltado, y quitando cualquier <script>, atributo onerror/onclick, o
# link javascript: que alguien intente colar.
_QUILL_TAGS_PERMITIDAS = ['p', 'br', 'strong', 'em', 'u', 's', 'span', 'ol', 'ul', 'li']
_QUILL_ATRIBUTOS_PERMITIDOS = {'span': ['style'], 'li': ['data-list']}
_QUILL_CSS_PERMITIDO = ['background-color', 'color']

if bleach and CSSSanitizer:
    _sanitizador_css = CSSSanitizer(allowed_css_properties=_QUILL_CSS_PERMITIDO)
else:
    _sanitizador_css = None


def _sanitizar_html_enriquecido(html_bruto):
    """Limpia el HTML que llega del editor de texto enriquecido antes de guardarlo. Si
    'bleach' no está disponible por alguna razón, se cae a texto plano (se escapan todas las
    etiquetas) — nunca se guarda HTML sin filtrar en la base de datos."""
    if not html_bruto:
        return html_bruto
    if not bleach:
        from markupsafe import escape
        return str(escape(html_bruto))
    return bleach.clean(
        html_bruto, tags=_QUILL_TAGS_PERMITIDAS, attributes=_QUILL_ATRIBUTOS_PERMITIDOS,
        css_sanitizer=_sanitizador_css, strip=True
    )


def _html_esta_vacio(html):
    """El editor de texto enriquecido (Quill) manda '<p><br></p>' cuando el usuario no
    escribió nada — sigue siendo una cadena "verdadera" en Python, así que una validación
    tipo `if contenido:` no detecta que en realidad está vacío. Esta función sí lo detecta,
    quitando TODAS las etiquetas y mirando si queda algo de texto real."""
    if not html:
        return True
    if bleach:
        texto_plano = bleach.clean(html, tags=[], attributes={}, strip=True)
    else:
        texto_plano = re.sub(r'<[^>]+>', '', html)
    return not texto_plano.strip()


def _html_a_texto_plano(html):
    """Igual que _html_esta_vacio pero devolviendo el texto en vez de un booleano — se usa
    para mandarle a la IA de clasificación (ver _clasificar_ticket_con_ia) solo el contenido
    real de la descripción, sin las etiquetas que agrega el editor Quill."""
    if not html:
        return ''
    if bleach:
        texto_plano = bleach.clean(html, tags=[], attributes={}, strip=True)
    else:
        texto_plano = re.sub(r'<[^>]+>', '', html)
    return texto_plano.strip()


def _mapa_nombres_usuarios():
    """Devuelve {usuario: nombre_para_mostrar} de TODOS los usuarios (activos o no), para
    resolver en bloque el alias/nombre real de quien publicó algo (comunicados, tickets, etc.)
    en vez de mostrar la cuenta de inicio de sesión cruda. Si un usuario no tiene 'nombre'
    guardado, o la cuenta ya no existe, se usa su propio usuario como respaldo — nunca deja el
    campo vacío."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT usuario, nombre FROM usuarios")
        mapa = {u[0]: (u[1] or u[0]) for u in cursor.fetchall()}
        conn.close()
        return mapa
    except Exception as e:
        print(f"⚠️ Error cargando el mapa de nombres de usuarios: {e}")
        return {}


def _nombre_para_mostrar(username, mapa_nombres):
    """Alias legible de 'username' según 'mapa_nombres' (ver _mapa_nombres_usuarios) — si no
    aparece ahí, se muestra el mismo username tal cual, para no dejar el campo vacío."""
    if not username:
        return username
    return mapa_nombres.get(username, username)


# 📞 Código de país por defecto para los números de contacto: la organización opera en
# Colombia y casi todos los teléfonos se guardan en formato local (10 dígitos, sin +57).
CODIGO_PAIS_WHATSAPP_DEFAULT = '57'


def _normalizar_telefono_whatsapp(telefono):
    """Convierte un teléfono guardado en cualquier formato local (con espacios, guiones,
    paréntesis, con o sin +57) al formato que exige wa.me: solo dígitos, con código de país.
    Devuelve None si no queda un número razonable (muy corto) para evitar generar un enlace
    de WhatsApp roto."""
    if not telefono:
        return None
    solo_digitos = re.sub(r'\D', '', telefono)
    if not solo_digitos:
        return None
    # Un celular colombiano local tiene 10 dígitos (empieza en 3) y no trae el código de país
    # todavía — se lo anteponemos. Si ya viene con el 57 delante (12 dígitos) se deja tal cual.
    if len(solo_digitos) == 10:
        solo_digitos = CODIGO_PAIS_WHATSAPP_DEFAULT + solo_digitos
    if len(solo_digitos) < 10:
        return None
    return solo_digitos


def _link_whatsapp_ticket(ticket, creador_info, nombre_agente):
    """Arma el enlace 'click to chat' de WhatsApp (wa.me) para contactar al solicitante de un
    ticket, con un mensaje ya redactado que el agente puede revisar y enviar desde WhatsApp
    Web/Desktop — Arkiv no manda el mensaje por sí solo, solo deja el borrador listo.

    El botón siempre se genera (para que el agente lo tenga disponible en todos los tickets,
    igual que en la referencia): si hay un teléfono de contacto válido, el chat se abre
    directamente con el solicitante; si no hay ninguno registrado, se arma el enlace sin
    número (wa.me/?text=...), que abre WhatsApp Web/Desktop con el mensaje ya redactado para
    que el agente elija el contacto manualmente."""
    telefono = _normalizar_telefono_whatsapp(ticket.get('telefono_contacto'))
    nombre_solicitante = (creador_info or {}).get('nombre') or ticket.get('creado_por') or 'usuario'
    # El mensaje nombra el tipo real del ticket (incidente/requerimiento) en vez de un genérico
    # "solicitud", igual a como lo redacta la referencia (Solvyx) con su "tu solicitud RQ-...").
    tipo_ticket = (ticket.get('tipo') or '').strip().lower()
    if tipo_ticket not in ('incidente', 'requerimiento'):
        tipo_ticket = 'solicitud'
    mensaje = (
        f"Hola {nombre_solicitante}, soy {nombre_agente} de Preventiva Salud SAS. "
        f"Te escribo respecto a tu {tipo_ticket} {ticket.get('codigo')} sobre \"{ticket.get('titulo')}\". "
        f"Estoy revisando el caso y quería coordinar contigo para resolverlo lo antes posible. "
        f"¿Cuándo te queda bien que conversemos?"
    )
    texto = urllib.parse.quote(mensaje)
    if telefono:
        return f"https://wa.me/{telefono}?text={texto}"
    return f"https://wa.me/?text={texto}"


# 🗓️ Motivos preestablecidos para justificar un corrimiento de la fecha límite de solución.
# Se muestran como opciones en el modal "Modificar fecha" del detalle del ticket, para que el
# equipo de soporte deje registrado explícitamente si el atraso depende de un tercero (proveedor,
# garantía, importación, etc.) y no necesariamente del analista o agente asignado.
MOTIVOS_MODIFICACION_SLA = [
    'Depende de un proveedor o tercero externo',
    'Se requiere repuesto, licencia o material no disponible',
    'Complejidad técnica mayor a la esperada',
    'El usuario solicitó reprogramar',
    'Otro',
]


def _codigo_ticket(tipo, id_ticket, fecha_creacion):
    """Código legible tipo 'IN-2026-000042' / 'RQ-2026-000042', solo para mostrar en la interfaz."""
    prefijo = 'RQ' if tipo == 'Requerimiento' else 'IN'
    anio = (fecha_creacion or '')[:4]
    if not anio.isdigit():
        anio = str(datetime.now().year)
    return f"{prefijo}-{anio}-{str(id_ticket).zfill(6)}"


def _subir_adjuntos_ticket(files):
    """Sube hasta MAX_ADJUNTOS_TICKET archivos (evidencias: fotos, capturas, PDFs, documentos)
    a Cloudinary y devuelve una lista de (url, nombre_original). Un archivo con extensión no
    permitida, o que falle al subir, simplemente se descarta: nunca debe tumbar la creación
    del ticket/comentario que lo acompaña."""
    subidos = []
    for file in (files or [])[:MAX_ADJUNTOS_TICKET]:
        if not file or not file.filename or not archivo_permitido(file.filename):
            continue
        try:
            ext = file.filename.rsplit('.', 1)[1].lower()
            if ext == 'pdf':
                upload_result = cloudinary.uploader.upload(
                    file, resource_type="image", format="pdf",
                    use_filename=True, unique_filename=True, timeout=60
                )
            elif ext in ['mp4', 'mov', 'webm', 'avi']:
                upload_result = cloudinary.uploader.upload(
                    file, resource_type="video",
                    use_filename=True, unique_filename=True, timeout=120
                )
            elif ext in ['zip', 'rar', '7z', 'tar', 'gz', 'txt', 'docx', 'xlsx', 'pptx']:
                upload_result = cloudinary.uploader.upload(
                    file, resource_type="raw",
                    use_filename=True, unique_filename=True, timeout=60
                )
            else:
                upload_result = cloudinary.uploader.upload(
                    file, resource_type="image",
                    use_filename=True, unique_filename=True, timeout=60
                )
            subidos.append((upload_result['secure_url'], file.filename))
        except Exception as e:
            print(f"⚠️ Error subiendo adjunto de ticket '{file.filename}': {e}")
    return subidos


def _guardar_adjuntos_ticket(cursor, db_type, ticket_id, comentario_id, subidos, usuario, fecha):
    q = "INSERT INTO tickets_adjuntos (ticket_id, comentario_id, url, nombre_original, subido_por, fecha) VALUES (%s, %s, %s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO tickets_adjuntos (ticket_id, comentario_id, url, nombre_original, subido_por, fecha) VALUES (?, ?, ?, ?, ?, ?)"
    for url, nombre in subidos:
        cursor.execute(q, (ticket_id, comentario_id, url, nombre, usuario, fecha))


def _guardar_adjuntos_inventario(cursor, db_type, activo_id, subidos, usuario, fecha):
    """Guarda en 'inventario_adjuntos' los archivos ya subidos a Cloudinary (mismo patrón que
    los adjuntos de tickets) para un activo del Inventario."""
    q = "INSERT INTO inventario_adjuntos (activo_id, url, nombre_original, subido_por, fecha) VALUES (%s, %s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO inventario_adjuntos (activo_id, url, nombre_original, subido_por, fecha) VALUES (?, ?, ?, ?, ?)"
    for url, nombre in subidos:
        cursor.execute(q, (activo_id, url, nombre, usuario, fecha))


@app.route('/tickets')
@login_required
def ver_tickets():
    q_estado = request.args.get('estado', '').strip()
    q_prioridad = request.args.get('prioridad', '').strip()
    q_categoria = request.args.get('categoria', '').strip()
    q_tipo = request.args.get('tipo', '').strip()
    q_area = request.args.get('area', '').strip()
    q_sede = request.args.get('sede', '').strip()
    q_busqueda = request.args.get('q', '').strip().lower()
    q_cumplimiento = request.args.get('cumplimiento', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()

    es_soporte = (session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)
    if es_soporte:
        # 🔔 Aviso perezoso de SLA por vencer/vencido — ver _revisar_alertas_sla(). Solo tiene
        # sentido correrlo cuando quien mira la lista es del equipo de soporte.
        _revisar_alertas_sla()

    categorias_config = _config_ticket_lista('categoria')
    areas_config = _config_ticket_lista('area')
    sedes_config = _config_ticket_lista('sede')
    nombres_categorias = [c['nombre'] for c in categorias_config] or CATEGORIAS_TICKET
    nombres_areas = [a['nombre'] for a in areas_config]
    nombres_sedes = [s['nombre'] for s in sedes_config]

    query = "SELECT id, titulo, descripcion, tipo, categoria, prioridad, estado, creado_por, asignado_a, fecha_creacion, fecha_actualizacion, sla_resolucion_limite, sla_resolucion_cumplida, area, sede, sla_pausado_desde FROM tickets WHERE COALESCE(eliminado, 0) = 0"
    params = []

    if not es_soporte:
        # Un usuario estándar solo ve las solicitudes que él mismo creó.
        query += " AND creado_por = %s" if db_type == 'postgres' else " AND creado_por = ?"
        params.append(session.get('username'))

    if q_estado in ESTADOS_TICKET:
        query += " AND estado = %s" if db_type == 'postgres' else " AND estado = ?"
        params.append(q_estado)
    if q_prioridad in PRIORIDADES_TICKET:
        query += " AND prioridad = %s" if db_type == 'postgres' else " AND prioridad = ?"
        params.append(q_prioridad)
    if q_categoria in nombres_categorias:
        query += " AND categoria = %s" if db_type == 'postgres' else " AND categoria = ?"
        params.append(q_categoria)
    if q_tipo in TIPOS_TICKET:
        query += " AND tipo = %s" if db_type == 'postgres' else " AND tipo = ?"
        params.append(q_tipo)
    if q_area in nombres_areas:
        query += " AND area = %s" if db_type == 'postgres' else " AND area = ?"
        params.append(q_area)
    if q_sede in nombres_sedes:
        query += " AND sede = %s" if db_type == 'postgres' else " AND sede = ?"
        params.append(q_sede)

    query += " ORDER BY (estado = 'Cerrado'), (estado = 'Resuelto'), id DESC"

    try:
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
    except Exception as e:
        print(f"Error consultando tickets: {e}")
        rows = []

    # 👤 Nombre completo por usuario (para mostrar "Creado por"/"Asignado a" en la tabla con
    # el nombre real de la persona en lugar de su usuario de inicio de sesión, un alias poco
    # reconocible como 'escobar' o 'tmira').
    cursor.execute("SELECT usuario, nombre FROM usuarios")
    nombres_usuarios = {u[0]: (u[1] or u[0]) for u in cursor.fetchall()}

    # 💻 Activos del inventario (no eliminados) para el selector "Activo relacionado" del
    # formulario de "Nueva Solicitud" — vincular el ticket a un equipo puntual del inventario.
    cursor.execute("SELECT id, nombre, tipo_activo FROM activos_inventario WHERE eliminado = 0 ORDER BY nombre ASC")
    activos_inventario = [{'id': a[0], 'nombre': a[1], 'tipo_activo': a[2]} for a in cursor.fetchall()]
    conn.close()

    tickets = []
    for r in rows:
        texto_full = f"{r[1]} {r[2]} {r[7]}".lower()
        if not q_busqueda or q_busqueda in texto_full:
            tipo_t = r[3] or 'Incidente'
            t = {
                'id': r[0], 'titulo': r[1], 'descripcion': r[2], 'tipo': tipo_t, 'categoria': r[4],
                'prioridad': r[5], 'estado': r[6], 'creado_por': r[7], 'asignado_a': r[8],
                'creado_por_nombre': nombres_usuarios.get(r[7], r[7]),
                'asignado_a_nombre': nombres_usuarios.get(r[8], r[8]) if r[8] else None,
                'fecha_creacion': r[9], 'fecha_actualizacion': r[10],
                'codigo': _codigo_ticket(tipo_t, r[0], r[9]),
                'sla_resolucion_limite': r[11], 'sla_resolucion_cumplida': r[12],
                'area': r[13], 'sede': r[14], 'sla_pausado_desde': r[15]
            }
            t['sla'] = _calcular_sla_ticket(t)
            t['cumplimiento'] = _bucket_cumplimiento_ticket(t)
            t['progreso'] = _progreso_ticket(t)
            tickets.append(t)

    # Conteos para las pestañas de cumplimiento de SLA (Vigentes/Próximos a vencer/Vencidos/
    # Cerrados), calculados ANTES de aplicar el filtro de pestaña, para que cada pestaña
    # muestre cuántos tickets hay en las demás sin perder los otros filtros ya aplicados.
    conteos_cumplimiento = {'vigente': 0, 'proximo_a_vencer': 0, 'vencido': 0, 'cerrado': 0}
    for t in tickets:
        conteos_cumplimiento[t['cumplimiento']] = conteos_cumplimiento.get(t['cumplimiento'], 0) + 1

    # 🔗 Tickets ya Resueltos/Cerrados que se pueden asociar a una solicitud nueva (ver
    # crear_ticket) — se arma ANTES del filtro de pestaña de arriba, sobre la misma lista que ya
    # respeta el permiso (un usuario estándar solo ve los suyos, ver filtro 'creado_por' arriba).
    tickets_cerrados = [
        {'id': t['id'], 'codigo': t['codigo'], 'titulo': t['titulo']}
        for t in tickets if t['estado'] in ('Resuelto', 'Cerrado')
    ]

    # 📂 La pestaña por defecto ("Activos") ya NO incluye Resueltos/Cerrados/Cancelados — esos
    # quedan disponibles únicamente en la pestaña "Cerrados" (el historial). Antes se mostraban
    # todos juntos (solo se ordenaban al final); ahora hay que entrar a "Cerrados" a propósito
    # para verlos, igual que se pidió en el seguimiento de Solvyx.
    total_tickets = len([t for t in tickets if t['cumplimiento'] != 'cerrado'])

    if q_cumplimiento in conteos_cumplimiento:
        tickets = [t for t in tickets if t['cumplimiento'] == q_cumplimiento]
    else:
        tickets = [t for t in tickets if t['cumplimiento'] != 'cerrado']

    # 📞 Prellena el número de contacto de "Nueva Solicitud" con el que el usuario ya tiene
    # registrado en su perfil (si tiene uno) — sigue siendo editable en el formulario, por si
    # para este ticket en particular hay que dejar otro número (p. ej. de un tercero).
    perfil_actual = _info_usuario(session.get('username'))
    telefono_usuario_actual = perfil_actual['telefono'] if perfil_actual else None

    # 🙋 Lista de usuarios para "¿Para quién es esta solicitud?" (ver crear_ticket) — solo la
    # necesita el equipo de soporte, para poder subir un caso a nombre de otra persona.
    usuarios_lista = []
    if es_soporte:
        usuarios_lista = sorted(
            [{'usuario': u, 'nombre': n} for u, n in nombres_usuarios.items() if u != session.get('username')],
            key=lambda x: x['nombre'].lower()
        )

    return render_template(
        'tickets.html', tickets=tickets, es_soporte=es_soporte,
        categorias=nombres_categorias, prioridades=PRIORIDADES_TICKET, estados=ESTADOS_TICKET,
        tipos=TIPOS_TICKET, tipos_info=TIPOS_TICKET_INFO,
        areas=nombres_areas, sedes=nombres_sedes, activos_inventario=activos_inventario,
        tickets_cerrados=tickets_cerrados, usuarios_lista=usuarios_lista,
        q_estado=q_estado, q_prioridad=q_prioridad, q_categoria=q_categoria, q_tipo=q_tipo, q_busqueda=q_busqueda,
        q_area=q_area, q_sede=q_sede,
        q_cumplimiento=q_cumplimiento, conteos_cumplimiento=conteos_cumplimiento, total_tickets=total_tickets,
        telefono_usuario_actual=telefono_usuario_actual, plantillas=_plantillas_ticket_activas(),
        ia_disponible=_ia_clasificacion_disponible()
    )


# 🤖 CLASIFICACIÓN DE TICKETS CON IA (Anthropic Claude) --------------------------------------
# Sugiere Tipo / Categoría / Prioridad a partir de lo que el usuario ya escribió en "Nueva
# Solicitud", antes de enviarla — igual que el "Clasificar con IA" de Solvyx que Tomás pidió
# imitar. Deliberadamente NO sugiere "Área": en Arkiv esa es el área física/organizacional de
# quien reporta (no una categoría de soporte), así que no tiene sentido que la IA la adivine.
# Requiere la variable de entorno ANTHROPIC_API_KEY en Render (ver _ia_clasificacion_disponible);
# sin ella, el botón simplemente no aparece — nunca bloquea la creación manual del ticket.
MODELO_IA_CLASIFICACION_TICKET = 'claude-haiku-4-5-20251001'


def _ia_clasificacion_disponible():
    return bool(anthropic) and bool(os.environ.get('ANTHROPIC_API_KEY'))


def _clasificar_ticket_con_ia(titulo, descripcion_texto, categorias_disponibles):
    """Le pide a Claude Haiku que sugiera tipo/categoría/prioridad para un ticket nuevo, a
    partir del título y la descripción en texto plano (ya sin el HTML del editor). Devuelve un
    dict {tipo, categoria, prioridad, confianza (0-100), explicacion} o None si la IA no está
    disponible o la respuesta no se pudo interpretar — nunca lanza, para que un problema con el
    proveedor de IA no le impida a nadie crear el ticket manualmente."""
    if not _ia_clasificacion_disponible():
        return None
    try:
        cliente = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
        categorias_txt = ', '.join(f'"{c}"' for c in categorias_disponibles)
        prompt = (
            "Eres un asistente que clasifica solicitudes de soporte TI para una IPS de salud "
            "en Colombia. Con base en el título y la descripción de abajo, responde ÚNICAMENTE "
            "un objeto JSON (sin texto adicional, sin bloque de código markdown) con "
            "exactamente estas claves:\n"
            '- "tipo": uno de "Incidente" o "Requerimiento" (Incidente = algo que dejó de '
            'funcionar o está fallando; Requerimiento = una solicitud nueva, como crear un '
            "acceso o instalar algo)\n"
            f'- "categoria": exactamente uno de estos valores, sin inventar otros: [{categorias_txt}]\n'
            '- "prioridad": uno de "Baja", "Media", "Alta" o "Urgente" (Urgente solo si '
            "impide totalmente trabajar o afecta la atención de pacientes)\n"
            '- "confianza": un número entero de 0 a 100\n'
            '- "explicacion": una sola frase corta en español explicando el porqué\n\n'
            f"Título: {titulo}\n"
            f"Descripción: {descripcion_texto}\n"
        )
        respuesta = cliente.messages.create(
            model=MODELO_IA_CLASIFICACION_TICKET,
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        texto = (respuesta.content[0].text or '').strip()
        # Por si Claude envuelve el JSON en ```json ... ``` a pesar de la instrucción de no
        # hacerlo — se limpia antes de parsear en vez de fallar.
        if texto.startswith('```'):
            texto = texto.strip('`')
            if texto.lower().startswith('json'):
                texto = texto[4:]
        datos = json.loads(texto.strip())

        tipo = datos.get('tipo') if datos.get('tipo') in TIPOS_TICKET else 'Incidente'
        categoria = datos.get('categoria')
        if categoria not in categorias_disponibles:
            categoria = categorias_disponibles[0] if categorias_disponibles else 'Otro'
        prioridad = datos.get('prioridad') if datos.get('prioridad') in PRIORIDADES_TICKET else 'Media'
        try:
            confianza = max(0, min(100, int(datos.get('confianza', 0))))
        except (TypeError, ValueError):
            confianza = 0
        explicacion = str(datos.get('explicacion', '')).strip()[:300]

        return {
            'tipo': tipo, 'categoria': categoria, 'prioridad': prioridad,
            'confianza': confianza, 'explicacion': explicacion,
        }
    except Exception as e:
        print(f"⚠️ Error clasificando ticket con IA: {e}")
        return None


@app.route('/tickets/clasificar_ia', methods=['POST'])
@login_required
def clasificar_ticket_ia():
    """Endpoint AJAX consultado por el botón 'Clasificar con IA' del modal de Nueva Solicitud
    (ver static/js/tickets.js). Devuelve JSON — nunca un 500 — para que si la IA falla, el
    usuario simplemente siga llenando el formulario a mano."""
    titulo = request.form.get('titulo', '').strip()
    descripcion_texto = _html_a_texto_plano(request.form.get('descripcion', ''))

    if len(titulo) < 3 or len(descripcion_texto) < 10:
        return jsonify({'ok': False, 'error': 'Escribe un título y una descripción un poco más completos.'}), 400

    nombres_categorias = [c['nombre'] for c in _config_ticket_lista('categoria')] or CATEGORIAS_TICKET
    resultado = _clasificar_ticket_con_ia(titulo, descripcion_texto, nombres_categorias)
    if not resultado:
        return jsonify({'ok': False, 'error': 'La IA no está disponible en este momento. Puedes clasificar el ticket manualmente.'}), 503
    return jsonify({'ok': True, **resultado})


@app.route('/tickets/crear', methods=['POST'])
@login_required
def crear_ticket():
    titulo = request.form.get('titulo', '').strip()
    # 📝 La descripción llega como HTML del editor de texto enriquecido (Quill) — se limpia
    # ANTES de guardarla, para no confiar en lo que manda el navegador.
    descripcion = _sanitizar_html_enriquecido(request.form.get('descripcion', '').strip())
    tipo = request.form.get('tipo', 'Incidente').strip()
    categoria = request.form.get('categoria', 'Otro').strip()
    prioridad = request.form.get('prioridad', 'Media').strip()
    area = request.form.get('area', '').strip()
    sede = request.form.get('sede', '').strip()
    telefono_contacto = request.form.get('telefono_contacto', '').strip() or None
    activo_id_raw = request.form.get('activo_id', '').strip()
    ticket_relacionado_id_raw = request.form.get('ticket_relacionado_id', '').strip()
    solicitante_real_raw = request.form.get('solicitante_real', '').strip()

    nombres_categorias = [c['nombre'] for c in _config_ticket_lista('categoria')] or CATEGORIAS_TICKET
    nombres_areas = [a['nombre'] for a in _config_ticket_lista('area')]
    nombres_sedes = [s['nombre'] for s in _config_ticket_lista('sede')]

    if tipo not in TIPOS_TICKET:
        tipo = 'Incidente'
    if categoria not in nombres_categorias:
        categoria = 'Otro'
    if prioridad not in PRIORIDADES_TICKET:
        prioridad = 'Media'
    # Área y Sede son opcionales: si el valor enviado no corresponde a una configuración
    # activa (o el equipo de soporte todavía no ha configurado ninguna), se guarda vacío.
    area = area if area in nombres_areas else None
    sede = sede if sede in nombres_sedes else None

    if titulo and descripcion and not _html_esta_vacio(descripcion):
        fecha_act = obtener_fecha_actual()
        usuario = session.get('username')
        archivos_subidos = _subir_adjuntos_ticket(request.files.getlist('adjuntos'))

        # 💻 Activo relacionado (opcional): validamos que exista y no esté eliminado antes de
        # vincularlo, para no dejar un ticket apuntando a un activo_id inválido.
        activo_id = None
        if activo_id_raw.isdigit():
            conn_chk, db_type_chk = get_db()
            cur_chk = conn_chk.cursor()
            ph_chk = '%s' if db_type_chk == 'postgres' else '?'
            cur_chk.execute(f"SELECT id FROM activos_inventario WHERE id = {ph_chk} AND eliminado = 0", (int(activo_id_raw),))
            if cur_chk.fetchone():
                activo_id = int(activo_id_raw)
            conn_chk.close()

        # 🔗 Ticket relacionado (opcional): un caso ya Resuelto/Cerrado que el usuario quiere
        # dejar referenciado (p. ej. un duplicado, o un caso nuevo que continúa uno anterior).
        # Se valida que exista, no esté eliminado y ya esté cerrado, igual que con activo_id. Un
        # usuario estándar solo puede asociar tickets QUE ÉL MISMO CREÓ (evita que referencie —
        # y así deje visible el título de— un ticket ajeno armando el POST a mano); el equipo de
        # soporte, que ya puede ver cualquier ticket, puede asociar cualquiera.
        es_soporte_creador = (session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)
        ticket_relacionado_id = None
        if ticket_relacionado_id_raw.isdigit():
            conn_chk2, db_type_chk2 = get_db()
            cur_chk2 = conn_chk2.cursor()
            ph_chk2 = '%s' if db_type_chk2 == 'postgres' else '?'
            if es_soporte_creador:
                cur_chk2.execute(
                    f"SELECT id FROM tickets WHERE id = {ph_chk2} AND COALESCE(eliminado, 0) = 0 "
                    f"AND estado IN ('Resuelto', 'Cerrado')", (int(ticket_relacionado_id_raw),)
                )
            else:
                cur_chk2.execute(
                    f"SELECT id FROM tickets WHERE id = {ph_chk2} AND COALESCE(eliminado, 0) = 0 "
                    f"AND estado IN ('Resuelto', 'Cerrado') AND creado_por = {ph_chk2}",
                    (int(ticket_relacionado_id_raw), session.get('username'))
                )
            if cur_chk2.fetchone():
                ticket_relacionado_id = int(ticket_relacionado_id_raw)
            conn_chk2.close()

        # 🙋 Solicitante real (opcional, solo lo puede fijar el equipo de soporte): cuando un
        # agente sube el caso a nombre de otro usuario (PQRS), este es quien debe poder
        # calificar el servicio al final — no el agente que lo está resolviendo. Se valida que
        # sea un usuario real y distinto de quien está creando el ticket.
        solicitante_real = None
        if es_soporte_creador and solicitante_real_raw and solicitante_real_raw != session.get('username'):
            conn_chk3, db_type_chk3 = get_db()
            cur_chk3 = conn_chk3.cursor()
            ph_chk3 = '%s' if db_type_chk3 == 'postgres' else '?'
            cur_chk3.execute(f"SELECT usuario FROM usuarios WHERE usuario = {ph_chk3}", (solicitante_real_raw,))
            if cur_chk3.fetchone():
                solicitante_real = solicitante_real_raw
            conn_chk3.close()

        horas_sla = SLA_HORAS_POR_PRIORIDAD.get(prioridad, SLA_HORAS_POR_PRIORIDAD['Media'])
        sla_respuesta_limite = _calcular_limite_sla(fecha_act, horas_sla['respuesta'])
        sla_resolucion_limite = _calcular_limite_sla(fecha_act, horas_sla['resolucion'])

        # 🤖 Asignación automática: si el equipo de soporte configuró un responsable para
        # esta categoría (o, si no hay, para el área) en /tickets/configuracion, el ticket
        # nace ya asignado a esa persona en vez de quedar "Sin asignar" hasta que alguien lo
        # tome manualmente. Prioridad: categoría primero, área como respaldo.
        asignado_auto = None
        for cfg in _config_ticket_lista('categoria'):
            if cfg['nombre'] == categoria and cfg.get('responsable'):
                asignado_auto = cfg['responsable']
                break
        if not asignado_auto and area:
            for cfg in _config_ticket_lista('area'):
                if cfg['nombre'] == area and cfg.get('responsable'):
                    asignado_auto = cfg['responsable']
                    break

        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            if db_type == 'postgres':
                q_ins = "INSERT INTO tickets (titulo, descripcion, tipo, categoria, prioridad, estado, creado_por, asignado_a, fecha_creacion, fecha_actualizacion, sla_respuesta_limite, sla_resolucion_limite, sla_modificaciones, area, sede, telefono_contacto, activo_id, ticket_relacionado_id, solicitante_real) VALUES (%s, %s, %s, %s, %s, 'Abierto', %s, %s, %s, %s, %s, %s, 0, %s, %s, %s, %s, %s, %s) RETURNING id"
                cursor.execute(q_ins, (titulo, descripcion, tipo, categoria, prioridad, usuario, asignado_auto, fecha_act, fecha_act, sla_respuesta_limite, sla_resolucion_limite, area, sede, telefono_contacto, activo_id, ticket_relacionado_id, solicitante_real))
                nuevo_id = cursor.fetchone()[0]
            else:
                q_ins = "INSERT INTO tickets (titulo, descripcion, tipo, categoria, prioridad, estado, creado_por, asignado_a, fecha_creacion, fecha_actualizacion, sla_respuesta_limite, sla_resolucion_limite, sla_modificaciones, area, sede, telefono_contacto, activo_id, ticket_relacionado_id, solicitante_real) VALUES (?, ?, ?, ?, ?, 'Abierto', ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?)"
                cursor.execute(q_ins, (titulo, descripcion, tipo, categoria, prioridad, usuario, asignado_auto, fecha_act, fecha_act, sla_respuesta_limite, sla_resolucion_limite, area, sede, telefono_contacto, activo_id, ticket_relacionado_id, solicitante_real))
                nuevo_id = cursor.lastrowid

            if archivos_subidos:
                _guardar_adjuntos_ticket(cursor, db_type, nuevo_id, None, archivos_subidos, usuario, fecha_act)

            conn.commit()
            detalle_log = f"Nuevo ticket [{tipo}]: '{titulo}' [{categoria} / {prioridad}]"
            if archivos_subidos:
                detalle_log += f" — {len(archivos_subidos)} adjunto(s)"
            if asignado_auto:
                detalle_log += f" — auto-asignado a {asignado_auto}"
            registrar_log(usuario, "Solicitud de Soporte Creada", detalle_log)

            # 📧🔔 Avisamos de la solicitud nueva: confirmación al que la creó, y aviso a quien
            # deba atenderla — si quedó auto-asignada a alguien puntual (arriba), solo a esa
            # persona recibe el correo/notificación (para no saturar a todo el equipo con algo
            # que ya tiene dueño); si no quedó asignada, se avisa a todo el equipo de soporte
            # activo (admins + agentes), igual que pidió Tomas.
            codigo = _codigo_ticket(tipo, nuevo_id, fecha_act)
            url_ticket = url_for('ver_ticket', ticket_id=nuevo_id)

            correo_creador = _correo_de_usuario(usuario)
            if correo_creador:
                asunto_creador = f"[Arkiv] Recibimos tu solicitud {codigo}"
                cuerpo_creador = (
                    f"Hola,\n\nRecibimos tu solicitud de soporte {codigo} ('{titulo}').\n\n"
                    f"Puedes ver el detalle y su avance ingresando a Arkiv, módulo Solicitudes TI.\n\n"
                    f"---\nEquipo de Soporte TI - Arkiv"
                )
                threading.Thread(target=enviar_correo_ticket, args=(correo_creador, asunto_creador, cuerpo_creador)).start()

            asunto_soporte = f"[Arkiv] Nueva solicitud {codigo}: '{titulo}'"
            cuerpo_soporte = (
                f"Se creó una nueva solicitud de soporte {codigo} ('{titulo}') — {categoria} / prioridad {prioridad}.\n"
                f"Creada por: {usuario}.\n\nPuedes verla ingresando a Arkiv, módulo Solicitudes TI.\n\n---\nArkiv"
            )
            if asignado_auto:
                correo_asignado = _correo_de_usuario(asignado_auto)
                if correo_asignado:
                    threading.Thread(target=enviar_correo_ticket, args=(correo_asignado, asunto_soporte, cuerpo_soporte)).start()
                crear_notificacion(asignado_auto, f"Nueva solicitud {codigo} asignada a ti: '{titulo}'", url=url_ticket)
            else:
                equipo = _equipo_soporte_activo()
                for miembro in equipo:
                    if miembro['correo']:
                        threading.Thread(target=enviar_correo_ticket, args=(miembro['correo'], asunto_soporte, cuerpo_soporte)).start()
                crear_notificacion_para_varios([m['usuario'] for m in equipo], f"Nueva solicitud {codigo}: '{titulo}'", url=url_ticket)
        except Exception as e:
            conn.rollback()
            print(f"Error creando ticket: {e}")
        conn.close()

    return redirect(url_for('ver_tickets'))


@app.route('/tickets/<int:ticket_id>')
@login_required
def ver_ticket(ticket_id):
    conn, db_type = get_db()
    cursor = conn.cursor()

    q_sel = "SELECT id, titulo, descripcion, tipo, categoria, prioridad, estado, creado_por, asignado_a, fecha_creacion, fecha_actualizacion, sla_respuesta_limite, sla_resolucion_limite, sla_respuesta_cumplida, sla_resolucion_cumplida, sla_modificaciones, calificacion, calificacion_fecha, area, sede, telefono_contacto, activo_id, sla_pausado_desde, ticket_relacionado_id, solicitante_real FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT id, titulo, descripcion, tipo, categoria, prioridad, estado, creado_por, asignado_a, fecha_creacion, fecha_actualizacion, sla_respuesta_limite, sla_resolucion_limite, sla_respuesta_cumplida, sla_resolucion_cumplida, sla_modificaciones, calificacion, calificacion_fecha, area, sede, telefono_contacto, activo_id, sla_pausado_desde, ticket_relacionado_id, solicitante_real FROM tickets WHERE id = ?"
    cursor.execute(q_sel, (ticket_id,))
    row = cursor.fetchone()

    if not row or not _puede_ver_ticket(row[7]):
        conn.close()
        return redirect(url_for('ver_tickets'))

    es_soporte = (session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)
    tipo_t = row[3] or 'Incidente'

    ticket = {
        'id': row[0], 'titulo': row[1], 'descripcion': row[2], 'tipo': tipo_t, 'categoria': row[4],
        'prioridad': row[5], 'estado': row[6], 'creado_por': row[7], 'asignado_a': row[8],
        'fecha_creacion': row[9], 'fecha_actualizacion': row[10],
        'codigo': _codigo_ticket(tipo_t, row[0], row[9]),
        'sla_respuesta_limite': row[11], 'sla_resolucion_limite': row[12],
        'sla_respuesta_cumplida': row[13], 'sla_resolucion_cumplida': row[14],
        'sla_modificaciones': row[15] or 0,
        'calificacion': row[16], 'calificacion_fecha': row[17],
        'area': row[18], 'sede': row[19], 'telefono_contacto': row[20], 'activo_id': row[21],
        'sla_pausado_desde': row[22], 'ticket_relacionado_id': row[23], 'solicitante_real': row[24]
    }
    ticket['sla'] = _calcular_sla_ticket(ticket)
    # 🙋 Beneficiario real de la solicitud: normalmente quien la creó, salvo que un agente la
    # haya subido a nombre de otra persona (ver crear_ticket) — ese es quien debe poder
    # calificar el servicio al cierre, nunca el agente que la resuelve (ver calificar_ticket()).
    ticket['beneficiario'] = ticket['solicitante_real'] or ticket['creado_por']

    # 💻 Si el ticket quedó vinculado a un activo del inventario, se muestra en el detalle
    # (nombre, tipo y estado actual) con enlace directo a Inventario.
    ticket['activo'] = None
    if ticket['activo_id']:
        cursor.execute(
            "SELECT id, nombre, tipo_activo, estado FROM activos_inventario WHERE id = %s" if db_type == 'postgres' else "SELECT id, nombre, tipo_activo, estado FROM activos_inventario WHERE id = ?",
            (ticket['activo_id'],)
        )
        fila_activo = cursor.fetchone()
        if fila_activo:
            ticket['activo'] = {'id': fila_activo[0], 'nombre': fila_activo[1], 'tipo_activo': fila_activo[2], 'estado': fila_activo[3]}

    # 🔗 Si el usuario asoció este ticket a uno ya Resuelto/Cerrado (ver crear_ticket), se
    # muestra aquí con enlace directo — puramente informativo.
    ticket['relacionado'] = None
    if ticket['ticket_relacionado_id']:
        cursor.execute(
            "SELECT id, titulo, tipo, estado, fecha_creacion FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT id, titulo, tipo, estado, fecha_creacion FROM tickets WHERE id = ?",
            (ticket['ticket_relacionado_id'],)
        )
        fila_rel = cursor.fetchone()
        if fila_rel:
            ticket['relacionado'] = {
                'id': fila_rel[0], 'titulo': fila_rel[1], 'estado': fila_rel[3],
                'codigo': _codigo_ticket(fila_rel[2] or 'Incidente', fila_rel[0], fila_rel[4])
            }

    # 🧑 La mayor cantidad de información posible sobre quién levantó la solicitud: nombre
    # completo, correo y teléfono registrados en su perfil de Arkiv (si los tiene). Si el
    # ticket no trae un teléfono de contacto propio, se usa el del perfil como respaldo.
    creador_info = _info_usuario(ticket['creado_por']) or {}
    if not ticket['telefono_contacto']:
        ticket['telefono_contacto'] = creador_info.get('telefono')

    # 👤 El nombre completo del agente asignado (no su usuario de inicio de sesión) se usa
    # para mostrar "Asignado a" de forma legible, igual que en el desplegable de arriba.
    if ticket['asignado_a']:
        info_asignado = _info_usuario(ticket['asignado_a'])
        ticket['asignado_a_nombre'] = (info_asignado or {}).get('nombre') or ticket['asignado_a']
    else:
        ticket['asignado_a_nombre'] = None

    # Adjuntos cargados junto con la solicitud original (no pertenecen a ningún comentario).
    q_adj_ticket = "SELECT url, nombre_original FROM tickets_adjuntos WHERE ticket_id = %s AND comentario_id IS NULL ORDER BY id ASC" if db_type == 'postgres' else "SELECT url, nombre_original FROM tickets_adjuntos WHERE ticket_id = ? AND comentario_id IS NULL ORDER BY id ASC"
    cursor.execute(q_adj_ticket, (ticket_id,))
    ticket['adjuntos'] = [{'url': a[0], 'nombre_original': a[1]} for a in cursor.fetchall()]

    if es_soporte:
        q_com = "SELECT id, autor, mensaje, tipo, fecha FROM tickets_comentarios WHERE ticket_id = %s ORDER BY id ASC" if db_type == 'postgres' else "SELECT id, autor, mensaje, tipo, fecha FROM tickets_comentarios WHERE ticket_id = ? ORDER BY id ASC"
    else:
        # Un usuario estándar no debe ver las notas internas que el equipo de soporte deja
        # únicamente para coordinarse entre ellos.
        q_com = "SELECT id, autor, mensaje, tipo, fecha FROM tickets_comentarios WHERE ticket_id = %s AND tipo != 'interno' ORDER BY id ASC" if db_type == 'postgres' else "SELECT id, autor, mensaje, tipo, fecha FROM tickets_comentarios WHERE ticket_id = ? AND tipo != 'interno' ORDER BY id ASC"
    cursor.execute(q_com, (ticket_id,))
    comentarios = [{'id': c[0], 'autor': c[1], 'mensaje': c[2], 'tipo': c[3], 'fecha': c[4], 'adjuntos': []} for c in cursor.fetchall()]

    ids_com = [c['id'] for c in comentarios]
    if ids_com:
        placeholder = '%s' if db_type == 'postgres' else '?'
        placeholders = ','.join([placeholder] * len(ids_com))
        q_adj_com = f"SELECT comentario_id, url, nombre_original FROM tickets_adjuntos WHERE comentario_id IN ({placeholders}) ORDER BY id ASC"
        cursor.execute(q_adj_com, tuple(ids_com))
        adjuntos_por_com = {}
        for com_id, url, nombre in cursor.fetchall():
            adjuntos_por_com.setdefault(com_id, []).append({'url': url, 'nombre_original': nombre})
        for c in comentarios:
            c['adjuntos'] = adjuntos_por_com.get(c['id'], [])

    agentes = []
    if es_soporte:
        # 👤 Se muestra el nombre completo de cada agente/admin en el desplegable "Asignar a"
        # (no el usuario de inicio de sesión, que suele ser un alias técnico poco reconocible).
        q_ag = "SELECT usuario, nombre FROM usuarios WHERE rol IN ('admin', 'agente') AND COALESCE(estado, 'activo') = 'activo' ORDER BY usuario ASC"
        cursor.execute(q_ag)
        agentes = [{'usuario': a[0], 'nombre': a[1] or a[0]} for a in cursor.fetchall()]

    # ✅ Tareas del ticket (subtareas internas para repartir el trabajo del caso) — solo
    # tienen sentido para el equipo operativo, igual que los comentarios internos.
    tareas = []
    if es_soporte:
        q_tareas = "SELECT id, asunto, descripcion, responsable, estado, fecha_limite, creado_por, fecha_creacion, fecha_completada FROM tickets_tareas WHERE ticket_id = %s ORDER BY (estado = 'completada') ASC, id ASC" if db_type == 'postgres' else "SELECT id, asunto, descripcion, responsable, estado, fecha_limite, creado_por, fecha_creacion, fecha_completada FROM tickets_tareas WHERE ticket_id = ? ORDER BY (estado = 'completada') ASC, id ASC"
        cursor.execute(q_tareas, (ticket_id,))
        nombres_agentes = {a['usuario']: a['nombre'] for a in agentes}
        for t in cursor.fetchall():
            responsable_usuario = t[3]
            tareas.append({
                'id': t[0], 'asunto': t[1], 'descripcion': t[2], 'responsable': responsable_usuario,
                'responsable_nombre': nombres_agentes.get(responsable_usuario, responsable_usuario),
                'estado': t[4] or 'pendiente', 'fecha_limite': t[5], 'creado_por': t[6],
                'fecha_creacion': t[7], 'fecha_completada': t[8],
                'vencida': bool(t[5] and t[4] not in ('completada', 'cancelada') and t[5] < datetime.now().strftime('%Y-%m-%d')),
            })

    conn.close()

    # 💬 Enlace "click to chat" de WhatsApp para contactar al solicitante desde el detalle del
    # ticket (solo el equipo de soporte lo ve — es una herramienta de gestión, no algo que el
    # propio solicitante necesite). Arkiv NO envía el mensaje: solo abre WhatsApp Web/Desktop
    # con un borrador ya redactado, el agente decide si lo edita y lo envía.
    whatsapp_url = None
    whatsapp_tiene_numero = False
    if es_soporte:
        perfil_agente = _info_usuario(session.get('username'))
        nombre_agente = (perfil_agente or {}).get('nombre') or session.get('username') or 'Soporte TI'
        whatsapp_url = _link_whatsapp_ticket(ticket, creador_info, nombre_agente)
        whatsapp_tiene_numero = bool(_normalizar_telefono_whatsapp(ticket.get('telefono_contacto')))

    return render_template(
        'ticket_detalle.html', ticket=ticket, comentarios=comentarios,
        es_soporte=es_soporte, agentes=agentes, creador_info=creador_info,
        estados=_estados_disponibles_ticket(ticket['estado']), prioridades=PRIORIDADES_TICKET,
        whatsapp_tiene_numero=whatsapp_tiene_numero,
        max_modificaciones_sla=MAX_MODIFICACIONES_SLA, motivos_sla=MOTIVOS_MODIFICACION_SLA,
        calificacion_max=CALIFICACION_MAX, session_username=session.get('username'),
        whatsapp_url=whatsapp_url, tareas=tareas
    )


# ✅ TAREAS DE UN TICKET ---------------------------------------------------------------------
# Subtareas internas para que el equipo de soporte reparta el trabajo de un mismo caso entre
# varias personas (por ejemplo: "Revisar con el proveedor", "Actualizar el servidor",
# "Confirmar con el usuario"), inspiradas en la pestaña "Tareas" de Aranda Service Desk pero
# simplificadas a lo que Arkiv necesita realmente. Son de uso exclusivamente interno: un
# usuario Estándar nunca las ve ni las gestiona, igual que pasa con los comentarios internos.

@app.route('/tickets/<int:ticket_id>/tareas/crear', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_tarea_ticket(ticket_id):
    asunto = (request.form.get('asunto') or '').strip()
    descripcion = (request.form.get('descripcion') or '').strip() or None
    responsable = (request.form.get('responsable') or '').strip() or None
    fecha_limite = (request.form.get('fecha_limite') or '').strip() or None

    if not asunto:
        return redirect(url_for('ver_ticket', ticket_id=ticket_id))

    conn, db_type = get_db()
    cursor = conn.cursor()
    # El ticket debe existir (evita crear una tarea "huérfana" con un ID inventado en la URL).
    q_existe = "SELECT id FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT id FROM tickets WHERE id = ?"
    cursor.execute(q_existe, (ticket_id,))
    if not cursor.fetchone():
        conn.close()
        return redirect(url_for('ver_tickets'))
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        q_ins = ("INSERT INTO tickets_tareas (ticket_id, asunto, descripcion, responsable, estado, fecha_limite, creado_por, fecha_creacion) VALUES (%s, %s, %s, %s, 'pendiente', %s, %s, %s)") if db_type == 'postgres' else \
                ("INSERT INTO tickets_tareas (ticket_id, asunto, descripcion, responsable, estado, fecha_limite, creado_por, fecha_creacion) VALUES (?, ?, ?, ?, 'pendiente', ?, ?, ?)")
        cursor.execute(q_ins, (ticket_id, asunto, descripcion, responsable, fecha_limite, session['username'], fecha_actual))
        conn.commit()
        registrar_log(session['username'], "Tarea de Ticket Creada", f"Ticket #{ticket_id}: tarea '{asunto}'" + (f" asignada a {responsable}" if responsable else ""))
        # 🔔 Si se le asignó a otra persona (no a quien la crea), se le avisa con una
        # notificación interna — igual que cuando se le asigna un ticket completo.
        if responsable and responsable != session['username']:
            crear_notificacion(
                responsable,
                f"Te asignaron la tarea \"{asunto}\" en el ticket #{ticket_id}.",
                url=url_for('ver_ticket', ticket_id=ticket_id), tipo='tarea'
            )
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error creando tarea en ticket {ticket_id}: {e}")
    conn.close()
    return redirect(url_for('ver_ticket', ticket_id=ticket_id))


@app.route('/tickets/<int:ticket_id>/tareas/<int:tarea_id>/estado', methods=['POST'])
@login_required
@agente_o_admin_required
def cambiar_estado_tarea_ticket(ticket_id, tarea_id):
    """Cambia el estado de una tarea del ticket: 'pendiente', 'en_progreso', 'completada' o
    'cancelada' (desplegable de estado en el detalle del ticket). 'cancelada' es para una tarea
    que ya no aplica —se distingue de 'completada' porque no se hizo el trabajo, solo dejó de
    ser necesario— y, junto con 'completada', es lo que actualizar_ticket() exige para poder
    marcar el ticket como Resuelto/Cerrado (ver ese bloqueo en app.py).

    Cada cambio queda en Logs (bitácora): antes solo se registraba crear/eliminar una tarea, no
    sus transiciones de estado, dejando un hueco en el rastro de auditoría de quién hizo qué y
    cuándo dentro de un caso."""
    nuevo_estado = (request.form.get('estado') or '').strip()
    # 🔁 Cuando el cambio se dispara desde la cola personal "Mis Tareas" (ver mis_tareas() más
    # abajo), conviene devolver al agente ahí mismo en vez de mandarlo al detalle del ticket —
    # si no, cada clic lo saca de su lista y tiene que volver a entrar para seguir marcando
    # tareas de otros casos.
    origen = request.form.get('origen', '')
    ver_todas = request.form.get('ver_todas', '')
    destino = url_for('mis_tareas', todas='1') if origen == 'mis_tareas' and ver_todas == '1' else \
              url_for('mis_tareas') if origen == 'mis_tareas' else url_for('ver_ticket', ticket_id=ticket_id)
    if nuevo_estado not in ('pendiente', 'en_progreso', 'completada', 'cancelada'):
        return redirect(destino)

    conn, db_type = get_db()
    cursor = conn.cursor()
    fecha_completada = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if nuevo_estado in ('completada', 'cancelada') else None
    try:
        q_sel = "SELECT asunto, estado FROM tickets_tareas WHERE id = %s AND ticket_id = %s" if db_type == 'postgres' else "SELECT asunto, estado FROM tickets_tareas WHERE id = ? AND ticket_id = ?"
        cursor.execute(q_sel, (tarea_id, ticket_id))
        fila_tarea = cursor.fetchone()
        if not fila_tarea:
            conn.close()
            return redirect(destino)
        asunto_tarea, estado_tarea_old = fila_tarea

        q_upd = ("UPDATE tickets_tareas SET estado = %s, fecha_completada = %s WHERE id = %s AND ticket_id = %s") if db_type == 'postgres' else \
                ("UPDATE tickets_tareas SET estado = ?, fecha_completada = ? WHERE id = ? AND ticket_id = ?")
        cursor.execute(q_upd, (nuevo_estado, fecha_completada, tarea_id, ticket_id))
        conn.commit()
        if nuevo_estado != estado_tarea_old:
            registrar_log(
                session['username'], "Estado de Tarea de Ticket Actualizado",
                f"Ticket #{ticket_id}: tarea '{asunto_tarea}' pasó de '{estado_tarea_old or 'pendiente'}' a '{nuevo_estado}'"
            )
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error actualizando estado de la tarea {tarea_id} (ticket {ticket_id}): {e}")
    conn.close()
    return redirect(destino)


@app.route('/tickets/<int:ticket_id>/tareas/<int:tarea_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_tarea_ticket(ticket_id, tarea_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_del = "DELETE FROM tickets_tareas WHERE id = %s AND ticket_id = %s" if db_type == 'postgres' else "DELETE FROM tickets_tareas WHERE id = ? AND ticket_id = ?"
        cursor.execute(q_del, (tarea_id, ticket_id))
        conn.commit()
        registrar_log(session['username'], "Tarea de Ticket Eliminada", f"Ticket #{ticket_id}: se eliminó la tarea #{tarea_id}")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error eliminando la tarea {tarea_id} (ticket {ticket_id}): {e}")
    conn.close()
    return redirect(url_for('ver_ticket', ticket_id=ticket_id))


@app.route('/tickets/mis_tareas')
@login_required
@agente_o_admin_required
def mis_tareas():
    """Cola de trabajo personal del agente: todas las tareas asignadas a él en CUALQUIER
    ticket, sin tener que recordar en cuáles casos le tocó algo (inspirado en la cola "Tareas"
    del menú lateral de Aranda Service Desk). Por defecto solo se muestran las tareas activas
    (pendiente/en_progreso); 'Ver completadas y canceladas' las agrega todas."""
    usuario = session.get('username')
    ver_todas = request.args.get('todas') == '1'

    conn, db_type = get_db()
    cursor = conn.cursor()
    ph = '%s' if db_type == 'postgres' else '?'
    filtro_estado = "" if ver_todas else "AND tt.estado NOT IN ('completada', 'cancelada')"
    q = (
        f"SELECT tt.id, tt.ticket_id, tt.asunto, tt.descripcion, tt.estado, tt.fecha_limite, "
        f"tt.fecha_completada, t.titulo, t.tipo, t.estado, t.fecha_creacion "
        f"FROM tickets_tareas tt JOIN tickets t ON t.id = tt.ticket_id "
        f"WHERE tt.responsable = {ph} AND COALESCE(t.eliminado, 0) = 0 {filtro_estado} "
        f"ORDER BY (tt.estado IN ('completada', 'cancelada')) ASC, "
        f"(tt.fecha_limite IS NULL OR tt.fecha_limite = '') ASC, tt.fecha_limite ASC, tt.id DESC"
    )
    cursor.execute(q, (usuario,))
    filas = cursor.fetchall()
    conn.close()

    hoy = datetime.now().strftime('%Y-%m-%d')
    tareas = []
    for f in filas:
        (tarea_id, ticket_id, asunto, descripcion, estado, fecha_limite, fecha_completada,
         ticket_titulo, ticket_tipo, ticket_estado, ticket_fecha_creacion) = f
        tareas.append({
            'id': tarea_id, 'ticket_id': ticket_id, 'asunto': asunto, 'descripcion': descripcion,
            'estado': estado or 'pendiente', 'fecha_limite': fecha_limite, 'fecha_completada': fecha_completada,
            'ticket_titulo': ticket_titulo, 'ticket_estado': ticket_estado,
            'ticket_codigo': _codigo_ticket(ticket_tipo or 'Incidente', ticket_id, ticket_fecha_creacion),
            'vencida': bool(fecha_limite and estado not in ('completada', 'cancelada') and fecha_limite < hoy),
        })

    return render_template('mis_tareas.html', es_soporte=True, tareas=tareas, ver_todas=ver_todas)


@app.route('/tickets/<int:ticket_id>/duplicar_datos')
@login_required
def duplicar_datos_ticket(ticket_id):
    """Datos mínimos de un ticket ya existente para prellenar el formulario de 'Nueva
    Solicitud' cuando el usuario elige 'Duplicar' desde el detalle (ver botón en
    ticket_detalle.html y el manejo de '?duplicar=' en tickets.html). Respeta el mismo permiso
    de visibilidad que ver_ticket(): solo quien puede ver el ticket original puede duplicarlo."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    q_sel = "SELECT titulo, descripcion, tipo, categoria, prioridad, area, sede, activo_id, creado_por, estado, fecha_creacion FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT titulo, descripcion, tipo, categoria, prioridad, area, sede, activo_id, creado_por, estado, fecha_creacion FROM tickets WHERE id = ?"
    cursor.execute(q_sel, (ticket_id,))
    row = cursor.fetchone()
    conn.close()

    if not row or not _puede_ver_ticket(row[8]):
        return jsonify({'error': 'No encontrado'}), 404

    titulo, descripcion, tipo, categoria, prioridad, area, sede, activo_id, creado_por, estado, fecha_creacion = row
    return jsonify({
        'titulo': titulo, 'descripcion': descripcion, 'tipo': tipo or 'Incidente',
        'categoria': categoria, 'prioridad': prioridad, 'area': area, 'sede': sede,
        'activo_id': activo_id,
        # Solo se sugiere auto-asociar el original si ya está Resuelto/Cerrado (ver
        # crear_ticket, que valida esto mismo del lado del servidor de todas formas).
        'ticket_relacionado_sugerido': ticket_id if estado in ('Resuelto', 'Cerrado') else None,
        'codigo': _codigo_ticket(tipo or 'Incidente', ticket_id, fecha_creacion)
    })


@app.route('/tickets/<int:ticket_id>/comentar', methods=['POST'])
@login_required
def comentar_ticket(ticket_id):
    # 📝 El mensaje llega como HTML del editor de texto enriquecido (Quill) — se limpia ANTES
    # de guardarlo, para no confiar en lo que manda el navegador.
    mensaje = _sanitizar_html_enriquecido(request.form.get('mensaje', '').strip())
    if _html_esta_vacio(mensaje):
        mensaje = ''
    es_admin = (session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)
    es_interno = es_admin and request.form.get('interno') == 'on'

    conn, db_type = get_db()
    cursor = conn.cursor()
    q_sel = "SELECT creado_por, estado, titulo, tipo, fecha_creacion, asignado_a FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT creado_por, estado, titulo, tipo, fecha_creacion, asignado_a FROM tickets WHERE id = ?"
    cursor.execute(q_sel, (ticket_id,))
    row = cursor.fetchone()

    if not row or not _puede_ver_ticket(row[0]):
        conn.close()
        return redirect(url_for('ver_tickets'))
    creado_por, estado_actual, titulo_ticket, tipo_ticket, fecha_creacion_ticket, asignado_actual = row

    # Un ticket cerrado ya no admite comentarios de quien lo creó (solo soporte TI podría
    # necesitar dejar una nota adicional sobre uno ya cerrado).
    if estado_actual == 'Cerrado' and not es_admin:
        conn.close()
        return redirect(url_for('ver_ticket', ticket_id=ticket_id))

    archivos_subidos = _subir_adjuntos_ticket(request.files.getlist('adjuntos'))

    if mensaje or archivos_subidos:
        fecha_act = obtener_fecha_actual()
        usuario = session.get('username')
        tipo_comentario = 'interno' if es_interno else 'comentario'
        mensaje_final = mensaje or "(adjuntó archivo(s) sin comentario)"
        try:
            if db_type == 'postgres':
                q_ins = "INSERT INTO tickets_comentarios (ticket_id, autor, mensaje, tipo, fecha) VALUES (%s, %s, %s, %s, %s) RETURNING id"
                cursor.execute(q_ins, (ticket_id, usuario, mensaje_final, tipo_comentario, fecha_act))
                nuevo_com_id = cursor.fetchone()[0]
            else:
                q_ins = "INSERT INTO tickets_comentarios (ticket_id, autor, mensaje, tipo, fecha) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(q_ins, (ticket_id, usuario, mensaje_final, tipo_comentario, fecha_act))
                nuevo_com_id = cursor.lastrowid

            if archivos_subidos:
                _guardar_adjuntos_ticket(cursor, db_type, ticket_id, nuevo_com_id, archivos_subidos, usuario, fecha_act)

            q_upd = "UPDATE tickets SET fecha_actualizacion = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE tickets SET fecha_actualizacion = ? WHERE id = ?"
            cursor.execute(q_upd, (fecha_act, ticket_id))
            conn.commit()
            etiqueta = "Comentario interno" if es_interno else "Comentario"
            registrar_log(usuario, "Comentario en Ticket", f"{etiqueta} agregado al ticket #{ticket_id}" + (f" ({len(archivos_subidos)} adjunto(s))" if archivos_subidos else ""))

            # 📧🔔 Avisamos al solicitante cuando alguien más (típicamente soporte TI) responde
            # su ticket. Las notas internas nunca generan este aviso: son solo para
            # coordinación entre el equipo de soporte.
            url_ticket = url_for('ver_ticket', ticket_id=ticket_id)
            codigo = _codigo_ticket(tipo_ticket or 'Incidente', ticket_id, fecha_creacion_ticket)
            if not es_interno and usuario != creado_por:
                correo_solicitante = _correo_de_usuario(creado_por)
                if correo_solicitante:
                    asunto = f"[Arkiv] Nueva respuesta en tu solicitud {codigo}"
                    cuerpo = (
                        f"Hola,\n\nTu solicitud de soporte {codigo} ('{titulo_ticket}') tiene una respuesta nueva "
                        f"de {usuario}.\n\nPuedes verla ingresando a Arkiv, módulo Solicitudes TI.\n\n---\n"
                        f"Equipo de Soporte TI - Arkiv"
                    )
                    threading.Thread(target=enviar_correo_ticket, args=(correo_solicitante, asunto, cuerpo)).start()
                crear_notificacion(creado_por, f"Nueva respuesta en tu solicitud {codigo}", url=url_ticket)
            # 🔔 Si quien comenta es el propio solicitante, avisamos (solo campanita, sin correo
            # adicional) a quien tenga el ticket asignado — o a todo el equipo si nadie lo ha
            # tomado — para que sepan que hay actividad nueva por atender.
            elif not es_interno and usuario == creado_por:
                if asignado_actual:
                    crear_notificacion(asignado_actual, f"{usuario} respondió la solicitud {codigo}", url=url_ticket)
                else:
                    crear_notificacion_para_varios([m['usuario'] for m in _equipo_soporte_activo()], f"{usuario} respondió la solicitud {codigo}", url=url_ticket)
        except Exception as e:
            conn.rollback()
            print(f"Error comentando ticket {ticket_id}: {e}")

    conn.close()
    return redirect(url_for('ver_ticket', ticket_id=ticket_id))


@app.route('/tickets/<int:ticket_id>/actualizar', methods=['POST'])
@login_required
@agente_o_admin_required
def actualizar_ticket(ticket_id):
    nuevo_estado = request.form.get('estado', '').strip()
    nueva_prioridad = request.form.get('prioridad', '').strip()
    nuevo_asignado = request.form.get('asignado_a', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT estado, prioridad, asignado_a, sla_respuesta_cumplida, sla_resolucion_cumplida, creado_por, titulo, tipo, fecha_creacion, sla_resolucion_limite, sla_pausado_desde FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT estado, prioridad, asignado_a, sla_respuesta_cumplida, sla_resolucion_cumplida, creado_por, titulo, tipo, fecha_creacion, sla_resolucion_limite, sla_pausado_desde FROM tickets WHERE id = ?"
        cursor.execute(q_sel, (ticket_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return redirect(url_for('ver_tickets'))

        (estado_old, prioridad_old, asignado_old, respuesta_cumplida_old, resolucion_cumplida_old,
         creado_por, titulo_ticket, tipo_ticket, fecha_creacion_ticket, resolucion_limite_old,
         pausado_desde_old) = row
        # 🔒 No se puede saltar directo a 'Resuelto' sin pasar antes por 'En Proceso', ni a
        # 'Cerrado' sin pasar antes por 'Resuelto'; una vez 'Cerrado' el ticket queda bloqueado
        # (debe ser el último estado). Se valida SIEMPRE del lado del servidor — el desplegable
        # de ticket_detalle.html ya solo ofrece las opciones permitidas (ver ver_ticket()), pero
        # esto es lo que de verdad lo hace cumplir aunque alguien arme el POST a mano. Misma
        # regla en _estados_disponibles_ticket(), para no tener dos criterios distintos.
        estado_final = nuevo_estado if nuevo_estado in _estados_disponibles_ticket(estado_old) else estado_old
        prioridad_final = nueva_prioridad if nueva_prioridad in PRIORIDADES_TICKET else prioridad_old
        asignado_final = nuevo_asignado or None

        # 🚧 Inspirado en la pestaña "Tareas" de Aranda Service Desk: un caso no debería poder
        # marcarse Resuelto/Cerrado mientras todavía tenga tareas internas sin terminar. Si el
        # agente intenta ese salto con tareas en 'pendiente'/'en_progreso', se bloquea el cambio
        # de estado (el resto de los campos del formulario —prioridad, asignado— sí se guardan)
        # y se le explica por qué con un mensaje flash. Una tarea que ya no aplica se marca
        # 'cancelada' (no 'completada') para no bloquear el cierre sin haberla hecho de verdad.
        if estado_final in ('Resuelto', 'Cerrado') and estado_final != estado_old:
            q_tareas_pend = ("SELECT COUNT(*) FROM tickets_tareas WHERE ticket_id = %s AND estado NOT IN ('completada', 'cancelada')") if db_type == 'postgres' else \
                            ("SELECT COUNT(*) FROM tickets_tareas WHERE ticket_id = ? AND estado NOT IN ('completada', 'cancelada')")
            cursor.execute(q_tareas_pend, (ticket_id,))
            tareas_pendientes = cursor.fetchone()[0]
            if tareas_pendientes:
                estado_final = estado_old
                flash(
                    f"No se pudo cambiar el estado a '{nuevo_estado}': el ticket todavía tiene "
                    f"{tareas_pendientes} tarea(s) sin completar o cancelar. Resuélvelas (o cancélalas) "
                    "en la sección Tareas de este ticket antes de continuar.",
                    "error"
                )

        # 🔒 Un ticket 'Cerrado' ya no se puede reasignar a otra persona: el caso quedó resuelto
        # y cerrado, así que cambiar después quién quedó como responsable no aporta nada y podía
        # usarse para mover casos cerrados de un agente a otro sin dejar rastro real de quién lo
        # atendió. Se ignora cualquier valor que llegue en 'asignado_a' y se conserva el que ya
        # tenía — igual que con el estado, se valida del lado del servidor sin importar si el
        # select de la plantilla también quedó deshabilitado.
        if estado_old == 'Cerrado':
            asignado_final = asignado_old

        # 👤 Si el ticket se resuelve/cierra sin que nadie quedara asignado (el agente cambió
        # el estado pero dejó el selector "Asignar a" en "Sin asignar"), se asigna
        # automáticamente a quien hizo el cambio. Sin esto, "Top agentes por solicitudes
        # resueltas" en Indicadores queda vacío aunque sí haya tickets resueltos, porque
        # ese indicador cuenta por 'asignado_a' — y de paso deja un registro correcto de
        # quién atendió realmente el caso.
        if estado_final in ('Resuelto', 'Cerrado') and not asignado_final:
            asignado_final = session.get('username')

        fecha_act = obtener_fecha_actual()

        # SLA de "primera respuesta": queda marcado la primera vez que el ticket sale de
        # 'Abierto' (y no se vuelve a tocar aunque después cambie de estado otra vez).
        respuesta_cumplida_final = respuesta_cumplida_old
        if estado_final != 'Abierto' and not respuesta_cumplida_final:
            respuesta_cumplida_final = fecha_act

        # SLA de "resolución": se marca al llegar a Resuelto/Cerrado. Si el ticket se
        # reabre (vuelve a Abierto/En Proceso), deja de contar como resuelto a tiempo.
        if estado_final in ('Resuelto', 'Cerrado'):
            resolucion_cumplida_final = resolucion_cumplida_old or fecha_act
        else:
            resolucion_cumplida_final = None

        # ⏸️ Pausa/reanudación del SLA de resolución para el estado 'Pendiente' (ver
        # _calcular_sla_ticket() y _progreso_ticket()). Al entrar se guarda desde cuándo quedó
        # pausado; al salir se corre el límite de resolución exactamente la cantidad de horas
        # que estuvo pausado, para que ese tiempo de espera no cuente en contra del agente, y se
        # limpia 'sla_alerta_nivel' (mismo patrón que modificar_sla_ticket) para que pueda volver
        # a avisar si el nuevo límite se acerca.
        pausa_iniciada = estado_final == 'Pendiente' and estado_old != 'Pendiente'
        pausa_reanudada = estado_old == 'Pendiente' and estado_final != 'Pendiente'
        resolucion_limite_final = resolucion_limite_old
        alerta_nivel_final = None
        limpiar_alerta = False
        if pausa_iniciada:
            pausado_desde_final = fecha_act
        elif pausa_reanudada:
            pausado_desde_final = None
            inicio_pausa = _parsear_fecha_ticket(pausado_desde_old)
            fin_pausa = _parsear_fecha_ticket(fecha_act)
            if inicio_pausa and fin_pausa and resolucion_limite_old:
                horas_pausa = max(0, (fin_pausa - inicio_pausa).total_seconds() / 3600)
                resolucion_limite_final = _calcular_limite_sla(resolucion_limite_old, horas_pausa)
                limpiar_alerta = True
        else:
            pausado_desde_final = pausado_desde_old

        if limpiar_alerta:
            q_upd = "UPDATE tickets SET estado = %s, prioridad = %s, asignado_a = %s, fecha_actualizacion = %s, sla_respuesta_cumplida = %s, sla_resolucion_cumplida = %s, sla_resolucion_limite = %s, sla_pausado_desde = %s, sla_alerta_nivel = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE tickets SET estado = ?, prioridad = ?, asignado_a = ?, fecha_actualizacion = ?, sla_respuesta_cumplida = ?, sla_resolucion_cumplida = ?, sla_resolucion_limite = ?, sla_pausado_desde = ?, sla_alerta_nivel = ? WHERE id = ?"
            cursor.execute(q_upd, (estado_final, prioridad_final, asignado_final, fecha_act, respuesta_cumplida_final, resolucion_cumplida_final, resolucion_limite_final, pausado_desde_final, alerta_nivel_final, ticket_id))
        else:
            q_upd = "UPDATE tickets SET estado = %s, prioridad = %s, asignado_a = %s, fecha_actualizacion = %s, sla_respuesta_cumplida = %s, sla_resolucion_cumplida = %s, sla_resolucion_limite = %s, sla_pausado_desde = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE tickets SET estado = ?, prioridad = ?, asignado_a = ?, fecha_actualizacion = ?, sla_respuesta_cumplida = ?, sla_resolucion_cumplida = ?, sla_resolucion_limite = ?, sla_pausado_desde = ? WHERE id = ?"
            cursor.execute(q_upd, (estado_final, prioridad_final, asignado_final, fecha_act, respuesta_cumplida_final, resolucion_cumplida_final, resolucion_limite_final, pausado_desde_final, ticket_id))

        cambios = []
        if estado_final != estado_old:
            cambios.append(f"estado: '{estado_old}' → '{estado_final}'")
        if prioridad_final != prioridad_old:
            cambios.append(f"prioridad: '{prioridad_old}' → '{prioridad_final}'")
        if asignado_final != asignado_old:
            cambios.append(f"asignado a: '{asignado_old or 'nadie'}' → '{asignado_final or 'nadie'}'")
        if pausa_reanudada and resolucion_limite_final != resolucion_limite_old:
            cambios.append("SLA de resolución extendido automáticamente por el tiempo en 'Pendiente'")

        usuario = session.get('username')
        if cambios:
            mensaje_sistema = f"{usuario} actualizó el ticket — " + "; ".join(cambios)
            q_ins = "INSERT INTO tickets_comentarios (ticket_id, autor, mensaje, tipo, fecha) VALUES (%s, %s, %s, 'sistema', %s)" if db_type == 'postgres' else "INSERT INTO tickets_comentarios (ticket_id, autor, mensaje, tipo, fecha) VALUES (?, ?, ?, 'sistema', ?)"
            cursor.execute(q_ins, (ticket_id, usuario, mensaje_sistema, fecha_act))

        conn.commit()

        if cambios:
            registrar_log(usuario, "Actualización de Ticket", f"Ticket #{ticket_id}: {'; '.join(cambios)}")

        # 📧 Avisamos por correo al solicitante cuando su ticket cambia de ESTADO (no en cada
        # cambio de prioridad/asignación, que es más una gestión interna del equipo de TI).
        if estado_final != estado_old and creado_por != usuario:
            codigo = _codigo_ticket(tipo_ticket or 'Incidente', ticket_id, fecha_creacion_ticket)
            correo_solicitante = _correo_de_usuario(creado_por)
            if correo_solicitante:
                asunto = f"[Arkiv] Tu solicitud {codigo} cambió a '{estado_final}'"
                cuerpo = (
                    f"Hola,\n\nTu solicitud de soporte {codigo} ('{titulo_ticket}') cambió de estado: "
                    f"'{estado_old}' → '{estado_final}'.\n\nPuedes ver el detalle completo ingresando a Arkiv, "
                    f"módulo Solicitudes TI.\n\n---\nEquipo de Soporte TI - Arkiv"
                )
                threading.Thread(target=enviar_correo_ticket, args=(correo_solicitante, asunto, cuerpo)).start()
            crear_notificacion(creado_por, f"Tu solicitud {codigo} cambió a '{estado_final}'", url=url_for('ver_ticket', ticket_id=ticket_id))
    except Exception as e:
        conn.rollback()
        print(f"Error actualizando ticket {ticket_id}: {e}")

    conn.close()
    return redirect(url_for('ver_ticket', ticket_id=ticket_id))


@app.route('/tickets/<int:ticket_id>/eliminar', methods=['POST'])
@login_required
@superadmin_required
def eliminar_ticket(ticket_id):
    """Baja lógica de un ticket (tickets.eliminado = 1): lo saca por completo de las listas,
    el panel de inicio y los indicadores — a diferencia de poner el estado en 'Cancelado', que
    deja el ticket visible como parte del historial. Pensada para los tickets de prueba,
    duplicados o creados por error que el equipo pidió poder limpiar antes de operar en serio
    (ver seguimiento de Solvyx). Solo el super-admin la tiene disponible, precisamente porque
    no se puede deshacer desde la interfaz.

    🛡️ Dos reglas adicionales para que esto no se use para borrar casos que ya alguien está
    atendiendo: (1) si el ticket ya tiene un responsable asignado, no se puede eliminar —
    borrarlo perdería el rastro de lo que ese agente ya gestionó; el único camino ahí es el
    flujo normal de estados. (2) si todavía NO tiene responsable, se exige justificar por qué
    se elimina (mínimo 10 caracteres); el motivo queda registrado en Auditoría y Logs."""
    motivo_eliminacion = request.form.get('motivo_eliminacion', '').strip()
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo, asignado_a FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT titulo, asignado_a FROM tickets WHERE id = ?"
        cursor.execute(q_sel, (ticket_id,))
        row = cursor.fetchone()
        if row:
            titulo, asignado_a = row[0], row[1]

            if asignado_a:
                conn.close()
                return redirect(url_for('ver_ticket', ticket_id=ticket_id))

            if len(motivo_eliminacion) < 10:
                conn.close()
                return redirect(url_for('ver_ticket', ticket_id=ticket_id))

            q_upd = "UPDATE tickets SET eliminado = 1 WHERE id = %s" if db_type == 'postgres' else "UPDATE tickets SET eliminado = 1 WHERE id = ?"
            cursor.execute(q_upd, (ticket_id,))
            conn.commit()
            registrar_log(session.get('username'), "Eliminación de Ticket", f"Se eliminó el ticket #{ticket_id} ('{titulo}') — Motivo: {motivo_eliminacion}")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error eliminando ticket {ticket_id}: {e}")
    conn.close()
    return redirect(url_for('ver_tickets'))


@app.route('/tickets/<int:ticket_id>/modificar_sla', methods=['POST'])
@login_required
@agente_o_admin_required
def modificar_sla_ticket(ticket_id):
    """Permite al equipo de soporte correr la fecha límite de SOLUCIÓN de un ticket (el
    'compromiso con el usuario'), siempre con un motivo y un tope de MAX_MODIFICACIONES_SLA
    veces — igual que en la mesa de ayuda externa que ya usa la organización."""
    nueva_fecha_raw = request.form.get('nueva_fecha', '').strip()  # <input type="datetime-local">
    motivo_categoria = request.form.get('motivo_categoria', '').strip()
    motivo_detalle = request.form.get('motivo', '').strip()
    if motivo_categoria not in MOTIVOS_MODIFICACION_SLA:
        motivo_categoria = 'Otro'
    # El registro guarda la categoría (p. ej. "Depende de un proveedor o tercero externo",
    # para dejar explícito que el atraso no depende del analista/agente) junto con el detalle
    # libre que el equipo de soporte escriba.
    motivo = f"{motivo_categoria} — {motivo_detalle}" if motivo_detalle else motivo_categoria

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT sla_modificaciones, estado FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT sla_modificaciones, estado FROM tickets WHERE id = ?"
        cursor.execute(q_sel, (ticket_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return redirect(url_for('ver_tickets'))

        mods_actuales = row[0] or 0
        # El detalle libre (motivo_detalle) sigue exigiendo mínimo 10 caracteres, igual que
        # antes; la categoría (motivo_categoria) es obligatoria por separado vía el <select>.
        if mods_actuales >= MAX_MODIFICACIONES_SLA or len(motivo_detalle) < 10 or not nueva_fecha_raw:
            conn.close()
            return redirect(url_for('ver_ticket', ticket_id=ticket_id))

        try:
            nueva_fecha_fmt = datetime.strptime(nueva_fecha_raw, "%Y-%m-%dT%H:%M").strftime(FORMATO_FECHA_TICKET)
        except Exception:
            conn.close()
            return redirect(url_for('ver_ticket', ticket_id=ticket_id))

        fecha_act = obtener_fecha_actual()
        usuario = session.get('username')

        # 🔔 Se limpia el nivel de alerta de SLA ya avisado: con la nueva fecha límite el ticket
        # vuelve a "vigente", y así puede volver a avisar más adelante si se acerca de nuevo.
        q_upd = "UPDATE tickets SET sla_resolucion_limite = %s, sla_modificaciones = sla_modificaciones + 1, fecha_actualizacion = %s, sla_alerta_nivel = NULL WHERE id = %s" if db_type == 'postgres' else "UPDATE tickets SET sla_resolucion_limite = ?, sla_modificaciones = sla_modificaciones + 1, fecha_actualizacion = ?, sla_alerta_nivel = NULL WHERE id = ?"
        cursor.execute(q_upd, (nueva_fecha_fmt, fecha_act, ticket_id))

        mensaje_sistema = f"{usuario} modificó la fecha límite de solución a {nueva_fecha_fmt} — Motivo: {motivo}"
        q_ins = "INSERT INTO tickets_comentarios (ticket_id, autor, mensaje, tipo, fecha) VALUES (%s, %s, %s, 'sistema', %s)" if db_type == 'postgres' else "INSERT INTO tickets_comentarios (ticket_id, autor, mensaje, tipo, fecha) VALUES (?, ?, ?, 'sistema', ?)"
        cursor.execute(q_ins, (ticket_id, usuario, mensaje_sistema, fecha_act))

        conn.commit()
        registrar_log(usuario, "SLA de Ticket Modificado", f"Ticket #{ticket_id}: nueva fecha límite de solución {nueva_fecha_fmt} ({mods_actuales + 1}/{MAX_MODIFICACIONES_SLA}) — {motivo}")
    except Exception as e:
        conn.rollback()
        print(f"Error modificando SLA del ticket {ticket_id}: {e}")

    conn.close()
    return redirect(url_for('ver_ticket', ticket_id=ticket_id))


@app.route('/tickets/<int:ticket_id>/calificar', methods=['POST'])
@login_required
def calificar_ticket(ticket_id):
    """El propio solicitante puede calificar de 1 a 5 estrellas su experiencia, una sola vez,
    y solo cuando el ticket ya está Resuelto o Cerrado — igual que en la mesa de ayuda externa."""
    try:
        calificacion = int(request.form.get('calificacion', '0'))
    except (TypeError, ValueError):
        calificacion = 0

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT creado_por, estado, calificacion, solicitante_real FROM tickets WHERE id = %s" if db_type == 'postgres' else "SELECT creado_por, estado, calificacion, solicitante_real FROM tickets WHERE id = ?"
        cursor.execute(q_sel, (ticket_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return redirect(url_for('ver_tickets'))

        creado_por, estado_actual, calificacion_previa, solicitante_real = row
        # 🙋 Quien califica es el BENEFICIARIO real de la solicitud, no necesariamente quien la
        # creó: si un agente la subió a nombre de otra persona (ver crear_ticket), es esa
        # persona quien debe poder calificar — así el agente que resuelve el caso no puede
        # calificarse a sí mismo (bug reportado por Tomás).
        beneficiario = solicitante_real or creado_por
        puede_calificar = (
            session.get('username') == beneficiario and
            estado_actual in ('Resuelto', 'Cerrado') and
            not calificacion_previa and
            CALIFICACION_MIN <= calificacion <= CALIFICACION_MAX
        )
        if puede_calificar:
            fecha_act = obtener_fecha_actual()
            q_upd = "UPDATE tickets SET calificacion = %s, calificacion_fecha = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE tickets SET calificacion = ?, calificacion_fecha = ? WHERE id = ?"
            cursor.execute(q_upd, (calificacion, fecha_act, ticket_id))
            conn.commit()
            registrar_log(session.get('username'), "Calificación de Ticket", f"Ticket #{ticket_id} calificado con {calificacion}/{CALIFICACION_MAX} estrellas")
    except Exception as e:
        conn.rollback()
        print(f"Error calificando ticket {ticket_id}: {e}")

    conn.close()
    return redirect(url_for('ver_ticket', ticket_id=ticket_id))


@app.route('/tickets/inicio')
@login_required
def inicio_tickets():
    """Página de aterrizaje propia del módulo Solicitudes TI (no la Bienvenida general de
    Arkiv): un resumen rápido distinto según el rol, más los tickets más recientes."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    es_soporte = (session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)
    usuario = session.get('username')
    ph = '%s' if db_type == 'postgres' else '?'

    resumen = {}
    try:
        if es_soporte:
            cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE asignado_a = {ph} AND estado IN ('Abierto', 'En Proceso') AND COALESCE(eliminado, 0) = 0", (usuario,))
            resumen['asignados_abiertos'] = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM tickets WHERE (asignado_a IS NULL OR asignado_a = '') AND estado IN ('Abierto', 'En Proceso') AND COALESCE(eliminado, 0) = 0")
            resumen['sin_asignar'] = cursor.fetchone()[0]
            cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE asignado_a = {ph} AND estado IN ('Resuelto', 'Cerrado') AND COALESCE(eliminado, 0) = 0", (usuario,))
            resumen['resueltos_por_mi'] = cursor.fetchone()[0]
            if session.get('rol') == 'admin':
                # 🚨 Cuántos tickets ASIGNADOS (a cualquier agente, no solo a mí) ya están
                # escalados por SLA vencido — ver _revisar_alertas_sla(). Solo tiene sentido
                # para el super-admin/admins: son los "supervisores" a quienes se escala.
                ahora_str = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None).strftime(FORMATO_FECHA_TICKET)
                cursor.execute(
                    f"SELECT COUNT(*) FROM tickets WHERE estado IN ('Abierto', 'En Proceso') AND COALESCE(eliminado, 0) = 0 "
                    f"AND asignado_a IS NOT NULL AND asignado_a != '' AND sla_resolucion_limite IS NOT NULL AND sla_resolucion_limite < {ph}",
                    (ahora_str,)
                )
                resumen['escalados'] = cursor.fetchone()[0]
        else:
            cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE creado_por = {ph} AND estado IN ('Abierto', 'En Proceso') AND COALESCE(eliminado, 0) = 0", (usuario,))
            resumen['mis_abiertos'] = cursor.fetchone()[0]
            cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE creado_por = {ph} AND estado IN ('Resuelto', 'Cerrado') AND COALESCE(eliminado, 0) = 0", (usuario,))
            resumen['mis_resueltos'] = cursor.fetchone()[0]
            cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE creado_por = {ph} AND COALESCE(eliminado, 0) = 0", (usuario,))
            resumen['mis_total'] = cursor.fetchone()[0]
    except Exception as e:
        print(f"Error calculando resumen de inicio de tickets: {e}")

    query = "SELECT id, titulo, tipo, categoria, prioridad, estado, creado_por, fecha_creacion FROM tickets WHERE COALESCE(eliminado, 0) = 0"
    params = []
    if not es_soporte:
        query += f" AND creado_por = {ph}"
        params.append(usuario)
    query += " ORDER BY id DESC LIMIT 5"

    recientes = []
    try:
        cursor.execute(query, tuple(params))
        for r in cursor.fetchall():
            tipo_t = r[2] or 'Incidente'
            recientes.append({
                'id': r[0], 'titulo': r[1], 'tipo': tipo_t, 'categoria': r[3], 'prioridad': r[4],
                'estado': r[5], 'creado_por': r[6], 'codigo': _codigo_ticket(tipo_t, r[0], r[7])
            })
    except Exception as e:
        print(f"Error consultando recientes de inicio de tickets: {e}")

    conn.close()
    return render_template('tickets_inicio.html', es_soporte=es_soporte, resumen=resumen, recientes=recientes)


# 📚 BASE DE CONOCIMIENTO (artículos con un documento adjunto, visibles para todos los
# roles del módulo de Tickets; solo el equipo de soporte TI los crea/edita/elimina).
@app.route('/tickets/conocimiento')
@login_required
def ver_conocimiento():
    q_busqueda = request.args.get('q', '').strip().lower()
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, titulo, descripcion, url_documento, nombre_archivo, vistas, fecha_creacion FROM conocimiento_articulos WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY id DESC")
        rows = cursor.fetchall()
    except Exception as e:
        print(f"Error consultando base de conocimiento: {e}")
        rows = []
    conn.close()

    articulos = []
    for r in rows:
        texto_full = f"{r[1]} {r[2] or ''}".lower()
        if not q_busqueda or q_busqueda in texto_full:
            articulos.append({
                'id': r[0], 'titulo': r[1], 'descripcion': r[2], 'url_documento': r[3],
                'nombre_archivo': r[4], 'vistas': r[5] or 0, 'fecha_creacion': r[6]
            })

    es_soporte = (session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)
    return render_template('conocimiento.html', articulos=articulos, es_soporte=es_soporte, q_busqueda=q_busqueda)


@app.route('/tickets/conocimiento/nuevo', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_conocimiento():
    titulo = request.form.get('titulo', '').strip()
    descripcion = request.form.get('descripcion', '').strip()
    archivo = request.files.get('documento')

    if titulo and archivo and archivo.filename and archivo_permitido(archivo.filename):
        subidos = _subir_adjuntos_ticket([archivo])
        if subidos:
            url_doc, nombre_doc = subidos[0]
            fecha_act = obtener_fecha_actual()
            usuario = session.get('username')
            conn, db_type = get_db()
            cursor = conn.cursor()
            try:
                q_ins = "INSERT INTO conocimiento_articulos (titulo, descripcion, url_documento, nombre_archivo, vistas, creado_por, fecha_creacion, estado) VALUES (%s, %s, %s, %s, 0, %s, %s, 'activo')" if db_type == 'postgres' else "INSERT INTO conocimiento_articulos (titulo, descripcion, url_documento, nombre_archivo, vistas, creado_por, fecha_creacion, estado) VALUES (?, ?, ?, ?, 0, ?, ?, 'activo')"
                cursor.execute(q_ins, (titulo, descripcion, url_doc, nombre_doc, usuario, fecha_act))
                conn.commit()
                registrar_log(usuario, "Artículo de Conocimiento Creado", f"'{titulo}' ({nombre_doc})")
            except Exception as e:
                conn.rollback()
                print(f"Error creando artículo de conocimiento: {e}")
            conn.close()

    return redirect(url_for('ver_conocimiento'))


@app.route('/tickets/conocimiento/<int:articulo_id>/editar', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_conocimiento(articulo_id):
    titulo = request.form.get('titulo', '').strip()
    descripcion = request.form.get('descripcion', '').strip()
    archivo = request.files.get('documento')

    if not titulo:
        return redirect(url_for('ver_conocimiento'))

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        if archivo and archivo.filename and archivo_permitido(archivo.filename):
            subidos = _subir_adjuntos_ticket([archivo])
            if subidos:
                url_doc, nombre_doc = subidos[0]
                q_upd = "UPDATE conocimiento_articulos SET titulo = %s, descripcion = %s, url_documento = %s, nombre_archivo = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE conocimiento_articulos SET titulo = ?, descripcion = ?, url_documento = ?, nombre_archivo = ? WHERE id = ?"
                cursor.execute(q_upd, (titulo, descripcion, url_doc, nombre_doc, articulo_id))
            else:
                q_upd = "UPDATE conocimiento_articulos SET titulo = %s, descripcion = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE conocimiento_articulos SET titulo = ?, descripcion = ? WHERE id = ?"
                cursor.execute(q_upd, (titulo, descripcion, articulo_id))
        else:
            q_upd = "UPDATE conocimiento_articulos SET titulo = %s, descripcion = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE conocimiento_articulos SET titulo = ?, descripcion = ? WHERE id = ?"
            cursor.execute(q_upd, (titulo, descripcion, articulo_id))
        conn.commit()
        registrar_log(session.get('username'), "Artículo de Conocimiento Editado", f"Artículo #{articulo_id}: '{titulo}'")
    except Exception as e:
        conn.rollback()
        print(f"Error editando artículo de conocimiento {articulo_id}: {e}")
    conn.close()
    return redirect(url_for('ver_conocimiento'))


@app.route('/tickets/conocimiento/<int:articulo_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_conocimiento(articulo_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_upd = "UPDATE conocimiento_articulos SET estado = 'eliminado' WHERE id = %s" if db_type == 'postgres' else "UPDATE conocimiento_articulos SET estado = 'eliminado' WHERE id = ?"
        cursor.execute(q_upd, (articulo_id,))
        conn.commit()
        registrar_log(session.get('username'), "Artículo de Conocimiento Eliminado", f"Artículo #{articulo_id}")
    except Exception as e:
        conn.rollback()
        print(f"Error eliminando artículo de conocimiento {articulo_id}: {e}")
    conn.close()
    return redirect(url_for('ver_conocimiento'))


@app.route('/tickets/conocimiento/<int:articulo_id>/abrir')
@login_required
def abrir_conocimiento(articulo_id):
    """Registra una vista y redirige al documento real (alojado en Cloudinary)."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    url_doc = None
    try:
        q_sel = "SELECT url_documento FROM conocimiento_articulos WHERE id = %s" if db_type == 'postgres' else "SELECT url_documento FROM conocimiento_articulos WHERE id = ?"
        cursor.execute(q_sel, (articulo_id,))
        row = cursor.fetchone()
        if row:
            url_doc = row[0]
            q_upd = "UPDATE conocimiento_articulos SET vistas = COALESCE(vistas, 0) + 1 WHERE id = %s" if db_type == 'postgres' else "UPDATE conocimiento_articulos SET vistas = COALESCE(vistas, 0) + 1 WHERE id = ?"
            cursor.execute(q_upd, (articulo_id,))
            conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Error registrando vista de artículo {articulo_id}: {e}")
    conn.close()

    if not url_doc:
        return redirect(url_for('ver_conocimiento'))
    return redirect(url_doc)


# ⚙️ CONFIGURACIÓN DE ÁREAS, SEDES Y CATEGORÍAS (solo equipo de soporte). Estos valores
# alimentan los desplegables al crear una solicitud y los filtros de la lista de tickets.
@app.route('/tickets/configuracion')
@login_required
@agente_o_admin_required
def configuracion_tickets():
    areas = _config_ticket_lista('area')
    sedes = _config_ticket_lista('sede')
    categorias = _config_ticket_lista('categoria')
    proveedores = _config_ticket_lista('proveedor')

    # 👤 Lista de usuarios activos para elegir el "responsable" de cada sede desde un
    # desplegable (en vez de escribir un nombre libre que podría no corresponder a
    # ninguna cuenta real del sistema).
    conn, db_type = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT usuario, nombre FROM usuarios WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY usuario ASC")
    usuarios_disponibles = [{'usuario': r[0], 'nombre': r[1]} for r in cursor.fetchall()]
    conn.close()

    return render_template('tickets_configuracion.html', es_soporte=True, areas=areas, sedes=sedes, categorias=categorias, proveedores=proveedores, usuarios_disponibles=usuarios_disponibles)


@app.route('/tickets/configuracion/nuevo', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_configuracion_ticket():
    tipo = request.form.get('tipo', '').strip()
    nombre = request.form.get('nombre', '').strip()
    # 📍 Dirección solo aplica a Sedes y Áreas. Responsable aplica a los tres tipos (se usa
    # para la asignación automática de tickets por categoría/área). NIT/Razón social solo
    # aplican a Proveedores.
    direccion = request.form.get('direccion', '').strip() or None
    responsable = request.form.get('responsable', '').strip() or None
    nit = request.form.get('nit', '').strip() or None
    razon_social = request.form.get('razon_social', '').strip() or None
    if tipo not in ('sede', 'area'):
        direccion = None
    if tipo not in ('sede', 'area', 'categoria'):
        responsable = None
    if tipo != 'proveedor':
        nit = None
        razon_social = None
    if tipo in ('area', 'sede', 'categoria', 'proveedor') and nombre:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            q = "INSERT INTO ticket_configuraciones (tipo, nombre, direccion, responsable, nit, razon_social) VALUES (%s, %s, %s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO ticket_configuraciones (tipo, nombre, direccion, responsable, nit, razon_social) VALUES (?, ?, ?, ?, ?, ?)"
            cursor.execute(q, (tipo, nombre, direccion, responsable, nit, razon_social))
            conn.commit()
            detalle = f"Se agregó {tipo} '{nombre}'"
            if tipo in ('sede', 'area') and (direccion or responsable):
                detalle += f" (dirección: {direccion or 'sin especificar'}, responsable: {responsable or 'sin asignar'})"
            if tipo == 'proveedor' and (nit or razon_social):
                detalle += f" (NIT: {nit or 'sin especificar'}, razón social: {razon_social or 'sin especificar'})"
            registrar_log(session.get('username'), "Configuración de Tickets", detalle)
        except Exception as e:
            conn.rollback()
            print(f"Error creando configuración de ticket: {e}")
        conn.close()
    return redirect(url_for('configuracion_tickets'))


@app.route('/tickets/configuracion/<int:config_id>/editar', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_configuracion_ticket(config_id):
    nombre = request.form.get('nombre', '').strip()
    if nombre:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            q_sel = "SELECT tipo FROM ticket_configuraciones WHERE id = %s" if db_type == 'postgres' else "SELECT tipo FROM ticket_configuraciones WHERE id = ?"
            cursor.execute(q_sel, (config_id,))
            fila = cursor.fetchone()
            tipo_actual = fila[0] if fila else None

            if tipo_actual in ('sede', 'area'):
                # 📍 Sedes y Áreas tienen dirección y responsable; sus formularios envían
                # estos dos campos junto con el nombre.
                direccion = request.form.get('direccion', '').strip() or None
                responsable = request.form.get('responsable', '').strip() or None
                q = "UPDATE ticket_configuraciones SET nombre = %s, direccion = %s, responsable = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE ticket_configuraciones SET nombre = ?, direccion = ?, responsable = ? WHERE id = ?"
                cursor.execute(q, (nombre, direccion, responsable, config_id))
                etiqueta_tipo = 'sede' if tipo_actual == 'sede' else 'área'
                detalle = f"Se editó la {etiqueta_tipo} '{nombre}' (dirección: {direccion or 'sin especificar'}, responsable: {responsable or 'sin asignar'})"
            elif tipo_actual == 'categoria':
                # 🤖 Categoría no tiene dirección, pero sí responsable (para la asignación
                # automática de tickets por categoría).
                responsable = request.form.get('responsable', '').strip() or None
                q = "UPDATE ticket_configuraciones SET nombre = %s, responsable = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE ticket_configuraciones SET nombre = ?, responsable = ? WHERE id = ?"
                cursor.execute(q, (nombre, responsable, config_id))
                detalle = f"Se editó la categoría '{nombre}' (responsable: {responsable or 'sin asignar'})"
            elif tipo_actual == 'proveedor':
                # 🚚 Proveedor tiene NIT y razón social en vez de dirección/responsable.
                nit = request.form.get('nit', '').strip() or None
                razon_social = request.form.get('razon_social', '').strip() or None
                q = "UPDATE ticket_configuraciones SET nombre = %s, nit = %s, razon_social = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE ticket_configuraciones SET nombre = ?, nit = ?, razon_social = ? WHERE id = ?"
                cursor.execute(q, (nombre, nit, razon_social, config_id))
                detalle = f"Se editó el proveedor '{nombre}' (NIT: {nit or 'sin especificar'}, razón social: {razon_social or 'sin especificar'})"
            else:
                q = "UPDATE ticket_configuraciones SET nombre = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE ticket_configuraciones SET nombre = ? WHERE id = ?"
                cursor.execute(q, (nombre, config_id))
                detalle = f"Se renombró configuración #{config_id} a '{nombre}'"

            conn.commit()
            registrar_log(session.get('username'), "Configuración de Tickets", detalle)
        except Exception as e:
            conn.rollback()
            print(f"Error editando configuración de ticket: {e}")
        conn.close()
    return redirect(url_for('configuracion_tickets'))


@app.route('/tickets/configuracion/<int:config_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_configuracion_ticket(config_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE ticket_configuraciones SET estado = 'eliminado' WHERE id = %s" if db_type == 'postgres' else "UPDATE ticket_configuraciones SET estado = 'eliminado' WHERE id = ?"
        cursor.execute(q, (config_id,))
        conn.commit()
        registrar_log(session.get('username'), "Configuración de Tickets", f"Se eliminó configuración #{config_id}")
    except Exception as e:
        conn.rollback()
        print(f"Error eliminando configuración de ticket: {e}")
    conn.close()
    return redirect(url_for('configuracion_tickets'))


# 📋 PLANTILLAS DE SOLICITUD (solo equipo de soporte administra; cualquier usuario logueado
# las usa): agilizan la creación de tickets recurrentes ("Solicitud de acceso a Kubapp",
# "Instalación de impresora nueva"...) prellenando tipo, título, categoría, prioridad, área,
# sede y descripción en el formulario de "Nueva Solicitud" — todo sigue siendo editable antes
# de enviar, la plantilla solo ahorra escribir lo mismo una y otra vez.
@app.route('/tickets/plantillas')
@login_required
@agente_o_admin_required
def ver_plantillas_ticket():
    conn, db_type = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre, tipo, categoria, prioridad, area, sede, titulo, descripcion FROM ticket_plantillas WHERE estado = 'activo' ORDER BY nombre ASC")
    plantillas = [
        {
            'id': r[0], 'nombre': r[1], 'tipo': r[2] or 'Incidente', 'categoria': r[3] or '',
            'prioridad': r[4] or 'Media', 'area': r[5] or '', 'sede': r[6] or '',
            'titulo': r[7], 'descripcion': r[8]
        }
        for r in cursor.fetchall()
    ]
    conn.close()

    nombres_categorias = [c['nombre'] for c in _config_ticket_lista('categoria')] or CATEGORIAS_TICKET
    nombres_areas = [a['nombre'] for a in _config_ticket_lista('area')]
    nombres_sedes = [s['nombre'] for s in _config_ticket_lista('sede')]

    return render_template(
        'tickets_plantillas.html', es_soporte=True, plantillas=plantillas,
        tipos=TIPOS_TICKET, categorias=nombres_categorias, prioridades=PRIORIDADES_TICKET,
        areas=nombres_areas, sedes=nombres_sedes
    )


def _validar_datos_plantilla_ticket(form):
    """Valida y normaliza los campos de una plantilla (creación o edición), reutilizando las
    mismas listas válidas de tipo/categoría/prioridad/área/sede que usa crear_ticket(), para que
    una plantilla nunca pueda guardar una combinación que el formulario de ticket no aceptaría."""
    nombre = form.get('nombre', '').strip()
    tipo = form.get('tipo', 'Incidente').strip()
    categoria = form.get('categoria', '').strip()
    prioridad = form.get('prioridad', 'Media').strip()
    area = form.get('area', '').strip()
    sede = form.get('sede', '').strip()
    titulo = form.get('titulo', '').strip()
    descripcion = _sanitizar_html_enriquecido(form.get('descripcion', '').strip())

    nombres_categorias = [c['nombre'] for c in _config_ticket_lista('categoria')] or CATEGORIAS_TICKET
    nombres_areas = [a['nombre'] for a in _config_ticket_lista('area')]
    nombres_sedes = [s['nombre'] for s in _config_ticket_lista('sede')]

    if tipo not in TIPOS_TICKET:
        tipo = 'Incidente'
    if categoria not in nombres_categorias:
        categoria = None
    if prioridad not in PRIORIDADES_TICKET:
        prioridad = 'Media'
    area = area if area in nombres_areas else None
    sede = sede if sede in nombres_sedes else None

    es_valido = bool(nombre) and bool(titulo) and not _html_esta_vacio(descripcion)
    return es_valido, {
        'nombre': nombre, 'tipo': tipo, 'categoria': categoria, 'prioridad': prioridad,
        'area': area, 'sede': sede, 'titulo': titulo, 'descripcion': descripcion
    }


@app.route('/tickets/plantillas/nuevo', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_plantilla_ticket():
    es_valido, datos = _validar_datos_plantilla_ticket(request.form)
    if es_valido:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            fecha_actual = obtener_fecha_actual()
            q = (
                "INSERT INTO ticket_plantillas (nombre, tipo, categoria, prioridad, area, sede, titulo, descripcion, estado, creado_por, fecha_creacion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'activo', %s, %s)"
                if db_type == 'postgres' else
                "INSERT INTO ticket_plantillas (nombre, tipo, categoria, prioridad, area, sede, titulo, descripcion, estado, creado_por, fecha_creacion) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'activo', ?, ?)"
            )
            cursor.execute(q, (
                datos['nombre'], datos['tipo'], datos['categoria'], datos['prioridad'], datos['area'],
                datos['sede'], datos['titulo'], datos['descripcion'], session.get('username'), fecha_actual
            ))
            conn.commit()
            registrar_log(session.get('username'), "Plantilla de Ticket", f"Se creó la plantilla '{datos['nombre']}'")
        except Exception as e:
            conn.rollback()
            print(f"Error creando plantilla de ticket: {e}")
        conn.close()
    return redirect(url_for('ver_plantillas_ticket'))


@app.route('/tickets/plantillas/<int:plantilla_id>/editar', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_plantilla_ticket(plantilla_id):
    es_valido, datos = _validar_datos_plantilla_ticket(request.form)
    if es_valido:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            q = (
                "UPDATE ticket_plantillas SET nombre = %s, tipo = %s, categoria = %s, prioridad = %s, area = %s, sede = %s, titulo = %s, descripcion = %s WHERE id = %s"
                if db_type == 'postgres' else
                "UPDATE ticket_plantillas SET nombre = ?, tipo = ?, categoria = ?, prioridad = ?, area = ?, sede = ?, titulo = ?, descripcion = ? WHERE id = ?"
            )
            cursor.execute(q, (
                datos['nombre'], datos['tipo'], datos['categoria'], datos['prioridad'], datos['area'],
                datos['sede'], datos['titulo'], datos['descripcion'], plantilla_id
            ))
            conn.commit()
            registrar_log(session.get('username'), "Plantilla de Ticket", f"Se editó la plantilla '{datos['nombre']}' (#{plantilla_id})")
        except Exception as e:
            conn.rollback()
            print(f"Error editando plantilla de ticket {plantilla_id}: {e}")
        conn.close()
    return redirect(url_for('ver_plantillas_ticket'))


@app.route('/tickets/plantillas/<int:plantilla_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_plantilla_ticket(plantilla_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE ticket_plantillas SET estado = 'eliminado' WHERE id = %s" if db_type == 'postgres' else "UPDATE ticket_plantillas SET estado = 'eliminado' WHERE id = ?"
        cursor.execute(q, (plantilla_id,))
        conn.commit()
        registrar_log(session.get('username'), "Plantilla de Ticket", f"Se eliminó la plantilla #{plantilla_id}")
    except Exception as e:
        conn.rollback()
        print(f"Error eliminando plantilla de ticket {plantilla_id}: {e}")
    conn.close()
    return redirect(url_for('ver_plantillas_ticket'))


# 📊 INDICADORES Y KPIS (solo equipo de soporte): panorama general del módulo de tickets —
# cumplimiento de SLA, distribución por prioridad/categoría/área/sede, satisfacción de los
# solicitantes y tendencia de solicitudes de los últimos 14 días — más exportación a Excel.
def _formatear_duracion_horas(horas):
    """Convierte una duración en horas (float) a un texto corto para mostrar en un KPI:
    '3.5 h' si es menos de un día, '2d 4h' si es un día o más. Devuelve None si no hay dato."""
    if horas is None:
        return None
    dias = int(horas // 24)
    if dias >= 1:
        horas_resto = round(horas - dias * 24)
        return f"{dias}d {horas_resto}h"
    return f"{horas:.1f} h"


def _filtro_sql_tickets(fecha_inicio=None, fecha_fin=None, agente=None, db_type='sqlite'):
    """Arma el WHERE + parámetros para filtrar tickets por rango de fecha de creación (inclusive,
    en formato 'YYYY-MM-DD') y/o por agente asignado. Se usa tanto para el listado principal de
    _calcular_indicadores_tickets como para el conteo del período anterior y el CSV de detalle,
    para no repetir esta lógica en cuatro lugares distintos."""
    ph = '%s' if db_type == 'postgres' else '?'
    condiciones = ["COALESCE(eliminado, 0) = 0"]
    parametros = []
    if fecha_inicio:
        condiciones.append(f"fecha_creacion >= {ph}")
        parametros.append(f"{fecha_inicio} 00:00:00")
    if fecha_fin:
        condiciones.append(f"fecha_creacion <= {ph}")
        parametros.append(f"{fecha_fin} 23:59:59")
    if agente:
        condiciones.append(f"asignado_a = {ph}")
        parametros.append(agente)
    return " AND ".join(condiciones), parametros


def _contar_tickets_por_rango(fecha_inicio, fecha_fin, agente=None):
    """Cuenta tickets creados en un rango de fechas puntual (usado para comparar el período
    filtrado contra el inmediatamente anterior) — consulta aparte porque ese rango queda FUERA
    del conjunto que ya cargó _calcular_indicadores_tickets cuando hay un filtro de fecha activo."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    where_sql, parametros = _filtro_sql_tickets(fecha_inicio, fecha_fin, agente, db_type)
    try:
        cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE {where_sql}", parametros)
        total = cursor.fetchone()[0]
    except Exception as e:
        print(f"Error contando tickets del período anterior ({fecha_inicio} a {fecha_fin}): {e}")
        total = 0
    conn.close()
    return total


MESES_ES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto',
            'septiembre', 'octubre', 'noviembre', 'diciembre']


def _resolver_filtros_indicadores():
    """Lee 'mes' (YYYY-MM), 'fecha_inicio'/'fecha_fin' (YYYY-MM-DD) y 'agente' desde
    request.args y los normaliza a un rango de fechas concreto. 'mes' tiene prioridad sobre
    fecha_inicio/fecha_fin si vienen ambos (evita que un rango manual "viejo" en la URL pise la
    selección del combo de mes). La usan la página de Indicadores, el Tablero Ejecutivo y las
    tres exportaciones, para que todas apliquen exactamente el mismo filtro."""
    mes = (request.args.get('mes') or '').strip()
    fecha_inicio = (request.args.get('fecha_inicio') or '').strip() or None
    fecha_fin = (request.args.get('fecha_fin') or '').strip() or None
    agente = (request.args.get('agente') or '').strip() or None
    if mes:
        try:
            anio_str, mes_str = mes.split('-')
            anio, mes_num = int(anio_str), int(mes_str)
            fecha_inicio = f"{anio:04d}-{mes_num:02d}-01"
            ultimo_dia = calendar.monthrange(anio, mes_num)[1]
            fecha_fin = f"{anio:04d}-{mes_num:02d}-{ultimo_dia:02d}"
        except Exception:
            mes = ''
    return fecha_inicio, fecha_fin, agente, mes


def _opciones_meses_filtro():
    """Últimos 12 meses (incluido el actual) para el selector 'Mes' del filtro de Indicadores:
    [{'valor': 'YYYY-MM', 'etiqueta': 'Enero 2026'}, ...] del más reciente al más antiguo."""
    hoy = datetime.now(ZONA_HORARIA_COLOMBIA)
    opciones = []
    anio, mes_num = hoy.year, hoy.month
    for _ in range(12):
        opciones.append({'valor': f"{anio:04d}-{mes_num:02d}", 'etiqueta': f"{MESES_ES[mes_num - 1].capitalize()} {anio}"})
        mes_num -= 1
        if mes_num == 0:
            mes_num = 12
            anio -= 1
    return opciones


def _opciones_agentes_filtro():
    """Agentes/administradores con al menos un ticket asignado (activo, no eliminado), para el
    selector 'Analista' del filtro de Indicadores — [{'valor': usuario, 'etiqueta': nombre}]
    ordenados por nombre para mostrar."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT asignado_a FROM tickets WHERE asignado_a IS NOT NULL AND asignado_a != '' "
            "AND COALESCE(eliminado, 0) = 0 ORDER BY asignado_a"
        )
        usuarios = [r[0] for r in cursor.fetchall()]
        conn.close()
    except Exception as e:
        print(f"Error cargando agentes para el filtro de indicadores: {e}")
        usuarios = []
    nombres = _mapa_nombres_usuarios()
    return sorted(
        ({'valor': u, 'etiqueta': _nombre_para_mostrar(u, nombres)} for u in usuarios),
        key=lambda x: (x['etiqueta'] or '').lower()
    )


def _calcular_indicadores_tickets(fecha_inicio=None, fecha_fin=None, agente=None):
    """Calcula todos los indicadores/KPIs de Tickets (por estado, prioridad, tipo, cumplimiento
    de SLA, categoría/área/sede, calificación promedio, top agentes y tendencia) — opcionalmente
    acotado a un rango de fechas de creación ('YYYY-MM-DD') y/o a un agente puntual (por
    'asignado_a'). Sin filtros se comporta exactamente como antes (últimos 14 días + comparativo
    contra la quincena anterior). Lo usan tanto la página de Indicadores como sus exportaciones a
    PDF/Excel/CSV, para no duplicar la lógica de agregación en varios lugares."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    where_sql, parametros = _filtro_sql_tickets(fecha_inicio, fecha_fin, agente, db_type)
    try:
        cursor.execute(
            f"SELECT id, tipo, categoria, prioridad, estado, creado_por, asignado_a, fecha_creacion, "
            f"sla_resolucion_limite, sla_resolucion_cumplida, calificacion, area, sede FROM tickets WHERE {where_sql}",
            parametros
        )
        rows = cursor.fetchall()
    except Exception as e:
        print(f"Error consultando indicadores de tickets: {e}")
        rows = []
    # ✅ Tareas completadas por agente: aproximación simple a las "Horas-Hombre" de Aranda
    # Service Desk sin necesitar registrar duración exacta (que Arkiv no pide al crear una
    # tarea) — cuenta cuántas tareas cerró cada persona, no cuánto tiempo le tomaron. Respeta el
    # mismo filtro de fecha (por fecha_completada) y de agente (por responsable) que el resto del
    # tablero, para que "Top agentes por tareas completadas" también se acote al período elegido.
    ph = '%s' if db_type == 'postgres' else '?'
    condiciones_tareas = ["estado = 'completada'", "responsable IS NOT NULL"]
    parametros_tareas = []
    if fecha_inicio:
        condiciones_tareas.append(f"fecha_completada >= {ph}")
        parametros_tareas.append(f"{fecha_inicio} 00:00:00")
    if fecha_fin:
        condiciones_tareas.append(f"fecha_completada <= {ph}")
        parametros_tareas.append(f"{fecha_fin} 23:59:59")
    if agente:
        condiciones_tareas.append(f"responsable = {ph}")
        parametros_tareas.append(agente)
    try:
        cursor.execute(
            f"SELECT responsable, COUNT(*) FROM tickets_tareas WHERE {' AND '.join(condiciones_tareas)} GROUP BY responsable",
            parametros_tareas
        )
        filas_tareas_agente = cursor.fetchall()
    except Exception as e:
        print(f"Error consultando tareas completadas por agente: {e}")
        filas_tareas_agente = []
    conn.close()

    tickets = []
    for r in rows:
        t = {
            'id': r[0], 'tipo': r[1] or 'Incidente', 'categoria': r[2], 'prioridad': r[3],
            'estado': r[4], 'creado_por': r[5], 'asignado_a': r[6], 'fecha_creacion': r[7],
            'sla_resolucion_limite': r[8], 'sla_resolucion_cumplida': r[9],
            'calificacion': r[10], 'area': r[11], 'sede': r[12]
        }
        t['sla'] = _calcular_sla_ticket(t)
        t['cumplimiento'] = _bucket_cumplimiento_ticket(t)
        tickets.append(t)

    total = len(tickets)
    por_estado = {e: 0 for e in ESTADOS_TICKET}
    por_prioridad = {p: 0 for p in PRIORIDADES_TICKET}
    por_tipo = {tp: 0 for tp in TIPOS_TICKET}
    por_cumplimiento = {'vigente': 0, 'proximo_a_vencer': 0, 'vencido': 0, 'cerrado': 0}
    por_categoria, por_area, por_sede, por_agente = {}, {}, {}, {}
    calificaciones = []
    conteo_por_fecha = {}
    # ⏱️ Duración real de resolución (no aproximada): 'sla_resolucion_cumplida' se graba la
    # primera vez que el ticket llega a Resuelto/Cerrado y se limpia si se reabre (ver
    # modificar_ticket()), así que a diferencia de 'fecha_actualizacion' sí es el momento exacto
    # en que se resolvió — resta contra 'fecha_creacion' para el tiempo promedio de resolución.
    duraciones_resolucion_horas = []

    for t in tickets:
        if t['estado'] in por_estado:
            por_estado[t['estado']] += 1
        if t['prioridad'] in por_prioridad:
            por_prioridad[t['prioridad']] += 1
        if t['tipo'] in por_tipo:
            por_tipo[t['tipo']] += 1
        por_cumplimiento[t['cumplimiento']] = por_cumplimiento.get(t['cumplimiento'], 0) + 1
        cat = t['categoria'] or 'Sin categoría'
        por_categoria[cat] = por_categoria.get(cat, 0) + 1
        area_n = t['area'] or 'Sin área'
        por_area[area_n] = por_area.get(area_n, 0) + 1
        sede_n = t['sede'] or 'Sin sede'
        por_sede[sede_n] = por_sede.get(sede_n, 0) + 1
        if t['calificacion']:
            calificaciones.append(t['calificacion'])
        if t['estado'] in ('Resuelto', 'Cerrado') and t['asignado_a']:
            por_agente[t['asignado_a']] = por_agente.get(t['asignado_a'], 0) + 1
        fecha_corta = (t['fecha_creacion'] or '')[:10]
        if fecha_corta:
            conteo_por_fecha[fecha_corta] = conteo_por_fecha.get(fecha_corta, 0) + 1
        if t.get('sla_resolucion_cumplida'):
            inicio_res = _parsear_fecha_ticket(t['fecha_creacion'])
            fin_res = _parsear_fecha_ticket(t['sla_resolucion_cumplida'])
            if inicio_res and fin_res and fin_res >= inicio_res:
                duraciones_resolucion_horas.append((fin_res - inicio_res).total_seconds() / 3600)

    hoy = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)
    filtro_fecha_activo = bool(fecha_inicio or fecha_fin)

    if filtro_fecha_activo:
        # 📅 Con un filtro de fecha activo, la tendencia deja de ser "últimos 14 días desde hoy"
        # y pasa a recorrer día por día el rango elegido — así el gráfico realmente muestra el
        # período que se está consultando. Si falta un extremo, se usa la fecha más antigua/nueva
        # que realmente aparece entre los tickets ya filtrados.
        fechas_en_rango = sorted(conteo_por_fecha.keys())
        ini_rango_str = fecha_inicio or (fechas_en_rango[0] if fechas_en_rango else hoy.strftime('%Y-%m-%d'))
        fin_rango_str = fecha_fin or (fechas_en_rango[-1] if fechas_en_rango else hoy.strftime('%Y-%m-%d'))
        try:
            ini_rango = datetime.strptime(ini_rango_str[:10], '%Y-%m-%d')
            fin_rango = datetime.strptime(fin_rango_str[:10], '%Y-%m-%d')
        except Exception:
            ini_rango = fin_rango = hoy

        dias_tendencia = []
        cursor_dia = ini_rango
        # 🛡️ Tope de seguridad: si alguien filtra un rango de años, el gráfico deja de sumar
        # puntos después de 366 días en vez de intentar dibujar miles de barras.
        while cursor_dia <= fin_rango and len(dias_tendencia) < 366:
            dia_str = cursor_dia.strftime('%Y-%m-%d')
            dias_tendencia.append({'fecha': dia_str, 'cantidad': conteo_por_fecha.get(dia_str, 0)})
            cursor_dia += timedelta(days=1)

        conteo_periodo_actual = sum(d['cantidad'] for d in dias_tendencia)
        duracion_periodo = (fin_rango - ini_rango).days + 1
        fin_anterior = ini_rango - timedelta(days=1)
        ini_anterior = fin_anterior - timedelta(days=duracion_periodo - 1)
        conteo_periodo_anterior = _contar_tickets_por_rango(
            ini_anterior.strftime('%Y-%m-%d'), fin_anterior.strftime('%Y-%m-%d'), agente
        )
        titulo_comparativo_periodo = 'período anterior'
    else:
        dias_tendencia = []
        for i in range(13, -1, -1):
            dia = (hoy - timedelta(days=i)).strftime('%Y-%m-%d')
            dias_tendencia.append({'fecha': dia, 'cantidad': conteo_por_fecha.get(dia, 0)})

        # 📈 Comparativo contra el período de 14 días inmediatamente anterior (días 14 a 27
        # atrás), para que el tablero no solo muestre "cuántas se crearon" sino si el volumen
        # está subiendo o bajando frente a la quincena previa — reutiliza 'conteo_por_fecha' (ya
        # cuenta TODOS los tickets por fecha de creación) en vez de otra consulta a la BD.
        conteo_periodo_actual = sum(d['cantidad'] for d in dias_tendencia)
        conteo_periodo_anterior = sum(
            conteo_por_fecha.get((hoy - timedelta(days=i)).strftime('%Y-%m-%d'), 0)
            for i in range(27, 13, -1)
        )
        titulo_comparativo_periodo = 'quincena anterior'

    if conteo_periodo_anterior > 0:
        variacion_pct_tendencia = round((conteo_periodo_actual - conteo_periodo_anterior) / conteo_periodo_anterior * 100, 1)
    else:
        variacion_pct_tendencia = None

    tiempo_promedio_resolucion_horas = (
        round(sum(duraciones_resolucion_horas) / len(duraciones_resolucion_horas), 1)
        if duraciones_resolucion_horas else None
    )
    tiempo_promedio_resolucion_texto = _formatear_duracion_horas(tiempo_promedio_resolucion_horas)

    promedio_calificacion = round(sum(calificaciones) / len(calificaciones), 2) if calificaciones else None
    # 👤 Igual que en el resto de Tickets, se muestra el nombre/alias real del agente, no su
    # usuario de inicio de sesión.
    nombres_usuarios_ind = _mapa_nombres_usuarios()
    top_agentes = [
        (_nombre_para_mostrar(u, nombres_usuarios_ind), cant)
        for u, cant in sorted(por_agente.items(), key=lambda x: x[1], reverse=True)[:5]
    ]
    # ✅ Mismo criterio que 'top_agentes' pero contando tareas internas completadas (ver
    # tickets_tareas) en vez de tickets resueltos/cerrados — aproximación a las "Horas-Hombre"
    # de Aranda Service Desk (ver comentario donde se consultó filas_tareas_agente).
    top_agentes_tareas = [
        (_nombre_para_mostrar(u, nombres_usuarios_ind), cant)
        for u, cant in sorted(filas_tareas_agente, key=lambda x: x[1], reverse=True)[:5]
    ]

    return {
        'total': total, 'por_estado': por_estado, 'por_prioridad': por_prioridad, 'por_tipo': por_tipo,
        'por_cumplimiento': por_cumplimiento, 'por_categoria': por_categoria,
        'por_area': por_area, 'por_sede': por_sede, 'dias_tendencia': dias_tendencia,
        'promedio_calificacion': promedio_calificacion, 'total_calificaciones': len(calificaciones),
        'top_agentes': top_agentes, 'top_agentes_tareas': top_agentes_tareas,
        'conteo_periodo_actual': conteo_periodo_actual, 'conteo_periodo_anterior': conteo_periodo_anterior,
        'variacion_pct_tendencia': variacion_pct_tendencia,
        'tiempo_promedio_resolucion_horas': tiempo_promedio_resolucion_horas,
        'tiempo_promedio_resolucion_texto': tiempo_promedio_resolucion_texto,
        'titulo_comparativo_periodo': titulo_comparativo_periodo,
    }


@app.route('/tickets/indicadores')
@login_required
@agente_o_admin_required
def indicadores_tickets():
    # 🔔 Aviso perezoso de SLA por vencer/vencido — ver _revisar_alertas_sla(). Esta página ya
    # está gateada a admin/agente, así que se corre sin condición adicional.
    _revisar_alertas_sla()

    fecha_inicio, fecha_fin, agente, mes = _resolver_filtros_indicadores()
    ind = _calcular_indicadores_tickets(fecha_inicio, fecha_fin, agente)
    return render_template(
        'tickets_indicadores.html', es_soporte=True,
        filtro_mes=mes, filtro_fecha_inicio=fecha_inicio or '', filtro_fecha_fin=fecha_fin or '',
        filtro_agente=agente or '', opciones_meses=_opciones_meses_filtro(), opciones_agentes=_opciones_agentes_filtro(),
        **ind
    )


@app.route('/tickets/indicadores/exportar_csv')
@login_required
@agente_o_admin_required
def exportar_indicadores_tickets():
    fecha_inicio, fecha_fin, agente, _mes = _resolver_filtros_indicadores()
    conn, db_type = get_db()
    cursor = conn.cursor()
    where_sql, parametros = _filtro_sql_tickets(fecha_inicio, fecha_fin, agente, db_type)
    cursor.execute(
        f"SELECT id, tipo, titulo, categoria, area, sede, prioridad, estado, creado_por, asignado_a, fecha_creacion, "
        f"fecha_actualizacion, calificacion FROM tickets WHERE {where_sql} ORDER BY id DESC",
        parametros
    )
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['CÓDIGO', 'TIPO', 'TÍTULO', 'CATEGORÍA', 'ÁREA', 'SEDE', 'PRIORIDAD', 'ESTADO', 'CREADO POR', 'ASIGNADO A', 'FECHA CREACIÓN', 'ÚLTIMA ACTUALIZACIÓN', 'CALIFICACIÓN'])

    for row in rows:
        tipo_t = row[1] or 'Incidente'
        codigo = _codigo_ticket(tipo_t, row[0], row[10])
        writer.writerow([codigo, tipo_t, row[2], row[3], row[4] or '', row[5] or '', row[6], row[7], row[8], row[9] or '', row[10], row[11], row[12] or ''])

    csv_bytes = '﻿' + output.getvalue()
    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y%m%d_%H%M")
    filename = f"Arkiv_Indicadores_Tickets_{fecha_filename}.csv"

    headers = {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    registrar_log(session.get('username'), "Exportación de Indicadores", f"Exportó {len(rows)} tickets a Excel/CSV")
    return Response(csv_bytes, headers=headers, status=200)


# 📊 Traducciones legibles para las tablas de los reportes de Indicadores (PDF/Excel) — los
# diccionarios internos usan claves técnicas ('proximo_a_vencer') que no se le muestran así a
# Tomas/gerencia en el reporte descargado.
ETIQUETAS_CUMPLIMIENTO_SLA = {
    'vigente': 'Vigente', 'proximo_a_vencer': 'Próximo a vencer',
    'vencido': 'Vencido', 'cerrado': 'Cerrado'
}


@app.route('/tickets/indicadores/exportar_pdf')
@login_required
@agente_o_admin_required
def exportar_indicadores_tickets_pdf():
    """Reporte de indicadores de Tickets en PDF — mismo contenido que se ve en pantalla, listo
    para imprimir o adjuntar en un correo a gerencia."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.legends import Legend

    fecha_inicio, fecha_fin, agente, _mes = _resolver_filtros_indicadores()
    ind = _calcular_indicadores_tickets(fecha_inicio, fecha_fin, agente)
    estilos = getSampleStyleSheet()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    ahora_txt = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d %H:%M')

    def tabla_conteos(titulo, datos, encabezados=('Categoría', 'Cantidad')):
        elementos = [Paragraph(titulo, estilos['Heading3'])]
        filas = [list(encabezados)] + [[k, str(v)] for k, v in datos]
        t = Table(filas, colWidths=[10 * cm, 4 * cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ]))
        elementos.append(t)
        elementos.append(Spacer(1, 0.5 * cm))
        return elementos

    # 📊 Los mismos gráficos que se ven en pantalla (Chart.js), redibujados aquí con
    # reportlab.graphics para que el PDF descargado no sea solo tablas — un vistazo visual
    # igual de rápido que el del dashboard, aunque en PDF no puede ser interactivo.
    def grafico_barras(labels, valores, color_hex, ancho=16 * cm, alto=6 * cm):
        if not labels:
            return None
        d = Drawing(ancho, alto)
        bc = VerticalBarChart()
        bc.x = 45
        bc.y = 45
        bc.height = alto - 60
        bc.width = ancho - 65
        bc.data = [valores]
        bc.categoryAxis.categoryNames = [str(l)[:14] for l in labels]
        bc.categoryAxis.labels.angle = 25
        bc.categoryAxis.labels.dy = -12
        bc.categoryAxis.labels.dx = -4
        bc.categoryAxis.labels.fontSize = 7
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 7
        bc.bars[0].fillColor = colors.HexColor(color_hex)
        bc.bars.strokeColor = None
        bc.barLabelFormat = '%d'
        bc.barLabels.fontSize = 7
        bc.barLabels.nudge = 8
        d.add(bc)
        return d

    def grafico_pie(labels, valores, colores_hex, ancho=16 * cm, alto=6 * cm):
        if sum(valores) == 0:
            return None
        d = Drawing(ancho, alto)
        pie = Pie()
        pie.x = 60
        pie.y = 20
        pie.width = 130
        pie.height = 130
        pie.data = valores
        pie.labels = [str(v) for v in valores]
        pie.simpleLabels = True
        pie.slices.strokeWidth = 0.75
        pie.slices.strokeColor = colors.white
        for i, c in enumerate(colores_hex):
            pie.slices[i].fillColor = colors.HexColor(c)
        d.add(pie)
        leyenda = Legend()
        leyenda.x = 230
        leyenda.y = 100
        leyenda.dx = 8
        leyenda.dy = 8
        leyenda.fontSize = 8
        leyenda.alignment = 'right'
        leyenda.colorNamePairs = [(colors.HexColor(c), l) for c, l in zip(colores_hex, labels)]
        d.add(leyenda)
        return d

    def grafico_tendencia(dias, ancho=16 * cm, alto=6 * cm):
        if not dias:
            return None
        d = Drawing(ancho, alto)
        lc = HorizontalLineChart()
        lc.x = 45
        lc.y = 30
        lc.height = alto - 50
        lc.width = ancho - 65
        lc.data = [[dia['cantidad'] for dia in dias]]
        lc.categoryAxis.categoryNames = [dia['fecha'][5:] for dia in dias]
        lc.categoryAxis.labels.angle = 30
        lc.categoryAxis.labels.dy = -12
        lc.categoryAxis.labels.fontSize = 6.5
        lc.valueAxis.valueMin = 0
        lc.valueAxis.labels.fontSize = 7
        lc.lines[0].strokeColor = colors.HexColor('#E67E00')
        lc.lines[0].strokeWidth = 2
        lc.lines[0].symbol = None
        d.add(lc)
        return d

    contenido = [
        Paragraph("Arkiv — Indicadores de Tickets de Soporte TI", estilos['Title']),
        Paragraph(f"Generado el {ahora_txt} (hora Colombia) — Total de solicitudes: {ind['total']}", estilos['Normal']),
        Spacer(1, 0.5 * cm),
    ]
    if ind['promedio_calificacion'] is not None:
        contenido.append(Paragraph(
            f"Calificación promedio de satisfacción: {ind['promedio_calificacion']}/5 "
            f"({ind['total_calificaciones']} calificaciones recibidas)", estilos['Normal']
        ))
        contenido.append(Spacer(1, 0.5 * cm))
    if ind['tiempo_promedio_resolucion_texto']:
        contenido.append(Paragraph(f"Tiempo promedio de resolución: {ind['tiempo_promedio_resolucion_texto']}", estilos['Normal']))
        contenido.append(Spacer(1, 0.3 * cm))
    if ind['variacion_pct_tendencia'] is not None:
        signo = '+' if ind['variacion_pct_tendencia'] >= 0 else ''
        contenido.append(Paragraph(
            f"Solicitudes creadas vs. quincena anterior: {signo}{ind['variacion_pct_tendencia']}% "
            f"({ind['conteo_periodo_actual']} vs {ind['conteo_periodo_anterior']})", estilos['Normal']
        ))
        contenido.append(Spacer(1, 0.5 * cm))

    dias_tendencia = ind['dias_tendencia']
    contenido.append(Paragraph("Solicitudes creadas — últimos 14 días", estilos['Heading3']))
    g = grafico_tendencia(dias_tendencia)
    if g:
        contenido += [g, Spacer(1, 0.3 * cm)]

    # 🎨 Mismos colores de marca que en el tablero en pantalla (ver tickets_indicadores.html):
    # Azul Real/variaciones y Cyan para lo estructural/operativo, Naranja para lo que implica
    # alerta o plazo, y '#f43f5e' (rose) reservado solo para "Vencidos" como única excepción.
    contenido += tabla_conteos("Por estado", ind['por_estado'].items(), ('Estado', 'Cantidad'))
    g = grafico_barras(list(ind['por_estado'].keys()), list(ind['por_estado'].values()), '#3D6BB5')
    if g:
        contenido += [g, Spacer(1, 0.5 * cm)]

    contenido += tabla_conteos("Por prioridad", ind['por_prioridad'].items(), ('Prioridad', 'Cantidad'))
    g = grafico_barras(list(ind['por_prioridad'].keys()), list(ind['por_prioridad'].values()), '#33B8E0')
    if g:
        contenido += [g, Spacer(1, 0.5 * cm)]

    contenido += tabla_conteos("Por tipo", ind['por_tipo'].items(), ('Tipo', 'Cantidad'))
    g = grafico_pie(list(ind['por_tipo'].keys()), list(ind['por_tipo'].values()), ['#1A4B9C', '#00A3DA'])
    if g:
        contenido += [g, Spacer(1, 0.5 * cm)]

    cumplimiento_items = list(ind['por_cumplimiento'].items())
    contenido += tabla_conteos(
        "Cumplimiento de SLA",
        [(ETIQUETAS_CUMPLIMIENTO_SLA.get(k, k), v) for k, v in cumplimiento_items],
        ('Estado de SLA', 'Cantidad')
    )
    g = grafico_pie(
        [ETIQUETAS_CUMPLIMIENTO_SLA.get(k, k) for k, v in cumplimiento_items],
        [v for k, v in cumplimiento_items],
        ['#00A3DA', '#E67E00', '#f43f5e', '#64748b']
    )
    if g:
        contenido += [g, Spacer(1, 0.5 * cm)]

    if ind['por_categoria']:
        cat_ordenada = sorted(ind['por_categoria'].items(), key=lambda x: x[1], reverse=True)
        contenido += tabla_conteos("Por categoría", cat_ordenada, ('Categoría', 'Cantidad'))
        g = grafico_barras([k for k, v in cat_ordenada], [v for k, v in cat_ordenada], '#6690D6')
        if g:
            contenido += [g, Spacer(1, 0.5 * cm)]
    if ind['por_area']:
        area_ordenada = sorted(ind['por_area'].items(), key=lambda x: x[1], reverse=True)
        contenido += tabla_conteos("Por área", area_ordenada, ('Área', 'Cantidad'))
        g = grafico_barras([k for k, v in area_ordenada], [v for k, v in area_ordenada], '#007FAE')
        if g:
            contenido += [g, Spacer(1, 0.5 * cm)]
    if ind['por_sede']:
        sede_ordenada = sorted(ind['por_sede'].items(), key=lambda x: x[1], reverse=True)
        contenido += tabla_conteos("Por sede", sede_ordenada, ('Sede', 'Cantidad'))
        g = grafico_barras([k for k, v in sede_ordenada], [v for k, v in sede_ordenada], '#1A4B9C')
        if g:
            contenido += [g, Spacer(1, 0.5 * cm)]
    if ind['top_agentes']:
        contenido += tabla_conteos("Top agentes (tickets resueltos/cerrados)", ind['top_agentes'], ('Agente', 'Tickets'))
    contenido += tabla_conteos("Tendencia — solicitudes creadas por día (últimos 14 días)", [(d['fecha'], d['cantidad']) for d in dias_tendencia], ('Fecha', 'Cantidad'))

    doc.build(contenido)
    buffer.seek(0)

    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y%m%d_%H%M")
    filename = f"Arkiv_Indicadores_Tickets_{fecha_filename}.pdf"
    registrar_log(session.get('username'), "Exportación de Indicadores", "Exportó los indicadores de Tickets a PDF")
    return Response(buffer.getvalue(), headers={
        'Content-Type': 'application/pdf',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }, status=200)


@app.route('/tickets/indicadores/exportar_xlsx')
@login_required
@agente_o_admin_required
def exportar_indicadores_tickets_xlsx():
    """Reporte de indicadores de Tickets en un libro de Excel real (con formato), a diferencia
    del CSV de detalle que ya existía — una hoja de Resumen con los KPIs y una de Detalle con
    el listado completo de tickets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties

    fecha_inicio, fecha_fin, agente, _mes = _resolver_filtros_indicadores()
    ind = _calcular_indicadores_tickets(fecha_inicio, fecha_fin, agente)
    wb = openpyxl.Workbook()

    relleno_encabezado = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    fuente_encabezado = Font(color='FFFFFF', bold=True)
    fuente_titulo = Font(bold=True, size=14)

    resumen = wb.active
    resumen.title = 'Resumen'
    fila = 1
    resumen.cell(row=fila, column=1, value='Arkiv — Indicadores de Tickets de Soporte TI').font = fuente_titulo
    fila += 1
    ahora_txt = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m-%d %H:%M')
    resumen.cell(row=fila, column=1, value=f"Generado el {ahora_txt} — Total de solicitudes: {ind['total']}")
    fila += 2
    if ind['promedio_calificacion'] is not None:
        resumen.cell(row=fila, column=1, value=(
            f"Calificación promedio de satisfacción: {ind['promedio_calificacion']}/5 "
            f"({ind['total_calificaciones']} calificaciones)"
        ))
        fila += 2
    if ind['tiempo_promedio_resolucion_texto']:
        resumen.cell(row=fila, column=1, value=f"Tiempo promedio de resolución: {ind['tiempo_promedio_resolucion_texto']}")
        fila += 2
    if ind['variacion_pct_tendencia'] is not None:
        signo = '+' if ind['variacion_pct_tendencia'] >= 0 else ''
        resumen.cell(row=fila, column=1, value=(
            f"Solicitudes creadas vs. quincena anterior: {signo}{ind['variacion_pct_tendencia']}% "
            f"({ind['conteo_periodo_actual']} vs {ind['conteo_periodo_anterior']})"
        ))
        fila += 2

    # 📊 Guardamos, por cada sección volcada, en qué filas quedó su encabezado y sus datos —
    # así los gráficos nativos de Excel que se arman más abajo pueden referenciar exactamente
    # esas celdas (Reference), en vez de duplicar los números a mano.
    rangos_secciones = {}

    def volcar_seccion(titulo, datos, encabezados):
        nonlocal fila
        resumen.cell(row=fila, column=1, value=titulo).font = Font(bold=True, size=12)
        fila += 1
        fila_encabezado = fila
        for col, texto in enumerate(encabezados, start=1):
            c = resumen.cell(row=fila, column=col, value=texto)
            c.fill = relleno_encabezado
            c.font = fuente_encabezado
        fila += 1
        fila_inicio_datos = fila
        datos = list(datos)
        for clave, valor in datos:
            resumen.cell(row=fila, column=1, value=clave)
            resumen.cell(row=fila, column=2, value=valor)
            fila += 1
        fila_fin_datos = fila - 1
        fila += 1
        rangos_secciones[titulo] = {
            'encabezado': fila_encabezado, 'inicio': fila_inicio_datos, 'fin': fila_fin_datos, 'filas': len(datos)
        }

    volcar_seccion('Por estado', ind['por_estado'].items(), ['Estado', 'Cantidad'])
    volcar_seccion('Por prioridad', ind['por_prioridad'].items(), ['Prioridad', 'Cantidad'])
    volcar_seccion('Por tipo', ind['por_tipo'].items(), ['Tipo', 'Cantidad'])
    volcar_seccion(
        'Cumplimiento de SLA',
        [(ETIQUETAS_CUMPLIMIENTO_SLA.get(k, k), v) for k, v in ind['por_cumplimiento'].items()],
        ['Estado de SLA', 'Cantidad']
    )
    if ind['por_categoria']:
        volcar_seccion('Por categoría', sorted(ind['por_categoria'].items(), key=lambda x: x[1], reverse=True), ['Categoría', 'Cantidad'])
    if ind['por_area']:
        volcar_seccion('Por área', sorted(ind['por_area'].items(), key=lambda x: x[1], reverse=True), ['Área', 'Cantidad'])
    if ind['por_sede']:
        volcar_seccion('Por sede', sorted(ind['por_sede'].items(), key=lambda x: x[1], reverse=True), ['Sede', 'Cantidad'])
    if ind['top_agentes']:
        volcar_seccion('Top agentes (tickets resueltos/cerrados)', ind['top_agentes'], ['Agente', 'Tickets'])
    volcar_seccion('Tendencia — últimos 14 días', [(d['fecha'], d['cantidad']) for d in ind['dias_tendencia']], ['Fecha', 'Cantidad'])

    resumen.column_dimensions['A'].width = 32
    resumen.column_dimensions['B'].width = 14

    # 📊 Hoja de "Gráficos": los mismos gráficos del dashboard (Chart.js), pero como gráficos
    # NATIVOS de Excel (BarChart/DoughnutChart/LineChart) — a diferencia de una imagen pegada,
    # estos son igual de interactivos que en el navegador: se puede pasar el mouse para ver el
    # valor exacto, ocultar series desde la leyenda, cambiar el tipo de gráfico, y si alguien
    # edita los números de la hoja "Resumen" el gráfico se actualiza solo.
    graficos = wb.create_sheet('Gráficos', 1)
    graficos.sheet_view.showGridLines = False
    fila_grafico = 1

    def agregar_grafico(chart, titulo):
        nonlocal fila_grafico
        chart.title = titulo
        chart.width = 18
        chart.height = 9
        chart.style = 10
        graficos.add_chart(chart, f'A{fila_grafico}')
        fila_grafico += 19

    r = rangos_secciones.get('Tendencia — últimos 14 días')
    if r and r['filas']:
        chart = LineChart()
        chart.grouping = 'standard'
        datos_ref = Reference(resumen, min_col=2, min_row=r['encabezado'], max_row=r['fin'])
        cats_ref = Reference(resumen, min_col=1, min_row=r['inicio'], max_row=r['fin'])
        chart.add_data(datos_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.y_axis.majorGridlines = None
        chart.series[0].graphicalProperties = GraphicalProperties(solidFill=None)
        chart.series[0].graphicalProperties.line.solidFill = 'E67E00'
        chart.series[0].graphicalProperties.line.width = 22000
        chart.series[0].smooth = False
        agregar_grafico(chart, 'Solicitudes creadas — últimos 14 días')

    r = rangos_secciones.get('Cumplimiento de SLA')
    if r and r['filas']:
        chart = DoughnutChart()
        datos_ref = Reference(resumen, min_col=2, min_row=r['encabezado'], max_row=r['fin'])
        cats_ref = Reference(resumen, min_col=1, min_row=r['inicio'], max_row=r['fin'])
        chart.add_data(datos_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        # 🎨 Mismos colores que el donut de Chart.js en pantalla, en el mismo orden
        # (vigente/próx. a vencer/vencido/cerrado) — paleta de marca Preventiva, con 'F43F5E'
        # (rose) como única excepción para el estado crítico "Vencido".
        colores_sla = ['00A3DA', 'E67E00', 'F43F5E', '64748B']
        chart.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(solidFill=c)) for i, c in enumerate(colores_sla[:r['filas']])
        ]
        agregar_grafico(chart, 'Cumplimiento de SLA')

    r = rangos_secciones.get('Por tipo')
    if r and r['filas']:
        chart = DoughnutChart()
        datos_ref = Reference(resumen, min_col=2, min_row=r['encabezado'], max_row=r['fin'])
        cats_ref = Reference(resumen, min_col=1, min_row=r['inicio'], max_row=r['fin'])
        chart.add_data(datos_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        colores_tipo = ['1A4B9C', '00A3DA', '6690D6']
        chart.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(solidFill=c)) for i, c in enumerate(colores_tipo[:r['filas']])
        ]
        agregar_grafico(chart, 'Por tipo')

    colores_barras = {
        'Por estado': '3D6BB5', 'Por prioridad': '33B8E0',
        'Por categoría': '6690D6', 'Por área': '007FAE', 'Por sede': '1A4B9C'
    }
    for titulo_seccion, color_hex in colores_barras.items():
        r = rangos_secciones.get(titulo_seccion)
        if r and r['filas']:
            chart = BarChart()
            chart.type = 'col'
            datos_ref = Reference(resumen, min_col=2, min_row=r['encabezado'], max_row=r['fin'])
            cats_ref = Reference(resumen, min_col=1, min_row=r['inicio'], max_row=r['fin'])
            chart.add_data(datos_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None
            chart.series[0].graphicalProperties = GraphicalProperties(solidFill=color_hex)
            agregar_grafico(chart, titulo_seccion)

    # 📋 Hoja de detalle: el mismo listado completo de tickets que ya ofrece la exportación CSV,
    # para tener ambas vistas (resumen y detalle) en un solo archivo.
    detalle = wb.create_sheet('Detalle')
    encabezados_detalle = ['CÓDIGO', 'TIPO', 'TÍTULO', 'CATEGORÍA', 'ÁREA', 'SEDE', 'PRIORIDAD', 'ESTADO', 'CREADO POR', 'ASIGNADO A', 'FECHA CREACIÓN', 'ÚLTIMA ACTUALIZACIÓN', 'CALIFICACIÓN']
    for col, texto in enumerate(encabezados_detalle, start=1):
        c = detalle.cell(row=1, column=col, value=texto)
        c.fill = relleno_encabezado
        c.font = fuente_encabezado

    conn, db_type = get_db()
    cursor = conn.cursor()
    where_sql, parametros = _filtro_sql_tickets(fecha_inicio, fecha_fin, agente, db_type)
    cursor.execute(
        f"SELECT id, tipo, titulo, categoria, area, sede, prioridad, estado, creado_por, asignado_a, fecha_creacion, "
        f"fecha_actualizacion, calificacion FROM tickets WHERE {where_sql} ORDER BY id DESC",
        parametros
    )
    rows = cursor.fetchall()
    conn.close()

    for i, row in enumerate(rows, start=2):
        tipo_t = row[1] or 'Incidente'
        codigo = _codigo_ticket(tipo_t, row[0], row[10])
        valores = [codigo, tipo_t, row[2], row[3], row[4] or '', row[5] or '', row[6], row[7], row[8], row[9] or '', row[10], row[11], row[12] or '']
        for col, valor in enumerate(valores, start=1):
            detalle.cell(row=i, column=col, value=valor)
    for letra, ancho in zip('ABCDEFGHIJKLM', [16, 12, 30, 18, 16, 16, 12, 12, 14, 14, 18, 18, 12]):
        detalle.column_dimensions[letra].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y%m%d_%H%M")
    filename = f"Arkiv_Indicadores_Tickets_{fecha_filename}.xlsx"
    registrar_log(session.get('username'), "Exportación de Indicadores", "Exportó los indicadores de Tickets a Excel")
    return Response(buffer.getvalue(), headers={
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }, status=200)


# 📈 TABLERO EJECUTIVO (solo Admin/gerencia): panorama de un vistazo de TODOS los módulos de
# Arkiv en una sola pantalla — a diferencia de /tickets/indicadores (que profundiza solo en
# Tickets, con gráficos y exportaciones a PDF/Excel/CSV, accesible también a agentes), este
# tablero da un resumen liviano de cada módulo con un enlace directo al detalle de cada uno.
# Reutiliza _calcular_indicadores_tickets() para no duplicar esa agregación.
def _calcular_tablero_ejecutivo(fecha_inicio=None, fecha_fin=None, agente=None):
    conn, db_type = get_db()
    cursor = conn.cursor()

    ind_tickets = _calcular_indicadores_tickets(fecha_inicio, fecha_fin, agente)
    total_tickets = ind_tickets['total']
    tickets_vencidos = ind_tickets['por_cumplimiento'].get('vencido', 0)
    pct_sla_cumplido = round((total_tickets - tickets_vencidos) / total_tickets * 100, 1) if total_tickets else None

    try:
        cursor.execute("SELECT rol, COUNT(*) FROM usuarios WHERE COALESCE(estado, 'activo') = 'activo' GROUP BY rol")
        usuarios_por_rol = {rol: cant for rol, cant in cursor.fetchall()}
    except Exception as e:
        print(f"⚠️ Error calculando usuarios por rol (tablero ejecutivo): {e}")
        usuarios_por_rol = {}
    total_usuarios_activos = sum(usuarios_por_rol.values())

    try:
        cursor.execute("SELECT COUNT(*), COALESCE(SUM(vistas), 0), COALESCE(SUM(descargas), 0) FROM galerias WHERE estado = 'activo'")
        total_instructivos, total_vistas, total_descargas = cursor.fetchone()
    except Exception as e:
        print(f"⚠️ Error calculando Gestor de Archivos (tablero ejecutivo): {e}")
        total_instructivos, total_vistas, total_descargas = 0, 0, 0

    try:
        cursor.execute("SELECT COUNT(*) FROM comunicados WHERE estado = 'activo'")
        total_comunicados = cursor.fetchone()[0]
    except Exception as e:
        print(f"⚠️ Error calculando Comunicados (tablero ejecutivo): {e}")
        total_comunicados = 0

    try:
        cursor.execute("SELECT COUNT(*) FROM credenciales WHERE COALESCE(estado, 'activo') != 'eliminado'")
        total_credenciales = cursor.fetchone()[0]
    except Exception as e:
        print(f"⚠️ Error calculando Bóveda de Accesos (tablero ejecutivo): {e}")
        total_credenciales = 0

    try:
        cursor.execute("SELECT COUNT(*) FROM credenciales_colaboradores WHERE estado = 'activo'")
        total_accesos_colaboradores = cursor.fetchone()[0]
    except Exception as e:
        print(f"⚠️ Error calculando Accesos de Colaboradores (tablero ejecutivo): {e}")
        total_accesos_colaboradores = 0

    try:
        cursor.execute("SELECT estado, COUNT(*) FROM activos_inventario WHERE COALESCE(eliminado, 0) = 0 GROUP BY estado")
        inventario_por_estado = {estado or 'Sin estado': cant for estado, cant in cursor.fetchall()}
    except Exception as e:
        print(f"⚠️ Error calculando Inventario (tablero ejecutivo): {e}")
        inventario_por_estado = {}
    total_activos_inventario = sum(inventario_por_estado.values())

    # 💰 Costos de Inventario: mismo cálculo que /tickets/inventario (_totales_costos_inventario),
    # pero sobre TODO el inventario activo, sin los filtros de sede/tipo que tiene esa vista — el
    # tablero ejecutivo es la foto general de la organización.
    try:
        cursor.execute("SELECT tipo_costo, costo_compra, costo_alquiler_mensual FROM activos_inventario WHERE COALESCE(eliminado, 0) = 0")
        totales_costos_inventario = _totales_costos_inventario([
            {'tipo_costo': tc, 'costo_compra': cc, 'costo_alquiler_mensual': cam}
            for tc, cc, cam in cursor.fetchall()
        ])
    except Exception as e:
        print(f"⚠️ Error calculando Costos de Inventario (tablero ejecutivo): {e}")
        totales_costos_inventario = _totales_costos_inventario([])

    # 📅 Vencimiento de Documentos: mismo criterio que /vencimientos (_bucket_vencimiento),
    # contando institucionales (galerias) y por empleado (documentos_empleado) juntos, para que
    # el tablero avise de un vistazo si hay algo por vencer sin tener que entrar al módulo.
    documentos_vencidos = 0
    documentos_proximos_vencer = 0
    try:
        cursor.execute(
            "SELECT fecha_vencimiento FROM galerias "
            "WHERE COALESCE(estado, 'activo') != 'eliminado' AND fecha_vencimiento IS NOT NULL AND fecha_vencimiento != ''"
        )
        fechas_vencimiento = [r[0] for r in cursor.fetchall()]
        cursor.execute(
            "SELECT fecha_vencimiento FROM documentos_empleado "
            "WHERE COALESCE(estado, 'activo') = 'activo' AND fecha_vencimiento IS NOT NULL AND fecha_vencimiento != ''"
        )
        fechas_vencimiento += [r[0] for r in cursor.fetchall()]
        for fecha_venc in fechas_vencimiento:
            bucket = _bucket_vencimiento(fecha_venc)
            if bucket == 'vencido':
                documentos_vencidos += 1
            elif bucket == 'proximo_a_vencer':
                documentos_proximos_vencer += 1
    except Exception as e:
        print(f"⚠️ Error calculando Vencimiento de Documentos (tablero ejecutivo): {e}")

    # 🔄 Certificaciones de devolución de activos registradas este mes calendario — módulo
    # reciente, para que gerencia vea de un vistazo cuánta rotación de equipos hubo.
    devoluciones_certificadas_mes = 0
    try:
        mes_actual = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y-%m')
        cursor.execute("SELECT fecha FROM inventario_devoluciones")
        devoluciones_certificadas_mes = sum(1 for (f,) in cursor.fetchall() if (f or '').startswith(mes_actual))
    except Exception as e:
        print(f"⚠️ Error calculando Certificación de Devoluciones (tablero ejecutivo): {e}")

    conn.close()

    return {
        'total_tickets': total_tickets,
        'tickets_abiertos': ind_tickets['por_estado'].get('Abierto', 0) + ind_tickets['por_estado'].get('En Proceso', 0),
        'tickets_vencidos': tickets_vencidos,
        'pct_sla_cumplido': pct_sla_cumplido,
        'promedio_calificacion': ind_tickets['promedio_calificacion'],
        'dias_tendencia': ind_tickets['dias_tendencia'],
        'usuarios_por_rol': usuarios_por_rol,
        'total_usuarios_activos': total_usuarios_activos,
        'total_instructivos': total_instructivos,
        'total_vistas': total_vistas,
        'total_descargas': total_descargas,
        'total_comunicados': total_comunicados,
        'total_credenciales': total_credenciales,
        'total_accesos_colaboradores': total_accesos_colaboradores,
        'inventario_por_estado': inventario_por_estado,
        'total_activos_inventario': total_activos_inventario,
        'totales_costos_inventario': totales_costos_inventario,
        'documentos_vencidos': documentos_vencidos,
        'documentos_proximos_vencer': documentos_proximos_vencer,
        'devoluciones_certificadas_mes': devoluciones_certificadas_mes,
        'variacion_pct_tendencia': ind_tickets['variacion_pct_tendencia'],
        'conteo_periodo_anterior': ind_tickets['conteo_periodo_anterior'],
        'tiempo_promedio_resolucion_texto': ind_tickets['tiempo_promedio_resolucion_texto'],
        'titulo_comparativo_periodo': ind_tickets['titulo_comparativo_periodo'],
    }


@app.route('/tablero-ejecutivo')
@login_required
@admin_required
def tablero_ejecutivo():
    # 🔎 Mismo filtro de fecha/mes/agente que /tickets/indicadores (ver _resolver_filtros_indicadores)
    # — así lo que Tomás/gerencia filtra en Indicadores se puede reflejar también aquí, acotando
    # los indicadores de Tickets del tablero al mismo período/agente elegido.
    fecha_inicio, fecha_fin, agente, mes = _resolver_filtros_indicadores()
    datos = _calcular_tablero_ejecutivo(fecha_inicio, fecha_fin, agente)
    return render_template(
        'tablero_ejecutivo.html',
        filtro_mes=mes, filtro_fecha_inicio=fecha_inicio or '', filtro_fecha_fin=fecha_fin or '',
        filtro_agente=agente or '', opciones_meses=_opciones_meses_filtro(), opciones_agentes=_opciones_agentes_filtro(),
        **datos
    )


# 📦 INVENTARIO DE ACTIVOS DE TI (solo equipo de soporte): registro de equipos/activos y a
# quién/qué sede están asignados. Reutiliza las Sedes configuradas en /tickets/configuracion.
@app.route('/tickets/inventario')
@login_required
@agente_o_admin_required
def ver_inventario():
    q_estado = request.args.get('estado', '').strip()
    q_tipo = request.args.get('tipo', '').strip()
    q_sede = request.args.get('sede', '').strip()
    q_area = request.args.get('area', '').strip()
    q_proveedor = request.args.get('proveedor', '').strip()
    q_busqueda = request.args.get('q', '').strip().lower()

    tipos_activo_catalogo = _catalogo_tipos_activo_activos()
    nombres_tipos_activo = [t['etiqueta'] for t in tipos_activo_catalogo] or TIPOS_ACTIVO

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nombre, tipo_activo, marca, modelo, numero_serie, estado, asignado_a, sede, area, proveedor, observaciones, fecha_creacion, creado_por, tipo_costo, costo_compra, costo_alquiler_mensual, firma_asignacion_url, es_biomedico FROM activos_inventario WHERE eliminado = 0 ORDER BY id DESC")
        rows = cursor.fetchall()
    except Exception as e:
        print(f"Error consultando inventario: {e}")
        rows = []

    # 📎 Adjuntos de todos los activos, agrupados por activo_id, en una sola consulta (en vez
    # de una consulta por activo dentro del ciclo).
    adjuntos_por_activo = {}
    try:
        cursor.execute("SELECT id, activo_id, url, nombre_original FROM inventario_adjuntos ORDER BY id ASC")
        for adj_id, activo_id, url, nombre_original in cursor.fetchall():
            adjuntos_por_activo.setdefault(activo_id, []).append({'id': adj_id, 'url': url, 'nombre_original': nombre_original})
    except Exception as e:
        print(f"Error consultando adjuntos de inventario: {e}")

    # 🔀 Trazabilidad: cadena de reemplazos de cada activo, en ambos sentidos — como activo
    # "saliente" (fue reemplazado por otro) y como activo "entrante" (reemplazó a otro).
    trazabilidad_por_activo = {}
    try:
        cursor.execute("""
            SELECT r.id, r.activo_anterior_id, r.activo_nuevo_id, r.motivo, r.notas, r.fecha_reemplazo,
                   r.estado_anterior_resultante, r.creado_por, r.fecha_creacion,
                   ant.nombre, nue.nombre
            FROM activos_reemplazos r
            LEFT JOIN activos_inventario ant ON r.activo_anterior_id = ant.id
            LEFT JOIN activos_inventario nue ON r.activo_nuevo_id = nue.id
            ORDER BY r.fecha_creacion DESC, r.id DESC
        """)
        for (rid, ant_id, nue_id, motivo, notas, fecha_reemplazo, estado_resultante, creado_por, fecha_creacion,
             ant_nombre, nue_nombre) in cursor.fetchall():
            trazabilidad_por_activo.setdefault(ant_id, []).append({
                'id': rid, 'direccion': 'saliente', 'motivo': motivo, 'notas': notas,
                'fecha_reemplazo': fecha_reemplazo, 'estado_resultante': estado_resultante,
                'creado_por': creado_por, 'fecha_creacion': fecha_creacion,
                'otro_id': nue_id, 'otro_nombre': nue_nombre
            })
            if nue_id:
                trazabilidad_por_activo.setdefault(nue_id, []).append({
                    'id': rid, 'direccion': 'entrante', 'motivo': motivo, 'notas': notas,
                    'fecha_reemplazo': fecha_reemplazo, 'estado_resultante': estado_resultante,
                    'creado_por': creado_por, 'fecha_creacion': fecha_creacion,
                    'otro_id': ant_id, 'otro_nombre': ant_nombre
                })
    except Exception as e:
        print(f"Error consultando trazabilidad de activos: {e}")

    # 🎫 Historial de tickets vinculados a cada activo (solicitudes de soporte que marcaron
    # "Activo relacionado" al crearse).
    tickets_por_activo = {}
    try:
        cursor.execute("SELECT id, titulo, tipo, estado, prioridad, fecha_creacion, activo_id FROM tickets WHERE activo_id IS NOT NULL AND COALESCE(eliminado, 0) = 0 ORDER BY id DESC")
        for tk_id, titulo, tipo_t, estado_t, prioridad_t, fecha_creacion_t, activo_id_t in cursor.fetchall():
            tickets_por_activo.setdefault(activo_id_t, []).append({
                'id': tk_id, 'titulo': titulo, 'tipo': tipo_t or 'Incidente', 'estado': estado_t,
                'prioridad': prioridad_t, 'fecha_creacion': fecha_creacion_t,
                'codigo': _codigo_ticket(tipo_t or 'Incidente', tk_id, fecha_creacion_t)
            })
    except Exception as e:
        print(f"Error consultando historial de tickets por activo: {e}")

    conn.close()

    activos_todos = [{
        'id': r[0], 'nombre': r[1], 'tipo_activo': r[2], 'marca': r[3], 'modelo': r[4],
        'numero_serie': r[5], 'estado': r[6], 'asignado_a': r[7], 'sede': r[8], 'area': r[9],
        'proveedor': r[10], 'observaciones': r[11], 'fecha_creacion': r[12], 'creado_por': r[13],
        'tipo_costo': r[14], 'costo_compra': r[15], 'costo_alquiler_mensual': r[16],
        'firma_asignacion_url': r[17], 'es_biomedico': bool(r[18]),
        'adjuntos': adjuntos_por_activo.get(r[0], []),
        'trazabilidad': trazabilidad_por_activo.get(r[0], []),
        'tickets_historial': tickets_por_activo.get(r[0], [])
    } for r in rows]
    total_activos_general = len(activos_todos)

    # 🏢 Distribución por sede: cuántos activos hay en cada sede y en qué estado están, sin
    # importar los filtros aplicados — le da al equipo de soporte una vista completa de la
    # organización de un vistazo, y cada fila enlaza a la vista filtrada por esa sede.
    distribucion_por_sede = {}
    for a in activos_todos:
        clave = a['sede'] or 'Sin sede asignada'
        fila = distribucion_por_sede.setdefault(clave, {'sede': clave, 'total': 0, 'estados': {e: 0 for e in ESTADOS_ACTIVO}})
        fila['total'] += 1
        fila['estados'][a['estado']] = fila['estados'].get(a['estado'], 0) + 1
    distribucion_por_sede = sorted(distribucion_por_sede.values(), key=lambda f: (f['sede'] == 'Sin sede asignada', -f['total'], f['sede']))

    # 📊 Los indicadores de arriba (Disponibles/Asignados/...) responden a la sede, tipo y
    # búsqueda ya elegidos —para que, al filtrar por una sede, se vea cuántos activos de ESA
    # sede están en cada estado— pero no al estado en sí, así el agente puede seguir comparando
    # los 4 estados entre ellos en vez de que el filtro los reduzca a uno solo.
    activos_en_contexto = activos_todos
    if q_sede:
        activos_en_contexto = [a for a in activos_en_contexto if (a['sede'] or '') == q_sede]
    if q_area:
        activos_en_contexto = [a for a in activos_en_contexto if (a['area'] or '') == q_area]
    if q_proveedor:
        activos_en_contexto = [a for a in activos_en_contexto if (a['proveedor'] or '') == q_proveedor]
    if q_tipo in nombres_tipos_activo:
        activos_en_contexto = [a for a in activos_en_contexto if a['tipo_activo'] == q_tipo]
    if q_busqueda:
        activos_en_contexto = [a for a in activos_en_contexto if q_busqueda in f"{a['nombre']} {a['marca'] or ''} {a['modelo'] or ''} {a['numero_serie'] or ''} {a['asignado_a'] or ''}".lower()]

    conteos_estado = {e: 0 for e in ESTADOS_ACTIVO}
    for a in activos_en_contexto:
        conteos_estado[a['estado']] = conteos_estado.get(a['estado'], 0) + 1
    total_activos = len(activos_en_contexto)

    # 💰 Costos de Inventario: responde a los mismos filtros de sede/área/proveedor/tipo/búsqueda
    # que los indicadores de arriba (conteos_estado), para que el total mostrado corresponda a lo
    # que el usuario está viendo en la tabla.
    totales_costos = _totales_costos_inventario(activos_en_contexto)

    # 📊 "Por tipo de activo": mismo criterio de contexto que arriba — responde a la sede y la
    # búsqueda ya elegidas, pero no al tipo en sí, para poder comparar los tipos entre ellos en
    # vez de que el propio filtro de tipo colapse el gráfico a una sola barra.
    activos_para_tipo = activos_todos
    if q_sede:
        activos_para_tipo = [a for a in activos_para_tipo if (a['sede'] or '') == q_sede]
    if q_busqueda:
        activos_para_tipo = [a for a in activos_para_tipo if q_busqueda in f"{a['nombre']} {a['marca'] or ''} {a['modelo'] or ''} {a['numero_serie'] or ''} {a['asignado_a'] or ''}".lower()]
    por_tipo_activo = {}
    for a in activos_para_tipo:
        tp = a['tipo_activo'] or 'Otro'
        por_tipo_activo[tp] = por_tipo_activo.get(tp, 0) + 1
    por_tipo_activo = dict(sorted(por_tipo_activo.items(), key=lambda kv: kv[1], reverse=True))

    activos = activos_en_contexto
    if q_estado in ESTADOS_ACTIVO:
        activos = [a for a in activos if a['estado'] == q_estado]

    sedes = _config_ticket_lista('sede')
    areas = _config_ticket_lista('area')
    proveedores = _config_ticket_lista('proveedor')
    error_placa = request.args.get('error_placa', '').strip()
    # 🔀 Candidatos para "Activo de reemplazo": todo el inventario activo (sin filtrar por la
    # vista actual), con los campos mínimos que necesita el buscador del modal.
    activos_para_reemplazo = [{
        'id': a['id'], 'nombre': a['nombre'], 'marca': a['marca'], 'modelo': a['modelo'], 'estado': a['estado']
    } for a in activos_todos]
    return render_template(
        'tickets_inventario.html', es_soporte=True, activos=activos,
        tipos_activo=nombres_tipos_activo, tipos_activo_catalogo=tipos_activo_catalogo,
        ICONOS_TIPO_ACTIVO=ICONOS_TIPO_ACTIVO,
        motivos_reemplazo=MOTIVOS_REEMPLAZO_ACTIVO, estados_resultantes_reemplazo=ESTADOS_RESULTANTES_REEMPLAZO,
        activos_para_reemplazo=activos_para_reemplazo,
        estados_activo=ESTADOS_ACTIVO, sedes=sedes, areas=areas, proveedores=proveedores,
        q_estado=q_estado, q_tipo=q_tipo, q_sede=q_sede, q_area=q_area, q_proveedor=q_proveedor, q_busqueda=q_busqueda,
        conteos_estado=conteos_estado, total_activos=total_activos,
        total_activos_general=total_activos_general, distribucion_por_sede=distribucion_por_sede,
        por_tipo_activo=por_tipo_activo, error_placa=error_placa,
        especialidades=_catalogo_especialidades_activas(),
        totales_costos=totales_costos
    )


# 📊 Columnas usadas tanto por la plantilla descargable de carga masiva como por el parser de
# importar_inventario_xlsx (posición por índice, no por nombre de encabezado — así una edición
# menor al texto de la cabecera no rompe la carga).
COLUMNAS_INVENTARIO_XLSX = [
    'Placa', 'Tipo de activo', 'Marca', 'Modelo', 'Número de serie', 'Estado', 'Asignado a',
    'Sede', 'Área', 'Proveedor', 'Tipo de costo (propio/alquilado)', 'Costo de compra',
    'Costo de alquiler mensual', 'Observaciones'
]


def _consultar_activos_inventario_filtrados(q_estado='', q_tipo='', q_sede='', q_area='', q_proveedor='', q_busqueda=''):
    """Misma lógica de filtrado que la tabla principal de ver_inventario, pero devolviendo solo
    las columnas base (sin adjuntos/trazabilidad/historial de tickets, que no hacen falta para
    exportar) — así "lo que ves en la vista filtrada es lo que exportas" en CSV/Excel/PDF."""
    nombres_tipos_activo_validos = [t['etiqueta'] for t in _catalogo_tipos_activo_activos()] or TIPOS_ACTIVO
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT id, nombre, tipo_activo, marca, modelo, numero_serie, estado, asignado_a, sede, area, "
            "proveedor, observaciones, fecha_creacion, creado_por, tipo_costo, costo_compra, costo_alquiler_mensual "
            "FROM activos_inventario WHERE eliminado = 0 ORDER BY id DESC"
        )
        rows = cursor.fetchall()
    except Exception as e:
        print(f"Error consultando inventario (exportación): {e}")
        rows = []
    conn.close()

    activos = [{
        'id': r[0], 'nombre': r[1], 'tipo_activo': r[2], 'marca': r[3], 'modelo': r[4],
        'numero_serie': r[5], 'estado': r[6], 'asignado_a': r[7], 'sede': r[8], 'area': r[9],
        'proveedor': r[10], 'observaciones': r[11], 'fecha_creacion': r[12], 'creado_por': r[13],
        'tipo_costo': r[14], 'costo_compra': r[15], 'costo_alquiler_mensual': r[16],
    } for r in rows]

    if q_sede:
        activos = [a for a in activos if (a['sede'] or '') == q_sede]
    if q_area:
        activos = [a for a in activos if (a['area'] or '') == q_area]
    if q_proveedor:
        activos = [a for a in activos if (a['proveedor'] or '') == q_proveedor]
    if q_tipo in nombres_tipos_activo_validos:
        activos = [a for a in activos if a['tipo_activo'] == q_tipo]
    if q_estado in ESTADOS_ACTIVO:
        activos = [a for a in activos if a['estado'] == q_estado]
    q_busqueda_norm = (q_busqueda or '').strip().lower()
    if q_busqueda_norm:
        activos = [a for a in activos if q_busqueda_norm in f"{a['nombre']} {a['marca'] or ''} {a['modelo'] or ''} {a['numero_serie'] or ''} {a['asignado_a'] or ''}".lower()]
    return activos


def _filtros_inventario_desde_query():
    return (
        request.args.get('estado', '').strip(), request.args.get('tipo', '').strip(),
        request.args.get('sede', '').strip(), request.args.get('area', '').strip(),
        request.args.get('proveedor', '').strip(), request.args.get('q', '').strip()
    )


@app.route('/tickets/inventario/plantilla_xlsx')
@login_required
@agente_o_admin_required
def inventario_plantilla_xlsx():
    """Plantilla descargable para la carga masiva: mismas columnas que espera
    importar_inventario_xlsx, con una fila de ejemplo para que quede claro el formato esperado
    (tipo_costo, por ejemplo, solo acepta 'propio' o 'alquilado')."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    ws.append(COLUMNAS_INVENTARIO_XLSX)
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
    ws.append([
        '15099', 'Portátil', 'Dell', 'Latitude 5420', 'ABC12345', 'Disponible', '',
        'Sede Principal', 'Sistemas', '', 'propio', 2500000, '',
        'Fila de ejemplo — bórrala antes de subir tu archivo.'
    ])
    anchos = [12, 16, 14, 18, 16, 14, 22, 16, 14, 16, 26, 16, 20, 40]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return Response(salida.read(), headers={
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename="Arkiv_Plantilla_Inventario.xlsx"'
    })


@app.route('/tickets/inventario/importar_xlsx', methods=['POST'])
@login_required
@agente_o_admin_required
def importar_inventario_xlsx():
    """Carga masiva de activos desde un archivo .xlsx con el formato de COLUMNAS_INVENTARIO_XLSX.
    Cada fila se valida con las MISMAS reglas que crear_activo (placa única y obligatoria, tipo/
    estado/sede/área/proveedor validados contra los catálogos, costos según _parsear_datos_costo_
    inventario) — las filas que fallan se omiten individualmente y se listan en el mensaje final,
    en vez de descartar todo el archivo por un solo error."""
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        flash("Selecciona un archivo .xlsx para cargar.", "error")
        return redirect(url_for('ver_inventario'))
    if not archivo.filename.lower().endswith('.xlsx'):
        flash("El archivo debe tener extensión .xlsx (Excel).", "error")
        return redirect(url_for('ver_inventario'))

    import openpyxl
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"No se pudo leer el archivo: {e}", "error")
        return redirect(url_for('ver_inventario'))

    nombres_tipos_activo_validos = [t['etiqueta'] for t in _catalogo_tipos_activo_activos()] or TIPOS_ACTIVO
    sedes_validas = [s['nombre'] for s in _config_ticket_lista('sede')]
    areas_validas = [a['nombre'] for a in _config_ticket_lista('area')]
    proveedores_validos = [p['nombre'] for p in _config_ticket_lista('proveedor')]

    fecha_act = obtener_fecha_actual()
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()

    q_dup = ("SELECT id FROM activos_inventario WHERE LOWER(nombre) = LOWER(%s) AND eliminado = 0" if db_type == 'postgres'
             else "SELECT id FROM activos_inventario WHERE LOWER(nombre) = LOWER(?) AND eliminado = 0")
    q_ins = ("INSERT INTO activos_inventario (nombre, tipo_activo, marca, modelo, numero_serie, estado, "
             "asignado_a, sede, area, proveedor, observaciones, fecha_creacion, creado_por, tipo_costo, "
             "costo_compra, costo_alquiler_mensual) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
             if db_type == 'postgres' else
             "INSERT INTO activos_inventario (nombre, tipo_activo, marca, modelo, numero_serie, estado, "
             "asignado_a, sede, area, proveedor, observaciones, fecha_creacion, creado_por, tipo_costo, "
             "costo_compra, costo_alquiler_mensual) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)")

    creados = 0
    omitidos = []
    placas_vistas_en_archivo = set()

    for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not fila or all(c is None or str(c).strip() == '' for c in fila):
            continue  # fila en blanco: se ignora sin contar como error

        valores = list(fila) + [None] * (14 - len(fila))
        nombre = str(valores[0]).strip() if valores[0] is not None else ''
        if not nombre:
            omitidos.append((num_fila, "sin número de placa"))
            continue
        if nombre.lower() in placas_vistas_en_archivo:
            omitidos.append((num_fila, f"placa '{nombre}' repetida más de una vez en este archivo"))
            continue

        cursor.execute(q_dup, (nombre,))
        if cursor.fetchone():
            omitidos.append((num_fila, f"la placa '{nombre}' ya existe en el inventario"))
            continue

        tipo_activo = str(valores[1]).strip() if valores[1] else 'Otro'
        if tipo_activo not in nombres_tipos_activo_validos:
            tipo_activo = 'Otro'
        marca = str(valores[2]).strip() if valores[2] else None
        modelo = str(valores[3]).strip() if valores[3] else None
        numero_serie = str(valores[4]).strip() if valores[4] else None
        estado = str(valores[5]).strip() if valores[5] else 'Disponible'
        if estado not in ESTADOS_ACTIVO:
            estado = 'Disponible'
        asignado_a = str(valores[6]).strip() if valores[6] else None
        sede = str(valores[7]).strip() if valores[7] else None
        sede = sede if sede in sedes_validas else None
        area = str(valores[8]).strip() if valores[8] else None
        area = area if area in areas_validas else None
        proveedor = str(valores[9]).strip() if valores[9] else None
        proveedor = proveedor if proveedor in proveedores_validos else None
        observaciones = str(valores[13]).strip() if valores[13] else None

        tipo_costo, costo_compra, costo_alquiler_mensual = _parsear_datos_costo_inventario({
            'tipo_costo': str(valores[10]).strip() if valores[10] else '',
            'costo_compra': str(valores[11]).strip() if valores[11] not in (None, '') else '',
            'costo_alquiler_mensual': str(valores[12]).strip() if valores[12] not in (None, '') else '',
        })

        try:
            cursor.execute(q_ins, (nombre, tipo_activo, marca, modelo, numero_serie, estado, asignado_a,
                                    sede, area, proveedor, observaciones, fecha_act, usuario,
                                    tipo_costo, costo_compra, costo_alquiler_mensual))
            conn.commit()
            creados += 1
            placas_vistas_en_archivo.add(nombre.lower())
        except Exception as e:
            conn.rollback()
            omitidos.append((num_fila, f"error al guardar: {e}"))

    conn.close()

    if creados:
        registrar_log(usuario, "Inventario de Activos", f"Carga masiva desde Excel: {creados} activo(s) registrado(s)")

    if creados and not omitidos:
        flash(f"Se cargaron {creados} activo(s) correctamente.", "exito")
    elif creados and omitidos:
        detalle = "\n".join(f"Fila {f}: {m}" for f, m in omitidos[:20])
        extra = f"\n(...y {len(omitidos) - 20} fila(s) más)" if len(omitidos) > 20 else ""
        flash(f"Se cargaron {creados} activo(s). Se omitieron {len(omitidos)} fila(s):\n{detalle}{extra}", "exito")
    else:
        detalle = "\n".join(f"Fila {f}: {m}" for f, m in omitidos[:20]) or "El archivo no tiene filas con datos."
        flash(f"No se cargó ningún activo. {detalle}", "error")

    return redirect(url_for('ver_inventario'))


@app.route('/tickets/inventario/exportar_csv')
@login_required
@agente_o_admin_required
def inventario_exportar_csv():
    activos = _consultar_activos_inventario_filtrados(*_filtros_inventario_desde_query())

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['PLACA', 'TIPO', 'MARCA', 'MODELO', 'N° SERIE', 'ESTADO', 'ASIGNADO A', 'SEDE', 'ÁREA',
                      'PROVEEDOR', 'TIPO DE COSTO', 'COSTO DE COMPRA', 'COSTO ALQUILER MENSUAL',
                      'OBSERVACIONES', 'FECHA DE CREACIÓN', 'CREADO POR'])
    for a in activos:
        writer.writerow([
            a['nombre'], a['tipo_activo'] or '', a['marca'] or '', a['modelo'] or '', a['numero_serie'] or '',
            a['estado'], a['asignado_a'] or '', a['sede'] or '', a['area'] or '', a['proveedor'] or '',
            a['tipo_costo'] or '', a['costo_compra'] if a['costo_compra'] is not None else '',
            a['costo_alquiler_mensual'] if a['costo_alquiler_mensual'] is not None else '',
            a['observaciones'] or '', a['fecha_creacion'] or '', a['creado_por'] or ''
        ])

    csv_bytes = '﻿' + output.getvalue()
    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y%m%d_%H%M')
    headers = {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': f'attachment; filename="Arkiv_Inventario_{fecha_filename}.csv"'
    }
    return Response(csv_bytes, headers=headers, status=200)


@app.route('/tickets/inventario/exportar_xlsx')
@login_required
@agente_o_admin_required
def inventario_exportar_xlsx():
    activos = _consultar_activos_inventario_filtrados(*_filtros_inventario_desde_query())

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    encabezados = ['Placa', 'Tipo', 'Marca', 'Modelo', 'N° Serie', 'Estado', 'Asignado a', 'Sede', 'Área',
                   'Proveedor', 'Tipo de costo', 'Costo de compra', 'Costo alquiler mensual',
                   'Observaciones', 'Fecha de creación', 'Creado por']
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
    for a in activos:
        ws.append([
            a['nombre'], a['tipo_activo'] or '', a['marca'] or '', a['modelo'] or '', a['numero_serie'] or '',
            a['estado'], a['asignado_a'] or '', a['sede'] or '', a['area'] or '', a['proveedor'] or '',
            a['tipo_costo'] or '', a['costo_compra'], a['costo_alquiler_mensual'],
            a['observaciones'] or '', a['fecha_creacion'] or '', a['creado_por'] or ''
        ])
    ws.freeze_panes = 'A2'
    anchos = [12, 16, 14, 18, 16, 14, 22, 16, 14, 16, 14, 16, 18, 32, 18, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y%m%d_%H%M')
    return Response(salida.read(), headers={
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="Arkiv_Inventario_{fecha_filename}.xlsx"'
    })


@app.route('/tickets/inventario/exportar_pdf')
@login_required
@agente_o_admin_required
def inventario_exportar_pdf():
    activos = _consultar_activos_inventario_filtrados(*_filtros_inventario_desde_query())
    totales = _totales_costos_inventario(activos)

    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    salida = io.BytesIO()
    doc = SimpleDocTemplate(salida, pagesize=landscape(letter), topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                             leftMargin=1 * cm, rightMargin=1 * cm)
    estilos = getSampleStyleSheet()
    elementos = [Paragraph("Arkiv &mdash; Inventario de Activos", estilos['Title']), Spacer(1, 0.25 * cm)]

    resumen = (
        f"{len(activos)} activo(s) &middot; Invertido en compra: ${totales['total_compra']:,.0f} "
        f"({totales['cantidad_propios']}) &middot; Alquiler mensual: ${totales['total_alquiler_mensual']:,.0f} "
        f"({totales['cantidad_alquilados']})"
    )
    elementos.append(Paragraph(resumen, estilos['Normal']))
    elementos.append(Spacer(1, 0.4 * cm))

    datos = [['Placa', 'Tipo', 'Marca/Modelo', 'Estado', 'Asignado a', 'Sede', 'Costo']]
    for a in activos:
        marca_modelo = ' '.join(filter(None, [a['marca'], a['modelo']])) or '-'
        if a['tipo_costo'] == 'propio' and a['costo_compra'] is not None:
            costo_texto = f"Compra: ${a['costo_compra']:,.0f}"
        elif a['tipo_costo'] == 'alquilado' and a['costo_alquiler_mensual'] is not None:
            costo_texto = f"Alquiler: ${a['costo_alquiler_mensual']:,.0f}/mes"
        else:
            costo_texto = '-'
        datos.append([a['nombre'], a['tipo_activo'] or '-', marca_modelo, a['estado'], a['asignado_a'] or '-',
                      a['sede'] or '-', costo_texto])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ea580c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    salida.seek(0)
    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime('%Y%m%d_%H%M')
    return Response(salida.read(), headers={
        'Content-Type': 'application/pdf',
        'Content-Disposition': f'attachment; filename="Arkiv_Inventario_{fecha_filename}.pdf"'
    })


def _parsear_datos_costo_inventario(form):
    """Lee y valida tipo_costo/costo_compra/costo_alquiler_mensual del formulario de crear/editar
    activo. Un activo es propio O alquilado, nunca las dos cosas a la vez: el campo que no aplica
    según 'tipo_costo' se descarta (queda en None) y por lo tanto no entra en
    _totales_costos_inventario, aunque el usuario haya escrito algo en ese otro campo antes de
    cambiar de opción en el formulario."""
    tipo_costo = (form.get('tipo_costo') or '').strip().lower()
    if tipo_costo not in ('propio', 'alquilado'):
        tipo_costo = None

    def _valor_numerico(valor_raw):
        valor_raw = (valor_raw or '').strip()
        if not valor_raw:
            return None
        try:
            numero = float(valor_raw)
        except ValueError:
            return None
        return numero if numero >= 0 else None

    costo_compra = _valor_numerico(form.get('costo_compra')) if tipo_costo == 'propio' else None
    costo_alquiler_mensual = _valor_numerico(form.get('costo_alquiler_mensual')) if tipo_costo == 'alquilado' else None
    return tipo_costo, costo_compra, costo_alquiler_mensual


def _totales_costos_inventario(activos):
    """Suma costo_compra de los activos 'propio' y costo_alquiler_mensual de los 'alquilado' —
    los que no tienen tipo_costo definido (activos creados antes de este cambio, o donde el campo
    simplemente se dejó vacío) no entran en ningún total. Recibe la lista de activos ya en el
    formato de diccionario que arma ver_inventario (con las claves 'tipo_costo', 'costo_compra' y
    'costo_alquiler_mensual')."""
    total_compra = 0.0
    cantidad_propios = 0
    total_alquiler_mensual = 0.0
    cantidad_alquilados = 0
    for a in activos:
        if a.get('tipo_costo') == 'propio' and a.get('costo_compra') is not None:
            total_compra += a['costo_compra']
            cantidad_propios += 1
        elif a.get('tipo_costo') == 'alquilado' and a.get('costo_alquiler_mensual') is not None:
            total_alquiler_mensual += a['costo_alquiler_mensual']
            cantidad_alquilados += 1
    return {
        'total_compra': total_compra, 'cantidad_propios': cantidad_propios,
        'total_alquiler_mensual': total_alquiler_mensual, 'cantidad_alquilados': cantidad_alquilados,
        'total_alquiler_anual': total_alquiler_mensual * 12,
    }


def _resolver_firma_para_asignacion(asignado_a):
    """Si 'asignado_a' viene en el formato 'Nombre (usuario)' que arma buscarAsignadoPorCedula/
    crearUsuarioRapido (ver tickets_inventario.html) y ese usuario tiene una firma registrada
    (capturada al crear su cuenta — ver 'usuarios.firma'), la devuelve para copiarla a
    'activos_inventario.firma_asignacion_url' como constancia de aceptación, sin pedirle que
    vuelva a firmar en cada asignación. Si el texto es libre (se escribió el nombre a mano, sin
    pasar por el buscador) o esa persona nunca registró firma, no hay nada que reutilizar — se
    resuelve del lado del servidor (no de lo que mande el formulario) para que no se pueda
    'inyectar' una firma ajena editando el campo oculto a mano."""
    if not asignado_a:
        return None
    coincidencia = re.search(r'\(([^()]+)\)\s*$', asignado_a)
    if not coincidencia:
        return None
    usuario = coincidencia.group(1).strip()
    if not usuario:
        return None
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "SELECT firma FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT firma FROM usuarios WHERE usuario = ?"
        cursor.execute(q, (usuario,))
        fila = cursor.fetchone()
    except Exception as e:
        print(f"⚠️ Error resolviendo firma para asignación ('{usuario}'): {e}")
        fila = None
    conn.close()
    return fila[0] if fila and fila[0] else None


RELACIONES_RESPONSABLE_ACTA = ('paciente', 'cuidador', 'otro')


def _registrar_acta_recibido_biomedico(activo_id, form, creador, conn, cursor, db_type):
    """Si el activo se marcó como biomédico ('es_biomedico') y el formulario trae los datos de
    la entrega a domicilio (quién recibió, dirección, firma), guarda un registro HISTÓRICO en
    'actas_recibido_biomedico' — a diferencia de 'firma_asignacion_url' en activos_inventario
    (que solo guarda la última firma y se sobreescribe en cada reasignación), aquí queda una fila
    nueva por cada entrega, con qué responder si el equipo se pierde o se daña más adelante.
    Devuelve (categoria, mensaje) listo para flash(), o (None, None) si esta vez no se intentó
    registrar ninguna acta (el checkbox no estaba marcado, o los campos del acta quedaron
    vacíos)."""
    if form.get('es_biomedico') not in ('on', '1', 'true'):
        return None, None

    nombre_responsable = (form.get('acta_nombre_responsable') or '').strip()
    relacion_responsable = (form.get('acta_relacion_responsable') or '').strip()
    documento_responsable = (form.get('acta_documento_responsable') or '').strip() or None
    telefono_contacto = (form.get('acta_telefono_contacto') or '').strip() or None
    direccion_entrega = (form.get('acta_direccion_entrega') or '').strip()
    firma_dataurl = form.get('acta_firma_dataurl') or ''

    if not any([nombre_responsable, direccion_entrega, firma_dataurl]):
        return None, None  # activo marcado como biomédico, pero no se diligenció ningún acta esta vez

    if not (nombre_responsable and direccion_entrega and firma_dataurl):
        return 'error', ("El activo se guardó, pero el ACTA DE RECIBIDO no: para registrarla hacen falta "
                          "el nombre de quien recibe, la dirección de entrega y su firma.")

    if relacion_responsable not in RELACIONES_RESPONSABLE_ACTA:
        relacion_responsable = 'otro'

    firma_url, error_firma = _subir_firma_desde_dataurl(firma_dataurl)
    if error_firma or not firma_url:
        return 'error', f"El activo se guardó, pero el ACTA DE RECIBIDO no: {error_firma or 'no se pudo guardar la firma.'}"

    try:
        q_ins = ("INSERT INTO actas_recibido_biomedico (activo_id, nombre_responsable, relacion_responsable, "
                 "documento_responsable, telefono_contacto, direccion_entrega, firma_url, fecha, creado_por) "
                 "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)" if db_type == 'postgres' else
                 "INSERT INTO actas_recibido_biomedico (activo_id, nombre_responsable, relacion_responsable, "
                 "documento_responsable, telefono_contacto, direccion_entrega, firma_url, fecha, creado_por) "
                 "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)")
        cursor.execute(q_ins, (activo_id, nombre_responsable, relacion_responsable, documento_responsable,
                                telefono_contacto, direccion_entrega, firma_url, obtener_fecha_actual(), creador))
        conn.commit()
        registrar_log(creador, "Inventario de Activos",
                      f"Acta de recibido registrada para el activo #{activo_id} (biomédico) — recibió: "
                      f"{nombre_responsable} ({relacion_responsable}), en {direccion_entrega}.")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error guardando acta de recibido biomédico (activo #{activo_id}): {e}")
        return 'error', "El activo se guardó, pero no se pudo registrar el acta de recibido. Intenta de nuevo."

    return 'exito', f"Acta de recibido registrada: {nombre_responsable} ({relacion_responsable}) en {direccion_entrega}."


@app.route('/tickets/inventario/<int:activo_id>/actas')
@login_required
@agente_o_admin_required
def listar_actas_recibido(activo_id):
    """JSON con el historial de actas de recibido de un activo biomédico — usado por el modal
    'Actas de recibido' en Inventario. A propósito devuelve TODAS las actas del activo (no solo
    la más reciente): cada una es evidencia de una entrega distinta a lo largo de la vida del
    equipo."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = ("SELECT id, nombre_responsable, relacion_responsable, documento_responsable, telefono_contacto, "
             "direccion_entrega, firma_url, fecha, creado_por FROM actas_recibido_biomedico "
             "WHERE activo_id = %s ORDER BY id DESC" if db_type == 'postgres' else
             "SELECT id, nombre_responsable, relacion_responsable, documento_responsable, telefono_contacto, "
             "direccion_entrega, firma_url, fecha, creado_por FROM actas_recibido_biomedico "
             "WHERE activo_id = ? ORDER BY id DESC")
        cursor.execute(q, (activo_id,))
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando actas de recibido del activo {activo_id}: {e}")
        filas = []
    conn.close()

    actas = [{
        'id': f[0], 'nombre_responsable': f[1], 'relacion_responsable': f[2], 'documento_responsable': f[3] or '',
        'telefono_contacto': f[4] or '', 'direccion_entrega': f[5], 'firma_url': f[6], 'fecha': f[7], 'creado_por': f[8]
    } for f in filas]
    return jsonify({'actas': actas})


@app.route('/tickets/inventario/actas/<int:acta_id>/pdf')
@login_required
@agente_o_admin_required
def acta_recibido_pdf(acta_id):
    """Genera el PDF formal del acta de recibido — el documento que queda como constancia física
    de que tal persona (paciente o cuidador) recibió tal equipo biomédico en tal dirección, con
    su firma, por si el equipo se pierde o se daña más adelante."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = ("SELECT a.nombre_responsable, a.relacion_responsable, a.documento_responsable, a.telefono_contacto, "
             "a.direccion_entrega, a.firma_url, a.fecha, a.creado_por, i.nombre, i.tipo_activo, i.marca, i.modelo, i.numero_serie "
             "FROM actas_recibido_biomedico a JOIN activos_inventario i ON i.id = a.activo_id WHERE a.id = %s"
             if db_type == 'postgres' else
             "SELECT a.nombre_responsable, a.relacion_responsable, a.documento_responsable, a.telefono_contacto, "
             "a.direccion_entrega, a.firma_url, a.fecha, a.creado_por, i.nombre, i.tipo_activo, i.marca, i.modelo, i.numero_serie "
             "FROM actas_recibido_biomedico a JOIN activos_inventario i ON i.id = a.activo_id WHERE a.id = ?")
        cursor.execute(q, (acta_id,))
        fila = cursor.fetchone()
    except Exception as e:
        print(f"⚠️ Error consultando acta de recibido {acta_id}: {e}")
        fila = None
    conn.close()

    if not fila:
        return redirect(url_for('ver_inventario'))

    (nombre_responsable, relacion_responsable, documento_responsable, telefono_contacto, direccion_entrega,
     firma_url, fecha, creado_por, placa, tipo_activo, marca, modelo, numero_serie) = fila

    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet

    salida = io.BytesIO()
    doc = SimpleDocTemplate(salida, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=2 * cm, rightMargin=2 * cm)
    estilos = getSampleStyleSheet()
    relaciones_texto = {'paciente': 'Paciente', 'cuidador': 'Cuidador/a', 'otro': 'Responsable'}

    elementos = [
        Paragraph("Arkiv &mdash; Acta de Recibido de Equipo Biomédico", estilos['Title']),
        Spacer(1, 0.3 * cm),
        Paragraph(
            "Constancia de entrega a domicilio de un equipo biomédico, con la firma de quien lo recibió, "
            "para efectos de trazabilidad en caso de pérdida o daño del equipo.", estilos['Normal']
        ),
        Spacer(1, 0.6 * cm),
    ]

    datos_equipo = [
        ['Placa / identificación', placa],
        ['Tipo de activo', tipo_activo or '-'],
        ['Marca / Modelo', ' '.join(filter(None, [marca, modelo])) or '-'],
        ['Número de serie', numero_serie or '-'],
    ]
    tabla_equipo = Table(datos_equipo, colWidths=[5.5 * cm, 10 * cm])
    tabla_equipo.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elementos += [Paragraph("Equipo entregado", estilos['Heading3']), tabla_equipo, Spacer(1, 0.5 * cm)]

    datos_responsable = [
        ['Recibió (nombre)', nombre_responsable],
        ['Relación con el paciente', relaciones_texto.get(relacion_responsable, relacion_responsable)],
        ['Documento de identidad', documento_responsable or '-'],
        ['Teléfono de contacto', telefono_contacto or '-'],
        ['Dirección de entrega', direccion_entrega],
        ['Fecha de entrega', fecha],
        ['Registrado por', creado_por],
    ]
    tabla_responsable = Table(datos_responsable, colWidths=[5.5 * cm, 10 * cm])
    tabla_responsable.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elementos += [Paragraph("Quién recibió el equipo", estilos['Heading3']), tabla_responsable, Spacer(1, 0.6 * cm)]

    elementos.append(Paragraph("Firma de quien recibe", estilos['Heading3']))
    firma_insertada = False
    if firma_url:
        try:
            with urllib.request.urlopen(firma_url, timeout=15) as resp:
                datos_imagen = resp.read()
            imagen_firma = Image(io.BytesIO(datos_imagen), width=7 * cm, height=3.5 * cm)
            imagen_firma.hAlign = 'LEFT'
            elementos.append(imagen_firma)
            firma_insertada = True
        except Exception as e:
            print(f"⚠️ No se pudo incrustar la firma del acta {acta_id} en el PDF: {e}")
    if not firma_insertada:
        elementos.append(Paragraph(f"(Firma registrada — ver imagen original: {firma_url})", estilos['Normal']))

    doc.build(elementos)
    salida.seek(0)
    return Response(salida.read(), headers={
        'Content-Type': 'application/pdf',
        'Content-Disposition': f'attachment; filename="Arkiv_Acta_Recibido_{acta_id}.pdf"'
    })


@app.route('/tickets/inventario/nuevo', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_activo():
    nombre = request.form.get('nombre', '').strip()
    tipo_activo = request.form.get('tipo_activo', 'Otro').strip()
    marca = request.form.get('marca', '').strip()
    modelo = request.form.get('modelo', '').strip()
    numero_serie = request.form.get('numero_serie', '').strip()
    estado = request.form.get('estado', 'Disponible').strip()
    asignado_a = request.form.get('asignado_a', '').strip()
    sede = request.form.get('sede', '').strip()
    area = request.form.get('area', '').strip()
    proveedor = request.form.get('proveedor', '').strip()
    observaciones = request.form.get('observaciones', '').strip()

    nombres_tipos_activo_validos = [t['etiqueta'] for t in _catalogo_tipos_activo_activos()] or TIPOS_ACTIVO
    if tipo_activo not in nombres_tipos_activo_validos:
        tipo_activo = 'Otro'
    if estado not in ESTADOS_ACTIVO:
        estado = 'Disponible'
    sedes_validas = [s['nombre'] for s in _config_ticket_lista('sede')]
    sede = sede if sede in sedes_validas else None
    areas_validas = [a['nombre'] for a in _config_ticket_lista('area')]
    area = area if area in areas_validas else None
    proveedores_validos = [p['nombre'] for p in _config_ticket_lista('proveedor')]
    proveedor = proveedor if proveedor in proveedores_validos else None
    tipo_costo, costo_compra, costo_alquiler_mensual = _parsear_datos_costo_inventario(request.form)
    firma_asignacion_url = _resolver_firma_para_asignacion(asignado_a)
    es_biomedico = request.form.get('es_biomedico') in ('on', '1', 'true')

    if nombre:
        fecha_act = obtener_fecha_actual()
        firma_asignacion_fecha = fecha_act if firma_asignacion_url else None
        usuario = session.get('username')
        conn, db_type = get_db()
        cursor = conn.cursor()
        nuevo_activo_id = None
        try:
            # 🏷️ El número de placa (columna "nombre" en la base) identifica un activo físico
            # puntual: no puede haber dos filas activas con la misma placa, o el inventario deja
            # de ser confiable (ya pasó — ver los "15011" duplicados que motivaron este cambio).
            q_dup = "SELECT id FROM activos_inventario WHERE LOWER(nombre) = LOWER(%s) AND eliminado = 0" if db_type == 'postgres' else "SELECT id FROM activos_inventario WHERE LOWER(nombre) = LOWER(?) AND eliminado = 0"
            cursor.execute(q_dup, (nombre,))
            if cursor.fetchone():
                conn.close()
                return redirect(url_for('ver_inventario', error_placa=nombre))

            q_ins = ("INSERT INTO activos_inventario (nombre, tipo_activo, marca, modelo, numero_serie, estado, asignado_a, sede, area, proveedor, observaciones, fecha_creacion, creado_por, tipo_costo, costo_compra, costo_alquiler_mensual, firma_asignacion_url, firma_asignacion_fecha, es_biomedico) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id"
                     if db_type == 'postgres' else
                     "INSERT INTO activos_inventario (nombre, tipo_activo, marca, modelo, numero_serie, estado, asignado_a, sede, area, proveedor, observaciones, fecha_creacion, creado_por, tipo_costo, costo_compra, costo_alquiler_mensual, firma_asignacion_url, firma_asignacion_fecha, es_biomedico) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)")
            cursor.execute(q_ins, (nombre, tipo_activo, marca or None, modelo or None, numero_serie or None, estado, asignado_a or None, sede, area, proveedor, observaciones or None, fecha_act, usuario, tipo_costo, costo_compra, costo_alquiler_mensual, firma_asignacion_url, firma_asignacion_fecha, es_biomedico))
            nuevo_activo_id = cursor.fetchone()[0] if db_type == 'postgres' else cursor.lastrowid
            conn.commit()
            registrar_log(usuario, "Inventario de Activos", f"Se registró el activo '{nombre}' [{tipo_activo}]")
        except Exception as e:
            conn.rollback()
            print(f"Error creando activo: {e}")
        if nuevo_activo_id:
            categoria_acta, mensaje_acta = _registrar_acta_recibido_biomedico(nuevo_activo_id, request.form, usuario, conn, cursor, db_type)
            if mensaje_acta:
                flash(mensaje_acta, categoria_acta)
        conn.close()
    return redirect(url_for('ver_inventario'))


@app.route('/tickets/inventario/<int:activo_id>/editar', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_activo(activo_id):
    nombre = request.form.get('nombre', '').strip()
    tipo_activo = request.form.get('tipo_activo', 'Otro').strip()
    marca = request.form.get('marca', '').strip()
    modelo = request.form.get('modelo', '').strip()
    numero_serie = request.form.get('numero_serie', '').strip()
    estado = request.form.get('estado', 'Disponible').strip()
    asignado_a = request.form.get('asignado_a', '').strip()
    sede = request.form.get('sede', '').strip()
    area = request.form.get('area', '').strip()
    proveedor = request.form.get('proveedor', '').strip()
    observaciones = request.form.get('observaciones', '').strip()

    nombres_tipos_activo_validos = [t['etiqueta'] for t in _catalogo_tipos_activo_activos()] or TIPOS_ACTIVO
    if tipo_activo not in nombres_tipos_activo_validos:
        tipo_activo = 'Otro'
    if estado not in ESTADOS_ACTIVO:
        estado = 'Disponible'
    sedes_validas = [s['nombre'] for s in _config_ticket_lista('sede')]
    sede = sede if sede in sedes_validas else None
    areas_validas = [a['nombre'] for a in _config_ticket_lista('area')]
    area = area if area in areas_validas else None
    proveedores_validos = [p['nombre'] for p in _config_ticket_lista('proveedor')]
    proveedor = proveedor if proveedor in proveedores_validos else None
    tipo_costo, costo_compra, costo_alquiler_mensual = _parsear_datos_costo_inventario(request.form)
    firma_asignacion_url = _resolver_firma_para_asignacion(asignado_a)
    firma_asignacion_fecha = obtener_fecha_actual() if firma_asignacion_url else None
    es_biomedico = request.form.get('es_biomedico') in ('on', '1', 'true')

    if nombre:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            # 🏷️ Misma validación de placa única que en crear_activo, excluyendo la propia fila
            # que se está editando (si no cambió la placa, no debe chocar consigo misma).
            q_dup = "SELECT id FROM activos_inventario WHERE LOWER(nombre) = LOWER(%s) AND eliminado = 0 AND id != %s" if db_type == 'postgres' else "SELECT id FROM activos_inventario WHERE LOWER(nombre) = LOWER(?) AND eliminado = 0 AND id != ?"
            cursor.execute(q_dup, (nombre, activo_id))
            if cursor.fetchone():
                conn.close()
                return redirect(url_for('ver_inventario', error_placa=nombre))

            q_upd = "UPDATE activos_inventario SET nombre = %s, tipo_activo = %s, marca = %s, modelo = %s, numero_serie = %s, estado = %s, asignado_a = %s, sede = %s, area = %s, proveedor = %s, observaciones = %s, tipo_costo = %s, costo_compra = %s, costo_alquiler_mensual = %s, firma_asignacion_url = %s, firma_asignacion_fecha = %s, es_biomedico = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE activos_inventario SET nombre = ?, tipo_activo = ?, marca = ?, modelo = ?, numero_serie = ?, estado = ?, asignado_a = ?, sede = ?, area = ?, proveedor = ?, observaciones = ?, tipo_costo = ?, costo_compra = ?, costo_alquiler_mensual = ?, firma_asignacion_url = ?, firma_asignacion_fecha = ?, es_biomedico = ? WHERE id = ?"
            cursor.execute(q_upd, (nombre, tipo_activo, marca or None, modelo or None, numero_serie or None, estado, asignado_a or None, sede, area, proveedor, observaciones or None, tipo_costo, costo_compra, costo_alquiler_mensual, firma_asignacion_url, firma_asignacion_fecha, es_biomedico, activo_id))
            conn.commit()
            registrar_log(session.get('username'), "Inventario de Activos", f"Se editó el activo #{activo_id} ('{nombre}')")
        except Exception as e:
            conn.rollback()
            print(f"Error editando activo {activo_id}: {e}")
        categoria_acta, mensaje_acta = _registrar_acta_recibido_biomedico(activo_id, request.form, session.get('username'), conn, cursor, db_type)
        if mensaje_acta:
            flash(mensaje_acta, categoria_acta)
        conn.close()
    return redirect(url_for('ver_inventario'))


@app.route('/tickets/inventario/<int:activo_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_activo(activo_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE activos_inventario SET eliminado = 1 WHERE id = %s" if db_type == 'postgres' else "UPDATE activos_inventario SET eliminado = 1 WHERE id = ?"
        cursor.execute(q, (activo_id,))
        conn.commit()
        registrar_log(session.get('username'), "Inventario de Activos", f"Se eliminó el activo #{activo_id}")
    except Exception as e:
        conn.rollback()
        print(f"Error eliminando activo {activo_id}: {e}")
    conn.close()
    return redirect(url_for('ver_inventario'))


# 🔀 REEMPLAZAR ACTIVO: registra en 'activos_reemplazos' el motivo, notas, fecha y qué pasa con
# el equipo saliente (queda en bodega/mantenimiento/baja/perdido); si se eligió un activo de
# reemplazo, ese activo "hereda" el responsable y la sede del que reemplaza, para no perder el
# hilo de a quién le está llegando el equipo nuevo. Inspirado en el flujo de Solvyx.
@app.route('/tickets/inventario/<int:activo_id>/reemplazar', methods=['POST'])
@login_required
@agente_o_admin_required
def reemplazar_activo(activo_id):
    motivo = request.form.get('motivo', '').strip()
    notas = request.form.get('notas', '').strip()
    activo_nuevo_id_raw = request.form.get('activo_nuevo_id', '').strip()
    fecha_reemplazo = request.form.get('fecha_reemplazo', '').strip()
    estado_resultante = request.form.get('estado_anterior_resultante', '').strip()

    motivos_validos = [m['clave'] for m in MOTIVOS_REEMPLAZO_ACTIVO]
    estados_resultantes_validos = [e['valor'] for e in ESTADOS_RESULTANTES_REEMPLAZO]
    if motivo not in motivos_validos or estado_resultante not in estados_resultantes_validos:
        return redirect(url_for('ver_inventario'))
    if not fecha_reemplazo:
        fecha_reemplazo = obtener_fecha_actual()

    activo_nuevo_id = int(activo_nuevo_id_raw) if activo_nuevo_id_raw.isdigit() else None
    if activo_nuevo_id == activo_id:
        activo_nuevo_id = None

    conn, db_type = get_db()
    cursor = conn.cursor()
    ph = '%s' if db_type == 'postgres' else '?'
    try:
        cursor.execute(f"SELECT nombre, asignado_a, sede, area FROM activos_inventario WHERE id = {ph} AND eliminado = 0", (activo_id,))
        fila_anterior = cursor.fetchone()
        if not fila_anterior:
            conn.close()
            return redirect(url_for('ver_inventario'))
        nombre_anterior, asignado_a_anterior, sede_anterior, area_anterior = fila_anterior

        usuario = session.get('username')
        fecha_creacion = obtener_fecha_actual()
        q_ins = f"INSERT INTO activos_reemplazos (activo_anterior_id, activo_nuevo_id, motivo, notas, fecha_reemplazo, estado_anterior_resultante, creado_por, fecha_creacion) VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph})"
        cursor.execute(q_ins, (activo_id, activo_nuevo_id, motivo, notas or None, fecha_reemplazo, estado_resultante, usuario, fecha_creacion))

        cursor.execute(f"UPDATE activos_inventario SET estado = {ph} WHERE id = {ph}", (estado_resultante, activo_id))

        detalle = f"Se reemplazó el activo #{activo_id} ('{nombre_anterior}') — motivo: {motivo}. Queda en estado '{estado_resultante}'."
        if activo_nuevo_id:
            # 🔁 El activo de reemplazo hereda responsable, sede y área del que sale de servicio,
            # para que el inventario refleje de una vez a quién/dónde quedó el equipo nuevo.
            estado_nuevo = 'Asignado' if asignado_a_anterior else 'Disponible'
            cursor.execute(
                f"UPDATE activos_inventario SET asignado_a = {ph}, sede = {ph}, area = {ph}, estado = {ph} WHERE id = {ph} AND eliminado = 0",
                (asignado_a_anterior, sede_anterior, area_anterior, estado_nuevo, activo_nuevo_id)
            )
            detalle += f" Reemplazado por el activo #{activo_nuevo_id}."

        conn.commit()
        registrar_log(usuario, "Inventario de Activos", detalle)
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error reemplazando activo {activo_id}: {e}")
    conn.close()
    return redirect(url_for('ver_inventario'))


# 🗂️ CATÁLOGO DE TIPOS DE ACTIVO (administrable desde el modal "Tipos de activo" de
# Inventario — key, etiqueta visible, ícono y orden). Reemplaza la lista fija TIPOS_ACTIVO.
@app.route('/tickets/inventario/tipos/crear', methods=['POST'])
@login_required
@admin_required
def crear_tipo_activo_catalogo():
    key = re.sub(r'[^A-Z0-9_]', '', request.form.get('key', '').strip().upper())
    etiqueta = request.form.get('etiqueta', '').strip()
    icono = (request.form.get('icono', '').strip() or 'box')
    if key and etiqueta:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT COALESCE(MAX(orden), -1) FROM tipos_activo_catalogo")
            siguiente_orden = cursor.fetchone()[0] + 1
            q = "INSERT INTO tipos_activo_catalogo (key, etiqueta, icono, orden) VALUES (%s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO tipos_activo_catalogo (key, etiqueta, icono, orden) VALUES (?, ?, ?, ?)"
            cursor.execute(q, (key, etiqueta, icono, siguiente_orden))
            conn.commit()
            registrar_log(session.get('username'), "Catálogo de Tipos de Activo", f"Se agregó el tipo '{etiqueta}' ({key})")
        except Exception as e:
            conn.rollback()
            print(f"⚠️ Error agregando tipo de activo '{etiqueta}': {e}")
        conn.close()
    return redirect(url_for('ver_inventario'))


@app.route('/tickets/inventario/tipos/<int:tipo_id>/eliminar', methods=['POST'])
@login_required
@admin_required
def eliminar_tipo_activo_catalogo(tipo_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE tipos_activo_catalogo SET estado = 'inactivo' WHERE id = %s" if db_type == 'postgres' else "UPDATE tipos_activo_catalogo SET estado = 'inactivo' WHERE id = ?"
        cursor.execute(q, (tipo_id,))
        conn.commit()
        registrar_log(session.get('username'), "Catálogo de Tipos de Activo", f"Se desactivó el tipo de activo ID {tipo_id}")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error desactivando tipo de activo {tipo_id}: {e}")
    conn.close()
    return redirect(url_for('ver_inventario'))


@app.route('/tickets/inventario/tipos/<int:tipo_id>/reordenar', methods=['POST'])
@login_required
@admin_required
def reordenar_tipo_activo_catalogo(tipo_id):
    """Sube o baja un tipo de activo en el orden de despliegue, intercambiando su 'orden' con
    el del vecino inmediato (entre los tipos activos) — las flechas ↑/↓ del modal."""
    direccion = request.form.get('direccion', '')
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, orden FROM tipos_activo_catalogo WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY orden ASC, etiqueta ASC")
        filas = cursor.fetchall()
        ids_ordenados = [f[0] for f in filas]
        if tipo_id in ids_ordenados:
            pos = ids_ordenados.index(tipo_id)
            pos_vecino = pos - 1 if direccion == 'up' else pos + 1
            if 0 <= pos_vecino < len(ids_ordenados):
                id_vecino = ids_ordenados[pos_vecino]
                orden_actual = filas[pos][1]
                orden_vecino = filas[pos_vecino][1]
                q = "UPDATE tipos_activo_catalogo SET orden = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE tipos_activo_catalogo SET orden = ? WHERE id = ?"
                cursor.execute(q, (orden_vecino, tipo_id))
                cursor.execute(q, (orden_actual, id_vecino))
                conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error reordenando tipo de activo {tipo_id}: {e}")
    conn.close()
    return redirect(url_for('ver_inventario'))


# 📎 ADJUNTAR ARCHIVOS A UN ACTIVO DEL INVENTARIO (facturas, fotos, garantías, manuales...).
# Reutiliza el mismo subidor de Cloudinary que ya usan los adjuntos de tickets.
@app.route('/tickets/inventario/<int:activo_id>/adjuntar', methods=['POST'])
@login_required
@agente_o_admin_required
def adjuntar_archivo_inventario(activo_id):
    archivos = request.files.getlist('adjuntos')
    subidos = _subir_adjuntos_ticket(archivos)

    if subidos:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            fecha_act = obtener_fecha_actual()
            usuario = session.get('username')
            _guardar_adjuntos_inventario(cursor, db_type, activo_id, subidos, usuario, fecha_act)
            conn.commit()
            registrar_log(usuario, "Inventario de Activos", f"Se adjuntaron {len(subidos)} archivo(s) al activo #{activo_id}")
        except Exception as e:
            conn.rollback()
            print(f"Error adjuntando archivos al activo {activo_id}: {e}")
        conn.close()
    return redirect(url_for('ver_inventario'))


@app.route('/tickets/inventario/adjunto/<int:adjunto_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_adjunto_inventario(adjunto_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "DELETE FROM inventario_adjuntos WHERE id = %s" if db_type == 'postgres' else "DELETE FROM inventario_adjuntos WHERE id = ?"
        cursor.execute(q, (adjunto_id,))
        conn.commit()
        registrar_log(session.get('username'), "Inventario de Activos", f"Se eliminó el adjunto #{adjunto_id}")
    except Exception as e:
        conn.rollback()
        print(f"Error eliminando adjunto de inventario {adjunto_id}: {e}")
    conn.close()
    return redirect(url_for('ver_inventario'))


# 🪪 CERTIFICACIÓN DE DEVOLUCIÓN DE ACTIVOS (retiro/liquidación de colaboradores) ---------
# Antes de dar de baja la cuenta de un colaborador en Gestión de Usuarios, Gestión Humana o
# TI deben certificar aquí que ya devolvió el PC u otro activo que tenía asignado en el
# Inventario (ver _activos_pendientes_devolucion / toggle_estado_usuario). Cada confirmación
# libera el activo (vuelve a 'Disponible') y queda registrada en 'inventario_devoluciones'
# como el certificado de devolución, exportable a CSV para Gestión Humana.
@app.route('/inventario/certificacion_devoluciones')
@login_required
@certificacion_devolucion_required
def certificacion_devoluciones():
    busqueda = (request.args.get('q') or '').strip().lower()
    conn, db_type = get_db()
    cursor = conn.cursor()
    pendientes = []
    try:
        cursor.execute("SELECT id, nombre, tipo_activo, marca, modelo, numero_serie, asignado_a, sede FROM activos_inventario WHERE estado = 'Asignado' AND COALESCE(eliminado, 0) = 0 ORDER BY asignado_a ASC")
        for aid, nombre, tipo_activo, marca, modelo, numero_serie, asignado_a, sede in cursor.fetchall():
            if busqueda and busqueda not in f"{asignado_a or ''} {nombre or ''}".lower():
                continue
            pendientes.append({
                'id': aid, 'nombre': nombre, 'tipo_activo': tipo_activo, 'marca': marca,
                'modelo': modelo, 'numero_serie': numero_serie, 'asignado_a': asignado_a, 'sede': sede
            })
    except Exception as e:
        print(f"⚠️ Error consultando pendientes de devolución: {e}")

    historial = []
    try:
        cursor.execute("""SELECT d.colaborador, a.nombre, a.tipo_activo, d.confirmado_por, d.fecha, d.observaciones
                           FROM inventario_devoluciones d JOIN activos_inventario a ON a.id = d.activo_id
                           ORDER BY d.id DESC LIMIT 200""")
        historial = [{'colaborador': r[0], 'nombre_activo': r[1], 'tipo_activo': r[2],
                      'confirmado_por': r[3], 'fecha': r[4], 'observaciones': r[5]} for r in cursor.fetchall()]
    except Exception as e:
        print(f"⚠️ Error consultando historial de devoluciones: {e}")
    conn.close()

    return render_template('inventario_certificacion.html', pendientes=pendientes, historial=historial,
                            busqueda=busqueda, es_soporte=(session.get('rol') in ROLES_CON_ACCESO_OPERATIVO))


@app.route('/inventario/<int:activo_id>/confirmar_devolucion', methods=['POST'])
@login_required
@certificacion_devolucion_required
def confirmar_devolucion_activo(activo_id):
    usuario = session.get('username')
    observaciones = (request.form.get('observaciones') or '').strip() or None
    conn, db_type = get_db()
    cursor = conn.cursor()
    ph = '%s' if db_type == 'postgres' else '?'
    try:
        cursor.execute(f"SELECT nombre, asignado_a, estado FROM activos_inventario WHERE id = {ph} AND COALESCE(eliminado, 0) = 0", (activo_id,))
        row = cursor.fetchone()
        if not row:
            flash("El activo indicado no existe o fue eliminado.", "error")
        elif row[2] != 'Asignado':
            flash("Este activo ya no figura como asignado; no hay una devolución pendiente que certificar.", "error")
        else:
            nombre_activo, colaborador, _ = row
            fecha_act = obtener_fecha_actual()
            q_ins = ("INSERT INTO inventario_devoluciones (activo_id, colaborador, confirmado_por, fecha, observaciones) VALUES (%s, %s, %s, %s, %s)"
                      if db_type == 'postgres' else
                      "INSERT INTO inventario_devoluciones (activo_id, colaborador, confirmado_por, fecha, observaciones) VALUES (?, ?, ?, ?, ?)")
            cursor.execute(q_ins, (activo_id, colaborador, usuario, fecha_act, observaciones))
            q_upd = f"UPDATE activos_inventario SET estado = {ph}, asignado_a = NULL WHERE id = {ph}"
            cursor.execute(q_upd, ('Disponible', activo_id))
            conn.commit()
            registrar_log(usuario, "Certificación de Devolución de Activo",
                          f"Se certificó la devolución de '{nombre_activo}' por parte de {colaborador}"
                          + (f" — nota: {observaciones}" if observaciones else ""))
            flash(f"Devolución de '{nombre_activo}' certificada. El activo vuelve a estar disponible en el Inventario.", "exito")
    except Exception as e:
        conn.rollback()
        print(f"Error certificando devolución del activo {activo_id}: {e}")
        flash("No se pudo registrar la certificación de devolución.", "error")
    conn.close()
    return redirect(url_for('certificacion_devoluciones'))


@app.route('/inventario/certificacion_devoluciones/exportar_csv')
@login_required
@certificacion_devolucion_required
def exportar_certificacion_devoluciones():
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""SELECT d.colaborador, a.nombre, a.tipo_activo, d.confirmado_por, d.fecha, d.observaciones
                           FROM inventario_devoluciones d JOIN activos_inventario a ON a.id = d.activo_id
                           ORDER BY d.id DESC""")
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error exportando certificación de devoluciones: {e}")
        filas = []
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['COLABORADOR', 'ACTIVO', 'TIPO_ACTIVO', 'CERTIFICADO_POR', 'FECHA', 'OBSERVACIONES'])
    for colaborador, nombre_activo, tipo_activo, confirmado_por, fecha, observaciones in filas:
        writer.writerow([colaborador, nombre_activo, tipo_activo or '', confirmado_por, fecha or '', observaciones or ''])
    csv_bytes = '﻿' + output.getvalue()
    headers = {'Content-Type': 'text/csv; charset=utf-8', 'Content-Disposition': 'attachment; filename="certificacion_devolucion_activos.csv"'}
    registrar_log(session.get('username'), "Exportación de Certificación de Devoluciones",
                  f"Se exportó el listado de {len(filas)} devolución(es) certificada(s)")
    return Response(csv_bytes, headers=headers, status=200)


# 📧 ENVÍO VÍA GMAIL APPS SCRIPT (PUERTO 443 HTTPS - SIN BLOQUEOS)
def enviar_correo_recuperacion(email_destino, usuario_nombre, codigo):
    try:
        cuerpo = f"Hola {usuario_nombre},\n\nTu código de verificación para restablecer tu contraseña en ARKIV es: {codigo}\n\nSi no solicitaste este cambio, por favor ignora este mensaje.\n---\nEquipo de Soporte - ARKIV System"
        
        payload = {
            "para": email_destino,
            "asunto": f"Código de Verificación - Gestor de Archivos",
            "cuerpo": cuerpo
        }

        if requests:
            res = requests.post(GMAIL_SCRIPT_URL, json=payload, timeout=15)
            print(f"✅ EXITO: Correo enviado a {email_destino} vía Google Script. Status: {res.status_code}")
            registrar_correo_log(email_destino, payload["asunto"], 'recuperacion', 'enviado')
            return True
        else:
            data_json = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(GMAIL_SCRIPT_URL, data=data_json, headers={'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=15) as response:
                res_text = response.read().decode('utf-8')
                print(f"✅ EXITO: Correo enviado a {email_destino} vía urllib. Respuesta: {res_text}")
                registrar_correo_log(email_destino, payload["asunto"], 'recuperacion', 'enviado')
                return True

    except Exception as e:
        print(f"❌ Error en envío vía Google Script: {e}")
        traceback.print_exc()
        # 🛡️ El detalle guardado es solo el mensaje de la excepción (p. ej. timeout, error HTTP);
        # el código de verificación vive únicamente en 'codigo' y jamás llega hasta aquí.
        registrar_correo_log(email_destino, "Código de Verificación - Gestor de Archivos", 'recuperacion', 'error', str(e)[:300])
        return False


# 👋📧 Correo de bienvenida al crear una cuenta nueva desde Gestión de Usuarios: le avisa al
# usuario su nombre de usuario y contraseña temporal, y le pide cambiarla en su primer ingreso.
# Se llama en un hilo aparte (threading.Thread) para no hacer esperar al admin que la creó.
# 🛡️ La contraseña SÍ viaja en el cuerpo del correo (es la única forma de que el usuario la
# reciba), pero NUNCA se guarda en correos_log — ahí solo queda el asunto genérico, igual que
# con el código de recuperación (ver registrar_correo_log).
def enviar_correo_bienvenida(email_destino, usuario, nombre_completo, password_temporal):
    if not email_destino:
        return False
    asunto = "Bienvenido a Arkiv - Tu cuenta fue creada"
    try:
        cuerpo = (
            f"Hola {nombre_completo},\n\n"
            f"Se creó tu cuenta en el sistema ARKIV. Estos son tus datos de acceso:\n\n"
            f"Usuario: {usuario}\n"
            f"Contraseña temporal: {password_temporal}\n\n"
            f"Por seguridad, te pedimos cambiar esta contraseña apenas inicies sesión por primera vez "
            f"(la puedes cambiar desde tu perfil, o con la opción '¿Olvidaste tu contraseña?' del login).\n\n"
            f"Ingresa aquí: https://gestor-archivos-mossoft-1.onrender.com/login\n\n"
            "Si no esperabas este correo, contacta a tu administrador.\n"
            "---\nEquipo de Soporte - ARKIV System"
        )
        payload = {"para": email_destino, "asunto": asunto, "cuerpo": cuerpo}
        if requests:
            res = requests.post(GMAIL_SCRIPT_URL, json=payload, timeout=15)
            print(f"✅ Correo de bienvenida enviado a {email_destino}. Status: {res.status_code}")
        else:
            data_json = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(GMAIL_SCRIPT_URL, data=data_json, headers={'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=15) as response:
                response.read()
            print(f"✅ Correo de bienvenida enviado a {email_destino} vía urllib.")
        registrar_correo_log(email_destino, asunto, 'bienvenida', 'enviado')
        return True
    except Exception as e:
        print(f"⚠️ Error enviando correo de bienvenida a {email_destino}: {e}")
        registrar_correo_log(email_destino, asunto, 'bienvenida', 'error', str(e)[:300])
        return False


# 🎫📧 Notificaciones por correo de Tickets: reutiliza el mismo webhook de Apps Script que ya
# usa la recuperación de contraseña. Se llama siempre en un hilo aparte (threading.Thread)
# desde las rutas de tickets para no hacer esperar al usuario a que el correo salga.
def enviar_correo_ticket(email_destino, asunto, cuerpo):
    if not email_destino:
        return False
    try:
        payload = {"para": email_destino, "asunto": asunto, "cuerpo": cuerpo}
        if requests:
            res = requests.post(GMAIL_SCRIPT_URL, json=payload, timeout=15)
            print(f"✅ Correo de ticket enviado a {email_destino}. Status: {res.status_code}")
        else:
            data_json = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(GMAIL_SCRIPT_URL, data=data_json, headers={'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=15) as response:
                response.read()
            print(f"✅ Correo de ticket enviado a {email_destino} vía urllib.")
        registrar_correo_log(email_destino, asunto, 'ticket', 'enviado')
        return True
    except Exception as e:
        print(f"⚠️ Error enviando correo de ticket a {email_destino}: {e}")
        registrar_correo_log(email_destino, asunto, 'ticket', 'error', str(e)[:300])
        return False

# 🔑 PASO 1: SOLICITAR CÓDIGO POR CORREO
@app.route('/recuperar', methods=['GET', 'POST'])
@limiter.limit("10 per minute", methods=["POST"])
def recuperar_clave():
    if request.method == 'POST':
        email_ingresado = request.form.get('email', '').strip().lower()
        print(f"📩 Solicitud de recuperación recibida para: '{email_ingresado}'")

        # 🛡️ Hallazgo QA H-06 (30/08/2026): igual que /login, exige reCAPTCHA para evitar que
        # un script automatizado pruebe correos en masa (enumeración) o sature de correos de
        # recuperación a una cuenta real.
        recaptcha_response = request.form.get('g-recaptcha-response')
        if not verificar_recaptcha(recaptcha_response):
            return render_template('recuperar.html', paso=1, error="Por favor, marca la casilla 'No soy un robot'.")

        conn, db_type = get_db()
        cursor = conn.cursor()
        query = "SELECT usuario FROM usuarios WHERE LOWER(TRIM(correo)) = %s" if db_type == 'postgres' else "SELECT usuario FROM usuarios WHERE LOWER(TRIM(correo)) = ?"
        cursor.execute(query, (email_ingresado,))
        user = cursor.fetchone()
        conn.close()

        if user:
            usuario_nombre = user[0]
            codigo_verificacion = str(random.randint(100000, 999999))

            session['reset_email'] = email_ingresado
            session['reset_user'] = usuario_nombre
            session['reset_code'] = codigo_verificacion
            # 🛡️ El código expira a los 10 minutos y se limita el número de intentos de
            # adivinarlo, para que no quede indefinidamente válido ni sea adivinable por fuerza bruta.
            session['reset_code_expira'] = datetime.now(timezone.utc).timestamp() + 600
            session['reset_intentos'] = 0

            threading.Thread(
                target=enviar_correo_recuperacion,
                args=(email_ingresado, usuario_nombre, codigo_verificacion)
            ).start()

            # 🛡️ El código NUNCA se escribe en el log: la tabla logs es visible para
            # cualquier admin en /logs, y dejarlo en texto plano permitiría a otro admin
            # secuestrar la recuperación de un tercero. Solo se envía por correo.
            registrar_log(usuario_nombre, "Solicitud de Código", f"Se generó un código de verificación para: {email_ingresado}")
            return render_template('recuperar.html', paso=2, email=email_ingresado)
        else:
            return render_template('recuperar.html', paso=1, error="El correo ingresado no está registrado en el sistema.")

    return render_template('recuperar.html', paso=1)

# 🔑 PASO 2: VALIDAR CÓDIGO Y CAMBIAR CONTRASEÑA
@app.route('/validar_codigo', methods=['POST'])
@limiter.limit("15 per minute")
def validar_codigo():
    codigo_ingresado = request.form.get('codigo', '').strip()
    nueva_pass = request.form.get('nueva_password', '').strip()

    codigo_correcto = session.get('reset_code')
    email_usuario = session.get('reset_email')
    nombre_usuario = session.get('reset_user')
    expira = session.get('reset_code_expira')

    if not codigo_correcto or not email_usuario:
        return render_template('recuperar.html', paso=1, error="La sesión expiró. Por favor solicita un nuevo código.")

    # 🛡️ Código vencido (10 minutos desde que se generó): obliga a pedir uno nuevo.
    if not expira or datetime.now(timezone.utc).timestamp() > expira:
        session.pop('reset_code', None)
        session.pop('reset_email', None)
        session.pop('reset_user', None)
        session.pop('reset_code_expira', None)
        session.pop('reset_intentos', None)
        return render_template('recuperar.html', paso=1, error="El código de verificación expiró. Solicita uno nuevo.")

    if codigo_ingresado != codigo_correcto:
        # 🛡️ Límite de intentos: tras 5 códigos incorrectos se invalida y hay que pedir uno nuevo,
        # para que un código de 6 dígitos no quede expuesto a fuerza bruta ilimitada.
        intentos = session.get('reset_intentos', 0) + 1
        session['reset_intentos'] = intentos
        if intentos >= 5:
            session.pop('reset_code', None)
            session.pop('reset_email', None)
            session.pop('reset_user', None)
            session.pop('reset_code_expira', None)
            session.pop('reset_intentos', None)
            return render_template('recuperar.html', paso=1, error="Demasiados intentos fallidos. Solicita un nuevo código.")
        return render_template('recuperar.html', paso=2, email=email_usuario, error="El código de verificación es incorrecto.")

    # 🔒 Hallazgo QA H-02: esta era la única ruta que asignaba una contraseña de cuenta sin
    # exigir la longitud mínima (sí se validaba en autoservicio, no aquí en recuperación).
    if not _password_cuenta_valida(nueva_pass):
        return render_template('recuperar.html', paso=2, email=email_usuario, error=f"La nueva contraseña debe tener al menos {LONGITUD_MINIMA_PASSWORD_CUENTA} caracteres.")

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        nuevo_hash = generate_password_hash(nueva_pass)
        # 🛡️ Se filtra por correo Y por el usuario que resolvió el paso 1: si en el futuro
        # dos cuentas volvieran a compartir el mismo correo, esto evita que una sola
        # recuperación cambie la clave de todas a la vez.
        # 🔒 Esta contraseña ya la eligió el propio usuario, así que se limpia
        # debe_cambiar_password (si venía marcada de una contraseña temporal asignada por un
        # admin, ya no aplica). También se limpia un eventual bloqueo por intentos fallidos
        # (ver UMBRAL_INTENTOS_FALLIDOS_LOGIN): probar el código que llegó a su correo ya
        # demuestra que es el dueño legítimo de la cuenta, así que este es uno de los dos
        # caminos válidos para recuperar el acceso (el otro es que un admin lo desbloquee).
        q_upd = "UPDATE usuarios SET password_hash = %s, debe_cambiar_password = FALSE, bloqueado_por_intentos = FALSE, intentos_fallidos_login = 0 WHERE LOWER(TRIM(correo)) = %s AND usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET password_hash = ?, debe_cambiar_password = 0, bloqueado_por_intentos = 0, intentos_fallidos_login = 0 WHERE LOWER(TRIM(correo)) = ? AND usuario = ?"
        cursor.execute(q_upd, (nuevo_hash, email_usuario, nombre_usuario))
        conn.commit()
        conn.close()

        session.pop('reset_code', None)
        session.pop('reset_email', None)
        session.pop('reset_user', None)
        session.pop('reset_code_expira', None)
        session.pop('reset_intentos', None)

        registrar_log(nombre_usuario, "Cambio Exitoso de Clave", "Se actualizó la clave vía código de verificación.")
        return render_template('recuperar.html', paso=1, exito="¡Contraseña actualizada con éxito! Ya puedes iniciar sesión.")

    except Exception as e:
        conn.rollback()
        conn.close()
        return render_template('recuperar.html', paso=2, email=email_usuario, error="Ocurrió un error al actualizar la contraseña.")

def _fondo_login_activo():
    """Imágenes/videos activos para el panel de marca del login (ver login_fondo_media),
    en el orden en que deben rotar. Se consulta en cada render de /login a propósito —el
    login no usa caché de plantillas— pero es una sola query liviana (pocas filas, sin
    joins) y así un cambio recién publicado por Comunicados se ve de inmediato, sin esperar
    a un reinicio del servidor. Si la tabla no existe todavía (despliegue viejo) o la query
    falla por lo que sea, se devuelve vacío y el login sencillamente no muestra el panel —
    nunca debe tumbar la página de inicio de sesión."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT tipo, url FROM login_fondo_media WHERE estado = 'activo' ORDER BY orden ASC, id ASC")
        filas = [{'tipo': t, 'url': u} for t, u in cursor.fetchall()]
        conn.close()
        return filas
    except Exception as e:
        print(f"⚠️ Error consultando el fondo del login (se omite el panel): {e}")
        return []


def _render_login(**kwargs):
    """Único punto por el que /login renderiza login.html: así el panel de marca
    (fondo_media) siempre viaja junto, sin tener que repetir la consulta en cada uno de los
    ~10 lugares donde /login puede devolver la plantilla (éxito, error de clave, cuenta
    bloqueada, etc.)."""
    return render_template('login.html', fondo_media=_fondo_login_activo(), **kwargs)


# 🔑 LOGIN
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("20 per minute", methods=["POST"])
def login():
    if request.method == 'POST':
        try:
            username = request.form.get('usuario') or request.form.get('username') or ''
            password = request.form.get('contrasena') or request.form.get('password') or ''
            recaptcha_response = request.form.get('g-recaptcha-response')

            if not verificar_recaptcha(recaptcha_response):
                return _render_login(error="Por favor, marca la casilla 'No soy un robot'.")

            try:
                conn, db_type = get_db()
                cursor = conn.cursor()
                # foto_perfil se agrega AL FINAL (índice 9) para no correr los índices existentes.
                query = "SELECT usuario, password_hash, rol, estado, tema, debe_cambiar_password, totp_habilitado, intentos_fallidos_login, bloqueado_por_intentos, foto_perfil FROM usuarios WHERE LOWER(TRIM(usuario)) = LOWER(TRIM(%s))" if db_type == 'postgres' else "SELECT usuario, password_hash, rol, estado, tema, debe_cambiar_password, totp_habilitado, intentos_fallidos_login, bloqueado_por_intentos, foto_perfil FROM usuarios WHERE LOWER(TRIM(usuario)) = LOWER(TRIM(?))"
                cursor.execute(query, (username,))
                user = cursor.fetchone()
                conn.close()

                clave_db = str(user[1] or '') if user else ''
                es_valida = False
                if user:
                    if clave_db.startswith('pbkdf2:') or clave_db.startswith('scrypt:'):
                        es_valida = check_password_hash(clave_db, password)
                    else:
                        # Contraseña vieja guardada en texto plano: valida por compatibilidad
                        # y de una vez la re-guarda ya hasheada para dejar de usar texto plano.
                        # compare_digest en vez de == : evita que el tiempo de respuesta filtre,
                        # carácter por carácter, cuánto de la contraseña coincidió (timing attack).
                        es_valida = hmac.compare_digest(clave_db, password)
                        if es_valida:
                            _migrar_password_a_hash(user[0], password)

                # 🔒 Mensaje de bloqueo por intentos fallidos. Igual que con "cuenta desactivada"
                # abajo, esto solo se revela a quien YA acertó la contraseña — así no se le
                # confirma a un tercero que adivina contraseñas si una cuenta está bloqueada o no.
                if es_valida and len(user) > 8 and user[8]:
                    registrar_log(user[0], "Inicio de Sesión Bloqueado", "Intento de acceso (con contraseña correcta) a una cuenta bloqueada por intentos fallidos.", ip=_obtener_ip_cliente(), dispositivo=_detectar_dispositivo(request.headers.get('User-Agent', '')))
                    return _render_login(error="Tu cuenta quedó bloqueada por varios intentos fallidos de inicio de sesión. Cámbiala desde \"¿Olvidaste tu clave?\" o pide a un administrador que la desbloquee.")

                if es_valida:
                    # Contraseña correcta y cuenta no bloqueada por intentos: se rompió la racha
                    # de fallos, así que el contador se reinicia (si tenía alguno acumulado).
                    if len(user) > 7 and user[7]:
                        _resetear_intentos_fallidos_login(user[0])

                    # ⚠️ user[3] es la columna "estado" (activo/inactivo). Un usuario bloqueado
                    # por un administrador no debe poder iniciar sesión aunque su clave sea correcta.
                    if (user[3] or 'activo') == 'inactivo':
                        registrar_log(user[0], "Inicio de Sesión Bloqueado", "Intento de acceso de una cuenta desactivada.", ip=_obtener_ip_cliente(), dispositivo=_detectar_dispositivo(request.headers.get('User-Agent', '')))
                        return _render_login(error="Tu cuenta ha sido desactivada. Contacta a un administrador.")

                    # 🔐 Si la cuenta tiene activada la verificación en dos pasos, la sesión NO se
                    # abre todavía: se guarda solo un marcador "pre-2FA" (session['logged_in']
                    # sigue sin existir, así que validar_instancia_y_sesion la deja pasar de largo)
                    # y se manda a /login/2fa a pedir el código de la app autenticadora.
                    if len(user) > 6 and user[6]:
                        session['pre_2fa_usuario'] = user[0]
                        session['pre_2fa_expira'] = (datetime.now(timezone.utc) + timedelta(minutes=10)).timestamp()
                        session.pop('pre_2fa_intentos', None)
                        return redirect(url_for('login_2fa'))

                    session.permanent = True
                    session['logged_in'] = True
                    session['username'] = user[0]
                    session['rol'] = user[2]
                    session['instance_id'] = SERVER_INSTANCE_ID
                    session['tema'] = user[4] or 'oscuro'
                    session['debe_cambiar_password'] = bool(user[5]) if len(user) > 5 else False
                    # 🔒 Hallazgo QA H-05: si llegamos aquí (no pasó por /login/2fa), esta cuenta
                    # todavía no tiene 2FA activo. Para admin/agente, eso ya no es opcional.
                    session['debe_activar_2fa'] = user[2] in ROLES_CON_ACCESO_OPERATIVO
                    # 🖼️ Foto de perfil: se cachea en la sesión para no consultar la BD en cada
                    # página solo para pintar el ícono del encabezado (ver partials/perfil_boton.html).
                    session['foto_perfil'] = user[9] if len(user) > 9 else None
                    registrar_log(user[0], "Inicio de Sesión", "Inicio de sesión exitoso", ip=_obtener_ip_cliente(), dispositivo=_detectar_dispositivo(request.headers.get('User-Agent', '')))
                    return redirect(url_for('bienvenida'))
                elif user:
                    # 🔒 Contraseña incorrecta en una cuenta que sí existe: cuenta como intento
                    # fallido hacia el bloqueo automático (ver UMBRAL_INTENTOS_FALLIDOS_LOGIN).
                    # El contador de intentos restantes se muestra en el propio mensaje de
                    # error, para que la persona sepa cuántas oportunidades le quedan antes de
                    # que se bloquee la cuenta.
                    bloqueada, intentos_restantes = _registrar_intento_fallido_login(user[0])
                    if bloqueada:
                        return _render_login(error="Contraseña incorrecta. Por seguridad, la cuenta quedó bloqueada tras varios intentos fallidos. Cámbiala desde \"¿Olvidaste tu clave?\" o pide a un administrador que la desbloquee.")
                    if intentos_restantes == 1:
                        return _render_login(error="Usuario o contraseña incorrectos. Te queda 1 intento antes de que la cuenta se bloquee por seguridad.")
                    return _render_login(error=f"Usuario o contraseña incorrectos. Te quedan {intentos_restantes} intentos antes de que la cuenta se bloquee por seguridad.")
            except Exception as db_err:
                print(f"Error consultando usuario en BD: {db_err}")

            return _render_login(error="Usuario o contraseña incorrectos.")

        except Exception as e:
            print(f"Error general en login: {e}")
            return _render_login(error="Ocurrió un error en el servidor. Por favor intenta de nuevo.")

    mensaje_expirado = "⚠️ Tu sesión ha expirado. Por favor ingresa nuevamente." if request.args.get('expirado') == '1' else None
    return _render_login(mensaje_expirado=mensaje_expirado)


@app.route('/login/2fa', methods=['GET', 'POST'])
@limiter.limit("15 per minute", methods=["POST"])
def login_2fa():
    """Segundo paso del inicio de sesión para las cuentas con 2FA activo. Solo es accesible tras
    validar usuario/contraseña en /login (que deja el marcador 'pre_2fa_usuario' en la sesión,
    con una ventana de 10 minutos para completarlo). No lleva @login_required porque, por
    definición, la sesión todavía no está autenticada en este punto."""
    usuario_pendiente = session.get('pre_2fa_usuario')
    expira = session.get('pre_2fa_expira')

    if not usuario_pendiente or not expira or datetime.now(timezone.utc).timestamp() > expira:
        session.pop('pre_2fa_usuario', None)
        session.pop('pre_2fa_expira', None)
        session.pop('pre_2fa_intentos', None)
        return redirect(url_for('login'))

    error = None
    if request.method == 'POST':
        codigo = re.sub(r'\s+', '', request.form.get('codigo', '') or '')
        codigo_respaldo = (request.form.get('codigo_respaldo') or '').strip()

        conn, db_type = get_db()
        cursor = conn.cursor()
        # foto_perfil se agrega AL FINAL (índice 6) para no correr los índices existentes.
        query = "SELECT usuario, rol, estado, tema, debe_cambiar_password, totp_secret, foto_perfil FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT usuario, rol, estado, tema, debe_cambiar_password, totp_secret, foto_perfil FROM usuarios WHERE usuario = ?"
        cursor.execute(query, (usuario_pendiente,))
        user = cursor.fetchone()
        conn.close()

        # 🛡️ Si la cuenta desapareció o fue bloqueada mientras esperaba el segundo paso, se
        # corta el flujo en vez de dejarla completar el login.
        if not user or (user[2] or 'activo') == 'inactivo':
            session.pop('pre_2fa_usuario', None)
            session.pop('pre_2fa_expira', None)
            session.pop('pre_2fa_intentos', None)
            return redirect(url_for('login'))

        es_valido = False
        via_respaldo = False
        if codigo_respaldo:
            es_valido = _verificar_codigo_respaldo_2fa(usuario_pendiente, codigo_respaldo)
            via_respaldo = es_valido
        elif codigo and user[5]:
            # 🔒 Hallazgo QA H-07: el secreto ya se guarda cifrado con Fernet (ver
            # _descifrar_totp_secret); esto también migra en el momento cualquier secreto
            # todavía en texto plano de antes de este cambio.
            secreto_totp = _descifrar_totp_secret(usuario_pendiente, user[5])
            es_valido = bool(secreto_totp) and pyotp.TOTP(secreto_totp).verify(codigo, valid_window=1)

        if es_valido:
            session.pop('pre_2fa_usuario', None)
            session.pop('pre_2fa_expira', None)
            session.pop('pre_2fa_intentos', None)
            session.permanent = True
            session['logged_in'] = True
            session['username'] = user[0]
            session['rol'] = user[1]
            session['instance_id'] = SERVER_INSTANCE_ID
            session['tema'] = user[3] or 'oscuro'
            session['debe_cambiar_password'] = bool(user[4])
            # 🔒 Hallazgo QA H-05: llegar aquí ya implica que la cuenta tiene 2FA activo.
            session['debe_activar_2fa'] = False
            # 🖼️ Foto de perfil: se cachea en la sesión (ver login() más arriba).
            session['foto_perfil'] = user[6] if len(user) > 6 else None
            detalle = "Inicio de sesión exitoso (código de respaldo 2FA)" if via_respaldo else "Inicio de sesión exitoso (verificación en dos pasos)"
            registrar_log(user[0], "Inicio de Sesión", detalle, ip=_obtener_ip_cliente(), dispositivo=_detectar_dispositivo(request.headers.get('User-Agent', '')))
            return redirect(url_for('bienvenida'))

        # 🛡️ Límite de intentos: 5 códigos incorrectos consecutivos obligan a volver a
        # ingresar usuario y contraseña, igual que en la recuperación de contraseña.
        intentos = session.get('pre_2fa_intentos', 0) + 1
        session['pre_2fa_intentos'] = intentos
        if intentos >= 5:
            registrar_log(usuario_pendiente, "Inicio de Sesión Bloqueado", "Demasiados intentos fallidos de verificación en dos pasos.", ip=_obtener_ip_cliente(), dispositivo=_detectar_dispositivo(request.headers.get('User-Agent', '')))
            session.pop('pre_2fa_usuario', None)
            session.pop('pre_2fa_expira', None)
            session.pop('pre_2fa_intentos', None)
            return _render_login(error="Demasiados intentos fallidos. Vuelve a iniciar sesión.")

        error = "El código ingresado no es válido. Verifica la hora de tu dispositivo o usa un código de respaldo."

    return render_template('login_2fa.html', error=error)

# 📊 RUTAS DE MÉTRICAS
@app.route('/incrementar_vista/<galeria_id>', methods=['POST'])
@csrf.exempt
@login_required
def incrementar_vista(galeria_id):
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "UPDATE galerias SET vistas = COALESCE(vistas, 0) + 1 WHERE id = %s" if db_type == 'postgres' else "UPDATE galerias SET vistas = COALESCE(vistas, 0) + 1 WHERE id = ?"
        cursor.execute(q, (galeria_id,))
        # 👁️ Además del contador total (arriba, que sigue sumando en cada vista), se guarda la
        # fecha de la PRIMERA vista de este usuario para este instructivo — para poder mostrar y
        # exportar quién lo ha visto (ver /galerias/<id>/vistas). ON CONFLICT DO NOTHING / INSERT
        # OR IGNORE: una vista repetida del mismo usuario no pisa la fecha original.
        usuario = session.get('username')
        if usuario:
            fecha_act = obtener_fecha_actual()
            if db_type == 'postgres':
                q_vista = ("INSERT INTO galerias_vistas (galeria_id, usuario, fecha) VALUES (%s, %s, %s) "
                           "ON CONFLICT (galeria_id, usuario) DO NOTHING")
            else:
                q_vista = "INSERT OR IGNORE INTO galerias_vistas (galeria_id, usuario, fecha) VALUES (?, ?, ?)"
            cursor.execute(q_vista, (galeria_id, usuario, fecha_act))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 200

@app.route('/incrementar_descarga/<galeria_id>', methods=['POST'])
@csrf.exempt
@login_required
def incrementar_descarga(galeria_id):
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        
        q_tit = "SELECT titulo FROM galerias WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM galerias WHERE id = ?"
        cursor.execute(q_tit, (galeria_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else galeria_id

        q = "UPDATE galerias SET descargas = COALESCE(descargas, 0) + 1 WHERE id = %s" if db_type == 'postgres' else "UPDATE galerias SET descargas = COALESCE(descargas, 0) + 1 WHERE id = ?"
        cursor.execute(q, (galeria_id,))
        conn.commit()
        conn.close()

        usuario_actual = session.get('username', 'Anónimo')
        registrar_log(usuario_actual, "Descarga de Archivo", f"El usuario descargó material del instructivo: '{titulo}'")

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 200


def _exportar_csv_personas(nombre_archivo, filas):
    """CSV compartido para exportar "quién vio/leyó algo": Nombre, Documento, Usuario, Fecha.
    Usado tanto por las vistas de instructivos (galerías) como por las lecturas de comunicados,
    para no repetir el mismo boilerplate de csv.writer/BOM/cabeceras dos veces. 'filas' es una
    lista de tuplas (nombre, cedula, usuario, fecha)."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['NOMBRE', 'DOCUMENTO', 'USUARIO', 'FECHA'])
    for nombre, cedula, usuario, fecha in filas:
        writer.writerow([nombre or usuario, cedula or '', usuario, fecha or ''])

    csv_bytes = '﻿' + output.getvalue()
    headers = {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': f'attachment; filename="{nombre_archivo}"'
    }
    return Response(csv_bytes, headers=headers, status=200)


@app.route('/galerias/<galeria_id>/vistas')
@login_required
@agente_o_admin_required
def vistas_galeria(galeria_id):
    """JSON con quién ha visto este instructivo (nombre, cédula y fecha de la primera vista) —
    usado por el modal "Ver vistas" en el Gestor de Instructivos. Restringido a agente/admin,
    igual que el resto de acciones de gestión de esta tarjeta."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = (
            "SELECT gv.usuario, COALESCE(u.nombre, gv.usuario), u.cedula, gv.fecha "
            "FROM galerias_vistas gv LEFT JOIN usuarios u ON u.usuario = gv.usuario "
            "WHERE gv.galeria_id = %s ORDER BY gv.fecha DESC"
        ) if db_type == 'postgres' else (
            "SELECT gv.usuario, COALESCE(u.nombre, gv.usuario), u.cedula, gv.fecha "
            "FROM galerias_vistas gv LEFT JOIN usuarios u ON u.usuario = gv.usuario "
            "WHERE gv.galeria_id = ? ORDER BY gv.fecha DESC"
        )
        cursor.execute(q, (galeria_id,))
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando vistas del instructivo {galeria_id}: {e}")
        filas = []
    conn.close()

    vistas = [{'usuario': f[0], 'nombre': f[1], 'cedula': f[2] or '', 'fecha': f[3]} for f in filas]
    return jsonify({'vistas': vistas})


@app.route('/galerias/<galeria_id>/vistas/exportar_csv')
@login_required
@agente_o_admin_required
def exportar_vistas_galeria(galeria_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_tit = "SELECT titulo FROM galerias WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM galerias WHERE id = ?"
        cursor.execute(q_tit, (galeria_id,))
        fila_tit = cursor.fetchone()
        titulo = fila_tit[0] if fila_tit else galeria_id

        q = (
            "SELECT COALESCE(u.nombre, gv.usuario), u.cedula, gv.usuario, gv.fecha "
            "FROM galerias_vistas gv LEFT JOIN usuarios u ON u.usuario = gv.usuario "
            "WHERE gv.galeria_id = %s ORDER BY gv.fecha DESC"
        ) if db_type == 'postgres' else (
            "SELECT COALESCE(u.nombre, gv.usuario), u.cedula, gv.usuario, gv.fecha "
            "FROM galerias_vistas gv LEFT JOIN usuarios u ON u.usuario = gv.usuario "
            "WHERE gv.galeria_id = ? ORDER BY gv.fecha DESC"
        )
        cursor.execute(q, (galeria_id,))
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error exportando vistas del instructivo {galeria_id}: {e}")
        filas = []
    conn.close()

    registrar_log(session.get('username'), "Exportación de Vistas de Instructivo", f"Exportó las vistas de '{titulo}'")
    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y%m%d_%H%M")
    nombre_limpio = re.sub(r'[^A-Za-z0-9_-]+', '_', titulo)[:60]
    return _exportar_csv_personas(f"Arkiv_Vistas_{nombre_limpio}_{fecha_filename}.csv", filas)


# 🛡️ Dominios de Cloudinary a los que este proxy tiene permitido conectarse.
# Cualquier otro destino se rechaza ANTES de hacer la petición saliente: sin esto,
# /pdf_proxy podía usarse como SSRF hacia servidores externos, y si esos servidores
# respondían 401, la app terminaba reenviando las credenciales reales de Cloudinary.
DOMINIOS_CLOUDINARY_PERMITIDOS = {'res.cloudinary.com', 'res-console.cloudinary.com'}


def _url_cloudinary_valida(url):
    try:
        partes = urllib.parse.urlparse(url)
    except Exception:
        return False
    return (
        partes.scheme == 'https'
        and (partes.hostname or '').lower() in DOMINIOS_CLOUDINARY_PERMITIDOS
    )


# 🚀 PROXY AUTENTICADO
@app.route('/pdf_proxy')
@login_required
def pdf_proxy():
    url_target = request.args.get('url')
    download_flag = request.args.get('download', '0')
    filename_custom = request.args.get('name', '')

    if not url_target:
        return "URL requerida", 400

    try:
        if not filename_custom:
            filename_custom = url_target.split('/')[-1]

        clean_url = url_target.replace('/fl_attachment/', '/').replace('/upload/fl_attachment/', '/upload/')

        # 🛡️ Bloquea SSRF: solo se permite reenviar la petición si el destino final
        # es realmente Cloudinary. Sin esta validación, cualquier usuario autenticado
        # podía apuntar "url" a un servidor propio y capturar las credenciales de
        # Cloudinary (ver comentario de DOMINIOS_CLOUDINARY_PERMITIDOS arriba).
        if not _url_cloudinary_valida(clean_url):
            usuario_actual = session.get('username', 'Anónimo')
            registrar_log(usuario_actual, "Intento de Proxy Bloqueado", f"URL no permitida (no es Cloudinary): {url_target[:200]}")
            return "URL no permitida: este proxy solo puede usarse para documentos alojados en Cloudinary.", 400

        if download_flag == '1':
            usuario_actual = session.get('username', 'Anónimo')
            registrar_log(usuario_actual, "Descarga de Documento", f"Archivo: '{filename_custom}'")

        disposition = 'attachment' if download_flag == '1' else 'inline'

        fname_lower = filename_custom.lower()
        if fname_lower.endswith('.png'):
            content_type = 'image/png'
        elif fname_lower.endswith(('.jpg', '.jpeg')):
            content_type = 'image/jpeg'
        elif fname_lower.endswith('.gif'):
            content_type = 'image/gif'
        elif fname_lower.endswith('.webp'):
            content_type = 'image/webp'
        elif fname_lower.endswith(('.mp4', '.mov', '.webm', '.avi')):
            content_type = 'video/mp4'
        elif fname_lower.endswith(('.zip', '.rar', '.7z', '.tar', '.gz')):
            content_type = 'application/zip'
        else:
            content_type = 'application/pdf'

        headers = {
            'Content-Type': content_type,
            'Content-Disposition': f'{disposition}; filename="{filename_custom}"'
        }

        # 🚿 STREAMING: para videos/PDFs grandes, no cargamos el archivo completo en
        # memoria (eso podía colgar el worker o agotar la RAM en Render con archivos
        # pesados, dejando la previsualización/descarga en blanco o con error).
        # Se transmite en bloques directamente desde Cloudinary hacia el navegador.
        if requests:
            upstream = requests.get(clean_url, timeout=(10, 60), stream=True)
            if upstream.status_code == 401:
                api_key = os.environ.get('CLOUDINARY_API_KEY')
                api_secret = os.environ.get('CLOUDINARY_API_SECRET')
                if api_key and api_secret:
                    upstream = requests.get(clean_url, auth=(api_key, api_secret), timeout=(10, 60), stream=True)

            if upstream.status_code >= 400:
                return f"Error obteniendo documento: el origen respondió {upstream.status_code}", 502

            content_length = upstream.headers.get('Content-Length')
            if content_length:
                headers['Content-Length'] = content_length

            def _generar():
                try:
                    for chunk in upstream.iter_content(chunk_size=65536):
                        if chunk:
                            yield chunk
                finally:
                    upstream.close()

            return Response(stream_with_context(_generar()), headers=headers, status=200)
        else:
            req = urllib.request.Request(clean_url)
            with urllib.request.urlopen(req) as response:
                content_data = response.read()
            return Response(content_data, headers=headers, status=200)
    except Exception as e:
        return f"Error obteniendo documento: {e}", 500

def _estado_rotacion_credencial(rotacion_dias, fecha_ultima_rotacion):
    """Calcula si a una credencial ya le toca rotar la contraseña, según su política opcional
    (rotacion_dias) y la fecha del último cambio. Devuelve None si no tiene política configurada
    — esa credencial simplemente no se hace seguimiento de rotación."""
    if not rotacion_dias:
        return None
    fecha_base = _parsear_fecha_ticket(fecha_ultima_rotacion)
    if not fecha_base:
        return {'vencida': False, 'dias_transcurridos': 0, 'dias_limite': rotacion_dias}
    ahora = datetime.now(ZONA_HORARIA_COLOMBIA).replace(tzinfo=None)
    dias_transcurridos = (ahora - fecha_base).days
    return {
        'vencida': dias_transcurridos >= rotacion_dias,
        'dias_transcurridos': dias_transcurridos,
        'dias_limite': rotacion_dias
    }


def _revisar_recordatorios_rotacion():
    """Recordatorio automático (una vez por ciclo) de rotación de contraseñas vencida: avisa por
    campanita/correo a todo el equipo de soporte activo. Se limpia (puede volver a avisar) en
    cuanto se cambia esa contraseña desde editar_credencial. Se llama de forma perezosa cada vez
    que se visita la bóveda o la Auditoría — nunca debe tumbar esa página si algo falla."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, titulo, rotacion_dias, fecha_ultima_rotacion FROM credenciales WHERE COALESCE(estado, 'activo') = 'activo' AND COALESCE(tipo_item, 'credencial') = 'credencial' AND rotacion_dias IS NOT NULL AND rotacion_recordatorio_fecha IS NULL")
        filas = cursor.fetchall()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error revisando recordatorios de rotación: {e}")
        return

    for cred_id, titulo, rotacion_dias, fecha_ultima_rotacion in filas:
        estado = _estado_rotacion_credencial(rotacion_dias, fecha_ultima_rotacion)
        if not estado or not estado['vencida']:
            continue

        equipo = _equipo_soporte_activo()
        url_cred = url_for('ver_credenciales')
        mensaje = f"🔐 Toca rotar la contraseña de '{titulo}' (política de {rotacion_dias} días)"
        asunto = f"[Arkiv] Rotación de contraseña pendiente: '{titulo}'"
        cuerpo = (
            f"La credencial '{titulo}' en la bóveda de Arkiv ya superó su política de rotación "
            f"({rotacion_dias} días sin cambiarse). Por seguridad, cambia esa contraseña y actualízala en Arkiv.\n\n---\nArkiv"
        )
        crear_notificacion_para_varios([m['usuario'] for m in equipo], mensaje, url=url_cred, tipo='credencial')
        for miembro in equipo:
            if miembro['correo']:
                threading.Thread(target=enviar_correo_ticket, args=(miembro['correo'], asunto, cuerpo)).start()

        try:
            conn, db_type = get_db()
            cursor = conn.cursor()
            fecha_act = obtener_fecha_actual()
            q = "UPDATE credenciales SET rotacion_recordatorio_fecha = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE credenciales SET rotacion_recordatorio_fecha = ? WHERE id = ?"
            cursor.execute(q, (fecha_act, cred_id))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"⚠️ Error marcando recordatorio de rotación de la credencial {cred_id}: {e}")


# 🔓 Lista corta de contraseñas extremadamente comunes/genéricas, para el análisis de
# seguridad local de la bóveda (ver _analizar_seguridad_credenciales). Todo en minúsculas.
CONTRASENAS_COMUNES_DEBILES = {
    '123456', '12345678', '123456789', '1234567890', '1234567', '12345',
    'password', 'contraseña', 'contrasena', 'qwerty', 'qwerty123', 'admin',
    'admin123', '111111', '000000', 'abc123', 'iloveyou', 'letmein', 'welcome',
    'monkey', 'dragon', 'password1', 'password123', 'clave123', 'clave1234',
    'usuario', 'root', 'toor', 'test', 'test123', '1q2w3e4r', 'qazwsx',
    'asdfgh', 'zxcvbn', 'cambiar123', 'temporal123',
}
UMBRAL_LONGITUD_CLAVE_DEBIL = 8


def _password_es_debil(password_plano):
    """Heurística local para marcar una contraseña como débil: muy corta o coincide con una de
    las contraseñas más comunes/genéricas conocidas. Deliberadamente simple y sin depender de
    ningún servicio externo (ver _analizar_seguridad_credenciales)."""
    if not password_plano or len(password_plano) < UMBRAL_LONGITUD_CLAVE_DEBIL:
        return True
    return password_plano.lower() in CONTRASENAS_COMUNES_DEBILES


def _analizar_seguridad_credenciales():
    """Revisa TODAS las credenciales activas de la bóveda buscando contraseñas débiles (muy
    cortas o genéricas) o repetidas (la misma clave usada en más de un ítem). Todo el análisis
    ocurre DENTRO del servidor: ninguna contraseña —ni siquiera cifrada o como hash— se envía a
    ningún servicio externo (nada de comprobar contra bases de filtraciones públicas tipo
    'Have I Been Pwned'): son credenciales reales de servidores/bases de datos de producción y
    no es prudente que salga ni un fragmento de ellas de Arkiv. Devuelve solo banderas (débil
    sí/no, repetida sí/no) por credencial — nunca la contraseña real, ni siquiera en este
    reporte interno de auditoría."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        # 📝 Las notas seguras (tipo_item = 'nota_segura') no tienen clave real — se excluyen
        # aquí para no marcarlas como "débil" (su password_cifrada queda vacío) ni agruparlas
        # entre sí como "repetida" solo por compartir ese vacío.
        cursor.execute("SELECT id, titulo, password_cifrada FROM credenciales WHERE COALESCE(estado, 'activo') = 'activo' AND COALESCE(tipo_item, 'credencial') = 'credencial'")
        filas = cursor.fetchall()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error analizando seguridad de la bóveda: {e}")
        return []

    resultado = {}
    por_password = {}
    for cred_id, titulo, pass_enc in filas:
        try:
            pass_plano = desencriptar_texto(pass_enc, cred_id)
        except Exception:
            continue
        resultado[cred_id] = {'id': cred_id, 'titulo': titulo, 'debil': _password_es_debil(pass_plano), 'repetida': False}
        por_password.setdefault(pass_plano, []).append(cred_id)

    for ids_repetidos in por_password.values():
        if len(ids_repetidos) > 1:
            for cred_id in ids_repetidos:
                resultado[cred_id]['repetida'] = True

    problemas = [r for r in resultado.values() if r['debil'] or r['repetida']]
    problemas.sort(key=lambda r: r['titulo'].lower())
    return problemas


def _compartidos_por_credencial():
    """Bóveda Fase 3 — trae de una sola consulta todas las comparticiones puntuales vigentes,
    agrupadas por credencial_id, para no hacer una consulta aparte por cada ítem al armar la
    lista de la bóveda o al validar permisos."""
    mapa = {}
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT credencial_id, usuario FROM credenciales_compartidas")
        for cred_id, usuario in cursor.fetchall():
            mapa.setdefault(cred_id, set()).add(usuario)
        conn.close()
    except Exception as e:
        print(f"⚠️ Error listando comparticiones de la bóveda: {e}")
    return mapa


def _usuarios_compartidos_credencial(cred_id):
    """Como _compartidos_por_credencial pero para UN solo ítem — usado en las rutas que ya
    conocen el ID (revelar, historial) y no necesitan traer la tabla completa."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT usuario FROM credenciales_compartidas WHERE credencial_id = %s" if db_type == 'postgres' else "SELECT usuario FROM credenciales_compartidas WHERE credencial_id = ?"
        cursor.execute(q, (cred_id,))
        usuarios = {r[0] for r in cursor.fetchall()}
        conn.close()
        return usuarios
    except Exception as e:
        print(f"⚠️ Error listando comparticiones de la credencial {cred_id}: {e}")
        return set()


def _puede_ver_credencial_item(propietario, visibilidad, username, rol, usuarios_compartidos=None):
    """Bóveda Fase 3 — un ítem 'equipo' (o sin visibilidad definida, es decir cualquier ítem
    creado antes de esta fase) lo sigue viendo cualquier admin/agente, exactamente como hasta
    ahora. Un ítem 'personal' solo lo ve su dueño, alguien con quien se compartió puntualmente,
    o un admin (que conserva acceso de auditoría sobre toda la bóveda por defecto)."""
    if (visibilidad or 'equipo') != 'personal':
        return True
    if rol == 'admin':
        return True
    if username and username == propietario:
        return True
    if usuarios_compartidos and username in usuarios_compartidos:
        return True
    return False


def _puede_gestionar_credencial_item(propietario, visibilidad, username, rol):
    """Bóveda Fase 3 — editar, eliminar, y administrar las comparticiones de un ítem 'personal'
    queda reservado a su dueño y a un admin: una compartición da acceso de solo consulta, no de
    edición (igual que compartir un ítem en 1Password/Bitwarden). Los ítems 'equipo' siguen
    gestionables por cualquier admin/agente, como siempre."""
    if (visibilidad or 'equipo') != 'personal':
        return True
    return rol == 'admin' or (username and username == propietario)


# 🔑 MÓDULO BÓVEDA DE CREDENCIALES
@app.route('/credenciales')
@login_required
@agente_o_admin_required
def ver_credenciales():
    # 🔐 Aviso perezoso de rotación de contraseñas vencida — ver _revisar_recordatorios_rotacion().
    _revisar_recordatorios_rotacion()

    q_busqueda = request.args.get('q', '').strip().lower()
    q_etiqueta = request.args.get('etiqueta', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()

    try:
        # 🔒 Ya NO se trae ni se descifra password_cifrada (ni contenido_seguro/campos_personalizados,
        # también cifrados) aquí: todo texto sensible se entrega a demanda, vía
        # /credenciales/<id>/revelar, para que la auditoría de accesos registre consultas reales
        # y no simplemente "abrió la página".
        cursor.execute("SELECT id, titulo, url_acceso, usuario_acceso, area, notas, fecha_creacion, rotacion_dias, fecha_ultima_rotacion, etiquetas, tipo_item, contenido_seguro, campos_personalizados, propietario, visibilidad FROM credenciales WHERE COALESCE(estado, 'activo') != 'eliminado' ORDER BY titulo ASC")
        rows = cursor.fetchall()
    except Exception:
        rows = []

    conn.close()

    # 🔐 Bóveda Fase 3 — quién soy y qué comparticiones existen, para filtrar los ítems
    # 'personal' que no me corresponden ANTES de que lleguen siquiera a la plantilla.
    username_actual = session.get('username')
    rol_actual = session.get('rol')
    es_admin_actual = rol_actual == 'admin'
    compartidos_map = _compartidos_por_credencial()

    lista_credenciales = []
    todas_las_etiquetas = set()
    for r in rows:
        try:
            c_id, servicio, url, usuario, categoria, notas, fecha, rotacion_dias, fecha_ultima_rotacion, etiquetas_texto, tipo_item, contenido_seguro, campos_personalizados, propietario, visibilidad = r
            tipo_item = tipo_item or 'credencial'
            visibilidad = visibilidad or 'equipo'
            usuarios_compartidos = compartidos_map.get(c_id, set())

            if not _puede_ver_credencial_item(propietario, visibilidad, username_actual, rol_actual, usuarios_compartidos):
                continue

            etiquetas = _lista_etiquetas(etiquetas_texto)
            todas_las_etiquetas.update(etiquetas)
            texto_full = f"{servicio} {usuario} {categoria} {notas} {etiquetas_texto or ''}".lower()
            if q_etiqueta and q_etiqueta.lower() not in [e.lower() for e in etiquetas]:
                continue
            if not q_busqueda or q_busqueda in texto_full:
                lista_credenciales.append({
                    'id': c_id,
                    'servicio': servicio,
                    'url': url or '',
                    'usuario': usuario,
                    'categoria': categoria or 'General',
                    'notas': notas or '',
                    'fecha': fecha,
                    'rotacion_dias': rotacion_dias,
                    'rotacion': _estado_rotacion_credencial(rotacion_dias, fecha_ultima_rotacion),
                    'etiquetas_texto': etiquetas_texto or '',
                    'etiquetas': etiquetas,
                    'tipo_item': tipo_item,
                    'es_nota': tipo_item == 'nota_segura',
                    'tiene_campos': bool(campos_personalizados),
                    'visibilidad': visibilidad,
                    'es_personal': visibilidad == 'personal',
                    'propietario': propietario or '',
                    'es_propio': bool(propietario) and propietario == username_actual,
                    'compartida_conmigo': visibilidad == 'personal' and propietario != username_actual and username_actual in usuarios_compartidos,
                    'puede_gestionar': _puede_gestionar_credencial_item(propietario, visibilidad, username_actual, rol_actual),
                })
        except Exception as e_row:
            # No dejar que una fila con datos inconsistentes tumbe toda la bóveda.
            print(f"⚠️ Error procesando credencial {r[0] if r else '?'}: {e_row}")
            continue

    # 🔐 Con quién se puede compartir un ítem personal: el resto del equipo de soporte activo.
    equipo_para_compartir = [m['usuario'] for m in _equipo_soporte_activo() if m['usuario'] != username_actual]

    return render_template(
        'credenciales.html', credenciales=lista_credenciales, q_busqueda=q_busqueda,
        q_etiqueta=q_etiqueta, todas_las_etiquetas=sorted(todas_las_etiquetas, key=str.lower),
        equipo_para_compartir=sorted(equipo_para_compartir, key=str.lower), es_admin_actual=es_admin_actual
    )


@app.route('/credenciales/<int:cred_id>/revelar', methods=['POST'])
@login_required
@agente_o_admin_required
def revelar_credencial(cred_id):
    """Descifra el contenido sensible de un ítem de la bóveda a demanda, y deja constancia en el
    log general (con credencial_id) de quién lo consultó/copió y cuándo — la base de la
    Auditoría de Accesos. Según 'accion' revela cosas distintas:
      - 'ver'/'copiar' (default): la clave (credencial normal) o el contenido (nota segura).
      - 'ver_campos': los campos personalizados del ítem (lista nombre/valor).
    Nunca revela más de un tipo de dato sensible por llamada."""
    accion = request.form.get('accion', 'ver')
    accion_legible = 'copió' if accion == 'copiar' else 'vio'
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT titulo, password_cifrada, tipo_item, contenido_seguro, campos_personalizados, propietario, visibilidad FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT titulo, password_cifrada, tipo_item, contenido_seguro, campos_personalizados, propietario, visibilidad FROM credenciales WHERE id = ?"
        cursor.execute(q, (cred_id,))
        row = cursor.fetchone()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error consultando credencial {cred_id} para revelar: {e}")
        return jsonify({'error': 'error interno'}), 500

    if not row:
        return jsonify({'error': 'no encontrada'}), 404

    titulo, pass_enc, tipo_item, contenido_cifrado, campos_cifrados, propietario, visibilidad = row
    tipo_item = tipo_item or 'credencial'

    # 🔐 Bóveda Fase 3 — un ítem 'personal' del que no soy dueño/admin/compartición no se revela,
    # sin importar que se conozca su ID directamente.
    if not _puede_ver_credencial_item(propietario, visibilidad, session.get('username'), session.get('rol'), _usuarios_compartidos_credencial(cred_id)):
        return jsonify({'error': 'no autorizado'}), 403

    if accion == 'ver_campos':
        campos_json = desencriptar_texto(campos_cifrados, cred_id) if campos_cifrados else ''
        registrar_log(session.get('username'), "Consulta de Credencial", f"Consultó los campos personalizados de '{titulo}' (ID {cred_id})", credencial_id=cred_id)
        return jsonify({'campos': _campos_personalizados_desde_json(campos_json)})

    if tipo_item == 'nota_segura':
        contenido_real = desencriptar_texto(contenido_cifrado, cred_id) if contenido_cifrado else ''
        registrar_log(session.get('username'), "Consulta de Credencial", f"Se {accion_legible} el contenido de la nota segura '{titulo}' (ID {cred_id})", credencial_id=cred_id)
        return jsonify({'contenido': contenido_real})

    pass_real = desencriptar_texto(pass_enc, cred_id)
    registrar_log(session.get('username'), "Consulta de Credencial", f"Se {accion_legible} la clave de '{titulo}' (ID {cred_id})", credencial_id=cred_id)
    return jsonify({'password': pass_real})


@app.route('/credenciales/<int:cred_id>/historial')
@login_required
@agente_o_admin_required
def historial_credencial(cred_id):
    """Lista (sin descifrar ninguna clave) las entradas del historial de contraseñas anteriores
    de una credencial — quién la cambió y cuándo. Cada clave puntual solo se descifra a demanda
    vía /credenciales/historial/<id>/revelar (auditado igual que revelar_credencial)."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_cred = "SELECT propietario, visibilidad FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT propietario, visibilidad FROM credenciales WHERE id = ?"
        cursor.execute(q_cred, (cred_id,))
        fila_cred = cursor.fetchone()
        if not fila_cred:
            conn.close()
            return jsonify({'error': 'no encontrada'}), 404
        propietario, visibilidad = fila_cred
        if not _puede_ver_credencial_item(propietario, visibilidad, session.get('username'), session.get('rol'), _usuarios_compartidos_credencial(cred_id)):
            conn.close()
            return jsonify({'error': 'no autorizado'}), 403

        q = "SELECT id, fecha_cambio, cambiado_por FROM credenciales_historial WHERE credencial_id = %s ORDER BY id DESC" if db_type == 'postgres' else "SELECT id, fecha_cambio, cambiado_por FROM credenciales_historial WHERE credencial_id = ? ORDER BY id DESC"
        cursor.execute(q, (cred_id,))
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando historial de credencial {cred_id}: {e}")
        filas = []
    conn.close()
    return jsonify({'historial': [{'id': h_id, 'fecha_cambio': fecha, 'cambiado_por': quien or '—'} for h_id, fecha, quien in filas]})


@app.route('/credenciales/historial/<int:historial_id>/revelar', methods=['POST'])
@login_required
@agente_o_admin_required
def revelar_historial_credencial(historial_id):
    """Descifra puntualmente UNA clave anterior del historial (nunca la lista completa de una
    vez), dejando el mismo rastro de auditoría que revelar_credencial."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT h.credencial_id, h.password_cifrada, h.fecha_cambio, c.titulo, c.propietario, c.visibilidad FROM credenciales_historial h JOIN credenciales c ON c.id = h.credencial_id WHERE h.id = %s" if db_type == 'postgres' else "SELECT h.credencial_id, h.password_cifrada, h.fecha_cambio, c.titulo, c.propietario, c.visibilidad FROM credenciales_historial h JOIN credenciales c ON c.id = h.credencial_id WHERE h.id = ?"
        cursor.execute(q, (historial_id,))
        row = cursor.fetchone()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error consultando historial {historial_id}: {e}")
        return jsonify({'error': 'error interno'}), 500

    if not row:
        return jsonify({'error': 'no encontrado'}), 404

    credencial_id, pass_enc, fecha_cambio, titulo, propietario, visibilidad = row
    if not _puede_ver_credencial_item(propietario, visibilidad, session.get('username'), session.get('rol'), _usuarios_compartidos_credencial(credencial_id)):
        return jsonify({'error': 'no autorizado'}), 403

    pass_real = desencriptar_texto(pass_enc)
    registrar_log(session.get('username'), "Consulta de Credencial", f"Consultó una clave anterior (del {fecha_cambio}) del historial de '{titulo}' (ID {credencial_id})", credencial_id=credencial_id)
    return jsonify({'password': pass_real})


@app.route('/credenciales/auditoria')
@login_required
@agente_o_admin_required
def auditoria_credenciales():
    """Panel de auditoría de la bóveda: estado de rotación de cada credencial activa (según su
    política opcional) y el historial reciente de consultas/copias de claves — quién, cuál
    credencial y cuándo."""
    conn, db_type = get_db()
    cursor = conn.cursor()

    try:
        # 📝 Las notas seguras no tienen clave que rotar, así que no aparecen en este panel.
        cursor.execute("SELECT id, titulo, area, rotacion_dias, fecha_ultima_rotacion FROM credenciales WHERE COALESCE(estado, 'activo') = 'activo' AND COALESCE(tipo_item, 'credencial') = 'credencial' ORDER BY titulo ASC")
        filas_cred = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando credenciales para auditoría: {e}")
        filas_cred = []

    rotacion = []
    for c_id, titulo, area, rotacion_dias, fecha_ultima_rotacion in filas_cred:
        rotacion.append({
            'id': c_id, 'titulo': titulo, 'area': area or 'General',
            'rotacion_dias': rotacion_dias,
            'estado': _estado_rotacion_credencial(rotacion_dias, fecha_ultima_rotacion)
        })
    # Vencidas primero, luego con política vigente, luego sin política configurada.
    rotacion.sort(key=lambda c: (0 if (c['estado'] and c['estado']['vencida']) else (1 if c['estado'] else 2), c['titulo']))

    try:
        cursor.execute("SELECT usuario, detalles, fecha, credencial_id FROM logs WHERE credencial_id IS NOT NULL ORDER BY id DESC LIMIT 200")
        filas_log = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando accesos a credenciales: {e}")
        filas_log = []
    conn.close()

    accesos = [{'usuario': u, 'detalles': d, 'fecha': f, 'credencial_id': cid} for u, d, f, cid in filas_log]

    seguridad = _analizar_seguridad_credenciales()

    return render_template('credenciales_auditoria.html', rotacion=rotacion, accesos=accesos, seguridad=seguridad)

@app.route('/credenciales/crear', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_credencial():
    servicio = request.form.get('servicio', '').strip()
    url = request.form.get('url', '').strip()
    usuario = request.form.get('usuario', '').strip()
    password = request.form.get('password', '').strip()
    categoria = request.form.get('categoria', 'General').strip()
    notas = request.form.get('notas', '').strip()
    rotacion_dias = _parsear_dias_rotacion(request.form.get('rotacion_dias', ''))
    etiquetas = _normalizar_etiquetas(request.form.get('etiquetas', ''))

    # 📝 Bóveda Fase 2 — tipo_item distingue una credencial normal de una nota segura (sin
    # usuario/clave reales: se guarda un placeholder en usuario_acceso y password_cifrada queda
    # vacío, ya que esas columnas siguen siendo NOT NULL). campos_personalizados aplica a
    # cualquiera de los dos tipos.
    tipo_item = request.form.get('tipo_item', 'credencial').strip()
    if tipo_item not in ('credencial', 'nota_segura'):
        tipo_item = 'credencial'
    es_nota = tipo_item == 'nota_segura'
    contenido_seguro = request.form.get('contenido_seguro', '').strip()
    campos_json = _serializar_campos_personalizados(request.form.getlist('campo_nombre[]'), request.form.getlist('campo_valor[]'))

    usuario_final = usuario or ('—' if es_nota else '')
    password_final = '' if es_nota else password

    # 🔐 Bóveda Fase 3 — 'equipo' (por defecto, el comportamiento de siempre: lo ve todo
    # admin/agente) o 'personal' (solo yo, quien admin, y a quien comparta puntualmente).
    visibilidad = request.form.get('visibilidad', 'equipo').strip()
    if visibilidad not in ('equipo', 'personal'):
        visibilidad = 'equipo'
    propietario = session.get('username')

    if servicio and (es_nota or (usuario_final and password_final)):
        try:
            pass_cifrada = encriptar_texto(password_final)
            contenido_cifrado = encriptar_texto(contenido_seguro)
            campos_cifrados = encriptar_texto(campos_json)
            fecha_act = obtener_fecha_actual()

            conn, db_type = get_db()
            cursor = conn.cursor()
            # 🔁 fecha_ultima_rotacion arranca igual a fecha_creacion: recién guardada, la
            # clave "acaba de rotarse" para efectos del recordatorio de rotación.
            q_ins = "INSERT INTO credenciales (titulo, url_acceso, usuario_acceso, password_cifrada, area, notas, fecha_creacion, estado, rotacion_dias, fecha_ultima_rotacion, etiquetas, tipo_item, contenido_seguro, campos_personalizados, propietario, visibilidad) VALUES (%s, %s, %s, %s, %s, %s, %s, 'activo', %s, %s, %s, %s, %s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO credenciales (titulo, url_acceso, usuario_acceso, password_cifrada, area, notas, fecha_creacion, estado, rotacion_dias, fecha_ultima_rotacion, etiquetas, tipo_item, contenido_seguro, campos_personalizados, propietario, visibilidad) VALUES (?, ?, ?, ?, ?, ?, ?, 'activo', ?, ?, ?, ?, ?, ?, ?, ?)"
            cursor.execute(q_ins, (servicio, url, usuario_final, pass_cifrada, categoria, notas, fecha_act, rotacion_dias, fecha_act, etiquetas, tipo_item, contenido_cifrado, campos_cifrados, propietario, visibilidad))
            conn.commit()
            conn.close()

            if es_nota:
                registrar_log(session['username'], "Guardado de Credencial", f"Se registró la nota segura '{servicio}'")
            else:
                registrar_log(session['username'], "Guardado de Credencial", f"Se registró el acceso para el aplicativo '{servicio}'")
        except Exception as e:
            print(f"⚠️ Error guardando credencial '{servicio}': {e}")

    return redirect(url_for('ver_credenciales'))

@app.route('/credenciales/editar/<int:cred_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_credencial(cred_id):
    servicio = request.form.get('servicio', '').strip()
    url = request.form.get('url', '').strip()
    usuario = request.form.get('usuario', '').strip()
    password = request.form.get('password', '').strip()
    categoria = request.form.get('categoria', 'General').strip()
    notas = request.form.get('notas', '').strip()
    rotacion_dias = _parsear_dias_rotacion(request.form.get('rotacion_dias', ''))
    etiquetas = _normalizar_etiquetas(request.form.get('etiquetas', ''))

    tipo_item = request.form.get('tipo_item', 'credencial').strip()
    if tipo_item not in ('credencial', 'nota_segura'):
        tipo_item = 'credencial'
    es_nota = tipo_item == 'nota_segura'
    contenido_seguro = request.form.get('contenido_seguro', '').strip()
    contenido_cifrado = encriptar_texto(contenido_seguro)
    campos_json = _serializar_campos_personalizados(request.form.getlist('campo_nombre[]'), request.form.getlist('campo_valor[]'))
    campos_cifrados = encriptar_texto(campos_json)
    if es_nota and not usuario:
        usuario = '—'

    visibilidad = request.form.get('visibilidad', 'equipo').strip()
    if visibilidad not in ('equipo', 'personal'):
        visibilidad = 'equipo'

    conn, db_type = get_db()
    cursor = conn.cursor()

    try:
        # 🔐 Bóveda Fase 3 — solo el dueño de un ítem 'personal' (o un admin) puede editarlo;
        # una compartición solo da acceso de consulta, no de edición.
        q_check = "SELECT propietario, visibilidad FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT propietario, visibilidad FROM credenciales WHERE id = ?"
        cursor.execute(q_check, (cred_id,))
        fila_actual = cursor.fetchone()
        if not fila_actual:
            conn.close()
            return redirect(url_for('ver_credenciales'))
        propietario_actual, visibilidad_actual = fila_actual
        if not _puede_gestionar_credencial_item(propietario_actual, visibilidad_actual, session.get('username'), session.get('rol')):
            conn.close()
            return jsonify({'error': 'no autorizado para editar este ítem personal'}), 403

        # 📜 Bóveda Fase 2 — si de verdad se está cambiando la clave de un ítem que YA era una
        # credencial (nunca para notas seguras, que no tienen clave real), la clave vieja se
        # guarda en el historial ANTES de sobrescribirla (ver historial_credencial /
        # revelar_historial_credencial).
        if password and not es_nota:
            q_old = "SELECT password_cifrada, COALESCE(tipo_item, 'credencial') FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT password_cifrada, COALESCE(tipo_item, 'credencial') FROM credenciales WHERE id = ?"
            cursor.execute(q_old, (cred_id,))
            fila_vieja = cursor.fetchone()
            if fila_vieja and fila_vieja[0] and fila_vieja[1] == 'credencial':
                q_hist = "INSERT INTO credenciales_historial (credencial_id, password_cifrada, fecha_cambio, cambiado_por) VALUES (%s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO credenciales_historial (credencial_id, password_cifrada, fecha_cambio, cambiado_por) VALUES (?, ?, ?, ?)"
                cursor.execute(q_hist, (cred_id, fila_vieja[0], obtener_fecha_actual(), session.get('username')))

        if password:
            # 🔁 Se cambió la clave de verdad: cuenta como una rotación — se actualiza
            # fecha_ultima_rotacion a ahora y se limpia el recordatorio ya enviado, para que
            # el próximo aviso de "toca rotar" se calcule desde este momento en adelante.
            pass_cifrada = encriptar_texto('' if es_nota else password)
            fecha_act = obtener_fecha_actual()
            q_upd = "UPDATE credenciales SET titulo=%s, url_acceso=%s, usuario_acceso=%s, password_cifrada=%s, area=%s, notas=%s, rotacion_dias=%s, fecha_ultima_rotacion=%s, rotacion_recordatorio_fecha=NULL, etiquetas=%s, tipo_item=%s, contenido_seguro=%s, campos_personalizados=%s, visibilidad=%s WHERE id=%s" if db_type == 'postgres' else "UPDATE credenciales SET titulo=?, url_acceso=?, usuario_acceso=?, password_cifrada=?, area=?, notas=?, rotacion_dias=?, fecha_ultima_rotacion=?, rotacion_recordatorio_fecha=NULL, etiquetas=?, tipo_item=?, contenido_seguro=?, campos_personalizados=?, visibilidad=? WHERE id=?"
            cursor.execute(q_upd, (servicio, url, usuario, pass_cifrada, categoria, notas, rotacion_dias, fecha_act, etiquetas, tipo_item, contenido_cifrado, campos_cifrados, visibilidad, cred_id))
        else:
            q_upd = "UPDATE credenciales SET titulo=%s, url_acceso=%s, usuario_acceso=%s, area=%s, notas=%s, rotacion_dias=%s, etiquetas=%s, tipo_item=%s, contenido_seguro=%s, campos_personalizados=%s, visibilidad=%s WHERE id=%s" if db_type == 'postgres' else "UPDATE credenciales SET titulo=?, url_acceso=?, usuario_acceso=?, area=?, notas=?, rotacion_dias=?, etiquetas=?, tipo_item=?, contenido_seguro=?, campos_personalizados=?, visibilidad=? WHERE id=?"
            cursor.execute(q_upd, (servicio, url, usuario, categoria, notas, rotacion_dias, etiquetas, tipo_item, contenido_cifrado, campos_cifrados, visibilidad, cred_id))

        conn.commit()
        registrar_log(session['username'], "Edición de Credencial", f"Se actualizó la credencial ID '{cred_id}' ({servicio})")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error editando credencial {cred_id}: {e}")

    conn.close()
    return redirect(url_for('ver_credenciales'))

@app.route('/credenciales/eliminar/<int:cred_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_credencial(cred_id):
    conn, db_type = get_db()
    cursor = conn.cursor()

    # 🔐 Bóveda Fase 3 — igual que editar: solo el dueño de un ítem 'personal' o un admin
    # pueden mandarlo a la papelera; una compartición no alcanza para eso.
    q_check = "SELECT propietario, visibilidad FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT propietario, visibilidad FROM credenciales WHERE id = ?"
    cursor.execute(q_check, (cred_id,))
    fila_actual = cursor.fetchone()
    if not fila_actual:
        conn.close()
        return redirect(url_for('ver_credenciales'))
    if not _puede_gestionar_credencial_item(fila_actual[0], fila_actual[1], session.get('username'), session.get('rol')):
        conn.close()
        return jsonify({'error': 'no autorizado para eliminar este ítem personal'}), 403

    q_upd = "UPDATE credenciales SET estado = 'eliminado' WHERE id = %s" if db_type == 'postgres' else "UPDATE credenciales SET estado = 'eliminado' WHERE id = ?"
    cursor.execute(q_upd, (cred_id,))
    conn.commit()
    conn.close()

    registrar_log(session['username'], "Eliminación de Credencial", f"Se envió a la papelera la credencial ID '{cred_id}'")
    return redirect(url_for('ver_credenciales'))


@app.route('/credenciales/<int:cred_id>/compartidos')
@login_required
@agente_o_admin_required
def compartidos_credencial(cred_id):
    """Lista con quién está compartido puntualmente un ítem personal — visible para cualquiera
    que ya pueda ver el ítem (dueño, compartidos, y admin), para que quede claro quién más
    tiene acceso."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    q_cred = "SELECT propietario, visibilidad FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT propietario, visibilidad FROM credenciales WHERE id = ?"
    cursor.execute(q_cred, (cred_id,))
    fila_cred = cursor.fetchone()
    if not fila_cred:
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404
    propietario, visibilidad = fila_cred

    q = "SELECT usuario, fecha_compartido FROM credenciales_compartidas WHERE credencial_id = %s ORDER BY fecha_compartido DESC" if db_type == 'postgres' else "SELECT usuario, fecha_compartido FROM credenciales_compartidas WHERE credencial_id = ? ORDER BY fecha_compartido DESC"
    cursor.execute(q, (cred_id,))
    filas = cursor.fetchall()
    conn.close()

    usuarios_compartidos = {u for u, _ in filas}
    if not _puede_ver_credencial_item(propietario, visibilidad, session.get('username'), session.get('rol'), usuarios_compartidos):
        return jsonify({'error': 'no autorizado'}), 403

    return jsonify({
        'propietario': propietario or '',
        'puede_gestionar': _puede_gestionar_credencial_item(propietario, visibilidad, session.get('username'), session.get('rol')),
        'compartidos': [{'usuario': u, 'fecha': f} for u, f in filas],
    })


@app.route('/credenciales/<int:cred_id>/compartir', methods=['POST'])
@login_required
@agente_o_admin_required
def compartir_credencial(cred_id):
    """Comparte puntualmente un ítem 'personal' con otro miembro activo del equipo de soporte
    (admin/agente) — solo el dueño del ítem o un admin pueden hacerlo. Un ítem 'equipo' no
    necesita esto: ya lo ve todo el mundo."""
    usuario_destino = (request.form.get('usuario') or '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()
    q_cred = "SELECT titulo, propietario, visibilidad FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT titulo, propietario, visibilidad FROM credenciales WHERE id = ?"
    cursor.execute(q_cred, (cred_id,))
    fila_cred = cursor.fetchone()
    if not fila_cred:
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404
    titulo, propietario, visibilidad = fila_cred

    if not _puede_gestionar_credencial_item(propietario, visibilidad, session.get('username'), session.get('rol')):
        conn.close()
        return jsonify({'error': 'no autorizado'}), 403
    if (visibilidad or 'equipo') != 'personal':
        conn.close()
        return jsonify({'error': "solo los ítems personales se pueden compartir puntualmente"}), 400

    # 🛡️ Solo se puede compartir con una cuenta activa del equipo de soporte (admin/agente) —
    # las mismas que ya tienen acceso a la Bóveda; no tendría sentido compartir con nadie más.
    equipo_valido = {m['usuario'] for m in _equipo_soporte_activo()}
    if not usuario_destino or usuario_destino not in equipo_valido or usuario_destino == propietario:
        conn.close()
        return jsonify({'error': 'usuario inválido'}), 400

    try:
        q_check = "SELECT id FROM credenciales_compartidas WHERE credencial_id = %s AND usuario = %s" if db_type == 'postgres' else "SELECT id FROM credenciales_compartidas WHERE credencial_id = ? AND usuario = ?"
        cursor.execute(q_check, (cred_id, usuario_destino))
        if not cursor.fetchone():
            q_ins = "INSERT INTO credenciales_compartidas (credencial_id, usuario, fecha_compartido, compartido_por) VALUES (%s, %s, %s, %s)" if db_type == 'postgres' else "INSERT INTO credenciales_compartidas (credencial_id, usuario, fecha_compartido, compartido_por) VALUES (?, ?, ?, ?)"
            cursor.execute(q_ins, (cred_id, usuario_destino, obtener_fecha_actual(), session.get('username')))
            conn.commit()
            registrar_log(session['username'], "Edición de Credencial", f"Compartió '{titulo}' (ID {cred_id}) con {usuario_destino}", credencial_id=cred_id)
        conn.close()
    except Exception as e:
        conn.close()
        print(f"⚠️ Error compartiendo credencial {cred_id}: {e}")
        return jsonify({'error': 'error interno'}), 500

    return jsonify({'ok': True})


@app.route('/credenciales/<int:cred_id>/descompartir', methods=['POST'])
@login_required
@agente_o_admin_required
def descompartir_credencial(cred_id):
    """Quita la compartición puntual de un ítem personal con un usuario dado — solo el dueño o
    un admin pueden hacerlo."""
    usuario_destino = (request.form.get('usuario') or '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()
    q_cred = "SELECT titulo, propietario, visibilidad FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT titulo, propietario, visibilidad FROM credenciales WHERE id = ?"
    cursor.execute(q_cred, (cred_id,))
    fila_cred = cursor.fetchone()
    if not fila_cred:
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404
    titulo, propietario, visibilidad = fila_cred

    if not _puede_gestionar_credencial_item(propietario, visibilidad, session.get('username'), session.get('rol')):
        conn.close()
        return jsonify({'error': 'no autorizado'}), 403

    try:
        q_del = "DELETE FROM credenciales_compartidas WHERE credencial_id = %s AND usuario = %s" if db_type == 'postgres' else "DELETE FROM credenciales_compartidas WHERE credencial_id = ? AND usuario = ?"
        cursor.execute(q_del, (cred_id, usuario_destino))
        conn.commit()
        conn.close()
        registrar_log(session['username'], "Edición de Credencial", f"Quitó la compartición de '{titulo}' (ID {cred_id}) con {usuario_destino}", credencial_id=cred_id)
    except Exception as e:
        conn.close()
        print(f"⚠️ Error quitando compartición de credencial {cred_id}: {e}")
        return jsonify({'error': 'error interno'}), 500

    return jsonify({'ok': True})


# 🪪 MÓDULO ALTAS Y BAJAS DE CREDENCIALES DE COLABORADORES — bitácora de qué credenciales se
# le crean a cada colaborador nuevo (o existente) en cada aplicativo institucional (KUBAPP,
# SAMI, Moodle, Wolkvox, Correo, Solvyx...), quién las gestionó/capacitó y por qué medio se
# entregaron. Es un registro por aplicativo (no por colaborador), para poder dar de baja el
# acceso a uno puntual sin afectar los demás que tenga esa misma persona. Distinto de la
# Bóveda de Accesos (que guarda credenciales de sistemas/servidores para el equipo de soporte).
MEDIOS_ENVIO_CREDENCIAL = ('WhatsApp', 'Correo', 'SMS', 'Chat Teams')


def _catalogo_aplicativos_activos():
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nombre FROM aplicativos_catalogo WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY nombre ASC")
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando catálogo de aplicativos: {e}")
        filas = []
    conn.close()
    return [{'id': f[0], 'nombre': f[1]} for f in filas]


def _catalogo_especialidades_activas():
    """Especialidades/áreas disponibles para asignar a un usuario (Medicina General, Enfermería,
    Administrativo...). Administrable desde /usuarios sin tocar código — ver crear_especialidad_catalogo
    y eliminar_especialidad_catalogo."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nombre FROM especialidades_catalogo WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY nombre ASC")
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando catálogo de especialidades: {e}")
        filas = []
    conn.close()
    return [{'id': f[0], 'nombre': f[1]} for f in filas]


def _catalogo_tipos_activo_activos():
    """Tipos de activo disponibles para Inventario (Portátil, Impresora, Servidor...),
    administrables desde /tickets/inventario/tipos sin tocar código — ver
    crear_tipo_activo_catalogo, editar_tipo_activo_catalogo, reordenar_tipo_activo_catalogo y
    eliminar_tipo_activo_catalogo. Antes era la lista fija TIPOS_ACTIVO."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, key, etiqueta, icono, orden FROM tipos_activo_catalogo WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY orden ASC, etiqueta ASC")
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando catálogo de tipos de activo: {e}")
        filas = []
    conn.close()
    return [{'id': f[0], 'key': f[1], 'etiqueta': f[2], 'icono': f[3] or 'box', 'orden': f[4] or 0} for f in filas]


def _catalogo_tipos_activo_todos():
    """Igual que _catalogo_tipos_activo_activos() pero incluye los inactivos — para el modal de
    administración, donde también se pueden reactivar."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, key, etiqueta, icono, orden, estado FROM tipos_activo_catalogo ORDER BY orden ASC, etiqueta ASC")
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando catálogo de tipos de activo: {e}")
        filas = []
    conn.close()
    return [{'id': f[0], 'key': f[1], 'etiqueta': f[2], 'icono': f[3] or 'box', 'orden': f[4] or 0, 'estado': f[5] or 'activo'} for f in filas]


@app.route('/credenciales/colaboradores')
@login_required
@agente_o_admin_required
def ver_credenciales_colaboradores():
    q_busqueda = request.args.get('q', '').strip().lower()
    f_aplicativo = request.args.get('aplicativo', '').strip()
    f_estado = request.args.get('estado', 'activo').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT id, colaborador, aplicativo, fecha_creacion, fecha_solicitud, analista_gestiona, "
            "solicitado_por, capacitado_por, medio_envio, estado, fecha_deshabilitacion, deshabilitado_por "
            "FROM credenciales_colaboradores ORDER BY id DESC"
        )
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error listando credenciales de colaboradores: {e}")
        filas = []
    conn.close()

    registros = []
    aplicativos_en_uso = set()
    for r in filas:
        (r_id, colaborador, aplicativo, fecha_creacion, fecha_solicitud, analista_gestiona,
         solicitado_por, capacitado_por, medio_envio, estado, fecha_deshabilitacion, deshabilitado_por) = r
        aplicativos_en_uso.add(aplicativo)

        if f_estado and f_estado != 'todos' and estado != f_estado:
            continue
        if f_aplicativo and aplicativo != f_aplicativo:
            continue
        if q_busqueda and q_busqueda not in f"{colaborador} {aplicativo} {solicitado_por or ''}".lower():
            continue

        registros.append({
            'id': r_id, 'colaborador': colaborador, 'aplicativo': aplicativo,
            'fecha_creacion': fecha_creacion, 'fecha_solicitud': fecha_solicitud,
            'analista_gestiona': analista_gestiona, 'solicitado_por': solicitado_por,
            'capacitado_por': capacitado_por, 'medio_envio': medio_envio, 'estado': estado,
            'fecha_deshabilitacion': fecha_deshabilitacion, 'deshabilitado_por': deshabilitado_por
        })

    conn2, db_type2 = get_db()
    cursor2 = conn2.cursor()
    try:
        cursor2.execute(
            "SELECT usuario, nombre FROM usuarios WHERE COALESCE(estado, 'activo') = 'activo' "
            "AND rol IN ('admin', 'agente') ORDER BY nombre ASC"
        )
        equipo_soporte = [{'usuario': r[0], 'nombre': r[1] or r[0]} for r in cursor2.fetchall()]
    except Exception as e:
        print(f"⚠️ Error listando equipo de soporte para altas de credenciales: {e}")
        equipo_soporte = []
    conn2.close()

    conn3, db_type3 = get_db()
    cursor3 = conn3.cursor()
    try:
        # 📋 Nombres de colaboradores ya registrados antes (en cualquier aplicativo), para que
        # el campo "Colaborador" del modal sugiera autocompletar en vez de obligar a retipear el
        # nombre completo cada vez que a la misma persona se le da de alta un acceso más. Sigue
        # siendo texto libre: si la persona es nueva, se puede escribir un nombre que no esté aquí.
        cursor3.execute("SELECT DISTINCT colaborador FROM credenciales_colaboradores ORDER BY colaborador ASC")
        colaboradores_existentes = [r[0] for r in cursor3.fetchall() if r[0]]
    except Exception as e:
        print(f"⚠️ Error listando colaboradores existentes para autocompletar: {e}")
        colaboradores_existentes = []
    conn3.close()

    return render_template(
        'credenciales_colaboradores.html', registros=registros,
        aplicativos=_catalogo_aplicativos_activos(), medios=MEDIOS_ENVIO_CREDENCIAL,
        q_busqueda=q_busqueda, f_aplicativo=f_aplicativo, f_estado=f_estado,
        equipo_soporte=equipo_soporte, colaboradores_existentes=colaboradores_existentes
    )


@app.route('/credenciales/colaboradores/crear', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_credencial_colaborador():
    """Da de alta una credencial por cada aplicativo seleccionado, para el mismo colaborador y
    con los mismos datos compartidos (contraseña, fechas, analista, capacitado por, medio de
    envío) — antes había que repetir todo el formulario un aplicativo a la vez. Sigue guardando
    UN registro por (colaborador, aplicativo) en credenciales_colaboradores (no se cambió ese
    esquema): esta ruta simplemente inserta varias filas en una sola transacción."""
    colaborador = request.form.get('colaborador', '').strip()
    aplicativos_sel = [a.strip() for a in request.form.getlist('aplicativos') if a.strip()]
    password = request.form.get('password', '').strip()
    fecha_creacion = request.form.get('fecha_creacion', '').strip() or None
    fecha_solicitud = request.form.get('fecha_solicitud', '').strip() or None
    analista_gestiona = request.form.get('analista_gestiona', '').strip() or None
    solicitado_por = request.form.get('solicitado_por', '').strip() or None
    capacitado_por = request.form.get('capacitado_por', '').strip() or None
    medio_envio = request.form.get('medio_envio', '').strip() or None
    if medio_envio not in MEDIOS_ENVIO_CREDENCIAL:
        medio_envio = None

    if colaborador and aplicativos_sel and password:
        try:
            pass_cifrada = encriptar_texto(password)
            fecha_act = obtener_fecha_actual()
            conn, db_type = get_db()
            cursor = conn.cursor()
            q_ins = (
                "INSERT INTO credenciales_colaboradores (colaborador, aplicativo, password_cifrada, fecha_creacion, "
                "fecha_solicitud, analista_gestiona, solicitado_por, capacitado_por, medio_envio, estado, "
                "fecha_registro, registrado_por) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'activo', %s, %s)"
                if db_type == 'postgres' else
                "INSERT INTO credenciales_colaboradores (colaborador, aplicativo, password_cifrada, fecha_creacion, "
                "fecha_solicitud, analista_gestiona, solicitado_por, capacitado_por, medio_envio, estado, "
                "fecha_registro, registrado_por) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'activo', ?, ?)"
            )
            for aplicativo in aplicativos_sel:
                cursor.execute(q_ins, (colaborador, aplicativo, pass_cifrada, fecha_creacion, fecha_solicitud,
                                        analista_gestiona, solicitado_por, capacitado_por, medio_envio,
                                        fecha_act, session.get('username')))
            conn.commit()
            conn.close()
            registrar_log(session.get('username'), "Alta de Credencial de Colaborador",
                          f"Se creó credencial de {', '.join(aplicativos_sel)} para {colaborador}")
        except Exception as e:
            print(f"⚠️ Error creando credencial de colaborador '{colaborador}'/{aplicativos_sel}: {e}")

    return redirect(url_for('ver_credenciales_colaboradores'))


@app.route('/credenciales/colaboradores/<int:reg_id>/editar', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_credencial_colaborador(reg_id):
    """Edita los datos de una fila existente de Altas de Credenciales (colaborador, aplicativo,
    fechas, analista, solicitante, capacitador, medio de envío). La contraseña es opcional: si el
    campo llega vacío se deja la clave actual sin tocar — mismo criterio que 'Editar Credencial'
    en la Bóveda de Accesos, para no obligar a reescribir una clave que no cambió."""
    colaborador = request.form.get('colaborador', '').strip()
    aplicativo = request.form.get('aplicativo', '').strip()
    password = request.form.get('password', '').strip()
    fecha_creacion = request.form.get('fecha_creacion', '').strip() or None
    fecha_solicitud = request.form.get('fecha_solicitud', '').strip() or None
    analista_gestiona = request.form.get('analista_gestiona', '').strip() or None
    solicitado_por = request.form.get('solicitado_por', '').strip() or None
    capacitado_por = request.form.get('capacitado_por', '').strip() or None
    medio_envio = request.form.get('medio_envio', '').strip() or None
    if medio_envio not in MEDIOS_ENVIO_CREDENCIAL:
        medio_envio = None

    if colaborador and aplicativo:
        try:
            conn, db_type = get_db()
            cursor = conn.cursor()
            campos_comunes = "colaborador = %s, aplicativo = %s, fecha_creacion = %s, fecha_solicitud = %s, analista_gestiona = %s, solicitado_por = %s, capacitado_por = %s, medio_envio = %s"
            valores_comunes = [colaborador, aplicativo, fecha_creacion, fecha_solicitud, analista_gestiona, solicitado_por, capacitado_por, medio_envio]
            if password:
                q = f"UPDATE credenciales_colaboradores SET {campos_comunes}, password_cifrada = %s WHERE id = %s"
                valores = valores_comunes + [encriptar_texto(password), reg_id]
            else:
                q = f"UPDATE credenciales_colaboradores SET {campos_comunes} WHERE id = %s"
                valores = valores_comunes + [reg_id]
            if db_type != 'postgres':
                q = q.replace('%s', '?')
            cursor.execute(q, valores)
            conn.commit()
            conn.close()
            registrar_log(session.get('username'), "Edición de Credencial de Colaborador",
                          f"Se editó la credencial de '{aplicativo}' de {colaborador} (ID {reg_id})")
        except Exception as e:
            print(f"⚠️ Error editando credencial de colaborador {reg_id}: {e}")

    return redirect(url_for('ver_credenciales_colaboradores'))


@app.route('/credenciales/colaboradores/<int:reg_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_credencial_colaborador(reg_id):
    """Elimina PERMANENTEMENTE una fila de Altas de Credenciales — a diferencia de
    'Deshabilitar' (que solo marca el acceso como inactivo y lo mantiene en el historial), esto
    borra el registro de la base de datos sin posibilidad de deshacerlo. El front pide
    confirmación antes de enviar el formulario."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q_sel = "SELECT colaborador, aplicativo FROM credenciales_colaboradores WHERE id = %s" if db_type == 'postgres' else "SELECT colaborador, aplicativo FROM credenciales_colaboradores WHERE id = ?"
        cursor.execute(q_sel, (reg_id,))
        row = cursor.fetchone()
        colaborador = row[0] if row else f"ID {reg_id}"
        aplicativo = row[1] if row else ""

        q_del = "DELETE FROM credenciales_colaboradores WHERE id = %s" if db_type == 'postgres' else "DELETE FROM credenciales_colaboradores WHERE id = ?"
        cursor.execute(q_del, (reg_id,))
        conn.commit()
        conn.close()
        registrar_log(session.get('username'), "Eliminación de Credencial de Colaborador",
                      f"Se eliminó permanentemente el acceso a '{aplicativo}' de {colaborador} (ID {reg_id})")
    except Exception as e:
        print(f"⚠️ Error eliminando credencial de colaborador {reg_id}: {e}")

    return redirect(url_for('ver_credenciales_colaboradores'))


@app.route('/credenciales/colaboradores/<int:reg_id>/deshabilitar', methods=['POST'])
@login_required
@agente_o_admin_required
def deshabilitar_credencial_colaborador(reg_id):
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        fecha_act = obtener_fecha_actual()
        q = (
            "UPDATE credenciales_colaboradores SET estado = 'deshabilitado', fecha_deshabilitacion = %s, deshabilitado_por = %s WHERE id = %s"
            if db_type == 'postgres' else
            "UPDATE credenciales_colaboradores SET estado = 'deshabilitado', fecha_deshabilitacion = ?, deshabilitado_por = ? WHERE id = ?"
        )
        cursor.execute(q, (fecha_act, session.get('username'), reg_id))
        conn.commit()
        conn.close()
        registrar_log(session.get('username'), "Baja de Credencial de Colaborador", f"Se deshabilitó la credencial ID {reg_id}")
    except Exception as e:
        print(f"⚠️ Error deshabilitando credencial de colaborador {reg_id}: {e}")
    return redirect(url_for('ver_credenciales_colaboradores'))


@app.route('/credenciales/colaboradores/<int:reg_id>/reactivar', methods=['POST'])
@login_required
@agente_o_admin_required
def reactivar_credencial_colaborador(reg_id):
    """Vuelve a 'activo' una fila de Altas de Credenciales que había quedado 'deshabilitado' —
    flujo simétrico al de deshabilitar. Limpia fecha_deshabilitacion/deshabilitado_por para que
    la fila quede exactamente como si nunca se hubiera dado de baja."""
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q_sel = "SELECT colaborador, aplicativo FROM credenciales_colaboradores WHERE id = %s" if db_type == 'postgres' else "SELECT colaborador, aplicativo FROM credenciales_colaboradores WHERE id = ?"
        cursor.execute(q_sel, (reg_id,))
        row = cursor.fetchone()
        colaborador = row[0] if row else f"ID {reg_id}"
        aplicativo = row[1] if row else ""

        q_upd = (
            "UPDATE credenciales_colaboradores SET estado = 'activo', fecha_deshabilitacion = NULL, deshabilitado_por = NULL WHERE id = %s"
            if db_type == 'postgres' else
            "UPDATE credenciales_colaboradores SET estado = 'activo', fecha_deshabilitacion = NULL, deshabilitado_por = NULL WHERE id = ?"
        )
        cursor.execute(q_upd, (reg_id,))
        conn.commit()
        conn.close()
        registrar_log(session.get('username'), "Reactivación de Credencial de Colaborador", f"Se reactivó el acceso a '{aplicativo}' de {colaborador} (ID {reg_id})")
    except Exception as e:
        print(f"⚠️ Error reactivando credencial de colaborador {reg_id}: {e}")
    return redirect(url_for('ver_credenciales_colaboradores'))


@app.route('/credenciales/colaboradores/<int:reg_id>/revelar', methods=['POST'])
@login_required
@agente_o_admin_required
def revelar_credencial_colaborador(reg_id):
    """Descifra a demanda la clave de una credencial de colaborador puntual y deja constancia
    en el log general de quién la consultó/copió — misma idea que /credenciales/<id>/revelar
    para la Bóveda de Accesos."""
    accion = request.form.get('accion', 'ver')
    accion_legible = 'copió' if accion == 'copiar' else 'vio'
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "SELECT colaborador, aplicativo, password_cifrada FROM credenciales_colaboradores WHERE id = %s" if db_type == 'postgres' else "SELECT colaborador, aplicativo, password_cifrada FROM credenciales_colaboradores WHERE id = ?"
        cursor.execute(q, (reg_id,))
        row = cursor.fetchone()
        conn.close()
    except Exception as e:
        print(f"⚠️ Error consultando credencial de colaborador {reg_id} para revelar: {e}")
        return jsonify({'error': 'error interno'}), 500

    if not row:
        return jsonify({'error': 'no encontrada'}), 404

    colaborador, aplicativo, pass_enc = row
    pass_real = desencriptar_texto(pass_enc)
    registrar_log(session.get('username'), "Consulta de Credencial de Colaborador",
                  f"Se {accion_legible} la clave de '{aplicativo}' de {colaborador} (ID {reg_id})")
    return jsonify({'password': pass_real})


@app.route('/credenciales/colaboradores/aplicativos/crear', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_aplicativo_catalogo():
    nombre = request.form.get('nombre', '').strip()
    if nombre:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            q = "INSERT INTO aplicativos_catalogo (nombre) VALUES (%s)" if db_type == 'postgres' else "INSERT INTO aplicativos_catalogo (nombre) VALUES (?)"
            cursor.execute(q, (nombre,))
            conn.commit()
            registrar_log(session.get('username'), "Catálogo de Aplicativos", f"Se agregó el aplicativo '{nombre}'")
        except Exception as e:
            conn.rollback()
            print(f"⚠️ Error agregando aplicativo '{nombre}': {e}")
        conn.close()
    return redirect(url_for('ver_credenciales_colaboradores'))


@app.route('/credenciales/colaboradores/aplicativos/<int:app_id>/eliminar', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_aplicativo_catalogo(app_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE aplicativos_catalogo SET estado = 'inactivo' WHERE id = %s" if db_type == 'postgres' else "UPDATE aplicativos_catalogo SET estado = 'inactivo' WHERE id = ?"
        cursor.execute(q, (app_id,))
        conn.commit()
        registrar_log(session.get('username'), "Catálogo de Aplicativos", f"Se desactivó el aplicativo ID {app_id}")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error desactivando aplicativo {app_id}: {e}")
    conn.close()
    return redirect(url_for('ver_credenciales_colaboradores'))


# ♻️ MÓDULO PAPELERA DE RECICLAJE
@app.route('/papelera')
@login_required
@agente_o_admin_required
def ver_papelera():
    conn, db_type = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT id, titulo, descripcion, fecha_subida, categoria, tipo FROM galerias WHERE estado = 'eliminado' ORDER BY fecha_subida DESC")
        eliminados = cursor.fetchall()
    except Exception:
        eliminados = []

    try:
        query_arch_elim = """
            SELECT a.id, COALESCE(a.filename, a.url_archivo), g.id, g.titulo, g.categoria
            FROM archivos a
            JOIN galerias g ON a.galeria_id = g.id
            WHERE a.estado = 'eliminado' AND COALESCE(g.estado, 'activo') != 'eliminado'
        """
        cursor.execute(query_arch_elim)
        archivos_eliminados = cursor.fetchall()
    except Exception:
        archivos_eliminados = []

    try:
        cursor.execute("SELECT id, titulo, usuario_acceso, area, fecha_creacion FROM credenciales WHERE estado = 'eliminado' ORDER BY id DESC")
        credenciales_eliminadas = cursor.fetchall()
    except Exception:
        credenciales_eliminadas = []

    try:
        cursor.execute("SELECT id, titulo, nivel, fecha, autor FROM comunicados WHERE estado = 'eliminado' ORDER BY id DESC")
        rows_com = cursor.fetchall()
    except Exception:
        rows_com = []

    # 👤 Igual que en el Muro de Comunicados: se muestra el alias/nombre real del autor.
    _nombres_papelera = _mapa_nombres_usuarios()
    comunicados_eliminados = [
        {'id': r[0], 'titulo': r[1], 'nivel': r[2], 'fecha': r[3], 'autor': _nombre_para_mostrar(r[4], _nombres_papelera)} for r in rows_com
    ]

    conn.close()
    return render_template(
        'papelera.html',
        eliminados=eliminados,
        archivos_eliminados=archivos_eliminados,
        credenciales_eliminadas=credenciales_eliminadas,
        comunicados_eliminados=comunicados_eliminados
    )

# 🔄 RESTAURAR CREDENCIAL
@app.route('/restaurar_credencial/<int:cred_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def restaurar_credencial(cred_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM credenciales WHERE id = ?"
        cursor.execute(q_sel, (cred_id,))
        row = cursor.fetchone()
        servicio = row[0] if row else f"ID {cred_id}"

        q_upd = "UPDATE credenciales SET estado = 'activo' WHERE id = %s" if db_type == 'postgres' else "UPDATE credenciales SET estado = 'activo' WHERE id = ?"
        cursor.execute(q_upd, (cred_id,))
        conn.commit()

        registrar_log(session['username'], "Restauración de Credencial", f"Se restauró el acceso '{servicio}' desde la papelera.")
    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('ver_papelera'))

# 💥 DESTRUIR CREDENCIAL
@app.route('/destruir_credencial/<int:cred_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def destruir_credencial(cred_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM credenciales WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM credenciales WHERE id = ?"
        cursor.execute(q_sel, (cred_id,))
        row = cursor.fetchone()
        servicio = row[0] if row else f"ID {cred_id}"

        q_del = "DELETE FROM credenciales WHERE id = %s" if db_type == 'postgres' else "DELETE FROM credenciales WHERE id = ?"
        cursor.execute(q_del, (cred_id,))
        conn.commit()

        registrar_log(session['username'], "Eliminación Permanente", f"Se destruyó permanentemente la credencial '{servicio}'.")
    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('ver_papelera'))

# 🔄 RESTAURAR INSTRUCTIVO COMPLETO
@app.route('/restaurar_galeria/<galeria_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def restaurar_galeria(galeria_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM galerias WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM galerias WHERE id = ?"
        cursor.execute(q_sel, (galeria_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else galeria_id

        q_upd = "UPDATE galerias SET estado = 'activo' WHERE id = %s" if db_type == 'postgres' else "UPDATE galerias SET estado = 'activo' WHERE id = ?"
        cursor.execute(q_upd, (galeria_id,))
        conn.commit()

        registrar_log(session['username'], "Restauración de Instructivo", f"El instructivo '{titulo}' fue restaurado desde la papelera.")
    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('ver_papelera'))

# 💥 BORRADO DEFINITIVO DE INSTRUCTIVO
@app.route('/destruir_galeria/<galeria_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def destruir_galeria(galeria_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM galerias WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM galerias WHERE id = ?"
        cursor.execute(q_sel, (galeria_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else galeria_id

        q_del1 = "DELETE FROM galerias WHERE id = %s" if db_type == 'postgres' else "DELETE FROM galerias WHERE id = ?"
        q_del2 = "DELETE FROM archivos WHERE galeria_id = %s" if db_type == 'postgres' else "DELETE FROM archivos WHERE galeria_id = ?"
        cursor.execute(q_del1, (galeria_id,))
        cursor.execute(q_del2, (galeria_id,))
        conn.commit()

        registrar_log(session['username'], "Eliminación Permanente", f"El instructivo '{titulo}' fue eliminado definitivamente del sistema.")
    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('ver_papelera'))

# 🗑️ BORRADO LÓGICO DE ARCHIVO INDIVIDUAL
@app.route('/eliminar_imagen/<galeria_id>/<path:filename>', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_imagen(galeria_id, filename):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM galerias WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM galerias WHERE id = ?"
        cursor.execute(q_sel, (galeria_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else galeria_id

        q_upd = "UPDATE archivos SET estado = 'eliminado' WHERE galeria_id = %s AND COALESCE(filename, url_archivo) = %s" if db_type == 'postgres' else "UPDATE archivos SET estado = 'eliminado' WHERE galeria_id = ? AND COALESCE(filename, url_archivo) = ?"
        cursor.execute(q_upd, (galeria_id, filename))
        conn.commit()

        nombre_limpio = filename.split('/')[-1] if 'http' in filename else filename
        registrar_log(session['username'], "Envío a Papelera (Archivo)", f"Se movió el archivo '{nombre_limpio}' del instructivo '{titulo}' a la papelera.")

    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('index'))

# 🔄 RESTAURAR ARCHIVO INDIVIDUAL
@app.route('/restaurar_archivo/<int:archivo_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def restaurar_archivo(archivo_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        query_info = """
            SELECT COALESCE(a.filename, a.url_archivo), g.titulo
            FROM archivos a
            JOIN galerias g ON a.galeria_id = g.id
            WHERE a.id = %s
        """ if db_type == 'postgres' else """
            SELECT COALESCE(a.filename, a.url_archivo), g.titulo
            FROM archivos a
            JOIN galerias g ON a.galeria_id = g.id
            WHERE a.id = ?
        """
        cursor.execute(query_info, (archivo_id,))
        row = cursor.fetchone()

        q_upd = "UPDATE archivos SET estado = 'activo' WHERE id = %s" if db_type == 'postgres' else "UPDATE archivos SET estado = 'activo' WHERE id = ?"
        cursor.execute(q_upd, (archivo_id,))
        conn.commit()

        if row:
            nombre_limpio = row[0].split('/')[-1] if 'http' in row[0] else row[0]
            registrar_log(session['username'], "Restauración de Archivo", f"Se reintegró el archivo '{nombre_limpio}' al instructivo '{row[1]}'.")

    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('ver_papelera'))

# 💥 DESTRUIR ARCHIVO INDIVIDUAL PERMANENTEMENTE
@app.route('/destruir_archivo/<int:archivo_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def destruir_archivo(archivo_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_del = "DELETE FROM archivos WHERE id = %s" if db_type == 'postgres' else "DELETE FROM archivos WHERE id = ?"
        cursor.execute(q_del, (archivo_id,))
        conn.commit()
        registrar_log(session['username'], "Eliminación Permanente (Archivo)", f"Se destruyó permanentemente un archivo adjunto ID '{archivo_id}'.")
    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('ver_papelera'))

def _crear_usuario_interno(datos, creador, conn, cursor, db_type):
    """Valida y crea una cuenta de usuario — lógica compartida entre el alta completa de
    Gestión de Usuarios y el alta rápida desde el modal de asignación de Inventario (ver
    crear_usuario_rapido_inventario), para no mantener dos copias de las mismas reglas de
    negocio (usuario único generado, contraseña mínima, correo válido, cédula sin duplicar).
    'datos' trae las claves primer_nombre/segundo_nombre/primer_apellido/segundo_apellido/
    password/email/telefono/cedula/especialidad/rol/firma_dataurl. Devuelve (error, nuevo_user,
    nombre_completo, firma_url) — error es None si todo salió bien; NO hace conn.close() (lo
    decide quien llama, según si sigue usando la conexión después)."""
    primer_nombre = (datos.get('primer_nombre') or '').strip()
    segundo_nombre = (datos.get('segundo_nombre') or '').strip()
    primer_apellido = (datos.get('primer_apellido') or '').strip()
    segundo_apellido = (datos.get('segundo_apellido') or '').strip()
    nuevo_pass = datos.get('password') or ''
    nuevo_email = (datos.get('email') or '').strip()
    nuevo_telefono = (datos.get('telefono') or '').strip() or None
    nueva_cedula = (datos.get('cedula') or '').strip() or None
    nueva_especialidad = (datos.get('especialidad') or '').strip() or None
    nuevo_rol = datos.get('rol') or 'estandar'
    if nuevo_rol not in ('admin', 'agente', 'estandar', 'gestion_humana'):
        nuevo_rol = 'estandar'
    if nuevo_rol == 'admin' and creador != 'admin':
        nuevo_rol = 'estandar'

    if not primer_nombre or not primer_apellido or not nuevo_pass or not nuevo_email or not nueva_especialidad:
        return "Nombre, primer apellido, correo, contraseña y especialidad son obligatorios para crear un usuario.", None, None, None

    if not _password_cuenta_valida(nuevo_pass):
        return f"La contraseña debe tener al menos {LONGITUD_MINIMA_PASSWORD_CUENTA} caracteres.", None, None, None

    if not _email_tiene_formato_valido(nuevo_email):
        return "El correo electrónico no tiene un formato válido (ej: nombre@dominio.com).", None, None, None

    if nueva_cedula:
        q_dup_cedula = "SELECT id FROM usuarios WHERE cedula = %s" if db_type == 'postgres' else "SELECT id FROM usuarios WHERE cedula = ?"
        cursor.execute(q_dup_cedula, (nueva_cedula,))
        if cursor.fetchone():
            return f"Ya existe un usuario registrado con la cédula {nueva_cedula}.", None, None, None

    # ✍️ Firma digital (opcional): si viene un data URL inválido o demasiado grande, sí se
    # bloquea el alta (igual que _subir_foto_perfil en /perfil) — pero no mandar firma alguna
    # nunca es un error, la firma se puede completar después.
    firma_url, error_firma = _subir_firma_desde_dataurl(datos.get('firma_dataurl'))
    if error_firma:
        return error_firma, None, None, None

    try:
        nuevo_user = _generar_username_unico(primer_nombre, primer_apellido, segundo_nombre, segundo_apellido)
        nombre_completo = ' '.join(p for p in [primer_nombre, segundo_nombre, primer_apellido, segundo_apellido] if p)
        nuevo_hash = generate_password_hash(nuevo_pass)
        q_ins = "INSERT INTO usuarios (usuario, password_hash, correo, rol, estado, nombre, telefono, cedula, especialidad, debe_cambiar_password, firma) VALUES (%s, %s, %s, %s, 'activo', %s, %s, %s, %s, TRUE, %s)" if db_type == 'postgres' else "INSERT INTO usuarios (usuario, password_hash, correo, rol, estado, nombre, telefono, cedula, especialidad, debe_cambiar_password, firma) VALUES (?, ?, ?, ?, 'activo', ?, ?, ?, ?, 1, ?)"
        cursor.execute(q_ins, (nuevo_user, nuevo_hash, nuevo_email, nuevo_rol, nombre_completo, nuevo_telefono, nueva_cedula, nueva_especialidad, firma_url))
        conn.commit()
        registrar_log(creador, "Creación de Usuario", f"Usuario '{nuevo_user}' ({nombre_completo}) [{nuevo_rol}]")

        # 👋 Correo de bienvenida (usuario + contraseña temporal), y notificación de campanita
        # dentro de Arkiv para su propia cuenta. El correo se manda en un hilo aparte para no
        # hacer esperar a quien está creando la cuenta.
        threading.Thread(
            target=enviar_correo_bienvenida,
            args=(nuevo_email, nuevo_user, nombre_completo, nuevo_pass)
        ).start()
        crear_notificacion(
            nuevo_user,
            f"¡Bienvenido a Arkiv, {primer_nombre}! Tu cuenta ya está activa. Revisa tu correo "
            f"({nuevo_email}) para conocer tu contraseña temporal y recuerda cambiarla en tu "
            f"primer inicio de sesión.",
            tipo='bienvenida'
        )
        return None, nuevo_user, nombre_completo, firma_url
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error creando usuario '{primer_nombre} {primer_apellido}': {e}")
        return "No se pudo crear el usuario. Verifica los datos e intenta de nuevo.", None, None, None


@app.route('/usuarios', methods=['GET', 'POST'])
@login_required
@admin_required
def gestion_usuarios():
    conn, db_type = get_db()
    cursor = conn.cursor()
    error = None
    form_data = None

    if request.method == 'POST':
        # 🛡️ Misma normalización de rol que aplicaba antes de factorizar la creación en
        # _crear_usuario_interno: se calcula aquí también (y no solo dentro del helper) para
        # que, si hay un error de validación, el formulario se vuelva a mostrar con el rol ya
        # normalizado (no con un valor crudo que un POST manual pudo haber mandado).
        nuevo_rol = request.form.get('rol', 'estandar')
        if nuevo_rol not in ('admin', 'agente', 'estandar', 'gestion_humana'):
            nuevo_rol = 'estandar'
        if nuevo_rol == 'admin' and session.get('username') != 'admin':
            nuevo_rol = 'estandar'
        form_data = {
            'primer_nombre': (request.form.get('primer_nombre') or '').strip(),
            'segundo_nombre': (request.form.get('segundo_nombre') or '').strip(),
            'primer_apellido': (request.form.get('primer_apellido') or '').strip(),
            'segundo_apellido': (request.form.get('segundo_apellido') or '').strip(),
            'email': (request.form.get('email') or '').strip(),
            'rol': nuevo_rol,
            'telefono': (request.form.get('telefono') or '').strip() or None,
            'cedula': (request.form.get('cedula') or '').strip() or None,
            'especialidad': (request.form.get('especialidad') or '').strip() or None,
        }
        error, nuevo_user, _nombre_completo, _firma_url = _crear_usuario_interno(
            {**form_data, 'password': request.form.get('password') or '', 'firma_dataurl': request.form.get('firma_dataurl')},
            session['username'], conn, cursor, db_type
        )
        if not error:
            conn.close()
            return redirect(url_for('gestion_usuarios', creado=nuevo_user))

    # 🛡️ La cuenta 'admin' queda oculta del listado para el resto de administradores: solo
    # la propia sesión de 'admin' la ve. El resto de admins no sabe que existe esta fila.
    if session.get('username') == 'admin':
        cursor.execute("SELECT id, usuario, correo, rol, estado, nombre, telefono, cedula, especialidad, totp_habilitado, bloqueado_por_intentos FROM usuarios ORDER BY id ASC")
    else:
        cursor.execute("SELECT id, usuario, correo, rol, estado, nombre, telefono, cedula, especialidad, totp_habilitado, bloqueado_por_intentos FROM usuarios WHERE usuario != 'admin' ORDER BY id ASC")
    lista_usuarios = cursor.fetchall()
    conn.close()
    usuario_creado = request.args.get('creado', '').strip()
    error_cedula = request.args.get('error_cedula', '').strip()
    # 🔒 Hallazgo QA H-02: avisa cuando editar_usuario descartó una contraseña por no cumplir
    # la longitud mínima (el resto de la edición sí se aplicó).
    error_pass_corta = request.args.get('error_pass_corta', '').strip()
    return render_template(
        'usuarios.html', usuarios=lista_usuarios, busqueda="", error=error, form_data=form_data,
        usuario_creado=usuario_creado, especialidades=_catalogo_especialidades_activas(),
        error_cedula=error_cedula, error_pass_corta=error_pass_corta,
        longitud_minima_password=LONGITUD_MINIMA_PASSWORD_CUENTA
    )


@app.route('/tickets/inventario/usuarios/crear_rapido', methods=['POST'])
@login_required
@agente_o_admin_required
def crear_usuario_rapido_inventario():
    """Alta rápida de usuario desde el modal de asignación de Inventario: si la persona a quien
    se le va a asignar un activo todavía no tiene cuenta en Arkiv, esto evita salir a Gestión de
    Usuarios y perder el formulario de asignación a medio llenar. Reutiliza exactamente las
    mismas reglas de _crear_usuario_interno (usuario generado, correo de bienvenida, etc.) — la
    única diferencia es que la contraseña temporal se genera sola (quien asigna el activo no
    necesita inventarla) y el rol queda fijo en 'estandar': este atajo es para dar de alta a un
    colaborador que va a recibir un equipo, no para crear cuentas de agente/admin."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    datos = {
        'primer_nombre': request.form.get('primer_nombre', ''),
        'segundo_nombre': request.form.get('segundo_nombre', ''),
        'primer_apellido': request.form.get('primer_apellido', ''),
        'segundo_apellido': request.form.get('segundo_apellido', ''),
        'email': request.form.get('email', ''),
        'telefono': request.form.get('telefono', ''),
        'cedula': request.form.get('cedula', ''),
        'especialidad': request.form.get('especialidad', ''),
        'rol': 'estandar',
        'password': secrets.token_urlsafe(12),
        'firma_dataurl': request.form.get('firma_dataurl'),
    }
    error, nuevo_user, nombre_completo, firma_url = _crear_usuario_interno(datos, session['username'], conn, cursor, db_type)
    conn.close()
    if error:
        return jsonify({'ok': False, 'error': error}), 400
    return jsonify({'ok': True, 'usuario': nuevo_user, 'nombre': nombre_completo, 'firma': firma_url})


@app.route('/usuarios/<usuario>/2fa/desactivar', methods=['POST'])
@login_required
@admin_required
def admin_desactivar_2fa(usuario):
    """Permite a un admin desactivar el 2FA de OTRA cuenta — recuperación para cuando esa
    persona perdió su dispositivo/app autenticadora y ya no puede generar códigos ni tiene a la
    mano sus códigos de respaldo. No requiere la contraseña de esa cuenta (el admin no la
    conoce); la protección aquí es que la acción exige sesión de administrador y queda auditada
    en el log."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    q_sel = "SELECT usuario, rol FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT usuario, rol FROM usuarios WHERE usuario = ?"
    cursor.execute(q_sel, (usuario,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return redirect(url_for('gestion_usuarios'))

    # 🛡️ Misma protección que ya existe para editar/eliminar: solo el super-admin puede tocar
    # la cuenta de otro admin (incluida la desactivación de su 2FA).
    if row[1] == 'admin' and session.get('username') != 'admin' and row[0] != session.get('username'):
        return redirect(url_for('gestion_usuarios'))

    _desactivar_2fa_cuenta(usuario)
    registrar_log(session['username'], "Desactivación Administrativa de 2FA", f"El administrador desactivó el 2FA de la cuenta '{usuario}' (recuperación por pérdida de dispositivo).")
    return redirect(url_for('gestion_usuarios'))


@app.route('/usuarios/buscar_cedula')
@login_required
@agente_o_admin_required
def buscar_usuario_por_cedula():
    """Busca una cuenta de Arkiv por número de cédula — usado desde el buscador rápido del
    Inventario de Activos para llenar 'Asignado a' sin tener que escribir el nombre completo de
    memoria. Antes lo podía consultar cualquier usuario logueado (incluido el rol Estándar);
    esto permitía enumerar nombres asociados a cédulas probando números al azar (hallazgo QA
    H-03, 30/08/2026). Ahora queda restringido a los roles operativos que de verdad lo usan:
    agentes y admins registrando/gestionando activos. Incluye 'firma' (URL de Cloudinary, o None
    si esa persona nunca registró una) para que el modal de asignar activo pueda reutilizarla
    como constancia de aceptación en vez de pedir que firme otra vez."""
    cedula = request.args.get('cedula', '').strip()
    if not cedula:
        return jsonify({'encontrado': False})

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "SELECT usuario, nombre, firma FROM usuarios WHERE cedula = %s AND COALESCE(estado, 'activo') = 'activo'" if db_type == 'postgres' else "SELECT usuario, nombre, firma FROM usuarios WHERE cedula = ? AND COALESCE(estado, 'activo') = 'activo'"
        cursor.execute(q, (cedula,))
        row = cursor.fetchone()
    except Exception as e:
        print(f"⚠️ Error buscando usuario por cédula: {e}")
        row = None
    conn.close()

    if not row:
        return jsonify({'encontrado': False})
    return jsonify({'encontrado': True, 'usuario': row[0], 'nombre': row[1] or row[0], 'firma': row[2]})


@app.route('/usuarios/buscar')
@login_required
@agente_o_admin_required
def buscar_usuarios():
    """Coincidencia PARCIAL por nombre, usuario o cédula (a diferencia de buscar_usuario_por_cedula,
    que exige la cédula completa) — usado por el campo "Colaborador" de Altas de Credenciales para
    sugerir cuentas ya registradas en Gestión de Usuarios mientras se escribe, sin obligar a que la
    persona ya tenga cuenta (si no aparece en las sugerencias, el campo sigue aceptando texto libre).
    Restringido a agente/admin, igual que buscar_usuario_por_cedula (hallazgo QA H-03: no debe poder
    consultarlo cualquier cuenta Estándar)."""
    q = (request.args.get('q') or '').strip().lower()
    if len(q) < 2:
        return jsonify({'resultados': []})

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT usuario, nombre, cedula FROM usuarios WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY nombre ASC"
        )
        filas = cursor.fetchall()
    except Exception as e:
        print(f"⚠️ Error buscando usuarios: {e}")
        filas = []
    conn.close()

    resultados = []
    for u_usuario, u_nombre, u_cedula in filas:
        if q in f"{u_nombre or ''} {u_usuario} {u_cedula or ''}".lower():
            resultados.append({'usuario': u_usuario, 'nombre': u_nombre or u_usuario, 'cedula': u_cedula or ''})
            if len(resultados) >= 10:
                break
    return jsonify({'resultados': resultados})


# ✏️ EDITAR USUARIO
@app.route('/editar_usuario/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def editar_usuario(usuario_id):
    nuevo_email = request.form.get('email', '').strip()
    nuevo_rol = request.form.get('rol', 'estandar').strip()
    if nuevo_rol not in ('admin', 'agente', 'estandar', 'gestion_humana'):
        nuevo_rol = 'estandar'
    nueva_pass = request.form.get('password', '').strip()
    nuevo_nombre = request.form.get('nombre', '').strip()
    nuevo_telefono = request.form.get('telefono', '').strip()
    nueva_cedula = request.form.get('cedula', '').strip()
    nueva_especialidad = request.form.get('especialidad', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT usuario, rol, nombre, telefono, cedula, especialidad, correo FROM usuarios WHERE id = %s" if db_type == 'postgres' else "SELECT usuario, rol, nombre, telefono, cedula, especialidad, correo FROM usuarios WHERE id = ?"
        cursor.execute(q_sel, (usuario_id,))
        row = cursor.fetchone()
        user_target = row[0] if row else None
        rol_target = row[1] if row else None
        # Si el admin deja el campo Nombre, Teléfono, Cédula o Especialidad vacío en el formulario
        # de edición, se conserva el valor que ya tenía (permite editar solo correo/rol/contraseña
        # sin borrar los demás datos).
        nombre_final = nuevo_nombre or (row[2] if row else None)
        telefono_final = nuevo_telefono or (row[3] if row else None)
        cedula_original = row[4] if row else None
        cedula_final = nueva_cedula or cedula_original
        especialidad_final = nueva_especialidad or (row[5] if row else None)

        # 📧 Igual que con los demás campos opcionales: si el correo que llegó no tiene un
        # formato válido (ver _email_tiene_formato_valido), NO se guarda un dato corrupto —
        # se conserva el correo que la cuenta ya tenía, en vez de tumbar toda la edición.
        correo_original = row[6] if row and len(row) > 6 else None
        if not _email_tiene_formato_valido(nuevo_email):
            nuevo_email = correo_original or nuevo_email

        if user_target is None:
            conn.close()
            return redirect(url_for('gestion_usuarios'))

        # 🪪 Solo se valida la unicidad de la cédula cuando el admin la está CAMBIANDO
        # activamente a un valor distinto del que este usuario ya tenía — así una pareja de
        # cuentas que ya compartía cédula desde antes de esta validación puede seguir
        # editando sus demás datos (correo, rol, teléfono...) sin quedar bloqueada por su
        # propio duplicado histórico. Solo se impide crear un duplicado NUEVO.
        if nueva_cedula and nueva_cedula != (cedula_original or ''):
            q_dup_cedula = "SELECT id FROM usuarios WHERE cedula = %s AND id != %s" if db_type == 'postgres' else "SELECT id FROM usuarios WHERE cedula = ? AND id != ?"
            cursor.execute(q_dup_cedula, (nueva_cedula, usuario_id))
            if cursor.fetchone():
                conn.close()
                return redirect(url_for('gestion_usuarios', error_cedula=nueva_cedula))

        es_superadmin = (session.get('username') == 'admin')

        # 🛡️ Solo la cuenta 'admin' (super-admin) puede editar los datos (correo, rol o
        # contraseña) de una cuenta con rol 'admin' o 'agente' — incluida la PROPIA cuenta de
        # quien edita. Esto protege a las cuentas admin/agente frente a sus pares (y frente a
        # sí mismas): un admin o agente comprometido o malicioso ya no puede tomar control de
        # otra cuenta admin/agente, ni cambiarse sus propias credenciales desde este panel.
        # Un admin/agente (no super-admin) solo puede editar cuentas con rol 'estandar'.
        if rol_target in ('admin', 'agente') and not es_superadmin:
            conn.close()
            return redirect(url_for('gestion_usuarios'))

        # 🛡️ Solo el super-admin puede ASCENDER a alguien a rol 'admin'. Sin esto, un admin
        # podría evadir la protección anterior ascendiendo a un usuario estándar a admin.
        if nuevo_rol == 'admin' and not es_superadmin:
            nuevo_rol = rol_target or 'estandar'

        # 🔒 Misma política de longitud mínima que el autoservicio (hallazgo QA H-02): si el
        # admin decide asignar una contraseña nueva desde aquí, no puede ser más corta que la
        # que se le exigiría a la propia persona en /perfil/cambiar_password. Se descarta solo
        # la contraseña (se conserva la que ya tenía); el resto de la edición sigue su curso, y
        # se avisa al admin con error_pass_corta para que no crea que sí quedó aplicada.
        pass_rechazada = bool(nueva_pass) and not _password_cuenta_valida(nueva_pass)
        if pass_rechazada:
            nueva_pass = ''

        if nueva_pass:
            nuevo_hash = generate_password_hash(nueva_pass)
            # 🔒 Igual que al crear el usuario: si el admin le asigna una contraseña nueva desde
            # aquí, se obliga a cambiarla en su próximo inicio de sesión.
            q_upd = "UPDATE usuarios SET correo = %s, rol = %s, password_hash = %s, nombre = %s, telefono = %s, cedula = %s, especialidad = %s, debe_cambiar_password = TRUE WHERE id = %s" if db_type == 'postgres' else "UPDATE usuarios SET correo = ?, rol = ?, password_hash = ?, nombre = ?, telefono = ?, cedula = ?, especialidad = ?, debe_cambiar_password = 1 WHERE id = ?"
            cursor.execute(q_upd, (nuevo_email, nuevo_rol, nuevo_hash, nombre_final, telefono_final, cedula_final, especialidad_final, usuario_id))
            detalle_log = f"Se actualizó correo, rol y CONTRASEÑA del usuario '{user_target}'"
        else:
            q_upd = "UPDATE usuarios SET correo = %s, rol = %s, nombre = %s, telefono = %s, cedula = %s, especialidad = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE usuarios SET correo = ?, rol = ?, nombre = ?, telefono = ?, cedula = ?, especialidad = ? WHERE id = ?"
            cursor.execute(q_upd, (nuevo_email, nuevo_rol, nombre_final, telefono_final, cedula_final, especialidad_final, usuario_id))
            detalle_log = f"Se actualizó correo y rol del usuario '{user_target}'"

        conn.commit()
        registrar_log(session['username'], "Edición de Usuario", detalle_log)
    except Exception as e:
        conn.rollback()
        pass_rechazada = False

    conn.close()
    if pass_rechazada:
        return redirect(url_for('gestion_usuarios', error_pass_corta=1))
    return redirect(url_for('gestion_usuarios'))


# 🩺 CATÁLOGO DE ESPECIALIDADES (administrable desde /usuarios, igual que el de aplicativos)
@app.route('/usuarios/especialidades/crear', methods=['POST'])
@login_required
@admin_required
def crear_especialidad_catalogo():
    nombre = request.form.get('nombre', '').strip()
    if nombre:
        conn, db_type = get_db()
        cursor = conn.cursor()
        try:
            q = "INSERT INTO especialidades_catalogo (nombre) VALUES (%s)" if db_type == 'postgres' else "INSERT INTO especialidades_catalogo (nombre) VALUES (?)"
            cursor.execute(q, (nombre,))
            conn.commit()
            registrar_log(session.get('username'), "Catálogo de Especialidades", f"Se agregó la especialidad '{nombre}'")
        except Exception as e:
            conn.rollback()
            print(f"⚠️ Error agregando especialidad '{nombre}': {e}")
        conn.close()
    return redirect(url_for('gestion_usuarios'))


@app.route('/usuarios/especialidades/<int:esp_id>/eliminar', methods=['POST'])
@login_required
@admin_required
def eliminar_especialidad_catalogo(esp_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE especialidades_catalogo SET estado = 'inactivo' WHERE id = %s" if db_type == 'postgres' else "UPDATE especialidades_catalogo SET estado = 'inactivo' WHERE id = ?"
        cursor.execute(q, (esp_id,))
        conn.commit()
        registrar_log(session.get('username'), "Catálogo de Especialidades", f"Se desactivó la especialidad ID {esp_id}")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error desactivando especialidad {esp_id}: {e}")
    conn.close()
    return redirect(url_for('gestion_usuarios'))


# ❌ ELIMINAR USUARIO
@app.route('/eliminar_usuario/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def eliminar_usuario(usuario_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT usuario, rol FROM usuarios WHERE id = %s" if db_type == 'postgres' else "SELECT usuario, rol FROM usuarios WHERE id = ?"
        cursor.execute(q_sel, (usuario_id,))
        row = cursor.fetchone()

        if not row:
            conn.close()
            return redirect(url_for('gestion_usuarios'))

        user_target, rol_target = row[0], row[1]

        # 🛡️ Nunca permitir eliminar la cuenta 'admin' (dejaría a todos sin acceso) ni la
        # propia cuenta con la que se inició sesión (evita un auto-eliminado accidental).
        if user_target == 'admin' or user_target == session.get('username'):
            conn.close()
            return redirect(url_for('gestion_usuarios'))

        # 🛡️ Solo la cuenta 'admin' (super-admin) puede eliminar a OTRA cuenta con rol
        # 'admin' o 'agente'. Sin esto, un admin al que ya se le bloquea desactivar a otro
        # admin/agente podría eludir esa protección eliminándolo directamente.
        if rol_target in ('admin', 'agente') and session.get('username') != 'admin':
            conn.close()
            return redirect(url_for('gestion_usuarios'))

        q_del = "DELETE FROM usuarios WHERE id = %s" if db_type == 'postgres' else "DELETE FROM usuarios WHERE id = ?"
        cursor.execute(q_del, (usuario_id,))
        conn.commit()

        registrar_log(session['username'], "Eliminación de Usuario", f"Se eliminó el usuario '{user_target}' del sistema")
    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('gestion_usuarios'))

# 🔒 BLOQUEAR / DESBLOQUEAR USUARIO
@app.route('/usuarios/toggle_estado/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def toggle_estado_usuario(usuario_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT usuario, estado, rol, nombre FROM usuarios WHERE id = %s" if db_type == 'postgres' else "SELECT usuario, estado, rol, nombre FROM usuarios WHERE id = ?"
        cursor.execute(q_sel, (usuario_id,))
        row = cursor.fetchone()

        if row:
            user_target, estado_actual, rol_target, nombre_target = row[0], (row[1] or 'activo'), row[2], row[3]
            # 🛡️ Nunca permitir bloquear la cuenta 'admin' (dejaría a todos sin acceso) ni la
            # propia cuenta con la que se inició sesión (evita un auto-bloqueo accidental).
            if user_target == 'admin' or user_target == session.get('username'):
                conn.close()
                return redirect(url_for('gestion_usuarios'))

            # 🛡️ Solo la cuenta 'admin' (super-admin) puede bloquear/desbloquear una
            # cuenta con rol 'admin' o 'agente' — incluida la PROPIA cuenta de quien actúa.
            if rol_target in ('admin', 'agente') and session.get('username') != 'admin':
                conn.close()
                return redirect(url_for('gestion_usuarios'))

            nuevo_estado = 'inactivo' if estado_actual == 'activo' else 'activo'

            # 🪪 Antes de bloquear (dar de baja) a un colaborador: si todavía tiene un PC u
            # otro activo del Inventario asignado a su nombre, no se deja bloquear la cuenta
            # hasta que Gestión Humana o TI certifiquen la devolución en
            # /inventario/certificacion_devoluciones. No aplica al desbloqueo.
            if nuevo_estado == 'inactivo':
                pendientes = _activos_pendientes_devolucion(nombre_target, user_target)
                if pendientes:
                    conn.close()
                    nombres_pendientes = ', '.join(f"{p['nombre']} ({p['tipo_activo']})" for p in pendientes)
                    flash(f"No se pudo bloquear a '{user_target}': todavía tiene activo(s) asignado(s) en el Inventario sin devolución confirmada: {nombres_pendientes}. "
                          "Gestión Humana o TI deben certificar la devolución en Inventario → Certificación de Devoluciones antes de continuar.", "error")
                    return redirect(url_for('gestion_usuarios'))

            q_upd = "UPDATE usuarios SET estado = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE usuarios SET estado = ? WHERE id = ?"
            cursor.execute(q_upd, (nuevo_estado, usuario_id))
            conn.commit()

            accion = "Bloqueo de Usuario" if nuevo_estado == 'inactivo' else "Desbloqueo de Usuario"
            registrar_log(session['username'], accion, f"El usuario '{user_target}' fue marcado como '{nuevo_estado}'")
    except Exception as e:
        conn.rollback()
        print(f"Error cambiando estado del usuario {usuario_id}: {e}")

    conn.close()
    return redirect(url_for('gestion_usuarios'))

# 🔓 DESBLOQUEAR CUENTA (BLOQUEADA AUTOMÁTICAMENTE POR INTENTOS FALLIDOS DE LOGIN)
# Distinto de toggle_estado_usuario de arriba: ese es el bloqueo MANUAL que decide un admin
# (columna 'estado'); este es el bloqueo AUTOMÁTICO que dispara el propio sistema de login
# tras varios fallos de contraseña seguidos (columna 'bloqueado_por_intentos', ver
# UMBRAL_INTENTOS_FALLIDOS_LOGIN). El otro camino para que se apague solo es que la propia
# persona recupere su clave por correo (ver validar_codigo).
@app.route('/usuarios/desbloquear_intentos/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def desbloquear_intentos_usuario(usuario_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT usuario FROM usuarios WHERE id = %s" if db_type == 'postgres' else "SELECT usuario FROM usuarios WHERE id = ?"
        cursor.execute(q_sel, (usuario_id,))
        row = cursor.fetchone()

        if row:
            q_upd = "UPDATE usuarios SET bloqueado_por_intentos = FALSE, intentos_fallidos_login = 0 WHERE id = %s" if db_type == 'postgres' else "UPDATE usuarios SET bloqueado_por_intentos = 0, intentos_fallidos_login = 0 WHERE id = ?"
            cursor.execute(q_upd, (usuario_id,))
            conn.commit()
            registrar_log(session['username'], "Desbloqueo por Intentos Fallidos", f"El administrador desbloqueó manualmente la cuenta '{row[0]}' (bloqueada por intentos fallidos de inicio de sesión).")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error desbloqueando por intentos al usuario {usuario_id}: {e}")

    conn.close()
    return redirect(url_for('gestion_usuarios'))

# 📑 RUTA /LOGS CON FILTROS
@app.route('/logs')
@login_required
@agente_o_admin_required
def ver_logs():
    q_usuario = request.args.get('usuario', '').strip()
    q_accion = request.args.get('accion', '').strip()
    q_busqueda = request.args.get('q', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT DISTINCT usuario FROM logs ORDER BY usuario ASC")
    lista_usuarios = [u[0] for u in cursor.fetchall() if u[0]]

    cursor.execute("SELECT DISTINCT accion FROM logs ORDER BY accion ASC")
    lista_acciones = [a[0] for a in cursor.fetchall() if a[0]]

    query = "SELECT usuario, accion, detalles, fecha FROM logs WHERE 1=1"
    params = []

    if q_usuario:
        query += " AND usuario = %s" if db_type == 'postgres' else " AND usuario = ?"
        params.append(q_usuario)

    if q_accion:
        query += " AND accion = %s" if db_type == 'postgres' else " AND accion = ?"
        params.append(q_accion)

    if q_busqueda:
        p_busq = f"%{q_busqueda}%"
        if db_type == 'postgres':
            query += " AND (detalles ILIKE %s OR fecha ILIKE %s)"
            params.extend([p_busq, p_busq])
        else:
            query += " AND (detalles LIKE ? OR fecha LIKE ?)"
            params.extend([p_busq, p_busq])

    # 📄 Paginación: esta bitácora crece sin límite (una fila por cada acción que registra
    # registrar_log en toda la app) y antes se traía la tabla completa en cada visita — con
    # meses de uso eso se vuelve una consulta cada vez más lenta y una página cada vez más
    # pesada de renderizar. El filtrado (usuario/acción/búsqueda) ya vive entero en el SQL de
    # arriba, así que paginar con LIMIT/OFFSET sobre esa misma consulta no esconde ningún
    # resultado: cambia página y sigue viendo exactamente los mismos filtros.
    POR_PAGINA_LOGS = 50
    try:
        pagina_actual = max(1, int(request.args.get('pagina', '1')))
    except ValueError:
        pagina_actual = 1

    query_conteo = query.replace(
        "SELECT usuario, accion, detalles, fecha FROM logs WHERE 1=1",
        "SELECT COUNT(*) FROM logs WHERE 1=1",
        1
    )
    cursor.execute(query_conteo, tuple(params))
    total_registros = cursor.fetchone()[0]
    total_paginas = max(1, math.ceil(total_registros / POR_PAGINA_LOGS))
    pagina_actual = min(pagina_actual, total_paginas)
    offset = (pagina_actual - 1) * POR_PAGINA_LOGS

    query += " ORDER BY id DESC LIMIT %s OFFSET %s" if db_type == 'postgres' else " ORDER BY id DESC LIMIT ? OFFSET ?"
    cursor.execute(query, tuple(params) + (POR_PAGINA_LOGS, offset))
    lista_logs = cursor.fetchall()
    conn.close()

    return render_template(
        'logs.html',
        logs=lista_logs,
        usuarios_opt=lista_usuarios,
        acciones_opt=lista_acciones,
        q_usuario=q_usuario,
        q_accion=q_accion,
        q_busqueda=q_busqueda,
        pagina_actual=pagina_actual,
        total_paginas=total_paginas,
        total_registros=total_registros,
        por_pagina=POR_PAGINA_LOGS
    )

# 📊 EXPORTAR AUDITORÍA A EXCEL / CSV
@app.route('/exportar_logs_csv')
@login_required
@agente_o_admin_required
def exportar_logs_csv():
    q_usuario = request.args.get('usuario', '').strip()
    q_accion = request.args.get('accion', '').strip()
    q_busqueda = request.args.get('q', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()

    query = "SELECT fecha, usuario, accion, detalles FROM logs WHERE 1=1"
    params = []

    if q_usuario:
        query += " AND usuario = %s" if db_type == 'postgres' else " AND usuario = ?"
        params.append(q_usuario)

    if q_accion:
        query += " AND accion = %s" if db_type == 'postgres' else " AND accion = ?"
        params.append(q_accion)

    if q_busqueda:
        p_busq = f"%{q_busqueda}%"
        if db_type == 'postgres':
            query += " AND (detalles ILIKE %s OR fecha ILIKE %s)"
            params.extend([p_busq, p_busq])
        else:
            query += " AND (detalles LIKE ? OR fecha LIKE ?)"
            params.extend([p_busq, p_busq])

    query += " ORDER BY id DESC"

    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['FECHA Y HORA', 'USUARIO', 'ACCIÓN', 'DETALLE DEL CAMBIO'])

    for row in rows:
        writer.writerow(row)

    csv_bytes = '\ufeff' + output.getvalue()
    
    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y%m%d_%H%M")
    filename = f"Arkiv_Auditoria_Logs_{fecha_filename}.csv"

    headers = {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

    return Response(csv_bytes, headers=headers, status=200)

# 📧 BITÁCORA DE CORREOS ENVIADOS (tickets + recuperación de clave)
# 🛡️ Solo muestra fecha, destinatario, asunto, tipo, estado y (si falló) el mensaje de la
# excepción de red/HTTP — nunca el cuerpo del correo ni el código de verificación, que no se
# guardan en ningún lado desde que se generan en /recuperar (ver registrar_correo_log).
@app.route('/logs/correos')
@login_required
@agente_o_admin_required
def ver_logs_correos():
    q_destinatario = request.args.get('destinatario', '').strip()
    q_tipo = request.args.get('tipo', '').strip()
    q_estado = request.args.get('estado', '').strip()
    q_busqueda = request.args.get('q', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT DISTINCT destinatario FROM correos_log ORDER BY destinatario ASC")
    destinatarios_opt = [d[0] for d in cursor.fetchall() if d[0]]

    cursor.execute("SELECT DISTINCT tipo FROM correos_log ORDER BY tipo ASC")
    tipos_opt = [t[0] for t in cursor.fetchall() if t[0]]

    query = "SELECT fecha, destinatario, asunto, tipo, estado, detalle_error FROM correos_log WHERE 1=1"
    params = []

    if q_destinatario:
        query += " AND destinatario = %s" if db_type == 'postgres' else " AND destinatario = ?"
        params.append(q_destinatario)

    if q_tipo:
        query += " AND tipo = %s" if db_type == 'postgres' else " AND tipo = ?"
        params.append(q_tipo)

    if q_estado:
        query += " AND estado = %s" if db_type == 'postgres' else " AND estado = ?"
        params.append(q_estado)

    if q_busqueda:
        p_busq = f"%{q_busqueda}%"
        if db_type == 'postgres':
            query += " AND (destinatario ILIKE %s OR asunto ILIKE %s)"
        else:
            query += " AND (destinatario LIKE ? OR asunto LIKE ?)"
        params.extend([p_busq, p_busq])

    query += " ORDER BY id DESC LIMIT 500"

    cursor.execute(query, tuple(params))
    lista_correos = cursor.fetchall()
    conn.close()

    return render_template(
        'logs_correos.html',
        correos=lista_correos,
        destinatarios_opt=destinatarios_opt,
        tipos_opt=tipos_opt,
        q_destinatario=q_destinatario,
        q_tipo=q_tipo,
        q_estado=q_estado,
        q_busqueda=q_busqueda
    )

# 📊 EXPORTAR BITÁCORA DE CORREOS A EXCEL / CSV
@app.route('/exportar_logs_correos_csv')
@login_required
@agente_o_admin_required
def exportar_logs_correos_csv():
    q_destinatario = request.args.get('destinatario', '').strip()
    q_tipo = request.args.get('tipo', '').strip()
    q_estado = request.args.get('estado', '').strip()
    q_busqueda = request.args.get('q', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()

    query = "SELECT fecha, destinatario, asunto, tipo, estado, detalle_error FROM correos_log WHERE 1=1"
    params = []

    if q_destinatario:
        query += " AND destinatario = %s" if db_type == 'postgres' else " AND destinatario = ?"
        params.append(q_destinatario)

    if q_tipo:
        query += " AND tipo = %s" if db_type == 'postgres' else " AND tipo = ?"
        params.append(q_tipo)

    if q_estado:
        query += " AND estado = %s" if db_type == 'postgres' else " AND estado = ?"
        params.append(q_estado)

    if q_busqueda:
        p_busq = f"%{q_busqueda}%"
        if db_type == 'postgres':
            query += " AND (destinatario ILIKE %s OR asunto ILIKE %s)"
        else:
            query += " AND (destinatario LIKE ? OR asunto LIKE ?)"
        params.extend([p_busq, p_busq])

    query += " ORDER BY id DESC"

    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['FECHA Y HORA', 'DESTINATARIO', 'ASUNTO', 'TIPO', 'ESTADO', 'DETALLE DEL ERROR'])

    for row in rows:
        writer.writerow(row)

    csv_bytes = '﻿' + output.getvalue()

    fecha_filename = datetime.now(ZONA_HORARIA_COLOMBIA).strftime("%Y%m%d_%H%M")
    filename = f"Arkiv_Log_Correos_{fecha_filename}.csv"

    headers = {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

    return Response(csv_bytes, headers=headers, status=200)

@app.route('/perfil/tema', methods=['POST'])
@login_required
def cambiar_tema():
    """Alterna entre los 3 temas del usuario que tiene la sesión abierta: claro, oscuro y
    'descanso' (variante sepia/tenue para descansar la vista — ver static/css/tema-descanso.css).
    Se guarda en su cuenta (no solo en el navegador) para que lo siga a donde inicie sesión — y
    de paso se actualiza la sesión actual para que se vea reflejado de inmediato, sin tener que
    volver a iniciar sesión."""
    nuevo_tema = request.form.get('tema', '').strip()
    if nuevo_tema not in ('claro', 'oscuro', 'descanso'):
        nuevo_tema = 'oscuro'
    try:
        conn, db_type = get_db()
        cursor = conn.cursor()
        q = "UPDATE usuarios SET tema = %s WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET tema = ? WHERE usuario = ?"
        cursor.execute(q, (nuevo_tema, session.get('username')))
        conn.commit()
        conn.close()
        session['tema'] = nuevo_tema
    except Exception as e:
        print(f"⚠️ Error guardando preferencia de tema: {e}")
    # Vuelve a la misma página desde la que se alternó el tema (con un respaldo razonable
    # por si el navegador no manda 'Referer').
    destino = request.referrer or url_for('bienvenida')
    return redirect(destino)


@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil_datos():
    """Página de autoservicio de datos personales: nombre completo, teléfono y foto de perfil.
    Deliberadamente NO incluye correo ni cédula — ambos siguen siendo de resorte exclusivo de un
    admin (Gestión de Usuarios), porque su edición ya tiene validaciones propias allí (formato de
    correo, duplicados de cédula) y son datos más sensibles para identidad/notificaciones. El
    usuario, el rol y el estado tampoco son editables aquí: solo se muestran de referencia."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    error = None
    exito = False

    if request.method == 'POST':
        nombre_nuevo = (request.form.get('nombre') or '').strip()
        telefono_nuevo = (request.form.get('telefono') or '').strip() or None

        # 🪪 Mínimo primer nombre + primer apellido (2 palabras). No se piden 4 campos
        # separados (primer/segundo nombre, primer/segundo apellido) como en el alta por admin
        # para no tener que adivinar cómo partir un nombre ya existente en esos 4 casilleros al
        # precargar el formulario — un nombre compuesto (p. ej. "María José Vélez De La Cruz")
        # se presta a particiones ambiguas. Un solo campo de texto, validado por cantidad de
        # palabras, evita ese problema sin perder la garantía mínima que pidió Tomás.
        if not nombre_nuevo or len(nombre_nuevo.split()) < 2:
            error = "Escribe tu nombre completo, con al menos nombre y apellido."

        foto_url = None
        if not error:
            foto_file = request.files.get('foto_perfil')
            if foto_file and foto_file.filename:
                foto_url, error = _subir_foto_perfil(foto_file)

        if not error:
            try:
                if foto_url:
                    q_upd = "UPDATE usuarios SET nombre = %s, telefono = %s, foto_perfil = %s WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET nombre = ?, telefono = ?, foto_perfil = ? WHERE usuario = ?"
                    cursor.execute(q_upd, (nombre_nuevo, telefono_nuevo, foto_url, session['username']))
                    session['foto_perfil'] = foto_url
                else:
                    # Sin foto nueva: no se toca la columna foto_perfil (se conserva la que ya había).
                    q_upd = "UPDATE usuarios SET nombre = %s, telefono = %s WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET nombre = ?, telefono = ? WHERE usuario = ?"
                    cursor.execute(q_upd, (nombre_nuevo, telefono_nuevo, session['username']))
                conn.commit()
                registrar_log(session['username'], "Actualización de Perfil", "El usuario actualizó sus datos de perfil" + (" (incluida foto)" if foto_url else ""))
                exito = True
            except Exception as e:
                conn.rollback()
                error = "No se pudieron guardar los cambios. Intenta de nuevo."

    q_sel = "SELECT nombre, telefono, correo, rol, foto_perfil, usuario FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT nombre, telefono, correo, rol, foto_perfil, usuario FROM usuarios WHERE usuario = ?"
    cursor.execute(q_sel, (session['username'],))
    row = cursor.fetchone()
    conn.close()

    return render_template(
        'perfil_datos.html',
        nombre=(row[0] if row else '') or '', telefono=(row[1] if row else '') or '',
        correo=row[2] if row else '', rol=row[3] if row else session.get('rol'),
        foto_perfil=row[4] if row else None, usuario=row[5] if row else session.get('username'),
        error=error, exito=exito,
    )


@app.route('/perfil/cambiar_password', methods=['GET', 'POST'])
@login_required
def cambiar_password_perfil():
    """Página de autoservicio para que un usuario cambie su propia contraseña. Cumple dos
    propósitos: (1) es la puerta obligatoria a la que 'validar_instancia_y_sesion' redirige a
    cualquier cuenta marcada con debe_cambiar_password=True (contraseña temporal asignada por
    un admin), y (2) queda disponible para que cualquier usuario la use quiera cuando quiera,
    tal como ya se lo promete el correo de bienvenida ("la puedes cambiar desde tu perfil")."""
    obligatorio = bool(session.get('debe_cambiar_password'))
    error = None

    if request.method == 'POST':
        password_actual = request.form.get('password_actual') or ''
        password_nueva = request.form.get('password_nueva') or ''
        password_confirmar = request.form.get('password_confirmar') or ''

        if not password_actual or not password_nueva or not password_confirmar:
            error = "Todos los campos son obligatorios."
        elif not _password_cuenta_valida(password_nueva):
            error = f"La nueva contraseña debe tener al menos {LONGITUD_MINIMA_PASSWORD_CUENTA} caracteres."
        elif password_nueva != password_confirmar:
            error = "La nueva contraseña y su confirmación no coinciden."

        if not error:
            conn, db_type = get_db()
            cursor = conn.cursor()
            try:
                q_sel = "SELECT password_hash FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT password_hash FROM usuarios WHERE usuario = ?"
                cursor.execute(q_sel, (session.get('username'),))
                row = cursor.fetchone()
                clave_db = str(row[0] or '') if row else ''

                es_valida = False
                if clave_db.startswith('pbkdf2:') or clave_db.startswith('scrypt:'):
                    es_valida = check_password_hash(clave_db, password_actual)
                else:
                    es_valida = hmac.compare_digest(clave_db, password_actual)

                if not es_valida:
                    error = "La contraseña actual no es correcta."
                elif password_actual == password_nueva:
                    error = "La nueva contraseña debe ser diferente a la actual."
                else:
                    nuevo_hash = generate_password_hash(password_nueva)
                    q_upd = "UPDATE usuarios SET password_hash = %s, debe_cambiar_password = FALSE WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET password_hash = ?, debe_cambiar_password = 0 WHERE usuario = ?"
                    cursor.execute(q_upd, (nuevo_hash, session.get('username')))
                    conn.commit()
                    session['debe_cambiar_password'] = False
                    registrar_log(session.get('username'), "Cambio de Contraseña", "El usuario cambió su propia contraseña.")
                    conn.close()
                    return redirect(url_for('bienvenida'))
            except Exception as e:
                conn.rollback()
                error = "No se pudo actualizar la contraseña. Intenta de nuevo."
            conn.close()

    return render_template('cambiar_password.html', obligatorio=obligatorio, error=error)


@app.route('/perfil/2fa', methods=['GET'])
@login_required
def perfil_2fa():
    """Página de autoservicio de verificación en dos pasos (2FA). Si la cuenta ya la tiene
    activa, muestra su estado y las opciones de desactivar/regenerar códigos de respaldo. Si no,
    genera (o reutiliza, mientras siga en la sesión) un secreto TOTP pendiente y muestra el
    código QR + el formulario para confirmarlo con el primer código de 6 dígitos."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    q = "SELECT totp_habilitado FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT totp_habilitado FROM usuarios WHERE usuario = ?"
    cursor.execute(q, (session['username'],))
    row = cursor.fetchone()
    conn.close()
    habilitado = bool(row[0]) if row else False

    # 🔒 Hallazgo QA H-05: si validar_instancia_y_sesion mandó para acá porque el rol exige 2FA,
    # se avisa con un banner (en vez de dejar que la persona se pregunte por qué no puede
    # navegar a ningún otro lado). rol_exige_2fa (independiente de si ya está activo en este
    # momento) también oculta la opción de autodesactivarlo para admin/agente: solo un admin
    # puede apagarlo, desde Gestión de Usuarios, para recuperación por pérdida de dispositivo.
    obligatorio = bool(session.get('debe_activar_2fa'))
    rol_exige_2fa = session.get('rol') in ROLES_CON_ACCESO_OPERATIVO

    if habilitado:
        return render_template('perfil_2fa.html', habilitado=True, error=request.args.get('error'), obligatorio=obligatorio, rol_exige_2fa=rol_exige_2fa)

    secreto_pendiente = session.get('totp_pendiente_secret')
    if not secreto_pendiente:
        secreto_pendiente = pyotp.random_base32()
        session['totp_pendiente_secret'] = secreto_pendiente

    otpauth_uri = pyotp.TOTP(secreto_pendiente).provisioning_uri(name=session['username'], issuer_name=NOMBRE_EMISOR_2FA)
    return render_template(
        'perfil_2fa.html', habilitado=False, secreto=secreto_pendiente, otpauth_uri=otpauth_uri,
        error=request.args.get('error'), obligatorio=obligatorio, rol_exige_2fa=rol_exige_2fa
    )


@app.route('/perfil/2fa/confirmar', methods=['POST'])
@login_required
def perfil_2fa_confirmar():
    """Confirma la activación: valida el primer código de 6 dígitos contra el secreto pendiente
    guardado en la sesión y, si coincide, lo persiste en la cuenta y genera los 10 códigos de
    respaldo (que se muestran UNA sola vez, en esta misma respuesta)."""
    secreto_pendiente = session.get('totp_pendiente_secret')
    codigo = re.sub(r'\s+', '', request.form.get('codigo', '') or '')

    if not secreto_pendiente:
        return redirect(url_for('perfil_2fa'))

    if not codigo or not pyotp.TOTP(secreto_pendiente).verify(codigo, valid_window=1):
        return redirect(url_for('perfil_2fa', error="El código ingresado no es válido. Inténtalo de nuevo."))

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        # 🔒 Hallazgo QA H-07: se guarda cifrado con Fernet, nunca en texto plano.
        q_upd = "UPDATE usuarios SET totp_secret = %s, totp_habilitado = TRUE WHERE usuario = %s" if db_type == 'postgres' else "UPDATE usuarios SET totp_secret = ?, totp_habilitado = 1 WHERE usuario = ?"
        cursor.execute(q_upd, (encriptar_texto(secreto_pendiente), session['username']))
        conn.commit()
        conn.close()
    except Exception as e:
        conn.rollback()
        conn.close()
        return redirect(url_for('perfil_2fa', error="No se pudo activar la verificación en dos pasos. Intenta de nuevo."))

    session.pop('totp_pendiente_secret', None)
    # 🔒 Hallazgo QA H-05: ya quedó activo, así que deja de exigirse en cada request.
    session['debe_activar_2fa'] = False
    codigos_respaldo = _generar_codigos_respaldo_2fa(session['username'])
    registrar_log(session['username'], "Activación de 2FA", "El usuario activó la verificación en dos pasos en su cuenta.")
    return render_template('perfil_2fa.html', habilitado=True, codigos_respaldo=codigos_respaldo, mostrar_codigos=True)


@app.route('/perfil/2fa/regenerar_respaldo', methods=['POST'])
@login_required
def perfil_2fa_regenerar_respaldo():
    """Invalida los códigos de respaldo actuales y genera 10 nuevos, mostrados una sola vez.
    Requiere reconfirmar la contraseña actual: son códigos que permiten iniciar sesión sin la
    app autenticadora, así que regenerarlos merece la misma verificación que desactivar el 2FA."""
    password_actual = request.form.get('password_actual') or ''
    conn, db_type = get_db()
    cursor = conn.cursor()
    q_sel = "SELECT password_hash, totp_habilitado FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT password_hash, totp_habilitado FROM usuarios WHERE usuario = ?"
    cursor.execute(q_sel, (session['username'],))
    row = cursor.fetchone()
    conn.close()

    if not row or not bool(row[1]):
        return redirect(url_for('perfil_2fa'))

    clave_db = str(row[0] or '')
    es_valida = check_password_hash(clave_db, password_actual) if (clave_db.startswith('pbkdf2:') or clave_db.startswith('scrypt:')) else hmac.compare_digest(clave_db, password_actual)
    if not es_valida:
        return render_template('perfil_2fa.html', habilitado=True, error="La contraseña ingresada no es correcta.", rol_exige_2fa=session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)

    codigos_respaldo = _generar_codigos_respaldo_2fa(session['username'])
    registrar_log(session['username'], "Regeneración de Códigos de Respaldo 2FA", "El usuario regeneró sus códigos de respaldo del 2FA.")
    return render_template('perfil_2fa.html', habilitado=True, codigos_respaldo=codigos_respaldo, mostrar_codigos=True)


@app.route('/perfil/2fa/desactivar', methods=['POST'])
@login_required
def perfil_2fa_desactivar():
    """Desactiva el 2FA de la propia cuenta. Requiere la contraseña actual: sin esta
    confirmación, cualquiera que dejara una sesión abierta podría apagar la protección sin
    saber la contraseña."""
    password_actual = request.form.get('password_actual') or ''
    conn, db_type = get_db()
    cursor = conn.cursor()
    q_sel = "SELECT password_hash FROM usuarios WHERE usuario = %s" if db_type == 'postgres' else "SELECT password_hash FROM usuarios WHERE usuario = ?"
    cursor.execute(q_sel, (session['username'],))
    row = cursor.fetchone()
    conn.close()

    clave_db = str(row[0] or '') if row else ''
    es_valida = check_password_hash(clave_db, password_actual) if (clave_db.startswith('pbkdf2:') or clave_db.startswith('scrypt:')) else hmac.compare_digest(clave_db, password_actual)

    if not es_valida:
        return render_template('perfil_2fa.html', habilitado=True, error="La contraseña ingresada no es correcta.", rol_exige_2fa=session.get('rol') in ROLES_CON_ACCESO_OPERATIVO)

    _desactivar_2fa_cuenta(session['username'])
    session.pop('totp_pendiente_secret', None)
    # 🔒 Hallazgo QA H-05: si el rol lo exige, apagar el 2FA no da un respiro — la próxima
    # petición ya vuelve a mandar a /perfil/2fa a reactivarlo (ver validar_instancia_y_sesion).
    if session.get('rol') in ROLES_CON_ACCESO_OPERATIVO:
        session['debe_activar_2fa'] = True
    registrar_log(session['username'], "Desactivación de 2FA", "El usuario desactivó la verificación en dos pasos en su cuenta.")
    return redirect(url_for('perfil_2fa'))


@app.route('/logout')
def logout():
    if session.get('username'):
        registrar_log(session['username'], "Cierre de Sesión", "Cierre de sesión de usuario", ip=_obtener_ip_cliente(), dispositivo=_detectar_dispositivo(request.headers.get('User-Agent', '')))
    session.clear()
    return redirect(url_for('login'))


# 🕵️ HISTORIAL DE SESIONES — autoservicio de seguridad: cualquier cuenta logueada ve sus propios
# inicios/cierres de sesión (fecha, IP, dispositivo) para detectar accesos que no reconoce. Un
# admin, además, puede elegir cualquier otra cuenta en el selector (mismo criterio de "ve todo"
# que ya tiene en /logs) — la cuenta 'admin' literal se oculta al resto de admins, igual que en
# gestion_usuarios(). Cualquier otra cuenta solo puede ver la suya: el parámetro ?usuario= se
# ignora por completo si quien consulta no es admin.
ETIQUETAS_ACCION_HISTORIAL_SESIONES = {
    'Inicio de Sesión': {'label': 'Inicio de sesión exitoso', 'icono': 'fa-right-to-bracket', 'color': 'emerald'},
    'Inicio de Sesión Bloqueado': {'label': 'Intento bloqueado (cuenta inactiva)', 'icono': 'fa-ban', 'color': 'rose'},
    'Cierre de Sesión': {'label': 'Cierre de sesión', 'icono': 'fa-arrow-right-from-bracket', 'color': 'slate'},
    'Cuenta Bloqueada por Intentos': {'label': 'Cuenta bloqueada por varios intentos fallidos', 'icono': 'fa-ban', 'color': 'rose'},
    'Desbloqueo por Intentos Fallidos': {'label': 'Un administrador desbloqueó la cuenta', 'icono': 'fa-unlock-keyhole', 'color': 'emerald'},
}

@app.route('/perfil/historial-sesiones')
@login_required
def historial_sesiones():
    usuario_sesion = session.get('username')
    es_admin = session.get('rol') == 'admin'

    conn, db_type = get_db()
    cursor = conn.cursor()

    lista_usuarios = []
    usuario_consultado = usuario_sesion
    if es_admin:
        cursor.execute("SELECT usuario FROM usuarios ORDER BY usuario ASC")
        todos = [u[0] for u in cursor.fetchall() if u[0]]
        lista_usuarios = todos if usuario_sesion == 'admin' else [u for u in todos if u != 'admin']
        usuario_param = request.args.get('usuario', '').strip()
        if usuario_param and (usuario_param in lista_usuarios or usuario_param == usuario_sesion):
            usuario_consultado = usuario_param

    marcador = '%s' if db_type == 'postgres' else '?'
    placeholders_accion = ','.join([marcador] * len(ETIQUETAS_ACCION_HISTORIAL_SESIONES))
    query = f"SELECT usuario, accion, ip, dispositivo, fecha FROM logs WHERE usuario = {marcador} AND accion IN ({placeholders_accion}) ORDER BY id DESC LIMIT 200"
    cursor.execute(query, tuple([usuario_consultado] + list(ETIQUETAS_ACCION_HISTORIAL_SESIONES.keys())))
    filas = cursor.fetchall()
    conn.close()

    eventos = []
    for u_fila, accion, ip, dispositivo, fecha in filas:
        meta = ETIQUETAS_ACCION_HISTORIAL_SESIONES.get(accion, {'label': accion, 'icono': 'fa-circle', 'color': 'slate'})
        eventos.append({
            'accion': accion, 'label': meta['label'], 'icono': meta['icono'], 'color': meta['color'],
            'ip': ip or '—', 'dispositivo': dispositivo or '—', 'fecha': fecha
        })

    return render_template(
        'historial_sesiones.html',
        eventos=eventos,
        es_admin=es_admin,
        usuarios_opt=lista_usuarios,
        usuario_consultado=usuario_consultado,
        es_propio=(usuario_consultado == usuario_sesion)
    )


# 🔔 NOTIFICACIONES (campanita) ------------------------------------------------------------
# Se generan en los mismos puntos donde ya sale un correo (ticket creado, comentado, cambio
# de estado — ver crear_ticket/comentar_ticket/actualizar_ticket) vía crear_notificacion() /
# crear_notificacion_para_varios(). Estas rutas son las que alimenta el ícono de campana que
# aparece en la barra de navegación de toda página autenticada.

@app.route('/notificaciones/resumen')
@login_required
def notificaciones_resumen():
    """JSON con el contador de no leídas y las últimas notificaciones del usuario en sesión —
    consultado periódicamente por la campanita en la barra de navegación.

    🗑️ Acepta ?vista=activas (por defecto) o ?vista=papelera: la campanita reutiliza esta
    misma ruta para pintar tanto la lista principal como la papelera, cambiando solo el
    'estado' que filtra. El contador de no leídas ('no_leidas') SIEMPRE se calcula sobre las
    activas — así el badge no cambia mientras alguien está mirando la papelera."""
    usuario = session.get('username')
    vista = 'papelera' if request.args.get('vista') == 'papelera' else 'activas'
    estado_filtro = 'archivada' if vista == 'papelera' else 'activa'
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_count = "SELECT COUNT(*) FROM notificaciones WHERE usuario = %s AND leida = 0 AND COALESCE(estado, 'activa') = 'activa'" if db_type == 'postgres' else "SELECT COUNT(*) FROM notificaciones WHERE usuario = ? AND leida = 0 AND COALESCE(estado, 'activa') = 'activa'"
        cursor.execute(q_count, (usuario,))
        no_leidas = cursor.fetchone()[0]

        q_lista = ("SELECT id, tipo, mensaje, url, leida, fecha FROM notificaciones WHERE usuario = %s AND COALESCE(estado, 'activa') = %s ORDER BY id DESC LIMIT 30" if db_type == 'postgres'
                   else "SELECT id, tipo, mensaje, url, leida, fecha FROM notificaciones WHERE usuario = ? AND COALESCE(estado, 'activa') = ? ORDER BY id DESC LIMIT 30")
        cursor.execute(q_lista, (usuario, estado_filtro))
        recientes = [{'id': r[0], 'tipo': r[1], 'mensaje': r[2], 'url': r[3], 'leida': bool(r[4]), 'fecha': r[5]} for r in cursor.fetchall()]
        conn.close()
        return {'no_leidas': no_leidas, 'recientes': recientes, 'vista': vista}
    except Exception as e:
        conn.close()
        print(f"⚠️ Error obteniendo resumen de notificaciones de '{usuario}': {e}")
        return {'no_leidas': 0, 'recientes': [], 'vista': vista}


@app.route('/notificaciones/<int:notif_id>/ir')
@login_required
def notificacion_ir(notif_id):
    """La campanita enlaza cada notificación acá en vez de directo a su URL destino: esto
    la marca como leída y de una vez redirige adonde corresponda (el ticket, etc.)."""
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()
    destino = url_for('bienvenida')
    try:
        q_sel = "SELECT usuario, url FROM notificaciones WHERE id = %s" if db_type == 'postgres' else "SELECT usuario, url FROM notificaciones WHERE id = ?"
        cursor.execute(q_sel, (notif_id,))
        row = cursor.fetchone()
        if row and row[0] == usuario:
            destino = row[1] or destino
            q_upd = "UPDATE notificaciones SET leida = 1 WHERE id = %s" if db_type == 'postgres' else "UPDATE notificaciones SET leida = 1 WHERE id = ?"
            cursor.execute(q_upd, (notif_id,))
            conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error marcando notificación {notif_id} como leída: {e}")
    conn.close()
    return redirect(destino)


@app.route('/notificaciones/marcar_todas_leidas', methods=['POST'])
@login_required
def notificaciones_marcar_todas_leidas():
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE notificaciones SET leida = 1 WHERE usuario = %s AND leida = 0" if db_type == 'postgres' else "UPDATE notificaciones SET leida = 1 WHERE usuario = ? AND leida = 0"
        cursor.execute(q, (usuario,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error marcando todas las notificaciones de '{usuario}' como leídas: {e}")
    conn.close()
    return ('', 204)


# 🗑️ PAPELERA DE NOTIFICACIONES ------------------------------------------------------------
# Dos niveles pensados para que nada se borre por accidente: "archivar" es reversible (saca la
# notificación de la lista principal y del badge, pero se conserva en /notificaciones/papelera
# para consultarla o restaurarla después) y "eliminar" es definitivo (solo se puede hacer
# DESDE la papelera, nunca desde la lista principal, como último candado de seguridad).

@app.route('/notificaciones/<int:notif_id>/archivar', methods=['POST'])
@login_required
def notificacion_archivar(notif_id):
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE notificaciones SET estado = 'archivada' WHERE id = %s AND usuario = %s" if db_type == 'postgres' else "UPDATE notificaciones SET estado = 'archivada' WHERE id = ? AND usuario = ?"
        cursor.execute(q, (notif_id, usuario))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error archivando notificación {notif_id} de '{usuario}': {e}")
    conn.close()
    return ('', 204)


@app.route('/notificaciones/<int:notif_id>/restaurar', methods=['POST'])
@login_required
def notificacion_restaurar(notif_id):
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE notificaciones SET estado = 'activa' WHERE id = %s AND usuario = %s" if db_type == 'postgres' else "UPDATE notificaciones SET estado = 'activa' WHERE id = ? AND usuario = ?"
        cursor.execute(q, (notif_id, usuario))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error restaurando notificación {notif_id} de '{usuario}': {e}")
    conn.close()
    return ('', 204)


@app.route('/notificaciones/<int:notif_id>/eliminar', methods=['POST'])
@login_required
def notificacion_eliminar(notif_id):
    """Borrado definitivo — solo permitido si la notificación ya está en la papelera
    ('archivada'), para que nadie borre algo sin verlo primero en la papelera."""
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "DELETE FROM notificaciones WHERE id = %s AND usuario = %s AND estado = 'archivada'" if db_type == 'postgres' else "DELETE FROM notificaciones WHERE id = ? AND usuario = ? AND estado = 'archivada'"
        cursor.execute(q, (notif_id, usuario))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error eliminando notificación {notif_id} de '{usuario}': {e}")
    conn.close()
    return ('', 204)


@app.route('/notificaciones/archivar_todas', methods=['POST'])
@login_required
def notificaciones_archivar_todas():
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "UPDATE notificaciones SET estado = 'archivada' WHERE usuario = %s AND COALESCE(estado, 'activa') = 'activa'" if db_type == 'postgres' else "UPDATE notificaciones SET estado = 'archivada' WHERE usuario = ? AND COALESCE(estado, 'activa') = 'activa'"
        cursor.execute(q, (usuario,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error archivando todas las notificaciones de '{usuario}': {e}")
    conn.close()
    return ('', 204)


@app.route('/notificaciones/papelera/vaciar', methods=['POST'])
@login_required
def notificaciones_papelera_vaciar():
    """Elimina definitivamente TODAS las notificaciones archivadas de la persona en sesión."""
    usuario = session.get('username')
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = "DELETE FROM notificaciones WHERE usuario = %s AND estado = 'archivada'" if db_type == 'postgres' else "DELETE FROM notificaciones WHERE usuario = ? AND estado = 'archivada'"
        cursor.execute(q, (usuario,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error vaciando la papelera de notificaciones de '{usuario}': {e}")
    conn.close()
    return ('', 204)


# 🔍 BUSCADOR GLOBAL — consultado por el modal de búsqueda disponible en la barra de
# navegación de toda la aplicación (ver partials/buscador.html y static/js/buscador.js).
# Cada categoría respeta EXACTAMENTE la misma regla de visibilidad que ya aplica su propia
# página (Solicitudes: solo las propias si no es soporte; Bóveda/Colaboradores/Usuarios: solo
# admin/agente o admin según corresponda; Gestor de Archivos: respeta 'visibilidad'), para que
# el buscador nunca muestre algo que esa cuenta no podría ver entrando al módulo directamente.
LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR = 6

@app.route('/buscar/api')
@login_required
def buscar_global_api():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify({'q': q, 'resultados': []})
    # 🔎 normalizar() quita tildes además de pasar a minúsculas (reutiliza el mismo helper que
    # ya se usa para generar nombres de usuario): así buscar "cedula" encuentra "Cédula" y
    # "razon social" encuentra "razón social", sin obligar a escribir el acento exacto.
    q_norm = normalizar(q)

    usuario = session.get('username')
    rol = session.get('rol')
    es_soporte = rol in ROLES_CON_ACCESO_OPERATIVO
    es_admin = rol == 'admin'
    # 🔁 Certificación de Devoluciones la puede ver también el rol "Gestión Humana", no solo
    # admin/agente — usa su propio set de roles (ROLES_CERTIFICACION_DEVOLUCION) en vez de
    # es_soporte para no ocultársela a Gestión Humana ni mostrársela a quien no debería verla.
    puede_ver_devoluciones = rol in ROLES_CERTIFICACION_DEVOLUCION

    resultados = []
    conn, db_type = get_db()
    cursor = conn.cursor()

    # --- Solicitudes TI (tickets) ---
    try:
        query = "SELECT id, titulo, descripcion, tipo, estado, fecha_creacion FROM tickets WHERE COALESCE(eliminado, 0) = 0"
        params = []
        if not es_soporte:
            query += " AND creado_por = %s" if db_type == 'postgres' else " AND creado_por = ?"
            params.append(usuario)
        query += " ORDER BY id DESC"
        cursor.execute(query, tuple(params))
        contador = 0
        for tid, titulo, descripcion, tipo, estado, fecha_creacion in cursor.fetchall():
            if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                break
            codigo = _codigo_ticket(tipo or 'Incidente', tid, fecha_creacion)
            # 🔎 Además de título/descripción, se puede buscar por el número de ticket: el
            # código completo ('IN-2026-000042'), por similitud (p. ej. '42' encuentra el
            # '000042' dentro del código, o encuentra el id 1042 por contener '42'), o el id
            # exacto — así el botón Buscar también sirve para ubicar un ticket puntual.
            texto_busqueda = normalizar(f"{titulo} {descripcion or ''} {codigo} {tid}")
            if q_norm in texto_busqueda:
                contador += 1
                resultados.append({
                    'categoria': 'Solicitudes TI',
                    'titulo': titulo,
                    'subtitulo': f"{codigo} · {estado}",
                    'url': url_for('ver_ticket', ticket_id=tid)
                })
    except Exception as e:
        print(f"⚠️ Error buscando en tickets (buscador global): {e}")

    # --- Comunicados (visibles para cualquier cuenta logueada) ---
    try:
        cursor.execute("SELECT id, titulo, contenido, nivel FROM comunicados WHERE estado = 'activo' ORDER BY id DESC")
        contador = 0
        for c_id, titulo, contenido, nivel in cursor.fetchall():
            if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                break
            if q_norm in normalizar(f"{titulo} {contenido or ''}"):
                contador += 1
                resultados.append({
                    'categoria': 'Comunicados',
                    'titulo': titulo,
                    'subtitulo': (nivel or '').capitalize(),
                    'url': url_for('ver_comunicados')
                })
    except Exception as e:
        print(f"⚠️ Error buscando en comunicados (buscador global): {e}")

    # --- Base de Conocimiento (visible para cualquier cuenta logueada) ---
    try:
        cursor.execute("SELECT id, titulo, descripcion, url_documento FROM conocimiento_articulos WHERE COALESCE(estado, 'activo') = 'activo' ORDER BY id DESC")
        contador = 0
        for a_id, titulo, descripcion, url_documento in cursor.fetchall():
            if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                break
            if q_norm in normalizar(f"{titulo} {descripcion or ''}"):
                contador += 1
                resultados.append({
                    'categoria': 'Base de Conocimiento',
                    'titulo': titulo,
                    'subtitulo': descripcion or '',
                    'url': url_documento or url_for('ver_conocimiento'),
                    'externo': bool(url_documento)
                })
    except Exception as e:
        print(f"⚠️ Error buscando en base de conocimiento (buscador global): {e}")

    # --- Gestor de Archivos (respeta 'visibilidad': un usuario Estándar no ve lo marcado 'admin') ---
    try:
        cursor.execute("SELECT id, titulo, descripcion, categoria, tags, visibilidad FROM galerias WHERE COALESCE(estado, 'activo') != 'eliminado'")
        contador = 0
        for g_id, titulo, descripcion, categoria, tags, visibilidad in cursor.fetchall():
            if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                break
            if (visibilidad or 'todos') == 'admin' and not es_soporte:
                continue
            if q_norm in normalizar(f"{titulo} {descripcion or ''} {tags or ''}"):
                contador += 1
                resultados.append({
                    'categoria': 'Gestor de Archivos',
                    'titulo': titulo,
                    'subtitulo': categoria or 'General',
                    'url': url_for('index', q=titulo)
                })
    except Exception as e:
        print(f"⚠️ Error buscando en galerías (buscador global): {e}")

    # --- Bóveda de Accesos (solo admin/agente — NUNCA se busca ni se expone la contraseña) ---
    if es_soporte:
        try:
            cursor.execute("SELECT id, titulo, usuario_acceso, area, notas FROM credenciales WHERE COALESCE(estado, 'activo') != 'eliminado'")
            contador = 0
            for c_id, titulo, usuario_acceso, area, notas in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                if q_norm in normalizar(f"{titulo} {usuario_acceso or ''} {area or ''} {notas or ''}"):
                    contador += 1
                    resultados.append({
                        'categoria': 'Bóveda de Accesos',
                        'titulo': titulo,
                        'subtitulo': area or 'General',
                        'url': url_for('ver_credenciales', q=titulo)
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en credenciales (buscador global): {e}")

        # --- Accesos de Colaboradores (registro de aplicativos entregados, solo admin/agente) ---
        try:
            cursor.execute("SELECT id, colaborador, aplicativo, solicitado_por FROM credenciales_colaboradores ORDER BY id DESC")
            contador = 0
            for r_id, colaborador, aplicativo, solicitado_por in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                if q_norm in normalizar(f"{colaborador} {aplicativo} {solicitado_por or ''}"):
                    contador += 1
                    resultados.append({
                        'categoria': 'Accesos de Colaboradores',
                        'titulo': colaborador,
                        'subtitulo': aplicativo,
                        'url': url_for('ver_credenciales_colaboradores', q=colaborador)
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en credenciales de colaboradores (buscador global): {e}")

        # --- Inventario de Activos (solo admin/agente): placa, tipo, marca, modelo, serie,
        # asignado a, sede, área y proveedor — todo lo que se puede filtrar/editar hoy en
        # /tickets/inventario queda cubierto aquí para que el buscador global sirva para
        # encontrar un equipo puntual sin tener que entrar al módulo primero.
        try:
            cursor.execute("SELECT id, nombre, tipo_activo, marca, modelo, numero_serie, estado, asignado_a, sede, area, proveedor FROM activos_inventario WHERE eliminado = 0 ORDER BY id DESC")
            contador = 0
            for (a_id, a_nombre, a_tipo, a_marca, a_modelo, a_serie, a_estado, a_asignado, a_sede, a_area, a_proveedor) in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                texto_busqueda = normalizar(f"{a_nombre} {a_tipo or ''} {a_marca or ''} {a_modelo or ''} {a_serie or ''} {a_estado or ''} {a_asignado or ''} {a_sede or ''} {a_area or ''} {a_proveedor or ''}")
                if q_norm in texto_busqueda:
                    contador += 1
                    subtitulo = ' · '.join([p for p in [a_tipo, a_marca, a_estado] if p])
                    resultados.append({
                        'categoria': 'Inventario de Activos',
                        'titulo': f"Placa {a_nombre}",
                        'subtitulo': subtitulo,
                        'url': url_for('ver_inventario', q=a_nombre)
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en inventario de activos (buscador global): {e}")

        # --- Proveedores (catálogo de /tickets/configuracion, solo admin/agente): busca por
        # nombre comercial, NIT o razón social. Antes de este bloque el buscador global no
        # cubría este catálogo en absoluto, así que escribir el NIT de un proveedor (que no
        # aparece en ningún otro texto del sistema) no encontraba nada aunque el proveedor sí
        # estuviera registrado — este bloque es lo que lo soluciona.
        try:
            q_proveedores_sql = ("SELECT id, nombre, nit, razon_social FROM ticket_configuraciones WHERE tipo = %s AND estado = 'activo' ORDER BY nombre ASC"
                                  if db_type == 'postgres' else
                                  "SELECT id, nombre, nit, razon_social FROM ticket_configuraciones WHERE tipo = ? AND estado = 'activo' ORDER BY nombre ASC")
            cursor.execute(q_proveedores_sql, ('proveedor',))
            contador = 0
            for (p_id, p_nombre, p_nit, p_razon_social) in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                texto_busqueda = normalizar(f"{p_nombre} {p_nit or ''} {p_razon_social or ''}")
                if q_norm in texto_busqueda:
                    contador += 1
                    subtitulo = ' · '.join([p for p in [f"NIT {p_nit}" if p_nit else None, p_razon_social] if p])
                    resultados.append({
                        'categoria': 'Proveedores',
                        'titulo': p_nombre,
                        'subtitulo': subtitulo,
                        'url': url_for('configuracion_tickets')
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en proveedores (buscador global): {e}")

        # --- Plantillas de Solicitud (solo admin/agente): agilizan crear tickets recurrentes;
        # se busca por nombre de la plantilla, tipo, categoría, título o descripción prellenada.
        try:
            cursor.execute("SELECT id, nombre, tipo, categoria, titulo, descripcion FROM ticket_plantillas WHERE estado = 'activo' ORDER BY nombre ASC")
            contador = 0
            for (pl_id, pl_nombre, pl_tipo, pl_categoria, pl_titulo, pl_descripcion) in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                texto_busqueda = normalizar(f"{pl_nombre} {pl_tipo or ''} {pl_categoria or ''} {pl_titulo or ''} {pl_descripcion or ''}")
                if q_norm in texto_busqueda:
                    contador += 1
                    resultados.append({
                        'categoria': 'Plantillas de Solicitud',
                        'titulo': pl_nombre,
                        'subtitulo': ' · '.join([p for p in [pl_tipo, pl_categoria] if p]),
                        'url': url_for('ver_plantillas_ticket')
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en plantillas de solicitud (buscador global): {e}")

        # --- Documentos de empleado con vencimiento (panel de Vencimiento de Documentos, solo
        # admin/agente): busca por título del documento, tipo de documento o el empleado dueño.
        # Los instructivos institucionales con vencimiento no se repiten aquí porque ya son
        # 'galerias' y quedan cubiertos por la categoría "Gestor de Archivos" de arriba.
        try:
            cursor.execute(
                "SELECT d.titulo, d.tipo_documento, d.usuario, d.fecha_vencimiento, COALESCE(u.nombre, d.usuario) "
                "FROM documentos_empleado d LEFT JOIN usuarios u ON u.usuario = d.usuario "
                "WHERE COALESCE(d.estado, 'activo') = 'activo'"
            )
            contador = 0
            for (dc_titulo, dc_tipo, dc_usuario, dc_fecha_venc, dc_nombre) in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                texto_busqueda = normalizar(f"{dc_titulo} {dc_tipo or ''} {dc_usuario or ''} {dc_nombre or ''}")
                if q_norm in texto_busqueda:
                    contador += 1
                    subtitulo = ' · '.join([p for p in [dc_tipo, dc_nombre or dc_usuario, f"Vence {dc_fecha_venc}" if dc_fecha_venc else None] if p])
                    resultados.append({
                        'categoria': 'Vencimiento de Documentos',
                        'titulo': dc_titulo,
                        'subtitulo': subtitulo,
                        'url': url_for('ver_vencimientos')
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en documentos de empleado (buscador global): {e}")

    # --- Certificación de Devoluciones (admin/agente/gestión humana): busca en el historial de
    # devoluciones ya confirmadas por colaborador, activo devuelto, quién la confirmó o las
    # observaciones dejadas — usa su propio permiso porque Gestión Humana entra aquí aunque no
    # sea 'soporte' (no ve tickets, inventario, etc.).
    if puede_ver_devoluciones:
        try:
            cursor.execute("""SELECT d.colaborador, a.nombre, a.tipo_activo, d.confirmado_por, d.observaciones
                               FROM inventario_devoluciones d JOIN activos_inventario a ON a.id = d.activo_id
                               ORDER BY d.id DESC""")
            contador = 0
            for (dv_colaborador, dv_activo_nombre, dv_tipo, dv_confirmado_por, dv_observaciones) in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                texto_busqueda = normalizar(f"{dv_colaborador} {dv_activo_nombre or ''} {dv_tipo or ''} {dv_confirmado_por or ''} {dv_observaciones or ''}")
                if q_norm in texto_busqueda:
                    contador += 1
                    resultados.append({
                        'categoria': 'Certificación de Devoluciones',
                        'titulo': f"Devolución de {dv_colaborador}",
                        'subtitulo': f"Placa {dv_activo_nombre}" if dv_activo_nombre else dv_tipo,
                        'url': url_for('certificacion_devoluciones', q=dv_colaborador)
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en certificación de devoluciones (buscador global): {e}")

    # --- Usuarios (solo rol admin — y la cuenta 'admin' literal se oculta al resto de admins) ---
    if es_admin:
        try:
            cursor.execute("SELECT usuario, nombre, correo, cedula FROM usuarios")
            contador = 0
            for u_usuario, u_nombre, u_correo, u_cedula in cursor.fetchall():
                if contador >= LIMITE_RESULTADOS_POR_CATEGORIA_BUSCADOR:
                    break
                if u_usuario == 'admin' and usuario != 'admin':
                    continue
                if q_norm in normalizar(f"{u_usuario} {u_nombre or ''} {u_correo or ''} {u_cedula or ''}"):
                    contador += 1
                    resultados.append({
                        'categoria': 'Usuarios',
                        'titulo': u_nombre or u_usuario,
                        'subtitulo': u_correo or u_usuario,
                        'url': url_for('gestion_usuarios')
                    })
        except Exception as e:
            print(f"⚠️ Error buscando en usuarios (buscador global): {e}")

    conn.close()
    return jsonify({'q': q, 'resultados': resultados})


@app.route('/')
def home():
    return redirect(url_for('bienvenida')) if session.get('logged_in') else redirect(url_for('login'))

@app.route('/bienvenida')
@login_required
def bienvenida():
    conn, db_type = get_db()
    cursor = conn.cursor()
    comunicado_fijado = None
    try:
        # "fijado = 1" literal fallaba en Postgres contra la columna BOOLEAN real; "true" funciona en ambos motores.
        query_fij = "SELECT id, titulo, contenido, nivel, imagen_url, fecha, autor FROM comunicados WHERE fijado = true AND estado = 'activo' ORDER BY id DESC LIMIT 1" if db_type == 'postgres' else "SELECT id, titulo, contenido, nivel, imagen_url, fecha, autor FROM comunicados WHERE fijado = 1 AND estado = 'activo' ORDER BY id DESC LIMIT 1"
        cursor.execute(query_fij)
        row = cursor.fetchone()
        if row:
            comunicado_fijado = {
                'id': row[0],
                'titulo': row[1],
                'contenido': row[2],
                'nivel': row[3],
                'imagen_url': row[4],
                'fecha': row[5],
                # 👤 Alias/nombre real de quien publicó, no su usuario de inicio de sesión.
                'autor': _nombre_para_mostrar(row[6], _mapa_nombres_usuarios())
            }
    except Exception:
        comunicado_fijado = None
    conn.close()

    # 👁️ Ver el comunicado fijado en la bienvenida también cuenta como "leído".
    if comunicado_fijado:
        _marcar_comunicado_leido(comunicado_fijado['id'], session.get('username'))

    return render_template('bienvenida.html', username=session.get('username'), rol=session.get('rol'), comunicado_fijado=comunicado_fijado)

@app.route('/gestor')
@login_required
def index():
    busqueda_raw = request.args.get('q', '').strip()
    cat_filtro = request.args.get('cat', '').strip()
    tipo_filtro = request.args.get('tipo', '').strip()
    formato_filtro = request.args.get('formato', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()

    # 👁️ Un usuario Estándar solo debe ver instructivos marcados como 'todos'; Admin y Agente
    # ven absolutamente todo (incluidos los marcados 'admin'), sin filtrar por visibilidad.
    puede_ver_todo = session.get('rol') in ROLES_CON_ACCESO_OPERATIVO

    # 🔔 Aviso perezoso de vencimiento de documentos — ver _revisar_alertas_vencimientos().
    # Solo tiene sentido para quien ya puede gestionar instructivos (evita que un usuario
    # Estándar dispare envíos de correo con cada visita al listado).
    if puede_ver_todo:
        _revisar_alertas_vencimientos()

    try:
        cursor.execute("SELECT id, titulo, descripcion, fecha_subida, categoria, tipo, tags, vistas, descargas, visibilidad, fecha_vencimiento FROM galerias WHERE COALESCE(estado, 'activo') != 'eliminado'")
        rows = cursor.fetchall()
    except Exception:
        try:
            conn.rollback()
            cursor.execute("SELECT id, titulo, descripcion, fecha_subida, categoria, tipo, tags FROM galerias WHERE COALESCE(estado, 'activo') != 'eliminado'")
            raw_rows = cursor.fetchall()
            rows = [r + (0, 0, 'todos', None) for r in raw_rows]
        except Exception:
            rows = []

    galerias = []
    sugerencias_titulos = []
    fecha_defecto = obtener_fecha_actual()

    STOP_WORDS = {'de', 'del', 'la', 'las', 'el', 'los', 'un', 'una', 'unos', 'unas', 'y', 'e', 'o', 'u', 'a', 'en', 'con', 'por', 'para'}

    palabras_clave = []
    if busqueda_raw:
        palabras_limpias = [normalizar(p) for p in busqueda_raw.split() if normalizar(p)]
        palabras_clave = [p for p in palabras_limpias if p not in STOP_WORDS]
        if not palabras_clave:
            palabras_clave = palabras_limpias

    for r in rows:
        galeria_id, titulo, descripcion, fecha = r[0], r[1], r[2], r[3]
        categoria = r[4] if len(r) > 4 and r[4] else 'General'
        tipo = r[5] if len(r) > 5 and r[5] else 'Instructivo'
        tags = r[6] if len(r) > 6 and r[6] else ''
        vistas = r[7] if len(r) > 7 and r[7] is not None else 0
        descargas = r[8] if len(r) > 8 and r[8] is not None else 0
        visibilidad = r[9] if len(r) > 9 and r[9] else 'todos'
        fecha_vencimiento = r[10] if len(r) > 10 else None

        # 👁️ Si es un usuario Estándar y este instructivo quedó marcado "Solo Admin", ni
        # siquiera entra a construir el item: no debe aparecer en su listado ni en las
        # sugerencias de búsqueda.
        if visibilidad == 'admin' and not puede_ver_todo:
            continue

        sugerencias_titulos.append(titulo)

        try:
            query_arch = "SELECT COALESCE(filename, url_archivo) FROM archivos WHERE galeria_id = %s AND COALESCE(estado, 'activo') != 'eliminado'" if db_type == 'postgres' else "SELECT COALESCE(filename, url_archivo) FROM archivos WHERE galeria_id = ? AND COALESCE(estado, 'activo') != 'eliminado'"
            cursor.execute(query_arch, (galeria_id,))
            archivos = [f[0] for f in cursor.fetchall()]
        except Exception as e_arch:
            # No dejar que un problema puntual (ej. tipos de dato inconsistentes
            # en galeria_id) tumbe toda la página de instructivos.
            print(f"⚠️ Error leyendo archivos de la galería {galeria_id}: {e_arch}")
            if db_type == 'postgres':
                conn.rollback()
            archivos = []

        item = {
            'id': galeria_id,
            'titulo': titulo,
            'descripcion': descripcion,
            'fecha': fecha or fecha_defecto,
            'categoria': categoria,
            'tipo': tipo,
            'tags': tags,
            'vistas': vistas,
            'descargas': descargas,
            'visibilidad': visibilidad,
            'fecha_vencimiento': fecha_vencimiento or '',
            'vencimiento_estado': _bucket_vencimiento(fecha_vencimiento),
            'archivos': archivos
        }

        texto_busqueda = normalizar(f"{titulo} {descripcion} {categoria} {tipo} {tags} {' '.join(archivos)}")

        if palabras_clave:
            coincide_busqueda = any(palabra in texto_busqueda for palabra in palabras_clave)
        else:
            coincide_busqueda = True

        coincide_cat = not cat_filtro or categoria == cat_filtro
        coincide_tipo = not tipo_filtro or tipo == tipo_filtro

        coincide_formato = True
        if formato_filtro == 'imagen':
            coincide_formato = any(any(ext in a.lower() for ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp']) or '/image/upload/' in a for a in archivos)
        elif formato_filtro == 'video':
            coincide_formato = any(any(ext in a.lower() for ext in ['.mp4', '.mov', '.webm', '.avi']) or '/video/upload/' in a for a in archivos)
        elif formato_filtro == 'pdf':
            coincide_formato = any('.pdf' in a.lower() or '.docx' in a.lower() or '.txt' in a.lower() or '/raw/upload/' in a for a in archivos)

        if coincide_busqueda and coincide_cat and coincide_tipo and coincide_formato:
            galerias.append(item)

    conn.close()
    return render_template('index.html', galerias=galerias, busqueda=busqueda_raw, cat_filtro=cat_filtro, tipo_filtro=tipo_filtro, formato_filtro=formato_filtro, sugerencias_titulos=list(set(sugerencias_titulos)), rol=session.get('rol'))

# 📦 SUBIDA DE ARCHIVOS (IMÁGENES, VIDEOS, DOCUMENTOS Y COMPRIMIDOS .ZIP/.RAR)
@app.route('/subir', methods=['POST'])
@login_required
@agente_o_admin_required
def subir_archivo():
    archivos = request.files.getlist('archivo')
    titulo = request.form.get('titulo', 'Sin título')
    descripcion = request.form.get('descripcion', '')
    categoria = request.form.get('categoria', 'General')
    tipo = request.form.get('tipo', 'Instructivo')
    tags = request.form.get('tags', '')
    # 👁️ Visibilidad del instructivo: 'todos' (cualquier usuario logueado) o 'admin' (solo
    # Admin/Agente). Cualquier otro valor recibido se descarta a favor de 'todos'.
    visibilidad = request.form.get('visibilidad', 'todos').strip()
    if visibilidad not in ('todos', 'admin'):
        visibilidad = 'todos'
    # 📅 Vencimiento (opcional): fecha a partir de la cual este instructivo se considera
    # vencido. Vacío = sin vencimiento (comportamiento igual al de siempre).
    fecha_vencimiento = (request.form.get('fecha_vencimiento') or '').strip() or None

    galeria_id = str(uuid.uuid4())[:8]
    fecha_actual = obtener_fecha_actual()
    
    archivos_guardados = []
    for file in archivos:
        if file and archivo_permitido(file.filename):
            try:
                ext = file.filename.rsplit('.', 1)[1].lower()

                if ext in ['mp4', 'mov', 'webm', 'avi']:
                    # 🛡️ Los videos van por upload_large (subida en fragmentos/"chunks"): el
                    # endpoint normal de Cloudinary (upload) rechaza con 413 "Request Entity
                    # Too Large" cualquier archivo que pase de ~100 MB, sin importar el plan
                    # de la cuenta. upload_large sube el archivo en pedazos y evita ese límite.
                    # ⚠️ Se le pasa file.stream (no el FileStorage de Flask/Werkzeug): upload_large
                    # hace "with file_io:" internamente, y FileStorage no implementa el protocolo
                    # de gestor de contexto (__enter__/__exit__) — el stream real sí. "filename"
                    # se pasa explícito porque el stream no trae el nombre original del archivo.
                    upload_result = cloudinary.uploader.upload_large(
                        file.stream,
                        resource_type="video",
                        filename=file.filename,
                        use_filename=True,
                        unique_filename=True,
                        chunk_size=6000000,
                        timeout=600
                    )
                elif ext == 'pdf':
                    upload_result = cloudinary.uploader.upload(
                        file,
                        resource_type="image",
                        format="pdf",
                        use_filename=True,
                        unique_filename=True,
                        timeout=60
                    )
                elif ext in ['zip', 'rar', '7z', 'tar', 'gz', 'txt', 'docx', 'xlsx', 'pptx']:
                    # 🛡️ Mismo límite de ~100 MB aplica a archivos "raw" (zip, comprimidos, etc.):
                    # se sube por chunks para evitar el mismo 413 con paquetes grandes.
                    upload_result = cloudinary.uploader.upload_large(
                        file.stream,
                        resource_type="raw",
                        filename=file.filename,
                        use_filename=True,
                        unique_filename=True,
                        chunk_size=6000000,
                        timeout=600
                    )
                else:
                    upload_result = cloudinary.uploader.upload(
                        file,
                        resource_type="image",
                        use_filename=True,
                        unique_filename=True,
                        timeout=60
                    )

                archivos_guardados.append((upload_result['secure_url'], file.filename))
            except Exception as e:
                # No dejar que un archivo con problema (ej. Cloudinary rechazándolo,
                # o sin credenciales configuradas) tumbe la subida completa del instructivo.
                print(f"⚠️ Error subiendo el archivo '{file.filename}' a Cloudinary: {e}")

    if archivos_guardados:
        try:
            conn, db_type = get_db()
            cursor = conn.cursor()
            # 'area' es NOT NULL en Neon sin valor por defecto: reutilizamos la categoría
            # elegida, ya que hoy no hay un campo separado de área en el formulario.
            q_galeria = "INSERT INTO galerias (id, titulo, descripcion, fecha_subida, categoria, area, tipo, tags, vistas, descargas, estado, visibilidad, fecha_vencimiento) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0, 0, 'activo', %s, %s)" if db_type == 'postgres' else "INSERT INTO galerias (id, titulo, descripcion, fecha_subida, categoria, area, tipo, tags, vistas, descargas, estado, visibilidad, fecha_vencimiento) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 'activo', ?, ?)"
            cursor.execute(q_galeria, (galeria_id, titulo, descripcion, fecha_actual, categoria, categoria, tipo, tags, visibilidad, fecha_vencimiento))

            # 'nombre_original' y 'url_archivo' son NOT NULL en Neon sin valor por defecto.
            # Se llenan junto con 'filename' (que se conserva por compatibilidad con lecturas existentes).
            q_archivo = "INSERT INTO archivos (galeria_id, filename, url_archivo, nombre_original, estado) VALUES (%s, %s, %s, %s, 'activo')" if db_type == 'postgres' else "INSERT INTO archivos (galeria_id, filename, url_archivo, nombre_original, estado) VALUES (?, ?, ?, ?, 'activo')"
            for url_arch, nombre_orig in archivos_guardados:
                cursor.execute(q_archivo, (galeria_id, url_arch, url_arch, nombre_orig))

            conn.commit()
            conn.close()
            registrar_log(session['username'], "Creación de Instructivo", f"Instructivo '{titulo}' [{categoria} / {tipo}] (visibilidad: {'solo Admin/Agente' if visibilidad == 'admin' else 'todos los usuarios'})")
        except Exception as e:
            print(f"⚠️ Error guardando el instructivo '{titulo}' en la base de datos: {e}")

    return redirect(url_for('index'))

@app.route('/editar_galeria/<galeria_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def editar_galeria(galeria_id):
    nuevo_titulo = (request.form.get('titulo') or '').strip()
    nueva_desc = (request.form.get('descripcion') or '').strip()
    nueva_cat = (request.form.get('categoria') or 'General').strip()
    nuevo_tipo = (request.form.get('tipo') or 'Instructivo').strip()
    nuevos_tags = (request.form.get('tags') or '').strip()
    nueva_visibilidad = (request.form.get('visibilidad') or 'todos').strip()
    if nueva_visibilidad not in ('todos', 'admin'):
        nueva_visibilidad = 'todos'
    # 📅 Vencimiento (opcional): vacío = sin vencimiento.
    nueva_fecha_vencimiento = (request.form.get('fecha_vencimiento') or '').strip() or None
    nuevos_archivos = request.files.getlist('nuevos_archivos')

    conn, db_type = get_db()
    cursor = conn.cursor()

    try:
        q_sel = "SELECT titulo, descripcion, categoria, tipo, tags, visibilidad, fecha_vencimiento FROM galerias WHERE id = %s" if db_type == 'postgres' else "SELECT titulo, descripcion, categoria, tipo, tags, visibilidad, fecha_vencimiento FROM galerias WHERE id = ?"
        cursor.execute(q_sel, (galeria_id,))
        antiguo = cursor.fetchone()

        cambios = []
        fecha_vencimiento_cambio = False
        if antiguo:
            tit_old = (antiguo[0] or '').strip()
            desc_old = (antiguo[1] or '').strip()
            cat_old = (antiguo[2] or 'General').strip()
            tipo_old = (antiguo[3] or 'Instructivo').strip()
            tags_old = (antiguo[4] or '').strip()
            visibilidad_old = (antiguo[5] or 'todos').strip()
            venc_old = (antiguo[6] or '').strip() or None

            if tit_old != nuevo_titulo:
                cambios.append(f"Título: '{tit_old}' ➔ '{nuevo_titulo}'")
            if desc_old != nueva_desc:
                cambios.append(f"Descripción: '{desc_old}' ➔ '{nueva_desc}'")
            if cat_old != nueva_cat:
                cambios.append(f"Categoría: '{cat_old}' ➔ '{nueva_cat}'")
            if tipo_old != nuevo_tipo:
                cambios.append(f"Tipo: '{tipo_old}' ➔ '{nuevo_tipo}'")
            if tags_old != nuevos_tags:
                cambios.append(f"Tags: '{tags_old}' ➔ '{nuevos_tags}'")
            if visibilidad_old != nueva_visibilidad:
                cambios.append(f"Visibilidad: '{visibilidad_old}' ➔ '{nueva_visibilidad}'")
            if venc_old != nueva_fecha_vencimiento:
                cambios.append(f"Vencimiento: '{venc_old or 'Sin definir'}' ➔ '{nueva_fecha_vencimiento or 'Sin definir'}'")
                fecha_vencimiento_cambio = True

        # 🔔 Si cambió la fecha de vencimiento, se limpia el nivel de alerta ya avisado para
        # que _revisar_alertas_vencimientos() pueda volver a avisar si la nueva fecha se
        # acerca otra vez — mismo patrón que extender_sla_ticket con sla_alerta_nivel.
        if fecha_vencimiento_cambio:
            q_upd = "UPDATE galerias SET titulo = %s, descripcion = %s, categoria = %s, tipo = %s, tags = %s, visibilidad = %s, fecha_vencimiento = %s, alerta_vencimiento_nivel = NULL WHERE id = %s" if db_type == 'postgres' else "UPDATE galerias SET titulo = ?, descripcion = ?, categoria = ?, tipo = ?, tags = ?, visibilidad = ?, fecha_vencimiento = ?, alerta_vencimiento_nivel = NULL WHERE id = ?"
        else:
            q_upd = "UPDATE galerias SET titulo = %s, descripcion = %s, categoria = %s, tipo = %s, tags = %s, visibilidad = %s, fecha_vencimiento = %s WHERE id = %s" if db_type == 'postgres' else "UPDATE galerias SET titulo = ?, descripcion = ?, categoria = ?, tipo = ?, tags = ?, visibilidad = ?, fecha_vencimiento = ? WHERE id = ?"
        cursor.execute(q_upd, (nuevo_titulo, nueva_desc, nueva_cat, nuevo_tipo, nuevos_tags, nueva_visibilidad, nueva_fecha_vencimiento, galeria_id))
        
        archivos_agregados = 0
        for file in nuevos_archivos:
            if file and archivo_permitido(file.filename):
                ext = file.filename.rsplit('.', 1)[1].lower()
                
                if ext in ['mp4', 'mov', 'webm', 'avi']:
                    # 🛡️ Igual que en /subir: upload_large evita el 413 de Cloudinary en
                    # archivos de más de ~100 MB, subiéndolos en fragmentos. Se pasa file.stream
                    # (no el FileStorage) porque upload_large necesita un objeto que soporte
                    # "with ... :", y FileStorage no lo soporta; "filename" se pasa explícito
                    # porque el stream no trae el nombre original del archivo.
                    upload_result = cloudinary.uploader.upload_large(
                        file.stream,
                        resource_type="video",
                        filename=file.filename,
                        use_filename=True,
                        unique_filename=True,
                        chunk_size=6000000,
                        timeout=600
                    )
                elif ext == 'pdf':
                    upload_result = cloudinary.uploader.upload(
                        file,
                        resource_type="image",
                        format="pdf",
                        use_filename=True,
                        unique_filename=True,
                        timeout=60
                    )
                elif ext in ['zip', 'rar', '7z', 'tar', 'gz', 'txt', 'docx', 'xlsx', 'pptx']:
                    upload_result = cloudinary.uploader.upload_large(
                        file.stream,
                        resource_type="raw",
                        filename=file.filename,
                        use_filename=True,
                        unique_filename=True,
                        chunk_size=6000000,
                        timeout=600
                    )
                else:
                    upload_result = cloudinary.uploader.upload(
                        file,
                        resource_type="image",
                        use_filename=True,
                        unique_filename=True,
                        timeout=60
                    )

                # 'nombre_original' y 'url_archivo' son NOT NULL en Neon sin valor por defecto.
                q_ins_arch = "INSERT INTO archivos (galeria_id, filename, url_archivo, nombre_original, estado) VALUES (%s, %s, %s, %s, 'activo')" if db_type == 'postgres' else "INSERT INTO archivos (galeria_id, filename, url_archivo, nombre_original, estado) VALUES (?, ?, ?, ?, 'activo')"
                cursor.execute(q_ins_arch, (galeria_id, upload_result['secure_url'], upload_result['secure_url'], file.filename))
                archivos_agregados += 1

        if archivos_agregados > 0:
            cambios.append(f"Archivos: +{archivos_agregados} nuevo(s)")

        conn.commit()

        if cambios:
            detalles_log = f"'{nuevo_titulo}' :: " + " | ".join(cambios)
        else:
            detalles_log = f"'{nuevo_titulo}' re-guardado sin cambios detectados"

        registrar_log(session['username'], "Edición de Galería", detalles_log)

    except Exception as e:
        conn.rollback()
        print(f"Error procesando edición en BD: {e}")

    conn.close()
    return redirect(url_for('index'))

# 🗑️ BORRADO LÓGICO DE INSTRUCTIVO
@app.route('/eliminar_galeria/<galeria_id>', methods=['POST'])
@login_required
@agente_o_admin_required
def eliminar_galeria(galeria_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT titulo FROM galerias WHERE id = %s" if db_type == 'postgres' else "SELECT titulo FROM galerias WHERE id = ?"
        cursor.execute(q_sel, (galeria_id,))
        row = cursor.fetchone()
        titulo = row[0] if row else galeria_id

        q_upd = "UPDATE galerias SET estado = 'eliminado' WHERE id = %s" if db_type == 'postgres' else "UPDATE galerias SET estado = 'eliminado' WHERE id = ?"
        cursor.execute(q_upd, (galeria_id,))
        conn.commit()

        registrar_log(session['username'], "Envío a Papelera", f"El instructivo '{titulo}' fue movido a la papelera de reciclaje.")
    except Exception as e:
        conn.rollback()

    conn.close()
    return redirect(url_for('index'))


# 📅 VENCIMIENTO DE DOCUMENTOS -------------------------------------------------------------
# Panel de reporte/filtro consolidado (institucionales + por empleado) y la gestión de
# documentos por empleado (subir/editar/eliminar), enlazada desde Gestión de Usuarios.

@app.route('/vencimientos')
@login_required
@agente_o_admin_required
def ver_vencimientos():
    """Panel consolidado de vencimiento de documentos: instructivos institucionales
    ('galerias') y documentos por empleado ('documentos_empleado') en una sola vista,
    filtrable por origen y por estado (vencido / próximo a vencer / vigente)."""
    _revisar_alertas_vencimientos()

    filtro_origen = request.args.get('origen', '').strip()
    filtro_estado = request.args.get('estado', '').strip()

    conn, db_type = get_db()
    cursor = conn.cursor()
    items = []

    try:
        cursor.execute(
            "SELECT id, titulo, fecha_vencimiento FROM galerias "
            "WHERE COALESCE(estado, 'activo') != 'eliminado' AND fecha_vencimiento IS NOT NULL AND fecha_vencimiento != ''"
        )
        for galeria_id, titulo, fecha_venc in cursor.fetchall():
            items.append({
                'origen': 'institucional', 'origen_label': 'Instructivo institucional',
                'titulo': titulo, 'detalle': '', 'fecha_vencimiento': fecha_venc,
                'estado': _bucket_vencimiento(fecha_venc) or 'vigente',
                'url': url_for('index'),
            })
    except Exception as e:
        print(f"⚠️ Error listando galerías para el panel de vencimientos: {e}")

    try:
        cursor.execute(
            "SELECT d.id, d.titulo, d.tipo_documento, d.usuario, d.fecha_vencimiento, COALESCE(u.nombre, d.usuario) "
            "FROM documentos_empleado d LEFT JOIN usuarios u ON u.usuario = d.usuario "
            "WHERE COALESCE(d.estado, 'activo') = 'activo' AND d.fecha_vencimiento IS NOT NULL AND d.fecha_vencimiento != ''"
        )
        for doc_id, titulo, tipo_doc, usuario_doc, fecha_venc, nombre_doc in cursor.fetchall():
            items.append({
                'origen': 'empleado', 'origen_label': 'Documento de empleado',
                'titulo': titulo, 'detalle': f"{tipo_doc} · {nombre_doc or usuario_doc}",
                'fecha_vencimiento': fecha_venc,
                'estado': _bucket_vencimiento(fecha_venc) or 'vigente',
                'url': url_for('gestion_usuarios'),
            })
    except Exception as e:
        print(f"⚠️ Error listando documentos de empleado para el panel de vencimientos: {e}")

    conn.close()

    if filtro_origen in ('institucional', 'empleado'):
        items = [i for i in items if i['origen'] == filtro_origen]
    if filtro_estado in ('vencido', 'proximo_a_vencer', 'vigente'):
        items = [i for i in items if i['estado'] == filtro_estado]

    orden_estado = {'vencido': 0, 'proximo_a_vencer': 1, 'vigente': 2}
    items.sort(key=lambda i: (orden_estado.get(i['estado'], 3), i['fecha_vencimiento']))

    total_vencidos = sum(1 for i in items if i['estado'] == 'vencido')
    total_proximos = sum(1 for i in items if i['estado'] == 'proximo_a_vencer')

    return render_template(
        'vencimientos.html', items=items, filtro_origen=filtro_origen, filtro_estado=filtro_estado,
        total_vencidos=total_vencidos, total_proximos=total_proximos, rol=session.get('rol')
    )


@app.route('/usuarios/<usuario>/documentos')
@login_required
@admin_required
def documentos_empleado_listar(usuario):
    """JSON con los documentos (no eliminados) de un empleado puntual — usado por el modal
    de 'Documentos' en Gestión de Usuarios para pintar la lista sin recargar la página."""
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q = ("SELECT id, tipo_documento, titulo, descripcion, url_archivo, nombre_original, fecha_emision, fecha_vencimiento FROM documentos_empleado "
             "WHERE usuario = %s AND COALESCE(estado, 'activo') = 'activo' ORDER BY id DESC") if db_type == 'postgres' else \
            ("SELECT id, tipo_documento, titulo, descripcion, url_archivo, nombre_original, fecha_emision, fecha_vencimiento FROM documentos_empleado "
             "WHERE usuario = ? AND COALESCE(estado, 'activo') = 'activo' ORDER BY id DESC")
        cursor.execute(q, (usuario,))
        docs = [{
            'id': r[0], 'tipo_documento': r[1], 'titulo': r[2], 'descripcion': r[3] or '',
            'url_archivo': r[4] or '', 'nombre_original': r[5] or '', 'fecha_emision': r[6] or '',
            'fecha_vencimiento': r[7] or '', 'estado_vencimiento': _bucket_vencimiento(r[7]) or 'vigente',
        } for r in cursor.fetchall()]
        conn.close()
        return {'documentos': docs}
    except Exception as e:
        conn.close()
        print(f"⚠️ Error listando documentos de '{usuario}': {e}")
        return {'documentos': []}


@app.route('/documentos_empleado/subir', methods=['POST'])
@login_required
@admin_required
def documentos_empleado_subir():
    usuario_doc = (request.form.get('usuario') or '').strip()
    tipo_documento = (request.form.get('tipo_documento') or '').strip() or 'Documento'
    titulo = (request.form.get('titulo') or '').strip()
    descripcion = (request.form.get('descripcion') or '').strip()
    fecha_emision = (request.form.get('fecha_emision') or '').strip() or None
    fecha_vencimiento = (request.form.get('fecha_vencimiento') or '').strip() or None
    archivo = request.files.get('archivo')

    if not usuario_doc or not titulo:
        return redirect(url_for('gestion_usuarios'))

    url_archivo, nombre_original = _subir_archivo_a_cloudinary(archivo)
    fecha_actual = obtener_fecha_actual()

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_ins = ("INSERT INTO documentos_empleado (usuario, tipo_documento, titulo, descripcion, url_archivo, nombre_original, fecha_emision, fecha_vencimiento, estado, creado_por, fecha_creacion) "
                 "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'activo', %s, %s)") if db_type == 'postgres' else \
                ("INSERT INTO documentos_empleado (usuario, tipo_documento, titulo, descripcion, url_archivo, nombre_original, fecha_emision, fecha_vencimiento, estado, creado_por, fecha_creacion) "
                 "VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'activo', ?, ?)")
        cursor.execute(q_ins, (usuario_doc, tipo_documento, titulo, descripcion, url_archivo or '', nombre_original or '', fecha_emision, fecha_vencimiento, session['username'], fecha_actual))
        conn.commit()
        registrar_log(session['username'], "Documento de Empleado Registrado", f"'{titulo}' ({tipo_documento}) para {usuario_doc}")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error guardando documento de empleado '{titulo}' para '{usuario_doc}': {e}")
    conn.close()

    return redirect(url_for('gestion_usuarios'))


@app.route('/documentos_empleado/editar/<int:doc_id>', methods=['POST'])
@login_required
@admin_required
def documentos_empleado_editar(doc_id):
    tipo_documento = (request.form.get('tipo_documento') or '').strip() or 'Documento'
    titulo = (request.form.get('titulo') or '').strip()
    descripcion = (request.form.get('descripcion') or '').strip()
    fecha_emision = (request.form.get('fecha_emision') or '').strip() or None
    fecha_vencimiento = (request.form.get('fecha_vencimiento') or '').strip() or None
    archivo = request.files.get('archivo')

    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_sel = "SELECT fecha_vencimiento, url_archivo, nombre_original FROM documentos_empleado WHERE id = %s" if db_type == 'postgres' else "SELECT fecha_vencimiento, url_archivo, nombre_original FROM documentos_empleado WHERE id = ?"
        cursor.execute(q_sel, (doc_id,))
        antiguo = cursor.fetchone()
        venc_old = (antiguo[0] or '').strip() if antiguo else None
        url_actual = antiguo[1] if antiguo else ''
        nombre_actual = antiguo[2] if antiguo else ''

        url_nueva, nombre_nuevo = _subir_archivo_a_cloudinary(archivo)
        if not url_nueva:
            url_nueva, nombre_nuevo = url_actual, nombre_actual

        # 🔔 Si cambió la fecha de vencimiento, se limpia el nivel de alerta para que pueda
        # volver a avisar si la nueva fecha se acerca otra vez.
        if venc_old != fecha_vencimiento:
            q_upd = ("UPDATE documentos_empleado SET tipo_documento = %s, titulo = %s, descripcion = %s, fecha_emision = %s, fecha_vencimiento = %s, url_archivo = %s, nombre_original = %s, alerta_nivel = NULL WHERE id = %s") if db_type == 'postgres' else \
                    ("UPDATE documentos_empleado SET tipo_documento = ?, titulo = ?, descripcion = ?, fecha_emision = ?, fecha_vencimiento = ?, url_archivo = ?, nombre_original = ?, alerta_nivel = NULL WHERE id = ?")
        else:
            q_upd = ("UPDATE documentos_empleado SET tipo_documento = %s, titulo = %s, descripcion = %s, fecha_emision = %s, fecha_vencimiento = %s, url_archivo = %s, nombre_original = %s WHERE id = %s") if db_type == 'postgres' else \
                    ("UPDATE documentos_empleado SET tipo_documento = ?, titulo = ?, descripcion = ?, fecha_emision = ?, fecha_vencimiento = ?, url_archivo = ?, nombre_original = ? WHERE id = ?")
        cursor.execute(q_upd, (tipo_documento, titulo, descripcion, fecha_emision, fecha_vencimiento, url_nueva, nombre_nuevo, doc_id))
        conn.commit()
        registrar_log(session['username'], "Documento de Empleado Editado", f"Documento #{doc_id}: '{titulo}'")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error editando documento de empleado {doc_id}: {e}")
    conn.close()

    return redirect(url_for('gestion_usuarios'))


@app.route('/documentos_empleado/eliminar/<int:doc_id>', methods=['POST'])
@login_required
@admin_required
def documentos_empleado_eliminar(doc_id):
    conn, db_type = get_db()
    cursor = conn.cursor()
    try:
        q_upd = "UPDATE documentos_empleado SET estado = 'eliminado' WHERE id = %s" if db_type == 'postgres' else "UPDATE documentos_empleado SET estado = 'eliminado' WHERE id = ?"
        cursor.execute(q_upd, (doc_id,))
        conn.commit()
        registrar_log(session['username'], "Documento de Empleado Eliminado", f"Documento #{doc_id}")
    except Exception as e:
        conn.rollback()
        print(f"⚠️ Error eliminando documento de empleado {doc_id}: {e}")
    conn.close()

    return redirect(url_for('gestion_usuarios'))


if __name__ == '__main__':
    # 🛡️ debug=False: este bloque no lo usa producción (Render arranca con gunicorn,
    # ver Procfile), pero si algún día el comando de arranque cambiara a "python app.py",
    # debug=True habilita el depurador interactivo de Werkzeug, que permite ejecutar
    # código Python arbitrario desde el navegador. Mejor dejarlo en False siempre.
    app.run(host='0.0.0.0', port=5000, debug=False)
